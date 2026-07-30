import io
import base64
import hmac
import json
import math
import os
import pickle
import re
import time
import traceback
import unicodedata
import uuid
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from html import escape
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

from ticket_system import (
    GitHubTicketStore,
    LocalTicketStore,
    MockJobAdapter,
    MockNotificationAdapter,
    PRIORITIES,
    PRIORITY_LABELS,
    ROLE_ADMIN,
    ROLE_BRAND,
    ROLE_OPERATOR,
    STATE_APPROVED,
    STATE_ASSIGNED,
    STATE_CANCELED,
    STATE_COMPLETED,
    STATE_COMPLETED_OBS,
    STATE_CORRECTED,
    STATE_DRAFT,
    STATE_DRY_RUN,
    STATE_FAILED,
    STATE_LABELS,
    STATE_LOADING,
    STATE_OBSERVED,
    STATE_PENDING,
    STATE_PREPARING,
    STATE_READY_EXECUTE,
    STATE_REJECTED,
    STATE_REQUEST_RECEIVED,
    STATE_REVIEW,
    STATE_WAITING_BRAND,
    STATE_VALIDATING,
    TicketConflictError,
    TicketError,
    TicketPermissionError,
    TicketService,
    TicketValidationError,
    file_sha256,
    ticket_age_hours,
    ticket_is_overdue,
)

from generate_columbia_matrixify import (
    SITE_CONFIGS,
    build_body_html as build_matrixify_body_html,
    build_columbia_matrixify,
    build_matrixify_updates,
    brand_display_name,
    brand_image_config,
    display_size_for_site,
    get_brand_config,
    image_candidates,
    input_brand_report,
    first_non_empty,
    is_internal_k_size,
    is_one_size,
    limit_words,
    normalize_brand_name,
    normalize_size as normalize_master_size,
    size_sort_key as master_size_sort_key,
    split_model_color,
    strip_html,
    is_zero_size,
    read_arti_source,
)

from engines.normalize import (
    ARTI_COLUMN_ALIASES_APP,
    SIZE_ORDER,
    SIZE_ORDER_GROUPS,
    _CATALOG_ALIAS_TO_ARTI_TARGET,
    _row_by_size_keys,
    _set_row_by_size_keys,
    _size_lookup_keys,
    clean_value,
    coalesce_duplicate_columns,
    expected_catalog_vendors,
    first_existing_column,
    first_row_value,
    format_datetime_lima,
    looks_like_mod_col,
    normalize_arti_columns_for_app,
    normalize_header,
    normalize_size,
    parse_iso_datetime,
    parse_publication_date,
    product_lookup_candidates,
    product_lookup_key,
    publication_date_from_row,
    repair_mojibake_dataframe,
    repair_mojibake_text,
    safe_float_value,
    safe_int_value,
    size_sort_key,
    slugify,
    variant_mod_col_candidates,
)
from engines.audit import (
    AuditError,
    AuditService,
    GitHubAuditStore,
    LocalAuditStore,
)
from engines.excel_io import (
    columbia_to_excel_bytes,
    dataframe_to_excel_bytes,
    read_excel,
    update_to_excel_bytes,
)

try:
    from catalog_rules import (
        CATALOG_FIELD_ALIASES,
        INPUT_COLUMNS as CATALOG_INPUT_COLUMNS,
        PRODUCT_TYPE_RULES,
        SIZE_GUIDE_RULES,
        aliases_for,
        build_catalog_handle,
        resolve_size_guide,
        validate_catalog_row,
    )
except Exception:
    CATALOG_FIELD_ALIASES = {}
    CATALOG_INPUT_COLUMNS = []
    PRODUCT_TYPE_RULES = []
    SIZE_GUIDE_RULES = []

    def aliases_for(field, fallback=None):
        return list(fallback or [])

    def build_catalog_handle(product_type="", gender="", brand="", mod_col=""):
        return ""

    def validate_catalog_row(row):
        return {"normalized": {}, "issues": [], "size_guide_decision": {}}

    def resolve_size_guide(brand="", category="", product_type="", gender="", age_group="", current_guide=""):
        return {
            "guide": clean_value(current_guide),
            "rule": "fallback",
            "match_level": "none",
            "warning": "No se pudo cargar catalog_rules.resolve_size_guide.",
            "status": "warning",
        }

try:
    import shopify_api as _shopify_api
except Exception as exc:
    _shopify_api = None
    _shopify_api_import_error = exc
else:
    _shopify_api_import_error = None


class _FallbackShopifyApiError(Exception):
    pass


def _missing_shopify_api_function(name):
    def _raise_missing(*args, **kwargs):
        detail = f" Detalle import: {_shopify_api_import_error}" if _shopify_api_import_error else ""
        raise RuntimeError(f"Falta actualizar shopify_api.py: no existe {name}.{detail}")

    return _raise_missing


def _shopify_attr(name, default=None):
    if _shopify_api is None:
        return default if default is not None else _missing_shopify_api_function(name)
    return getattr(_shopify_api, name, default if default is not None else _missing_shopify_api_function(name))


DEFAULT_API_VERSION = _shopify_attr("DEFAULT_API_VERSION", "2026-04")
ShopifyApiError = _shopify_attr("ShopifyApiError", _FallbackShopifyApiError)
fetch_metaobject_definitions = _shopify_attr("fetch_metaobject_definitions")
fetch_metaobjects = _shopify_attr("fetch_metaobjects")
fetch_product_options_and_variants = _shopify_attr("fetch_product_options_and_variants")
fetch_products = _shopify_attr("fetch_products")
file_create = _shopify_attr("file_create")
inventory_item_update = _shopify_attr("inventory_item_update", None)
inventory_activate = _shopify_attr("inventory_activate", None)
inventory_item_active_locations = _shopify_attr("inventory_item_active_locations", None)
fetch_locations = _shopify_attr("fetch_locations", None)
metafields_set = _shopify_attr("metafields_set")
normalize_shop_domain = _shopify_attr("normalize_shop_domain")
product_create = _shopify_attr("product_create")
product_create_media = _shopify_attr("product_create_media")
product_delete_media = _shopify_attr("product_delete_media")
product_options_reorder = _shopify_attr("product_options_reorder")
publishable_publish = _shopify_attr("publishable_publish")
product_set_files = _shopify_attr("product_set_files")
product_update = _shopify_attr("product_update")
product_variants_bulk_create = _shopify_attr("product_variants_bulk_create")
product_variants_bulk_update = _shopify_attr("product_variants_bulk_update", None)
product_variants_bulk_reorder = _shopify_attr("product_variants_bulk_reorder")
staged_upload_image = _shopify_attr("staged_upload_image")
test_connection = _shopify_attr("test_connection")
wait_file_statuses = _shopify_attr("wait_file_statuses")
wait_media_statuses = _shopify_attr("wait_media_statuses")
fetch_metaobjects_for_definition = _shopify_attr("fetch_metaobjects_for_definition", None)
fetch_metafield_definition = _shopify_attr("fetch_metafield_definition", None)

try:
    from centry_static_masters import (
        CENTRY_CATEGORY_FULL,
        CENTRY_CATEGORY_GENDER_TYPE,
        CENTRY_CODEX_CATEGORIES,
        CENTRY_DIMENSIONS,
    )
except ImportError:
    CENTRY_CATEGORY_FULL = []
    CENTRY_CATEGORY_GENDER_TYPE = []
    CENTRY_CODEX_CATEGORIES = []
    CENTRY_DIMENSIONS = []


APP_TITLE = "Catálogo Control Center"
DEFAULT_ARTI_PATH = "data/arti.xlsx"
DEFAULT_ARTI_CSV_PATH = "data/arti.csv"
DEFAULT_ARTI_ZIP_PATH = "data/arti.zip"
DEFAULT_MATRIXIFY_PATH = "data/matrixify_modelo.xlsx"
DEFAULT_ECOMM_WAREHOUSES_PATH = Path("data/bodegas_ecomm.xlsx")
DEFAULT_CENTRY_CATEGORY_PATHS = [
    Path("data/base_categorias_centry.xlsx"),
]
DEFAULT_CENTRY_DIMENSIONS_PATHS = [
    Path("data/dimensiones_productos.xlsx"),
]
DEFAULT_CENTRY_CODEX_CATEGORY_PATHS = [
    Path("data/centry_codex_categorias.xlsx"),
]
DEFAULT_PRODUCT_MASTER_TABLE = "forus-analitica-prod-datalake.bronze.stg_pe_central_arti"
FORUS_LOGO_PATH = Path("assets/forus_logo.png")
SHOPIFY_LOGO_PATH = Path("assets/shopify_logo.png")
KPI_AUTO_REFRESH_SECONDS = 15 * 60
OUTPUT_DIR = Path("outputs")
KPI_CACHE_DIR = OUTPUT_DIR / "kpi_cache"
SYNC_JOB_DIR = OUTPUT_DIR / "sync_jobs"
KPI_CACHE_VERSION = "2026-07-04-missing-input-enriched-v1"

DEFAULT_ECOMM_SITE_WAREHOUSES = {
    "columbiape": ["320", "145", "143", "142", "139", "130", "114", "113", "112", "111", "96", "88", "84", "83", "59", "52", "46", "19", "18", "2"],
    "hushpuppiespe": ["320", "129", "111", "97", "96", "88", "46", "44", "43", "30", "23", "19", "18", "16", "8", "7"],
    "rockfordpe": ["320", "145", "143", "142", "139", "130", "129", "122", "114", "113", "112", "111", "97", "96", "88", "84", "83", "59", "52", "44", "43", "30", "23", "19", "16", "8", "7", "2"],
    "vanspe": ["320", "152", "151", "150", "149"],
}
DEFAULT_ECOMM_STOCK_SECURITY = {
    "114": 0, "88": 0, "84": 0, "113": 0, "320": 0, "111": 0, "8": 0, "44": 0, "46": 0,
    "97": 0, "43": 0, "112": 0, "59": 0, "52": 0, "2": 0, "83": 0, "16": 1, "122": 0,
    "18": 0, "7": 0, "30": 0, "19": 0, "23": 0, "96": 1, "130": 0, "129": 0, "139": 0,
    "142": 0, "143": 0, "145": 0, "149": 0, "150": 0, "152": 0, "151": 0,
}

MATRIXIFY_COLUMNS = [
    "Command",
    "Handle",
    "Title",
    "Body HTML",
    "Vendor",
    "Type",
    "Tags",
    "Status",
    "Published",
    "Option1 Name",
    "Option1 Value",
    "Option2 Name",
    "Option2 Value",
    "Variant SKU",
    "Variant Barcode",
    "Variant Price",
    "Variant Compare At Price",
    "Variant Inventory Qty",
    "Variant Inventory Tracker",
    "Variant Inventory Policy",
    "Variant Fulfillment Service",
    "Variant Requires Shipping",
    "Variant Taxable",
    "Variant Weight",
    "Variant Weight Unit",
    "Image Src",
    "Image Position",
    "Metafield: custom.estilo [single_line_text_field]",
    "Metafield: custom.color [single_line_text_field]",
]

for catalog_field, arti_targets in _CATALOG_ALIAS_TO_ARTI_TARGET.items():
    for target in arti_targets:
        if target not in ARTI_COLUMN_ALIASES_APP:
            continue
        for alias in CATALOG_FIELD_ALIASES.get(catalog_field, []):
            if alias not in ARTI_COLUMN_ALIASES_APP[target]:
                ARTI_COLUMN_ALIASES_APP[target].append(alias)


def read_uploaded_excel_cached(uploaded_file, state_prefix, sheet_name=0):
    if not uploaded_file:
        st.session_state.pop(f"{state_prefix}_fingerprint", None)
        st.session_state.pop(f"{state_prefix}_df", None)
        return None
    fingerprint = uploaded_file_fingerprint(uploaded_file)
    if (
        st.session_state.get(f"{state_prefix}_fingerprint") == fingerprint
        and st.session_state.get(f"{state_prefix}_df") is not None
    ):
        return st.session_state[f"{state_prefix}_df"]
    try:
        uploaded_file.seek(0)
    except Exception:
        pass
    try:
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name, dtype=object)
    except ValueError as exc:
        if isinstance(sheet_name, str) and "Worksheet named" in str(exc):
            try:
                uploaded_file.seek(0)
            except Exception:
                pass
            sheets = pd.read_excel(uploaded_file, sheet_name=None, dtype=object)
            exact_sheet = next((name for name in sheets if clean_value(name).lower() == sheet_name.lower()), "")
            if exact_sheet:
                df = sheets[exact_sheet]
            else:
                non_empty_sheets = [
                    sheet_df.dropna(how="all")
                    for sheet_df in sheets.values()
                    if isinstance(sheet_df, pd.DataFrame) and not sheet_df.dropna(how="all").empty
                ]
                df = non_empty_sheets[0] if non_empty_sheets else pd.DataFrame()
        else:
            raise
    if isinstance(df, dict):
        df = next(iter(df.values()), pd.DataFrame())
    df = df.dropna(how="all")
    st.session_state[f"{state_prefix}_fingerprint"] = fingerprint
    st.session_state[f"{state_prefix}_df"] = df
    return df


def get_bigquery_config():
    config = {}
    try:
        if "bigquery" in st.secrets:
            config.update(dict(st.secrets["bigquery"]))
        if "gcp_service_account" in st.secrets:
            config["service_account_info"] = dict(st.secrets["gcp_service_account"])
    except Exception:
        return {}

    service_account_json = config.pop("service_account_json", None)
    if service_account_json and "service_account_info" not in config:
        config["service_account_info"] = json.loads(service_account_json)
    return config


def is_bigquery_configured(config):
    if not config:
        return False
    enabled = str(config.get("enabled", "true")).strip().lower()
    if enabled in ("0", "false", "no", "off"):
        return False
    has_query = bool(str(config.get("query", "")).strip())
    service_project = ""
    if isinstance(config.get("service_account_info"), dict):
        service_project = str(config["service_account_info"].get("project_id", "")).strip()
    has_project = bool(str(config.get("project_id", "")).strip() or service_project)
    table = str(config.get("table", "")).strip()
    has_full_table = table.count(".") == 2
    has_split_table = has_project and bool(str(config.get("dataset", "")).strip() and table)
    has_table = has_full_table or has_split_table
    return has_query or has_table


def get_shopify_config(site_key):
    config = {}
    try:
        shopify_sites = st.secrets.get("shopify_sites", {})
        if site_key in shopify_sites:
            config.update(dict(shopify_sites[site_key]))
        inventory_config = dict(st.secrets.get("inventory", {}))
        inventory_locations_config = dict(st.secrets.get("inventory_locations", {}))
        for source in (inventory_config, inventory_locations_config):
            for field in ("inventory_location_ids", "inventory_locations", "location_ids"):
                if config.get(field):
                    continue
                config[field] = (
                    source.get(f"{site_key}_{field}")
                    or source.get(site_key)
                    or source.get(field)
                    or config.get(field)
                )
    except Exception:
        return {}

    return {
        "shop_domain": normalize_shop_domain(config.get("shop_domain") or config.get("domain")),
        "client_id": clean_value(config.get("client_id")),
        "client_secret": clean_value(config.get("client_secret")),
        "admin_access_token": clean_value(
            config.get("admin_access_token") or config.get("access_token") or config.get("token")
        ),
        "api_version": clean_value(config.get("api_version")) or DEFAULT_API_VERSION,
        "inventory_location_ids": clean_value(config.get("inventory_location_ids")),
        "inventory_locations": clean_value(config.get("inventory_locations")),
        "location_ids": clean_value(config.get("location_ids")),
    }


def is_shopify_configured(config):
    has_token = bool(config.get("admin_access_token"))
    has_client_credentials = bool(config.get("client_id") and config.get("client_secret"))
    return bool(config.get("shop_domain") and (has_token or has_client_credentials))


def session_shopify_products(site_key, shopify_config, force_refresh=False):
    cache_key = f"shopify_products_cache_{clean_value(site_key)}"
    meta_key = f"{cache_key}_meta"
    cache_meta = {
        "shop_domain": clean_value(shopify_config.get("shop_domain")),
        "api_version": clean_value(shopify_config.get("api_version")),
    }
    if (
        not force_refresh
        and st.session_state.get(meta_key) == cache_meta
        and st.session_state.get(cache_key) is not None
    ):
        return st.session_state[cache_key]
    products = fetch_products(shopify_config)
    st.session_state[cache_key] = products
    st.session_state[meta_key] = cache_meta
    st.session_state[f"{cache_key}_loaded_at"] = datetime.now(timezone.utc).isoformat()
    return products


def clear_shopify_products_cache(site_key):
    cache_key = f"shopify_products_cache_{clean_value(site_key)}"
    for key in (cache_key, f"{cache_key}_meta", f"{cache_key}_loaded_at"):
        st.session_state.pop(key, None)


def read_arti_for_app(brand_config):
    arti_df, source = read_arti_source(
        bigquery_config=get_bigquery_config(),
        allow_local_fallback=False,
        brand_config=brand_config,
    )
    arti_df = normalize_arti_columns_for_app(arti_df).dropna(how="all")
    arti_df, ean_source = enrich_arti_barcodes_from_bigquery_table(arti_df, get_bigquery_config())
    if ean_source:
        source = f"{source} + {ean_source}"
    return arti_df, source


def session_arti_for_app(brand_config, force_refresh=False):
    site_key = clean_value(brand_config.get("site_key"))
    cache_key = f"arti_cache_{site_key}"
    source_key = f"{cache_key}_source"
    meta_key = f"{cache_key}_meta"
    bigquery_config = get_bigquery_config()
    cache_meta = {
        "site_key": site_key,
        "table": clean_value(bigquery_config.get("table")),
        "dataset": clean_value(bigquery_config.get("dataset")),
        "project_id": clean_value(bigquery_config.get("project_id")),
        "query": clean_value(bigquery_config.get("query")),
    }
    if (
        not force_refresh
        and st.session_state.get(meta_key) == cache_meta
        and st.session_state.get(cache_key) is not None
    ):
        return st.session_state[cache_key], st.session_state.get(source_key, "BigQuery")
    try:
        arti_df, source = read_arti_for_app(brand_config)
    except Exception:
        for key in (cache_key, source_key, meta_key):
            st.session_state.pop(key, None)
        raise
    st.session_state[cache_key] = arti_df
    st.session_state[source_key] = source
    st.session_state[meta_key] = cache_meta
    return arti_df, source


def _bigquery_table_id_from_config(config):
    config = dict(config or {})
    table = clean_value(config.get("table"))
    if not table:
        return ""
    if table.count(".") == 2:
        return table
    project_id = clean_value(config.get("project_id"))
    dataset = clean_value(config.get("dataset"))
    if project_id and dataset:
        return f"{project_id}.{dataset}.{table}"
    return ""


def _bigquery_product_master_table_id(config):
    config = dict(config or {})
    table = clean_value(
        config.get("product_master_table")
        or config.get("maestro_productos_table")
        or config.get("ean_table")
        or config.get("barcode_table")
        or DEFAULT_PRODUCT_MASTER_TABLE
    )
    if not table:
        return ""
    if table.count(".") == 2:
        return table
    project_id = clean_value(config.get("product_master_project_id") or config.get("project_id"))
    dataset = clean_value(config.get("product_master_dataset") or config.get("dataset"))
    if project_id and dataset:
        return f"{project_id}.{dataset}.{table}"
    return ""


def _bigquery_schema_columns(client, table_id):
    if not table_id:
        return []
    return [field.name for field in client.get_table(table_id).schema]


def _bigquery_barcode_candidate_columns(columns):
    schema_df = pd.DataFrame(columns=list(columns or []))
    exact = first_existing_column(schema_df, ARTI_COLUMN_ALIASES_APP["CodBarras"])
    candidates = [exact] if exact else []
    tokens = ("ean", "barra", "barras", "barcode", "bar_code", "upc", "gtin")
    for column in columns or []:
        normalized = normalize_header(column)
        if any(token in normalized for token in tokens):
            candidates.append(column)
    return list(dict.fromkeys([column for column in candidates if column]))


def _bigquery_sku_candidate_column(columns):
    schema_df = pd.DataFrame(columns=list(columns or []))
    exact = first_existing_column(schema_df, ARTI_COLUMN_ALIASES_APP["CODINT_MA"])
    if exact:
        return exact
    tokens = ("codint", "sku", "id_producto", "idproducto")
    for column in columns or []:
        normalized = normalize_header(column)
        if any(token in normalized for token in tokens):
            return column
    return ""


def _coalesce_barcode_candidates(df, candidate_columns):
    if df is None or df.empty:
        return df
    result = normalize_arti_columns_for_app(df).copy()
    if "CodBarras" not in result.columns:
        result["CodBarras"] = ""
    for column in candidate_columns or []:
        if column not in result.columns:
            continue
        fill_mask = result["CodBarras"].map(clean_value).eq("") & result[column].map(clean_value).ne("")
        if fill_mask.any():
            result.loc[fill_mask, "CodBarras"] = result.loc[fill_mask, column]
    return result


def bigquery_barcode_schema_diagnostics(bigquery_config):
    table_id = _bigquery_product_master_table_id(bigquery_config) or _bigquery_table_id_from_config(bigquery_config)
    if not table_id:
        return "Maestro Productos BigQuery no configurado"
    try:
        from google.cloud import bigquery
        from google.oauth2 import service_account
    except ImportError:
        return "No pude inspeccionar schema BigQuery: falta google-cloud-bigquery"

    try:
        config = dict(bigquery_config or {})
        project_id = clean_value(config.get("project_id"))
        credentials_info = config.get("service_account_info")
        credentials = None
        if credentials_info:
            credentials = service_account.Credentials.from_service_account_info(dict(credentials_info))
            project_id = project_id or credentials.project_id
        job_project_id = clean_value(config.get("job_project_id")) or project_id
        client = bigquery.Client(project=job_project_id or None, credentials=credentials)
        schema_columns = _bigquery_schema_columns(client, table_id)
        sku_col = _bigquery_sku_candidate_column(schema_columns)
        barcode_candidates = _bigquery_barcode_candidate_columns(schema_columns)
        preview = ", ".join(barcode_candidates[:12]) or "ninguna"
        return f"Schema {table_id}: sku={sku_col or 'NO detectado'}; columnas EAN/barra={preview}"
    except Exception as exc:
        return f"No pude inspeccionar schema Maestro Productos: {type(exc).__name__}: {exc}"


def enrich_arti_barcodes_from_bigquery_table(arti_df, bigquery_config):
    if arti_df is None or arti_df.empty or "CODINT_MA" not in arti_df.columns:
        return arti_df, ""
    result = normalize_arti_columns_for_app(arti_df).copy()
    missing_mask = result["CodBarras"].map(clean_value) == ""
    missing_skus = sorted({clean_value(value) for value in result.loc[missing_mask, "CODINT_MA"] if clean_value(value)})
    if not missing_skus:
        return result, ""

    table_id = _bigquery_product_master_table_id(bigquery_config) or _bigquery_table_id_from_config(bigquery_config)
    product_master_query = clean_value(
        (bigquery_config or {}).get("product_master_query")
        or (bigquery_config or {}).get("maestro_productos_query")
        or (bigquery_config or {}).get("ean_query")
        or (bigquery_config or {}).get("barcode_query")
    )
    if not table_id and not product_master_query:
        return result, ""

    try:
        from google.cloud import bigquery
        from google.oauth2 import service_account
    except ImportError:
        return result, ""

    try:
        config = dict(bigquery_config or {})
        project_id = clean_value(config.get("project_id"))
        credentials_info = config.get("service_account_info")
        credentials = None
        if credentials_info:
            credentials = service_account.Credentials.from_service_account_info(dict(credentials_info))
            project_id = project_id or credentials.project_id
        job_project_id = clean_value(config.get("job_project_id")) or project_id
        client = bigquery.Client(project=job_project_id or None, credentials=credentials)
        if product_master_query:
            wrapped_query = product_master_query.rstrip().rstrip(";")
            query = f"""
            SELECT *
            FROM ({wrapped_query})
            WHERE CAST(CODINT_MA AS STRING) IN UNNEST(@skus)
            """
        else:
            schema_columns = _bigquery_schema_columns(client, table_id)
            sku_col = _bigquery_sku_candidate_column(schema_columns)
            barcode_cols = _bigquery_barcode_candidate_columns(schema_columns)[:20]
            if not sku_col or not barcode_cols:
                return result, ""
            barcode_selects = [
                f"CAST(`{column}` AS STRING) AS `__EAN_CAND_{index}`"
                for index, column in enumerate(barcode_cols)
            ]
            query = f"""
            SELECT
              CAST(`{sku_col}` AS STRING) AS CODINT_MA,
              {", ".join(barcode_selects)}
            FROM `{table_id}`
            WHERE CAST(`{sku_col}` AS STRING) IN UNNEST(@skus)
            """
        job_config = bigquery.QueryJobConfig(
            use_legacy_sql=False,
            query_parameters=[bigquery.ArrayQueryParameter("skus", "STRING", missing_skus)],
        )
        lookup_df = client.query(query, job_config=job_config, location=clean_value(config.get("location")) or None).to_dataframe()
        ean_alias_columns = [column for column in lookup_df.columns if str(column).startswith("__EAN_CAND_")]
        lookup_df = _coalesce_barcode_candidates(lookup_df, ean_alias_columns)
        lookup = {
            clean_value(row.get("CODINT_MA")): clean_value(row.get("CodBarras"))
            for _, row in lookup_df.iterrows()
            if clean_value(row.get("CODINT_MA")) and clean_value(row.get("CodBarras"))
        }
        if not lookup:
            return result, ""
        fill_mask = result["CodBarras"].map(clean_value) == ""
        result.loc[fill_mask, "CodBarras"] = result.loc[fill_mask, "CODINT_MA"].map(lambda sku: lookup.get(clean_value(sku), ""))
        filled = safe_int_value((result.loc[fill_mask, "CodBarras"].map(clean_value) != "").sum())
        source_label = "Maestro Productos BigQuery" if product_master_query or _bigquery_product_master_table_id(bigquery_config) else "EAN tabla BigQuery"
        return result, f"{source_label} ({filled:,})" if filled else ""
    except Exception:
        return result, ""


def arti_barcode_diagnostics(arti_df):
    if arti_df is None or arti_df.empty:
        return "ARTI vacio"
    df = coalesce_duplicate_columns(arti_df).copy()
    pieces = []
    for alias in ARTI_COLUMN_ALIASES_APP["CodBarras"]:
        column = first_existing_column(df, [alias])
        if column is None:
            continue
        non_empty = safe_int_value(df[column].map(clean_value).ne("").sum())
        pieces.append(f"{column}: {non_empty:,}")
    if not pieces:
        return "No llego ninguna columna tipo EAN/barcode en el query"
    return "Columnas EAN detectadas - " + " | ".join(dict.fromkeys(pieces))


def detect_input_columns(df):
    return {
        "style": first_existing_column(df, ["style", "estilo", "modelo", "codigo", "sku padre", "parent sku", "item"]),
        "title": first_existing_column(df, ["title", "titulo", "producto", "descripcion", "description", "nombre"]),
        "vendor": first_existing_column(df, ["vendor", "marca", "brand"]),
        "type": first_existing_column(df, ["type", "tipo", "categoria", "category"]),
        "color": first_existing_column(df, ["color", "colour"]),
        "price": first_existing_column(df, ["price", "precio", "precio venta", "variant price", "pvp"]),
        "barcode": first_existing_column(df, ["barcode", "ean", "upc", "codigo barra", "codigo de barra"]),
        "sku": first_existing_column(df, ["sku", "variant sku", "codigo sku"]),
        "image": first_existing_column(df, ["image src", "imagen", "image", "url imagen", "foto"]),
        "tags": first_existing_column(df, ["tags", "etiquetas"]),
        "body": first_existing_column(df, ["body html", "descripcion larga", "body", "detalle"]),
    }


def detect_arti_columns(df):
    return {
        "style": first_existing_column(df, ["style", "estilo", "modelo", "codigo", "sku padre", "parent sku", "item"]),
        "size": first_existing_column(df, ["size", "talla", "tallas"]),
        "sku": first_existing_column(df, ["sku", "variant sku", "codigo sku", "codigo"]),
        "barcode": first_existing_column(df, ["barcode", "ean", "upc", "codigo barra", "codigo de barra"]),
        "color": first_existing_column(df, ["color", "colour"]),
    }


def build_arti_lookup(arti_df, arti_cols):
    if not arti_cols["style"] or not arti_cols["size"]:
        return {}

    lookup = defaultdict(list)
    for _, row in arti_df.iterrows():
        style = clean_value(row.get(arti_cols["style"]))
        size = normalize_size(row.get(arti_cols["size"]))
        if not style or not size:
            continue

        item = {
            "size": size,
            "sku": clean_value(row.get(arti_cols["sku"])) if arti_cols["sku"] else "",
            "barcode": clean_value(row.get(arti_cols["barcode"])) if arti_cols["barcode"] else "",
            "color": clean_value(row.get(arti_cols["color"])) if arti_cols["color"] else "",
        }
        lookup[style.upper()].append(item)

    for style, rows in lookup.items():
        seen = set()
        unique_rows = []
        for item in sorted(rows, key=lambda value: size_sort_key(value["size"])):
            key = item["size"]
            if key in seen:
                continue
            seen.add(key)
            unique_rows.append(item)
        lookup[style] = unique_rows

    return lookup


def manual_sizes_from_text(value):
    text = clean_value(value)
    if not text:
        return []
    parts = re.split(r"[,;/|]+", text)
    return sorted({normalize_size(part) for part in parts if normalize_size(part)}, key=size_sort_key)


def row_value(row, column_name):
    if not column_name:
        return ""
    return clean_value(row.get(column_name))


COMMERCIAL_INPUT_TEMPLATE_VERSION = "CCC_INPUT_MARCA_V7_2026-07-22"
COMMERCIAL_INPUT_MAX_ROWS = 5000
COMMERCIAL_INPUT_REQUIRED_COLUMNS = [
    "Mod-Col",
    "Marca",
    "Genero",
    "Clase",
    "Tipo de prenda",
    "Color Comercial",
    "Color web/filtro",
    "Nombre de Producto",
    "Descripcion",
    "Caracteristicas",
    "Materiales",
    "Cuidados",
]
COMMERCIAL_INPUT_TEXT_LIST_COLUMNS = [
    "Caracteristicas",
    "Materiales",
    "Cuidados",
    "Tecnologia",
    "Tags adicionales",
]
COMMERCIAL_INPUT_INVALID_TEXTS = {
    "",
    "-",
    ".",
    "0",
    "00",
    "000",
    "n/a",
    "na",
    "null",
    "none",
    "sin informacion",
    "sininformacion",
    "pendiente",
    "por completar",
    "porcompletar",
    "tbd",
}

COMMERCIAL_BRAND_ALLOWED_CLASSES = {
    "COLUMBIA": ["Calzado", "Vestuario", "Accesorios"],
    "ROCKFORD": ["Calzado", "Vestuario", "Accesorios"],
    "VANS": ["Calzado", "Vestuario", "Accesorios"],
    "MOUNTAIN HARDWEAR": ["Vestuario", "Accesorios"],
    "PATAGONIA": ["Vestuario", "Accesorios"],
    "SOREL": ["Calzado"],
    "HUSH PUPPIES": ["Calzado", "Accesorios"],
    "HUSH PUPPIES KIDS": ["Calzado", "Vestuario", "Accesorios"],
    "ACCESORIOS HP": ["Accesorios"],
    "KEDS": ["Calzado"],
}

# El input sigue siendo comercial: estos extras cambian por marca, mientras
# siblings, logos, guias, categorias tecnicas y relaciones los arma la app.
COMMERCIAL_BRAND_INPUT_PROFILES = {
    "COLUMBIA": {
        "site_profile": "Columbia.pe",
        "extra_columns": ["Tecnologia", "Pais de fabricacion"],
        "technology_type": "list.single_line_text_field",
        "technology_example": "Omni-Tech|Omni-Shield",
    },
    "ROCKFORD": {
        "site_profile": "Rockford.pe",
        "extra_columns": ["Tecnologia", "Pais de fabricacion"],
        "technology_type": "single_line_text_field",
        "technology_example": "Libre de arrugas",
    },
    "MOUNTAIN HARDWEAR": {
        "site_profile": "Rockford.pe",
        "extra_columns": ["Tecnologia", "Pais de fabricacion"],
        "technology_type": "single_line_text_field",
        "technology_example": "Tecnologia comercial si aplica",
    },
    "PATAGONIA": {
        "site_profile": "Rockford.pe",
        "extra_columns": ["Tecnologia", "Pais de fabricacion"],
        "technology_type": "single_line_text_field",
        "technology_example": "Tecnologia comercial si aplica",
    },
    "SOREL": {
        "site_profile": "Rockford.pe",
        "extra_columns": ["Tecnologia", "Pais de fabricacion"],
        "technology_type": "single_line_text_field",
        "technology_example": "Tecnologia comercial si aplica",
    },
    "HUSH PUPPIES": {
        "site_profile": "HushPuppies.pe",
        "extra_columns": ["Tecnologia", "Categoria de Tecnologia", "Estilo", "Pais de fabricacion"],
        "technology_type": "single_line_text_field",
        "technology_example": "Bounce Plus",
    },
    "HUSH PUPPIES KIDS": {
        "site_profile": "HushPuppies.pe",
        "extra_columns": ["Tecnologia", "Categoria de Tecnologia", "Estilo", "Pais de fabricacion"],
        "technology_type": "single_line_text_field",
        "technology_example": "Bounce Plus",
    },
    "ACCESORIOS HP": {
        "site_profile": "HushPuppies.pe",
        "extra_columns": ["Tecnologia", "Categoria de Tecnologia", "Estilo", "Pais de fabricacion"],
        "technology_type": "single_line_text_field",
        "technology_example": "Tecnologia comercial si aplica",
    },
    "KEDS": {
        "site_profile": "HushPuppies.pe",
        "extra_columns": ["Tecnologia", "Categoria de Tecnologia", "Estilo", "Pais de fabricacion"],
        "technology_type": "single_line_text_field",
        "technology_example": "Tecnologia comercial si aplica",
    },
    "VANS": {
        "site_profile": "Vans.pe",
        "extra_columns": ["Tecnologia", "Codigo de referencia"],
        "technology_type": "single_line_text_field",
        "technology_example": "ComfyCush",
    },
}

COMMERCIAL_BRAND_DISPLAY_NAMES = {
    "COLUMBIA": "Columbia",
    "ROCKFORD": "Rockford",
    "MOUNTAIN HARDWEAR": "Mountain Hardwear",
    "PATAGONIA": "Patagonia",
    "SOREL": "Sorel",
    "HUSH PUPPIES": "Hush Puppies",
    "HUSH PUPPIES KIDS": "Hush Puppies Kids",
    "ACCESORIOS HP": "Accesorios HP",
    "KEDS": "Keds",
    "VANS": "Vans",
}

COMMERCIAL_CLASS_EXAMPLES = {
    "Calzado": {
        "Mod-Col": "EJEMPLO-CALZADO-001",
        "Genero": "Hombre",
        "Tipo de prenda": "Zapatillas",
        "Color Comercial": "Negro/Blanco",
        "Color web/filtro": "Negro",
        "Nombre": "Zapatilla Hombre",
        "Descripcion": "Zapatilla liviana para uso urbano y actividades diarias con ajuste comodo.",
        "Caracteristicas": "Suela flexible|Ajuste seguro|Plantilla confortable",
        "Materiales": "Capellada textil|Suela de caucho",
        "Cuidados": "Limpiar con pano humedo|Secar a la sombra",
        "Tags": "Urbano|Lanzamiento",
    },
    "Vestuario": {
        "Mod-Col": "EJEMPLO-VESTUARIO-001",
        "Genero": "Mujer",
        "Tipo de prenda": "Casacas",
        "Color Comercial": "Azul",
        "Color web/filtro": "Azul",
        "Nombre": "Casaca Mujer",
        "Descripcion": "Casaca respirable para proteger de la lluvia ligera y acompanar salidas outdoor.",
        "Caracteristicas": "Repelente al agua|Capucha ajustable|Bolsillos laterales",
        "Materiales": "Exterior: 100% poliester",
        "Cuidados": "Lavar en ciclo suave|No usar lejia",
        "Tags": "Outdoor|Uso diario",
    },
    "Accesorios": {
        "Mod-Col": "EJEMPLO-ACCESORIO-001",
        "Genero": "Unisex",
        "Tipo de prenda": "Gorros",
        "Color Comercial": "Beige",
        "Color web/filtro": "Beige",
        "Nombre": "Gorro Unisex",
        "Descripcion": "Gorro comodo para uso diario con construccion suave y facil de combinar.",
        "Caracteristicas": "Tejido suave|Uso diario|Ajuste comodo",
        "Materiales": "100% acrilico",
        "Cuidados": "Lavar a mano|No planchar",
        "Tags": "Regalo|Uso diario",
    },
}


def _input_norm_key(value):
    text = unicodedata.normalize("NFKD", clean_value(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def commercial_allowed_classes_for_brand(brand_name):
    brand_key = normalize_brand_name(brand_name)
    if "HUSH PUPPIES" in brand_key and "KIDS" in brand_key:
        return COMMERCIAL_BRAND_ALLOWED_CLASSES["HUSH PUPPIES KIDS"]
    if brand_key in COMMERCIAL_BRAND_ALLOWED_CLASSES:
        return COMMERCIAL_BRAND_ALLOWED_CLASSES[brand_key]
    display_key = normalize_brand_name(brand_display_name(brand_name, brand_name))
    return COMMERCIAL_BRAND_ALLOWED_CLASSES.get(display_key, ["Calzado", "Vestuario", "Accesorios"])


def commercial_brand_display_name(brand_name):
    brand_key = normalize_brand_name(brand_name)
    return COMMERCIAL_BRAND_DISPLAY_NAMES.get(
        brand_key,
        brand_display_name(brand_name, clean_value(brand_name)),
    )


def commercial_input_profile_for_brand(brand_name):
    brand_key = normalize_brand_name(brand_name)
    if "HUSH PUPPIES" in brand_key and "KIDS" in brand_key:
        brand_key = "HUSH PUPPIES KIDS"
    profile = COMMERCIAL_BRAND_INPUT_PROFILES.get(brand_key)
    if profile:
        return dict(profile)
    display_key = normalize_brand_name(brand_display_name(brand_name, brand_name))
    profile = COMMERCIAL_BRAND_INPUT_PROFILES.get(display_key)
    if profile:
        return dict(profile)
    return {
        "site_profile": clean_value(brand_name),
        "extra_columns": [],
        "technology_type": "single_line_text_field",
        "technology_example": "",
    }


def commercial_product_type_rules_for_brand(brand_name):
    allowed_classes = {_input_norm_key(item) for item in commercial_allowed_classes_for_brand(brand_name)}
    rows = []
    for rule in PRODUCT_TYPE_RULES:
        category = clean_value(rule.get("category"))
        if category and _input_norm_key(category) not in allowed_classes:
            continue
        rows.append(rule)
    return rows


def configured_commercial_brands():
    brands = {}
    for config in SITE_CONFIGS.values():
        for brand in config.get("allowed_arti_brands", []):
            display = commercial_brand_display_name(brand)
            brands[normalize_brand_name(brand)] = display
    for config in SITE_CONFIGS.values():
        display = commercial_brand_display_name(config.get("label"))
        if display:
            brands.setdefault(normalize_brand_name(config.get("label")), display)
    return [brands[key] for key in sorted(brands)]


def sites_for_commercial_brand(brand_name):
    brand_key = normalize_brand_name(brand_name)
    sites = []
    for site_key, config in SITE_CONFIGS.items():
        allowed = {normalize_brand_name(value) for value in config.get("allowed_arti_brands", [])}
        label_key = normalize_brand_name(config.get("label"))
        if brand_key in allowed or brand_key == label_key:
            site_config = dict(config)
            site_config["site_key"] = site_key
            sites.append(site_config)
    return sites


def publication_column_for_site(site_label):
    label = clean_value(site_label).upper()
    label = label.replace(".", "_").replace("-", "_")
    label = re.sub(r"[^A-Z0-9_]+", "_", label)
    label = re.sub(r"_+", "_", label).strip("_")
    return f"PUBLICAR_{label}"


def commercial_input_metafields_for_brand(brand_name):
    brand_key = normalize_brand_name(brand_name)
    profile = commercial_input_profile_for_brand(brand_name)

    def metafield_row(name, key, data_type="single_line_text_field", appears="NO", rule="La app lo calcula automaticamente.", owner="Catalog Control Center"):
        return {
            "Nombre visible": name,
            "Namespace": "custom",
            "Key": key,
            "Tipo de dato": data_type,
            "Responsable": owner,
            "Aparece en input": appears,
            "Regla": rule,
        }

    common = [
        metafield_row("Marca", "marca", rule="Se toma de la marca seleccionada; Brand no debe cambiarla."),
        metafield_row("Codigo Modelo Color", "codigo_modelo_color", "id", rule="Se toma de Mod-Col y se usa como llave principal."),
        metafield_row("Tipo", "tipo", appears="SI", rule="Se toma de Tipo de prenda y se pluraliza segun el diccionario web.", owner="Brand / Catalog Control Center"),
        metafield_row("Guia de tallas", "guia_de_tallas", "page_reference", rule="La app la asigna por marca, clase, genero y tipo; Brand no la llena."),
        metafield_row("Categoria", "categoria", rule="La app la deriva de Clase y del diccionario por sitio."),
        metafield_row("Sub Categoria", "sub_categoria", rule="La app la deriva del tipo de prenda normalizado."),
        metafield_row("Nombre corto", "nombre_corto", rule="La app lo deriva de Nombre de Producto."),
        metafield_row("Descripcion corta", "descripcion_corta", rule="La app la deriva de Descripcion."),
        metafield_row("Genero", "genero", rule="La app normaliza el Genero informado y ARTI."),
        metafield_row("Grupo Color", "grupo_color", rule="La app lo obtiene del color web/filtro y sus equivalencias."),
        metafield_row("Color Forus", "color_forus", rule="La app lo obtiene del color web/filtro; no usa el codigo de color."),
        metafield_row("Siblings", "siblings", rule="La app relaciona automaticamente los colores del mismo modelo."),
        metafield_row("Siblings color", "siblings_color", rule="La app lo genera con el color visible del producto."),
        metafield_row("Materialidad", "materialidad", appears="SI", rule="Se toma de Materiales sin borrar valores si el input viene vacio.", owner="Brand / Catalog Control Center"),
    ]
    if "Tecnologia" in profile.get("extra_columns", []):
        common.append(
            metafield_row(
                "Tecnologia",
                "tecnologia",
                profile.get("technology_type", "single_line_text_field"),
                appears="SI",
                rule="Brand informa nombres separados por | cuando hay mas de uno; la app normaliza el valor segun el sitio.",
                owner="Brand / Catalog Control Center",
            )
        )
    if brand_key == "COLUMBIA":
        common.append(
            metafield_row(
                "Logo",
                "logo",
                "list.metaobject_reference",
                rule="La app resuelve los GID desde Tecnologia; Brand nunca llena logos ni GID.",
            )
        )
    if profile.get("site_profile") in {"HushPuppies.pe", "Rockford.pe"}:
        common.append(metafield_row("Logo", "logo", rule="Campo tecnico del sitio; Brand no lo llena y un vacio no borra el valor actual."))
    if "Pais de fabricacion" in profile.get("extra_columns", []):
        common.append(
            metafield_row(
                "Pais de Fabricacion",
                "pais_de_fabricacion",
                appears="SI",
                rule="Se toma del input cuando Brand lo conoce; vacio no borra el valor existente.",
                owner="Brand",
            )
        )
    if "Categoria de Tecnologia" in profile.get("extra_columns", []):
        common.extend(
            [
                metafield_row("Categoria de Tecnologia", "categoria_de_tecnologia", appears="SI", rule="Clasificacion comercial de la tecnologia Hush Puppies si aplica.", owner="Brand"),
                metafield_row("Estilo", "estilo", appears="SI", rule="Estilo comercial Hush Puppies si aplica; vacio no borra el valor existente.", owner="Brand"),
            ]
        )
    if profile.get("site_profile") == "Vans.pe":
        common.extend(
            [
                metafield_row("Composicion", "composicion", "multi_line_text_field", appears="SI", rule="Se toma de Materiales y tambien alimenta el Body HTML.", owner="Brand / Catalog Control Center"),
                metafield_row("Codigo de referencia", "codigo_de_referencia", appears="SI", rule="Referencia comercial Vans si existe; vacio no borra el valor actual.", owner="Brand"),
                metafield_row("Productos relacionados", "productos_relacionados", rule="La app conserva o genera relaciones; Brand no llena IDs ni referencias."),
                metafield_row("Sibling", "sibling", rule="La app conserva o genera la relacion; Brand no la llena."),
            ]
        )
    return pd.DataFrame(common)


def commercial_input_columns_for_brand(brand_name):
    base_columns = [
        "Mod-Col",
        "Marca",
        "Genero",
        "Clase",
        "Tipo de prenda",
        "Color Comercial",
        "Color web/filtro",
        "Nombre de Producto",
        "Descripcion",
        "Caracteristicas",
        "Materiales",
        "Cuidados",
    ]
    profile = commercial_input_profile_for_brand(brand_name)
    base_columns.extend(profile.get("extra_columns", []))
    base_columns.extend(["Tags adicionales", "Fecha publicacion"])
    site_columns = [publication_column_for_site(site["site_label"]) for site in sites_for_commercial_brand(brand_name)]
    return base_columns + site_columns


def _commercial_values_rows(brand_name):
    sites = sites_for_commercial_brand(brand_name)
    allowed_classes = commercial_allowed_classes_for_brand(brand_name)
    profile = commercial_input_profile_for_brand(brand_name)
    values = []
    for item in ["Hombre", "Mujer", "Unisex", "Nino", "Nina", "Bebe"]:
        values.append({"Lista": "Genero", "Valor": item, "Marca": brand_name, "Observacion": ""})
    for item in ["Adulto", "Kids", "Junior", "Bebe"]:
        values.append({"Lista": "Grupo de edad", "Valor": item, "Marca": brand_name, "Observacion": ""})
    for item in allowed_classes:
        values.append({"Lista": "Clase", "Valor": item, "Marca": brand_name, "Observacion": ""})
        values.append({"Lista": "Categoria", "Valor": item, "Marca": brand_name, "Observacion": ""})
    for item in ["SI", "NO"]:
        values.append({"Lista": "Publicacion sitio", "Valor": item, "Marca": brand_name, "Observacion": "Obligatorio por sitio."})
    for rule in commercial_product_type_rules_for_brand(brand_name):
        values.append({"Lista": "Tipo de prenda", "Valor": rule.get("plural") or rule.get("normalized"), "Marca": brand_name, "Observacion": rule.get("category", "")})
    for rule in SIZE_GUIDE_RULES:
        guide = clean_value(rule.get("guide"))
        if guide:
            values.append({"Lista": "Guia de talla", "Valor": guide, "Marca": rule.get("brand", ""), "Observacion": rule.get("family", "")})
    if normalize_brand_name(brand_name) == "COLUMBIA":
        for item in ["Omni-Tech", "Omni-Heat Infinity", "Omni-Shield", "Omni-Grip", "OutDry", "Techlite", "Thermarator"]:
            values.append({"Lista": "Tecnologia", "Valor": item, "Marca": "Columbia", "Observacion": "Solo si aplica."})
    elif profile.get("technology_example"):
        values.append(
            {
                "Lista": "Tecnologia",
                "Valor": profile["technology_example"].split("|")[0],
                "Marca": brand_name,
                "Observacion": "Ejemplo comercial; no limita otros valores validos del sitio.",
            }
        )
    for site in sites:
        values.append({"Lista": "Sitios asociados", "Valor": site["site_label"], "Marca": brand_name, "Observacion": publication_column_for_site(site["site_label"])})
    return pd.DataFrame(values)


def _commercial_dictionary_rows(brand_name):
    columns = commercial_input_columns_for_brand(brand_name)
    allowed_classes = commercial_allowed_classes_for_brand(brand_name)
    profile = commercial_input_profile_for_brand(brand_name)
    auto_tag_fields = "marca, Mod-Col, tipo de prenda, color y clase"
    if "Tecnologia" in columns:
        auto_tag_fields += ", ademas de las tecnologias informadas"
    auto_tags_note = (
        "La app agrega automaticamente los tags tradicionales desde ARTI/BigQuery/Shopify: "
        f"{auto_tag_fields}. Brand no debe escribirlos aqui."
    )
    descriptions = {
        "Mod-Col": "Codigo modelo-color real. Es la llave principal del producto.",
        "Marca": "Marca del producto. La plantilla la deja fija para evitar cargas cruzadas.",
        "Genero": "Genero comercial si Brand lo conoce. La app puede reconciliar con ARTI.",
        "Clase": f"Clase comercial permitida para esta marca: {', '.join(allowed_classes)}.",
        "Tipo de prenda": "Tipo web pluralizado cuando corresponda. La app avisa si detecta un tipo nuevo.",
        "Color Comercial": "Color comercial del producto. La app puede usarlo para lectura interna.",
        "Color web/filtro": "Color visible/filtro web, no codigo de color.",
        "Nombre de Producto": "Titulo comercial que se enviara a Shopify como Product.title.",
        "Descripcion": "Texto comercial base. La app arma el Body HTML; Brand no escribe HTML.",
        "Caracteristicas": "Beneficios o bullets separados por |.",
        "Materiales": "Materiales/composicion separados por |. Se usa para Body HTML y metafields si aplica.",
        "Cuidados": "Cuidados separados por |. Se usa para Body HTML.",
        "Tecnologia": (
            "Tecnologias separadas solo por |. La app actualiza custom.tecnologia y, para Columbia, resuelve tambien los logos sin pedir GID al Brand."
            if profile.get("technology_type") == "list.single_line_text_field"
            else "Tecnologia comercial del producto. Si excepcionalmente hay mas de una, separarlas solo por |."
        ),
        "Categoria de Tecnologia": "Clasificacion comercial de la tecnologia Hush Puppies. Informar un solo valor si aplica.",
        "Estilo": "Estilo comercial Hush Puppies si aplica. Vacio no borra el valor actual de Shopify.",
        "Pais de fabricacion": "Pais de fabricacion si la marca dispone del dato. Vacio no borra Shopify.",
        "Codigo de referencia": "Codigo o referencia comercial Vans si existe. Vacio no borra Shopify.",
        "Tags adicionales": f"Solo tags comerciales adicionales separados por |. {auto_tags_note}",
        "Fecha publicacion": "Fecha sugerida si aplica. Puede quedar vacia.",
    }
    examples = {
        "Mod-Col": "2092991-NRY",
        "Marca": brand_display_name(brand_name, brand_name),
        "Genero": "Mujer",
        "Clase": allowed_classes[0] if allowed_classes else "Vestuario",
        "Tipo de prenda": "Casacas",
        "Color Comercial": "Black",
        "Color web/filtro": "Negro",
        "Nombre de Producto": "Casaca Impermeable Mujer Arcadia II",
        "Descripcion": "Casaca impermeable y respirable para lluvia diaria.",
        "Caracteristicas": "Impermeable|Respirable|Capucha ajustable",
        "Materiales": "Exterior: 100% poliester|Forro: malla respirable",
        "Cuidados": "Lavar con agua fria|No usar blanqueador",
        "Tecnologia": profile.get("technology_example", ""),
        "Categoria de Tecnologia": "Confort",
        "Estilo": "Urbano",
        "Pais de fabricacion": "Vietnam",
        "Codigo de referencia": "VN000Y7HBKA",
        "Tags adicionales": "Outdoor|Uso diario|Nueva temporada",
        "Fecha publicacion": "2026-08-01 09:00",
    }
    rows = []
    for column in columns:
        required = column in COMMERCIAL_INPUT_REQUIRED_COLUMNS or column.startswith("PUBLICAR_")
        is_site = column.startswith("PUBLICAR_")
        rows.append(
            {
                "Nombre exacto": column,
                "Nombre visible": column,
                "Descripcion": (
                    "SI se considera para publicar en este sitio; NO se mantiene apagado/no se publica para este sitio."
                    if is_site
                    else descriptions.get(column, f"Campo comercial {column}.")
                ),
                "Responsable de llenado": "Brand" if column not in {"Marca"} and not is_site else ("Usuario eCommerce" if is_site else "Catalog Control Center"),
                "Tipo de dato": "Fecha" if column == "Fecha publicacion" else "Texto",
                "Formato permitido": "SI/NO" if is_site else ("yyyy-mm-dd hh:mm" if column == "Fecha publicacion" else "Texto limpio"),
                "Obligatorio": "SI" if required else "NO",
                "Longitud minima": 150 if column == "Descripcion" else "",
                "Longitud recomendada": "300-1000" if column == "Descripcion" else ("30-80" if column == "Nombre de Producto" else ""),
                "Longitud maxima": 5000 if column == "Descripcion" else "",
                "Valores permitidos": "SI|NO" if is_site else "",
                "Separador aplicable": "|" if column in COMMERCIAL_INPUT_TEXT_LIST_COLUMNS else "",
                "Ejemplo correcto": examples.get(column, "NO" if is_site else ""),
                "Ejemplo incorrecto": "vacio" if required else "",
                "Regla de validacion": "No vacio; SI/NO" if is_site else "Validar contra diccionario si aplica.",
                "Transformacion realizada por la aplicacion": "Genera Body HTML desde campos comerciales." if column in {"Descripcion", "Caracteristicas", "Materiales", "Cuidados"} else "Normaliza espacios y equivalencias.",
                "Campo de Shopify": {
                    "Nombre de Producto": "Product.title",
                    "Descripcion": "Product.bodyHtml",
                    "Caracteristicas": "Product.bodyHtml",
                    "Materiales": "Product.bodyHtml/custom.materialidad",
                    "Cuidados": "Product.bodyHtml",
                    "Tecnologia": "custom.tecnologia/custom.logo",
                    "Categoria de Tecnologia": "custom.categoria_de_tecnologia",
                    "Estilo": "custom.estilo",
                    "Pais de fabricacion": "custom.pais_de_fabricacion",
                    "Codigo de referencia": "custom.codigo_de_referencia",
                    "Tipo de prenda": "Product.productType/custom.tipo",
                }.get(column, "Publication" if is_site else "Auxiliar"),
                "Namespace": "custom" if column in {"Tecnologia", "Categoria de Tecnologia", "Estilo", "Pais de fabricacion", "Codigo de referencia", "Tipo de prenda", "Materiales"} else "",
                "Key": {
                    "Tecnologia": "tecnologia/logo",
                    "Categoria de Tecnologia": "categoria_de_tecnologia",
                    "Estilo": "estilo",
                    "Pais de fabricacion": "pais_de_fabricacion",
                    "Codigo de referencia": "codigo_de_referencia",
                    "Tipo de prenda": "tipo",
                    "Materiales": "materialidad",
                }.get(column, ""),
                "Comportamiento si esta vacio": "Bloquea" if required else "No actualiza ni borra informacion existente.",
                "Nivel de error": "Bloqueo" if required else "Advertencia",
            }
        )
    return pd.DataFrame(rows)


def _commercial_examples_df(brand_name):
    brand_label = commercial_brand_display_name(brand_name)
    columns = commercial_input_columns_for_brand(brand_name)
    allowed_classes = commercial_allowed_classes_for_brand(brand_name)
    profile = commercial_input_profile_for_brand(brand_name)
    site_values = {column: "NO" for column in columns if column.startswith("PUBLICAR_")}
    if site_values:
        first_site_col = next(iter(site_values))
        site_values[first_site_col] = "SI"
    rows = []
    for class_name in allowed_classes:
        example = COMMERCIAL_CLASS_EXAMPLES.get(class_name, {})
        rows.append(
            {
                "Mod-Col": example.get("Mod-Col", f"EJEMPLO-{class_name.upper()}-001"),
                "Marca": brand_label,
                "Genero": example.get("Genero", "Unisex"),
                "Clase": class_name,
                "Tipo de prenda": example.get("Tipo de prenda", ""),
                "Color Comercial": example.get("Color Comercial", ""),
                "Color web/filtro": example.get("Color web/filtro", ""),
                "Nombre de Producto": f"{example.get('Nombre', class_name)} {brand_label} Ejemplo",
                "Descripcion": example.get("Descripcion", ""),
                "Caracteristicas": example.get("Caracteristicas", ""),
                "Materiales": example.get("Materiales", ""),
                "Cuidados": example.get("Cuidados", ""),
                "Tecnologia": (
                    "Techlite" if class_name == "Calzado" and normalize_brand_name(brand_name) == "COLUMBIA"
                    else "Omni-Shield" if class_name == "Vestuario" and normalize_brand_name(brand_name) == "COLUMBIA"
                    else profile.get("technology_example", "")
                ),
                "Categoria de Tecnologia": "Confort" if profile.get("site_profile") == "HushPuppies.pe" else "",
                "Estilo": "Urbano" if profile.get("site_profile") == "HushPuppies.pe" else "",
                "Pais de fabricacion": "Vietnam",
                "Codigo de referencia": "VN000Y7HBKA" if profile.get("site_profile") == "Vans.pe" else "",
                "Tags adicionales": example.get("Tags", ""),
                "Fecha publicacion": "",
                **site_values,
            }
        )
    return pd.DataFrame(rows).reindex(columns=columns)


def _commercial_input_blank_df(brand_name, rows=100):
    columns = commercial_input_columns_for_brand(brand_name)
    df = pd.DataFrame([{column: "" for column in columns} for _ in range(rows)])
    df["Marca"] = commercial_brand_display_name(brand_name)
    for column in columns:
        if column.startswith("PUBLICAR_"):
            df[column] = "NO"
    return df


def _commercial_brand_fill_guide_df(brand_name):
    """Return only the columns and wording that a Brand needs to see."""
    columns = commercial_input_columns_for_brand(brand_name)
    allowed_classes = commercial_allowed_classes_for_brand(brand_name)
    examples_df = _commercial_examples_df(brand_name)
    example_row = examples_df.iloc[0].to_dict() if not examples_df.empty else {}
    descriptions = {
        "Mod-Col": "Código real del modelo y color.",
        "Marca": "Marca del producto. Esta columna ya viene completada.",
        "Genero": "Hombre, Mujer, Unisex, Niño, Niña o Bebé.",
        "Clase": f"Selecciona una opción: {', '.join(allowed_classes)}.",
        "Tipo de prenda": "Tipo de producto, por ejemplo Casacas, Zapatillas o Gorros.",
        "Color Comercial": "Nombre comercial del color informado por la marca.",
        "Color web/filtro": "Color simple que verá el cliente, por ejemplo Negro, Azul o Beige.",
        "Nombre de Producto": "Nombre comercial completo del producto.",
        "Descripcion": "Descripción comercial principal del producto.",
        "Caracteristicas": "Beneficios principales. Si hay varios, sepáralos únicamente con |.",
        "Materiales": "Materiales o composición. Si hay varios, sepáralos únicamente con |.",
        "Cuidados": "Recomendaciones de cuidado. Si hay varias, sepáralas únicamente con |.",
        "Tecnologia": "Tecnologías comerciales. Si hay varias, sepáralas únicamente con |.",
        "Categoria de Tecnologia": "Categoría comercial de la tecnología, si aplica.",
        "Estilo": "Estilo comercial del producto, si aplica.",
        "Pais de fabricacion": "País donde fue fabricado el producto, si se conoce.",
        "Codigo de referencia": "Código de referencia comercial, si aplica.",
        "Tags adicionales": "Palabras comerciales adicionales. Si hay varias, sepáralas únicamente con |.",
        "Fecha publicacion": (
            "Fecha y hora de publicación en formato DD/MM/AAAA HH:MM, hora de Lima. "
            "Ejemplo: 01/08/2026 09:00. Déjala vacía para publicar de inmediato."
        ),
    }
    rows = []
    for column in columns:
        is_site = column.startswith("PUBLICAR_")
        required = column in COMMERCIAL_INPUT_REQUIRED_COLUMNS or is_site
        if is_site:
            site_name = column.removeprefix("PUBLICAR_").replace("_", ".").title()
            description = f"Escribe SI para incluir el producto en {site_name}; escribe NO si no corresponde."
            example = "SI"
        elif column == "Fecha publicacion":
            description = descriptions[column]
            example = "01/08/2026 09:00"
        else:
            description = descriptions.get(column, f"Completa la información de {column}.")
            example = clean_value(example_row.get(column))
        rows.append(
            {
                "Columna": column,
                "Qué debes completar": description,
                "Ejemplo": example,
                "¿Es obligatorio?": "SI" if required else "NO",
            }
        )
    return pd.DataFrame(rows)


def build_brand_commercial_input_workbook(brand_name):
    """Build the three-sheet, Brand-facing commercial workbook."""
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    brand_label = commercial_brand_display_name(brand_name)
    columns = commercial_input_columns_for_brand(brand_label)
    site_columns = [column for column in columns if column.startswith("PUBLICAR_")]
    allowed_classes = commercial_allowed_classes_for_brand(brand_label)
    examples_df = _commercial_examples_df(brand_label)
    guide_df = _commercial_brand_fill_guide_df(brand_label)
    blank_df = _commercial_input_blank_df(brand_label)
    guide_lookup = guide_df.set_index("Columna")["Qué debes completar"].to_dict()

    sheets = {
        "PARA_COMPLETAR": blank_df,
        "EJEMPLO": examples_df,
        "COMO_LLENAR": guide_df,
    }
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            repair_mojibake_dataframe(df).to_excel(writer, index=False, sheet_name=sheet_name)

        wb = writer.book
        navy = "063B73"
        blue = "0B78D0"
        pale_blue = "EAF4FF"
        pale_green = "EAF7EF"
        white = "FFFFFF"
        text_color = "10233F"
        thin = Side(style="thin", color="D7E2EE")

        for ws in wb.worksheets:
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A2"
            ws.protection.sheet = False
            ws.auto_filter.ref = ws.dimensions
            ws.row_dimensions[1].height = 38
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=navy)
                cell.font = Font(bold=True, color=white, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(bottom=Side(style="medium", color=blue))
            for row_index in range(2, ws.max_row + 1):
                ws.row_dimensions[row_index].height = 34
                row_fill = PatternFill("solid", fgColor=white if row_index % 2 == 0 else "F7FAFD")
                for cell in ws[row_index]:
                    cell.fill = row_fill
                    cell.font = Font(color=text_color, size=10)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = Border(bottom=thin)

        widths = {
            "Mod-Col": 22,
            "Marca": 20,
            "Genero": 16,
            "Clase": 17,
            "Tipo de prenda": 24,
            "Color Comercial": 22,
            "Color web/filtro": 22,
            "Nombre de Producto": 38,
            "Descripcion": 52,
            "Caracteristicas": 46,
            "Materiales": 42,
            "Cuidados": 42,
            "Tecnologia": 32,
            "Categoria de Tecnologia": 28,
            "Estilo": 22,
            "Pais de fabricacion": 22,
            "Codigo de referencia": 24,
            "Tags adicionales": 38,
            "Fecha publicacion": 22,
        }
        for sheet_name in ("PARA_COMPLETAR", "EJEMPLO"):
            ws = wb[sheet_name]
            for index, column in enumerate(columns, start=1):
                ws.column_dimensions[get_column_letter(index)].width = widths.get(column, 24 if not column.startswith("PUBLICAR_") else 24)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(2, ws.max_row)}"

        input_ws = wb["PARA_COMPLETAR"]
        input_ws.sheet_properties.tabColor = blue
        positions = {cell.value: cell.column for cell in input_ws[1]}
        for column in columns:
            column_index = positions[column]
            header = input_ws.cell(1, column_index)
            header.comment = Comment(guide_lookup.get(column, "Completa esta columna."), "Catalog Control Center")
            header.comment.width = 430
            header.comment.height = 190
            if column == "Marca":
                for row_index in range(2, input_ws.max_row + 1):
                    input_ws.cell(row_index, column_index).fill = PatternFill("solid", fgColor=pale_blue)
            elif column not in COMMERCIAL_INPUT_REQUIRED_COLUMNS and not column.startswith("PUBLICAR_"):
                for row_index in range(2, input_ws.max_row + 1):
                    input_ws.cell(row_index, column_index).fill = PatternFill("solid", fgColor=pale_green)

        si_no_validation = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
        input_ws.add_data_validation(si_no_validation)
        for column in site_columns:
            letter = get_column_letter(positions[column])
            si_no_validation.add(f"{letter}2:{letter}{COMMERCIAL_INPUT_MAX_ROWS + 1}")

        for target_column, values in (
            ("Genero", ["Hombre", "Mujer", "Unisex", "Nino", "Nina", "Bebe"]),
            ("Clase", allowed_classes),
        ):
            if target_column not in positions or not values:
                continue
            validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
            input_ws.add_data_validation(validation)
            letter = get_column_letter(positions[target_column])
            validation.add(f"{letter}2:{letter}{COMMERCIAL_INPUT_MAX_ROWS + 1}")

        if "Fecha publicacion" in positions:
            date_letter = get_column_letter(positions["Fecha publicacion"])
            for row_index in range(2, COMMERCIAL_INPUT_MAX_ROWS + 2):
                input_ws[f"{date_letter}{row_index}"].number_format = "dd/mm/yyyy hh:mm"

        example_ws = wb["EJEMPLO"]
        example_ws.sheet_properties.tabColor = "22A06B"
        for row_index in range(2, example_ws.max_row + 1):
            example_ws.row_dimensions[row_index].height = 66

        guide_ws = wb["COMO_LLENAR"]
        guide_ws.sheet_properties.tabColor = "F3B61F"
        guide_ws.column_dimensions["A"].width = 30
        guide_ws.column_dimensions["B"].width = 86
        guide_ws.column_dimensions["C"].width = 54
        guide_ws.column_dimensions["D"].width = 18
        for row_index in range(2, guide_ws.max_row + 1):
            description_length = len(clean_value(guide_ws.cell(row_index, 2).value))
            example_length = len(clean_value(guide_ws.cell(row_index, 3).value))
            guide_ws.row_dimensions[row_index].height = max(
                52,
                18 * max(2, math.ceil(description_length / 78), math.ceil(example_length / 48)),
            )
            required_cell = guide_ws.cell(row_index, 4)
            required_cell.font = Font(
                bold=True,
                color="A61B1B" if required_cell.value == "SI" else "3A556F",
            )

    buffer.seek(0)
    return buffer


def _commercial_find_column(df, aliases):
    normalized = {normalize_header(column): column for column in df.columns}
    for alias in aliases:
        key = normalize_header(alias)
        if key in normalized:
            return normalized[key]
    return None


def build_body_html_from_commercial_row(row):
    def parts(value):
        raw = clean_value(value)
        if not raw:
            return []
        pieces = re.split(r"[|]", raw)
        return [escape(clean_value(piece)) for piece in pieces if clean_value(piece)]

    description = clean_value(row.get("Descripcion"))
    sections = []
    if description:
        sections.append(f"<section><h3>Descripcion</h3><p>{escape(description)}</p></section>")
    for title, column in [
        ("Caracteristicas", "Caracteristicas"),
        ("Materiales", "Materiales"),
        ("Cuidados", "Cuidados"),
    ]:
        if column == "Materiales":
            items = parts(first_non_empty(row.get("Materiales"), row.get("Materiales o composicion"), row.get("Composicion")))
        else:
            items = parts(row.get(column))
        if items:
            sections.append(f"<section><h3>{title}</h3><ul>{''.join(f'<li>{item}</li>' for item in items)}</ul></section>")
    return "\n".join(sections)


def validate_brand_commercial_input(uploaded_file, brand_name):
    try:
        xls = pd.ExcelFile(uploaded_file)
    except Exception as exc:
        return pd.DataFrame(), pd.DataFrame([{"Fila": "", "Campo": "Archivo", "Estado": "Bloqueado", "Mensaje": f"No se pudo leer Excel: {exc}"}]), pd.DataFrame()
    sheet_name = "INPUT_COMERCIAL" if "INPUT_COMERCIAL" in xls.sheet_names else xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=sheet_name, dtype=object).dropna(how="all")
    df = repair_mojibake_dataframe(df)
    if df.empty:
        return df, pd.DataFrame([{"Fila": "", "Campo": "Archivo", "Estado": "Bloqueado", "Mensaje": "El input no tiene filas reales."}]), pd.DataFrame()
    brand_label = commercial_brand_display_name(brand_name)
    expected_columns = commercial_input_columns_for_brand(brand_label)
    site_columns = [column for column in expected_columns if column.startswith("PUBLICAR_")]
    allowed_classes = commercial_allowed_classes_for_brand(brand_label)
    allowed_class_keys = {_input_norm_key(value) for value in allowed_classes}
    report_rows = []
    preview_rows = []
    column_aliases = {
        "Nombre de Producto": ["Nombre de Producto", "Nombre web o Title", "Title", "Titulo", "Título", "Nombre Producto"],
        "Materiales": ["Materiales", "Materiales o composicion", "Materiales o composición", "Composicion", "Composición"],
        "Tecnologia": ["Tecnologia", "Tecnología", "Tecnologias", "Tecnologías", "Technology"],
        "Categoria de Tecnologia": ["Categoria de Tecnologia", "Categoría de Tecnología", "Categoria Tecnologia", "Technology Category"],
        "Pais de fabricacion": ["Pais de fabricacion", "País de fabricación", "Pais Fabricacion", "Country of origin", "Country"],
        "Codigo de referencia": ["Codigo de referencia", "Código de referencia", "Codigo referencia", "Reference code", "Referencia"],
        "Tags adicionales": ["Tags adicionales", "Tags sugeridos", "Tags extra", "Tags"],
        "Color Comercial": ["Color Comercial", "Color comercial", "Color Comercial ", "Color"],
        "Color web/filtro": ["Color web/filtro", "Color Web", "Color Forus", "Color visible", "Color filtro"],
    }
    col_map = {column: _commercial_find_column(df, column_aliases.get(column, [column])) for column in expected_columns}
    for required in COMMERCIAL_INPUT_REQUIRED_COLUMNS + site_columns:
        if not col_map.get(required):
            report_rows.append({"Fila": "", "Mod-Col": "", "Campo": required, "Valor original": "", "Valor normalizado": "", "Estado": "Bloqueado", "Mensaje": "Falta columna obligatoria.", "Accion recomendada": "Usar formato vigente por marca."})
    if report_rows:
        return df, pd.DataFrame(report_rows), pd.DataFrame()
    seen = set()
    for idx, row in df.iterrows():
        excel_row = int(idx) + 2
        normalized = {}
        for column in expected_columns:
            source_col = col_map.get(column)
            normalized[column] = clean_value(row.get(source_col)) if source_col else ""
        mod_col = normalized.get("Mod-Col")
        if not mod_col or mod_col.upper().startswith("EJEMPLO-"):
            continue
        row_status = "Listo"
        row_messages = []
        row_brand = normalize_brand_name(normalized.get("Marca"))
        expected_brand = normalize_brand_name(brand_label)
        if row_brand != expected_brand:
            row_status = "Bloqueado"
            row_messages.append(f"Marca del archivo ({normalized.get('Marca')}) no corresponde a {brand_label}.")
        for column in COMMERCIAL_INPUT_REQUIRED_COLUMNS:
            value = normalized.get(column)
            if _input_norm_key(value) in COMMERCIAL_INPUT_INVALID_TEXTS:
                row_status = "Bloqueado"
                row_messages.append(f"{column} obligatorio vacio o invalido.")
        for column in site_columns:
            value = normalized.get(column).upper()
            if value not in {"SI", "NO"}:
                row_status = "Bloqueado"
                row_messages.append(f"{column} debe ser SI o NO.")
        class_value = normalized.get("Clase")
        if class_value and _input_norm_key(class_value) not in allowed_class_keys:
            row_status = "Bloqueado"
            row_messages.append(f"Clase '{class_value}' no permitida para {brand_label}. Permitidas: {', '.join(allowed_classes)}.")
        publication_date = normalized.get("Fecha publicacion")
        if publication_date:
            normalized_publication_date = parse_publication_date(publication_date)
            if parse_iso_datetime(normalized_publication_date) is None:
                row_status = "Bloqueado"
                row_messages.append(
                    "Fecha publicacion debe usar DD/MM/AAAA HH:MM. Ejemplo: 01/08/2026 09:00."
                )
        key = (mod_col, normalized.get("Nombre de Producto"))
        if key in seen:
            row_status = "Con advertencia" if row_status == "Listo" else row_status
            row_messages.append("Mod-Col duplicado en el input; revisar si corresponde a variantes o duplicidad.")
        seen.add(key)
        type_decision = validate_catalog_row(
            {
                "Mod-Col": mod_col,
                "Marca": normalized.get("Marca"),
                "Genero": normalized.get("Genero"),
                "Categoria": normalized.get("Clase"),
                "Tipo de prenda": normalized.get("Tipo de prenda"),
                "Guia de tallas": "",
                "Talla": "M",
                "SKU": "VALIDACION",
            }
        )
        for issue in type_decision.get("issues", []):
            if issue.get("field") in {"Tipo de prenda", "Guia de tallas"}:
                level = clean_value(issue.get("level"))
                if level == "bloqueo":
                    row_status = "Bloqueado"
                elif row_status == "Listo":
                    row_status = "Con advertencia"
                row_messages.append(f"{issue.get('field')}: {issue.get('message')}")
        for list_column in COMMERCIAL_INPUT_TEXT_LIST_COLUMNS:
            value = normalized.get(list_column)
            if "||" in value:
                if row_status == "Listo":
                    row_status = "Con advertencia"
                row_messages.append(f"{list_column} contiene separadores vacios ||.")
            if value and any(separator in value for separator in [",", ";", "\n", "\r"]):
                if row_status == "Listo":
                    row_status = "Con advertencia"
                row_messages.append(
                    f"{list_column} debe usar solo | para separar valores. "
                    "La app convierte | en listas compatibles con Shopify y bullets del Body HTML."
                )
        description = normalized.get("Descripcion")
        visible_len = len(strip_html(description))
        if description and visible_len < 150:
            if row_status == "Listo":
                row_status = "Con advertencia"
            row_messages.append("Descripcion bajo 150 caracteres visibles.")
        handle = build_catalog_handle(
            product_type=normalized.get("Tipo de prenda"),
            gender=normalized.get("Genero"),
            brand=normalized.get("Marca"),
            mod_col=mod_col,
        )
        body_html = build_body_html_from_commercial_row(normalized)
        preview_rows.append(
            {
                "Fila": excel_row,
                "Mod-Col": mod_col,
                "Marca": normalized.get("Marca"),
                "Title propuesto": normalized.get("Nombre de Producto"),
                "Handle sugerido": handle,
                "Body HTML generado": body_html,
                "Sitios SI": ", ".join(column for column in site_columns if normalized.get(column).upper() == "SI"),
                "Estado": row_status,
                "Mensaje": " | ".join(row_messages),
            }
        )
        if row_messages:
            for message in row_messages:
                report_rows.append(
                    {
                        "Fila": excel_row,
                        "Mod-Col": mod_col,
                        "Campo": "Validacion",
                        "Valor original": "",
                        "Valor normalizado": "",
                        "Estado": row_status,
                        "Mensaje": message,
                        "Accion recomendada": "Corregir input o aprobar valor nuevo en diccionario.",
                    }
                )
    preview_df = pd.DataFrame(preview_rows)
    report_df = pd.DataFrame(report_rows)
    summary_df = pd.DataFrame(
        [
            {"Indicador": "Filas analizadas", "Valor": len(preview_df)},
            {"Indicador": "Modelos-color", "Valor": preview_df["Mod-Col"].nunique() if not preview_df.empty else 0},
            {"Indicador": "Registros validos", "Valor": int(preview_df["Estado"].eq("Listo").sum()) if not preview_df.empty else 0},
            {"Indicador": "Registros con advertencias", "Valor": int(preview_df["Estado"].eq("Con advertencia").sum()) if not preview_df.empty else 0},
            {"Indicador": "Registros bloqueados", "Valor": int(preview_df["Estado"].eq("Bloqueado").sum()) if not preview_df.empty else 0},
            {"Indicador": "Valores nuevos / revisar", "Valor": len(report_df) if not report_df.empty else 0},
        ]
    )
    return preview_df, report_df, summary_df


def render_commercial_input_center(download_only=False, forced_brands=None, actor=None):
    brand_options = list(forced_brands or configured_commercial_brands())
    if not brand_options:
        st.error("Tu usuario no tiene marcas autorizadas. Solicita al administrador configurar app_auth.brands en Secrets.")
        return
    if st.session_state.get("commercial_input_brand") not in brand_options:
        st.session_state["commercial_input_brand"] = brand_options[0]
    with st.container(key="commercial_input_download_panel"):
        st.markdown(
            """
            <div class="commercial-input-heading">
                <p>Formato por marca</p>
                <h2>Descargar input comercial</h2>
                <span>Un formato simple, ejemplos completos y una guía breve</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        selected_brand = st.selectbox("Marca", brand_options, key="commercial_input_brand")
        sites = sites_for_commercial_brand(selected_brand)
        if not sites:
            st.warning("No encontre sitios asociados para esta marca en la configuracion actual.")
        else:
            site_rules = pd.DataFrame(
                [
                    {
                        "Sitio": site.get("site_label"),
                        "Cómo indicarlo": publication_column_for_site(site.get("site_label")),
                        "SI": "Incluir",
                        "NO": "No incluir",
                    }
                    for site in sites
                ]
            )
            st.dataframe(site_rules, use_container_width=True, hide_index=True)

        workbook_bytes = build_brand_commercial_input_workbook(selected_brand)
        file_date = datetime.now().strftime("%Y%m%d")
        file_brand = re.sub(r"[^A-Za-z0-9]+", "_", selected_brand).strip("_").upper()
        st.download_button(
            f"Descargar input editable de {selected_brand}",
            data=workbook_bytes,
            file_name=f"Input_Catalogo_{file_brand}_{file_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_brand_commercial_input",
        )
        st.markdown(
            """
            <div class="commercial-input-note">
                <strong>Completa únicamente las columnas del formato.</strong>
                Para separar varias características, materiales, cuidados, tecnologías o palabras adicionales usa solo <strong>|</strong>.
                Todo lo demás se prepara automáticamente.
            </div>
            """,
            unsafe_allow_html=True,
        )

    if download_only:
        return

    with st.container(key="commercial_input_validate_panel"):
        st.markdown('<div class="commercial-input-heading"><p>Revisión antes de crear</p><h2>Validar input comercial</h2></div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("Subir input comercial completado", type=["xlsx", "xls"], key="validate_brand_commercial_input")
        if uploaded is not None:
            current_upload_hash = file_sha256(uploaded.getvalue())
            validated_hash = clean_value(st.session_state.get("brand_input_validated_hash"))
            validated_brand = clean_value(st.session_state.get("brand_input_validated_brand"))
            if validated_hash and (validated_hash != current_upload_hash or validated_brand != selected_brand):
                for stale_key in [
                    "brand_input_preview_df",
                    "brand_input_report_df",
                    "brand_input_summary_df",
                    "brand_input_validated_bytes",
                    "brand_input_validated_name",
                    "brand_input_validated_hash",
                    "brand_input_validated_brand",
                ]:
                    st.session_state.pop(stale_key, None)
        if uploaded is not None and st.button("Analizar input comercial", type="primary", key="analyze_brand_commercial_input"):
            input_bytes = uploaded.getvalue()
            preview_df, report_df, summary_df = validate_brand_commercial_input(uploaded, selected_brand)
            st.session_state["brand_input_preview_df"] = preview_df
            st.session_state["brand_input_report_df"] = report_df
            st.session_state["brand_input_summary_df"] = summary_df
            st.session_state["brand_input_validated_bytes"] = input_bytes
            st.session_state["brand_input_validated_name"] = clean_value(uploaded.name) or "input_comercial.xlsx"
            st.session_state["brand_input_validated_hash"] = file_sha256(input_bytes)
            st.session_state["brand_input_validated_brand"] = selected_brand
        summary_df = st.session_state.get("brand_input_summary_df")
        preview_df = st.session_state.get("brand_input_preview_df")
        report_df = st.session_state.get("brand_input_report_df")
        if isinstance(summary_df, pd.DataFrame) and not summary_df.empty:
            cols = st.columns(min(4, len(summary_df)))
            for idx, row in summary_df.iterrows():
                cols[idx % len(cols)].metric(clean_value(row.get("Indicador")), int(row.get("Valor", 0)))
            if isinstance(preview_df, pd.DataFrame) and not preview_df.empty:
                st.subheader("Vista previa")
                st.dataframe(preview_df, use_container_width=True, hide_index=True)
            if isinstance(report_df, pd.DataFrame) and not report_df.empty:
                st.subheader("Errores y advertencias")
                st.dataframe(report_df, use_container_width=True, hide_index=True)
            st.download_button(
                "Descargar reporte de validacion",
                data=(validation_report_bytes := dataframe_to_excel_bytes(
                    {
                        "Resumen": summary_df,
                        "Vista previa": preview_df if isinstance(preview_df, pd.DataFrame) else pd.DataFrame(),
                        "Errores y advertencias": report_df if isinstance(report_df, pd.DataFrame) else pd.DataFrame(),
                    }
                )),
                file_name=f"reporte_validacion_input_{file_brand}_{file_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_brand_input_validation_report",
            )
            summary_map = {
                clean_value(row.get("Indicador")): int(row.get("Valor", 0) or 0)
                for _, row in summary_df.iterrows()
            }
            blocked = int(summary_map.get("Registros bloqueados", 0))
            products = int(summary_map.get("Filas analizadas", 0))
            model_colors = int(summary_map.get("Modelos-color", 0))
            same_validation = (
                st.session_state.get("brand_input_validated_brand") == selected_brand
                and st.session_state.get("brand_input_validated_hash")
                == file_sha256(st.session_state.get("brand_input_validated_bytes", b""))
            )
            if blocked:
                st.error(f"La solicitud no puede enviarse: existen {blocked} registros bloqueados.")
            elif products <= 0:
                st.warning("No existen productos válidos para crear una solicitud.")
            elif actor and actor.get("role") in {ROLE_BRAND, ROLE_ADMIN} and same_validation:
                st.markdown("### Enviar al equipo de catálogo")
                ticket_comment = st.text_area(
                    "Comentario para Operaciones",
                    key="brand_ticket_comment",
                    placeholder="Campaña, fecha requerida o contexto que deba conocer el equipo.",
                )
                priority_label = st.selectbox(
                    "Prioridad",
                    [PRIORITY_LABELS[key] for key in PRIORITIES],
                    index=list(PRIORITIES).index("normal"),
                    key="brand_ticket_priority",
                )
                confirmed = st.checkbox(
                    "Confirmo que revisé la vista previa y que esta versión está lista para revisión interna.",
                    key="brand_ticket_confirmed",
                )
                if st.button(
                    "Enviar solicitud de carga",
                    type="primary",
                    disabled=not confirmed,
                    key="submit_catalog_ticket",
                ):
                    try:
                        service, backend = get_ticket_service()
                        selected_publication_columns = set()
                        if isinstance(preview_df, pd.DataFrame) and "Sitios SI" in preview_df.columns:
                            for value in preview_df["Sitios SI"].dropna().astype(str):
                                selected_publication_columns.update(
                                    clean_value(item) for item in value.split(",") if clean_value(item)
                                )
                        selected_sites = [
                            clean_value(site.get("site_label"))
                            for site in sites
                            if publication_column_for_site(site.get("site_label")) in selected_publication_columns
                        ]
                        warnings = []
                        if isinstance(report_df, pd.DataFrame) and not report_df.empty:
                            warning_rows = report_df[report_df["Estado"].astype(str).ne("Bloqueado")]
                            warnings = warning_rows.get("Mensaje", pd.Series(dtype=object)).dropna().astype(str).tolist()
                        priority = next(
                            key for key, label in PRIORITY_LABELS.items() if label == priority_label
                        )
                        ticket = service.create_ticket(
                            actor,
                            brand=selected_brand,
                            sites=selected_sites,
                            filename=st.session_state.get("brand_input_validated_name", "input_comercial.xlsx"),
                            input_bytes=st.session_state.get("brand_input_validated_bytes", b""),
                            report_bytes=validation_report_bytes,
                            template_version="2026.07",
                            load_type="complete",
                            summary={
                                "products": products,
                                "model_colors": model_colors,
                                "variants": 0,
                                "new_products": 0,
                                "updated_products": 0,
                                "blocked": blocked,
                                "warnings": int(summary_map.get("Registros con advertencias", 0)),
                            },
                            warnings=warnings,
                            comment=ticket_comment,
                            model_colors=preview_df.get("Mod-Col", pd.Series(dtype=object)).dropna().astype(str).tolist(),
                            priority=priority,
                        )
                        st.session_state["selected_catalog_ticket"] = ticket["code"]
                        st.success(f"Solicitud {ticket['code']} creada y enviada a Operaciones.")
                        if backend == "local":
                            st.info("Prueba local: el ticket se guardó en outputs/catalog_tickets. En producción configura el backend GitHub.")
                    except TicketError as exc:
                        st.error(str(exc))


CENTRY_BASE_COLUMNS = [
    "Nombre del Producto",
    "Marca",
    "Descripcion",
    "Listado de caracterÃ­sticas",
    "GarantÃ­a",
    "Alto del paquete",
    "Ancho del paquete",
    "Largo del paquete",
    "Peso del paquete",
    "CategorÃ­a",
    "Precio",
    "SKU del producto",
    "SKU de la variante",
    "CÃ³digo de barra variante (EAN/UPC/ISBN)",
    "Color",
    "Talla",
    "CondiciÃ³n del Producto",
    "AÃ±o de temporada",
    "Temporada",
    "GÃ©nero",
    "URL imagen principal",
    "URL imagen 2",
    "URL imagen 3",
    "URL imagen 4",
    "URL imagen 5",
    "Estado",
    "Incluir en Falabella Global Peru / FalabellaGlobalProduction",
]
CENTRY_FOOTWEAR_COLUMNS = [
    "Material principal - Calzado (Falabella GSC PerÃº)",
    "GÃ©nero - Calzado (Falabella GSC PerÃº)",
    "Tipo - Calzado (Falabella GSC PerÃº)",
    "Altura de la plataforma - Calzado (Falabella GSC PerÃº)",
    "Altura del taco - Calzado (Falabella GSC PerÃº)",
    "Material de la plantilla - Calzado (Falabella GSC PerÃº)",
    "Material de la suela - Calzado (Falabella GSC PerÃº)",
    "Horma - Calzado (Falabella GSC PerÃº)",
    "Tipo de caÃ±a - Calzado (Falabella GSC PerÃº)",
    "Forma de la punta - Calzado (Falabella GSC PerÃº)",
    "Material del forro - Calzado (Falabella GSC PerÃº)",
    "Material de la suela (MercadoLibre PerÃº)",
    "Con suela antideslizante (MercadoLibre PerÃº)",
    "Tipo de caÃ±a (MercadoLibre PerÃº)",
    "Tipo de calzado (MercadoLibre PerÃº)",
    "Edad (MercadoLibre PerÃº)",
    "Forma de la punta (MercadoLibre PerÃº)",
    "Material del calzado (MercadoLibre PerÃº)",
    "Materiales del exterior (MercadoLibre PerÃº)",
    "Materiales del interior(MercadoLibre PerÃº)",
    "Tipo de taco (MercadoLibre PerÃº)",
]
CENTRY_APPAREL_COLUMNS = [
    "GÃ©nero de vestuario - Ropa y accesorios (Falabella GSC PerÃº)",
    "Material de vestuario - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de cierre - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de prenda para la parte superior - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de camisa/blusa/polo/camiseta - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de chaqueta/chaleco - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de pantalÃ³n largo/corto - Ropa y accesorios (Falabella GSC PerÃº)",
    "ComposiciÃ³n - Ropa y accesorios (Falabella GSC PerÃº)",
    "Largo de mangas - Ropa y accesorios (Falabella GSC PerÃº)",
    "DiseÃ±o - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de cuello - Tipo de cuello - Ropa y accesorios (Falabella GSC PerÃº)",
]
CENTRY_MARKETPLACE_EXTRA_COLUMNS = [
    "cccc",
    "Incluir en MercadoLibrePe / MercadoLibrePe",
    "Es inflamable (MercadoLibre PerÃº)",
    "Material del accesorio - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de ropa para la cabeza - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de ropa para el cuello - Ropa y accesorios (Falabella GSC PerÃº)",
    "ComposiciÃ³n (MercadoLibre PerÃº)",
    "Material principal - Material de vestuario - Ropa y accesorios (Falabella GSC PerÃº)",
    "Contenido del paquete - Package content - Almacenamiento (Falabella GSC PerÃº)",
    "DiseÃ±o de la tela (MercadoLibre PerÃº)",
    "TamaÃ±o del bolso maleta - Ropa y accesorios (Falabella GSC PerÃº)",
    "Cantidad de bolsillos interiores - Cantidad de bolsillos interiores - Ropa y accesorios (Falabella GSC PerÃº)",
    "Material de la maleta - Material de la maleta - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de maleta - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de bolso mochila funda - Ropa y accesorios (Falabella GSC PerÃº)",
    "Con monedero (MercadoLibre PerÃº)",
    "Tipo de sombrero (MercadoLibre PerÃº)",
    "Con cierre ajustable (MercadoLibre PerÃº)",
    "Es reversible (MercadoLibre PerÃº)",
    "Uso de la cartera/mochila/bolsa - Uso de la cartera mochila bolsa - Ropa y accesorios (Falabella GSC PerÃº)",
    "Tipo de cartera (MercadoLibre PerÃº)",
    "Con cierre (MercadoLibre PerÃº)",
    "INFORMACIÃ“N ADICIONAL",
    "Unnamed: 52",
    "Unnamed: 53",
    "Unnamed: 54",
    "Unnamed: 55",
    "Unnamed: 56",
]
CENTRY_TAIL_COLUMNS = ["Clase", "GuÃ­a de tallas", "Base de categorÃ­a", "Cod Mod Col Talla", "Mod", "Col", "Tal"]
CENTRY_COLUMNS = list(dict.fromkeys(CENTRY_BASE_COLUMNS + CENTRY_MARKETPLACE_EXTRA_COLUMNS + CENTRY_FOOTWEAR_COLUMNS + CENTRY_APPAREL_COLUMNS + CENTRY_TAIL_COLUMNS))
CENTRY_SIAL_COLUMNS = [
    "Mod", "Col", "Tal", "Product Name ", "Product Bullets", "Product Description", "Image URL",
    "Product Weight", "Product Length", "Product Width", "Product Height",
    "Package Weight", "Package Length", "Package Width", "Package Height",
    "Boost ", "Talla Web ", "Color Web", "Categoria ", "Sub Categoria", "Genero",
    "Estilo ", "Colecciones ", "Temporada ", "Modelo", "Marca", "Tecnologias ",
    "Caracteristicas", "Tipo de Boardshort", "Tipo de Bikini", "Iniciativas",
    "Tipo de Material", "1", "Tipo de Prenda", "Adicional 2 ", "Adicional 3 ",
    "Adicional 4 ", "Adicional 5 ", "Adicional 6 ", "Adicional 7 ", "Adicional 8 ",
    "Adicional 9 ", "Adicional 10", "Mod-Col", "Sku - Sial",
    "Nuevo o Actualizar (Rockford.pe)", "Sku - Supermall.pe", "Porduct Id - Supermall.pe",
]


def centry_value(value, fallback=""):
    text = clean_value(value)
    if not text or text.upper() in {"#N/D", "#ND", "#N/A", "NAN", "NONE", "NULL"}:
        return fallback
    return text


def centry_is_footwear(row):
    haystack = " ".join(
        centry_value(row.get(column)).lower()
        for column in ("Type", "Tags", "Title", "CategorÃ­a", "Categoria ")
    )
    return any(word in haystack for word in ("calzado", "zapatilla", "zapato", "botin", "bota", "sandalia"))


def centry_gender(row):
    text = " ".join(centry_value(row.get(column)).lower() for column in ("Title", "Tags", "Body HTML", "Type"))
    if "unisex" in text:
        return "Unisex"
    if "mujer" in text or "femenino" in text:
        return "Femenino"
    if "hombre" in text or "masculino" in text:
        return "Masculino"
    if "niÃ±o" in text or "nino" in text or "kids" in text:
        return "NiÃ±os"
    return ""


def centry_category(row):
    product_type = centry_value(row.get("Type"))
    gender = centry_gender(row)
    if centry_is_footwear(row):
        branch = "Calzados Masculinos" if gender == "Masculino" else "Calzados Femeninos" if gender == "Femenino" else "Calzados"
        return f"Calzados / {branch} / {product_type or 'Zapatillas'}"
    branch = "Ropa Masculina" if gender == "Masculino" else "Ropa Femenina" if gender == "Femenino" else "Ropa"
    return f"Vestuario / {branch} / {product_type or 'Prendas'}"


def centry_master_key(*values):
    return normalize_header(" ".join(clean_value(value) for value in values if clean_value(value)))


def centry_master_gender(value):
    text = centry_value(value).lower()
    if text == "unisex":
        return "Unisex"
    if text in ("masculino", "hombre", "men"):
        return "Hombre"
    if text in ("femenino", "mujer", "women"):
        return "Mujer"
    if "ni" in text or "kid" in text:
        return "NiÃ±os"
    return centry_value(value)


def first_existing_path(paths):
    for path in paths:
        if Path(path).exists():
            return Path(path)
    return None


@st.cache_data(show_spinner=False)
def load_centry_category_lookup():
    path = first_existing_path(DEFAULT_CENTRY_CATEGORY_PATHS)
    by_full_key = {}
    by_gender_type = {}
    for item in CENTRY_CATEGORY_FULL:
        product_type = clean_value(item.get("product_type"))
        gender = centry_master_gender(item.get("gender"))
        brand = clean_value(item.get("brand"))
        if product_type and gender and brand:
            key = centry_master_key(product_type, gender, brand)
            record = {
                "category": clean_value(item.get("category")),
                "class": clean_value(item.get("class")),
                "base": clean_value(item.get("base")) or f"{product_type}{gender}{brand}",
            }
            if key not in by_full_key or clean_value(record.get("category")):
                by_full_key[key] = record
    for item in CENTRY_CATEGORY_GENDER_TYPE:
        gender = centry_master_gender(item.get("gender"))
        product_type = clean_value(item.get("product_type"))
        category = clean_value(item.get("category"))
        if gender and product_type and category:
            by_gender_type[centry_master_key(product_type, gender)] = {
                "category": category,
                "class": "Vestuario" if category.lower().startswith("vestuario") else "",
                "base": clean_value(item.get("base")) or f"{gender}{product_type}",
            }
    if path is None:
        return {"path": "", "by_full_key": by_full_key, "by_gender_type": by_gender_type}
    try:
        categories = pd.read_excel(path, sheet_name="CategorÃ­as", dtype=object).dropna(how="all")
        for _, row in categories.iterrows():
            product_type = clean_value(row.get("Tipo de Producto"))
            gender = centry_master_gender(row.get("GÃ©nero"))
            brand = clean_value(row.get("Marca"))
            record = {
                "category": clean_value(row.get("CategorÃ­a")),
                "class": clean_value(row.get("Clase")),
                "base": clean_value(row.get("Unnamed: 5")) or f"{product_type}{gender}{brand}",
            }
            if product_type and gender and brand:
                key = centry_master_key(product_type, gender, brand)
                if key not in by_full_key or clean_value(record.get("category")):
                    by_full_key[key] = record
        all_sheet = pd.read_excel(path, sheet_name="all", dtype=object).dropna(how="all")
        for _, row in all_sheet.iterrows():
            gender = centry_master_gender(row.get("GÃ©nero"))
            product_type = clean_value(row.get("Unnamed: 2"))
            category = clean_value(row.get("CategorÃ­a"))
            if gender and product_type and category:
                by_gender_type[centry_master_key(product_type, gender)] = {
                    "category": category,
                    "class": "Vestuario" if category.lower().startswith("vestuario") else "",
                    "base": clean_value(row.get("Unnamed: 3")) or f"{gender}{product_type}",
                }
    except Exception:
        return {"path": str(path), "by_full_key": {}, "by_gender_type": {}}
    return {"path": str(path), "by_full_key": by_full_key, "by_gender_type": by_gender_type}


@st.cache_data(show_spinner=False)
def load_centry_codex_category_lookup():
    path = first_existing_path(DEFAULT_CENTRY_CODEX_CATEGORY_PATHS)
    by_type = {}
    by_class_type = {}
    for item in CENTRY_CODEX_CATEGORIES:
        product_type = clean_value(item.get("product_type"))
        class_name = clean_value(item.get("class"))
        category_name = clean_value(item.get("category"))
        base = clean_value(item.get("base")) or f"{class_name}{product_type}"
        if not product_type or not class_name:
            continue
        category = f"{class_name} / {category_name}" if category_name and category_name != class_name else class_name
        record = {"category": category, "class": class_name, "base": base}
        by_type.setdefault(centry_master_key(product_type), record)
        by_class_type.setdefault(centry_master_key(class_name, product_type), record)
    if path is None:
        return {"path": "", "by_type": by_type, "by_class_type": by_class_type}
    try:
        df = pd.read_excel(path, dtype=object).dropna(how="all")
        for _, row in df.iterrows():
            product_type = clean_value(row.get("Tipo de Producto"))
            class_name = clean_value(row.get("Clase"))
            category_name = clean_value(row.get("CategorÃ­a"))
            base = clean_value(row.get("Llave")) or f"{class_name}{product_type}"
            if not product_type or not class_name:
                continue
            category = f"{class_name} / {category_name}" if category_name and category_name != class_name else class_name
            record = {"category": category, "class": class_name, "base": base}
            by_type.setdefault(centry_master_key(product_type), record)
            by_class_type.setdefault(centry_master_key(class_name, product_type), record)
    except Exception:
        return {"path": str(path), "by_type": {}, "by_class_type": {}}
    return {"path": str(path), "by_type": by_type, "by_class_type": by_class_type}


@st.cache_data(show_spinner=False)
def load_centry_dimension_lookup():
    path = first_existing_path(DEFAULT_CENTRY_DIMENSIONS_PATHS)
    lookup = {}
    for item in CENTRY_DIMENSIONS:
        product_type = clean_value(item.get("product_type"))
        if product_type:
            lookup.setdefault(
                centry_master_key(product_type),
                {
                    "height": clean_value(item.get("height")),
                    "width": clean_value(item.get("width")),
                    "length": clean_value(item.get("length")),
                    "weight": clean_value(item.get("weight")),
                },
            )
    if path is None:
        return {"path": "", "lookup": lookup}
    try:
        xl = pd.ExcelFile(path)
        for sheet in xl.sheet_names:
            df = pd.read_excel(path, sheet_name=sheet, dtype=object).dropna(how="all")
            if "Tipo de Producto" not in df.columns:
                continue
            for _, row in df.iterrows():
                product_type = clean_value(row.get("Tipo de Producto"))
                if not product_type:
                    continue
                height = first_non_empty(row.get("Product Height"), row.get("Unnamed: 5"), row.get("Alto (cm)"))
                width = first_non_empty(row.get("Product Width"), row.get("Unnamed: 6"), row.get("Ancho (cm)"))
                length = first_non_empty(row.get("Product Length"), row.get("Unnamed: 7"), row.get("Largo (cm)"))
                weight = first_non_empty(row.get("Product Weight (gr)"), row.get("Unnamed: 8"), row.get("Peso (gr)"))
                lookup.setdefault(
                    centry_master_key(product_type),
                    {"height": height, "width": width, "length": length, "weight": weight},
                )
    except Exception:
        return {"path": str(path), "lookup": {}}
    return {"path": str(path), "lookup": lookup}


def centry_lookup_category(product_type, gender, vendor, fallback_category):
    lookup = load_centry_category_lookup()
    gender_master = centry_master_gender(gender)
    type_gender_key = centry_master_key(product_type, gender_master)
    record = lookup["by_full_key"].get(centry_master_key(product_type, gender_master, vendor))
    if record is None:
        record = lookup["by_gender_type"].get(type_gender_key)
    elif not clean_value(record.get("category")):
        type_record = lookup["by_gender_type"].get(type_gender_key)
        if type_record and clean_value(type_record.get("category")):
            record = {**record, "category": type_record.get("category")}
    if record is None:
        codex_lookup = load_centry_codex_category_lookup()
        fallback_class = clean_value(fallback_category).split("/", 1)[0].strip()
        record = codex_lookup["by_class_type"].get(centry_master_key(fallback_class, product_type))
        if record is None:
            record = codex_lookup["by_type"].get(centry_master_key(product_type))
    return record or {"category": fallback_category, "class": "", "base": f"{product_type}{gender_master}{vendor}"}


def centry_lookup_dimensions(product_type, fallback):
    lookup = load_centry_dimension_lookup()["lookup"]
    record = lookup.get(centry_master_key(product_type))
    return record or fallback


def centry_sial_dimension_package(row, product_type, fallback):
    dimensions = centry_lookup_dimensions(product_type, fallback)
    aliases = {
        "weight": ("Product Weight", "Package Weight", "Peso del paquete", "Peso (gr)", "Product Weight (gr)"),
        "length": ("Product Length", "Package Length", "Largo del paquete", "Largo (cm)"),
        "width": ("Product Width", "Package Width", "Ancho del paquete", "Ancho (cm)"),
        "height": ("Product Height", "Package Height", "Alto del paquete", "Alto (cm)"),
    }
    package = {}
    for key, columns in aliases.items():
        package[key] = first_non_empty(*(row.get(column) for column in columns), dimensions.get(key), fallback[key])
    return package


def centry_split_images(value):
    parts = [centry_value(part) for part in re.split(r"[;\n]+", centry_value(value)) if centry_value(part)]
    return parts[:5]


def centry_mod_col_from_row(row):
    mod_col = first_non_empty(
        row.get("Metafield: custom.codigo_modelo_color [id]"),
        row.get("Mod-Col"),
        row.get("COD MOD COL"),
    ).upper()
    if mod_col:
        return mod_col
    sku = centry_value(row.get("Variant SKU"))
    return sku.rsplit("-", 1)[0].upper() if "-" in sku else sku.upper()


def centry_output_is_accessory(row):
    text = normalize_header(
        " ".join(
            clean_value(row.get(column))
            for column in ("Clase", "CategorÃ­a", "Categoria ", "Type", "Sub Categoria")
            if column in row.index
        )
    )
    return "accesorio" in text or "accesorios" in text


def centry_output_blocks_zero_size(row):
    text = normalize_header(
        " ".join(
            clean_value(row.get(column))
            for column in ("Clase", "CategorÃ­a", "Categoria ", "Type", "Sub Categoria")
            if column in row.index
        )
    )
    return "calzado" in text or "vestuario" in text


def centry_display_size(value, force_one_size=False):
    if force_one_size or is_one_size(value):
        return "O/S"
    return first_non_empty(normalize_master_size(value), centry_value(value))


def build_centry_arti_lookup(arti_df):
    lookup = {"by_sku": {}, "by_mod_size": {}}
    if arti_df is None or arti_df.empty:
        return lookup
    df = normalize_arti_columns_for_app(arti_df).copy()
    for column in ("CODINT_MA", "COD MOD COL", "Mod-Col", "TALNUM_MA", "CodBarras", "ColorNombre"):
        if column not in df.columns:
            df[column] = ""
    for _, row in df.iterrows():
        sku = clean_value(row.get("CODINT_MA"))
        mod_col = first_non_empty(row.get("COD MOD COL"), row.get("Mod-Col"))
        raw_size = clean_value(row.get("TALNUM_MA"))
        display_size = normalize_master_size(raw_size)
        barcode = clean_value(row.get("CodBarras"))
        color_name = clean_value(row.get("ColorNombre"))
        if not barcode and not raw_size:
            continue
        item = {"barcode": barcode, "raw_size": raw_size, "display_size": display_size, "color_name": color_name}
        if sku:
            lookup["by_sku"].setdefault(sku.upper(), item)
        if mod_col:
            for size_key in {raw_size, display_size}:
                size_key = clean_value(size_key)
                if size_key:
                    lookup["by_mod_size"].setdefault((mod_col.upper(), size_key.upper()), item)
    return lookup


def centry_arti_item_for_row(row, arti_lookup, current_mod_col, raw_size):
    if not arti_lookup:
        return {}
    sku = centry_value(row.get("Variant SKU")).upper()
    if sku and sku in arti_lookup.get("by_sku", {}):
        return arti_lookup["by_sku"][sku]
    mod_col = clean_value(current_mod_col).upper()
    for size_key in (raw_size, normalize_master_size(raw_size), row.get("Option1 Value")):
        size_key = clean_value(size_key).upper()
        if mod_col and size_key and (mod_col, size_key) in arti_lookup.get("by_mod_size", {}):
            return arti_lookup["by_mod_size"][(mod_col, size_key)]
    return {}


def centry_size_guide(gender, class_name, vendor):
    gender_map = {
        "Masculino": "Hombre",
        "Femenino": "Mujer",
        "NiÃ±os": "Ninos",
        "NiÃ±as": "Ninas",
        "Unisex": "Unisex",
    }
    class_map = {"Calzado": "calzado", "Vestuario": "vestuario", "Accesorios": "accesorios"}
    gender_text = gender_map.get(centry_value(gender), centry_value(gender))
    class_text = class_map.get(centry_value(class_name), centry_value(class_name).lower())
    vendor_text = centry_value(vendor).replace(" ", "")
    return f"{gender_text}{class_text}{vendor_text}" if gender_text and class_text and vendor_text else ""


def centry_tag_value(row, *labels):
    text = " | ".join(
        centry_value(row.get(column))
        for column in ("Tags", "Listado de caracterÃ­sticas", "Product Bullets")
        if column in row.index
    )
    if not text:
        return ""
    for label in labels:
        pattern = rf"{re.escape(label)}\s*[:|]\s*(.*?)(?:\s*\|\s*|,\s{{2,}}|$)"
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            value = clean_value(match.group(1)).strip(" ,;")
            if value:
                return value
    return ""


def centry_material_from_row(row):
    return first_non_empty(
        centry_tag_value(row, "Material principal", "Material"),
        row.get("Metafield: custom.materialidad [single_line_text_field]"),
        row.get("Material"),
        row.get("ComposiciÃ³n"),
    )


def centry_composition_from_row(row):
    return first_non_empty(
        centry_tag_value(row, "ComposiciÃ³n", "Composicion"),
        row.get("Metafield: custom.materialidad [single_line_text_field]"),
        row.get("ComposiciÃ³n"),
    )


def centry_age_from_gender(gender):
    normalized = centry_master_gender(gender)
    if normalized in ("NiÃ±os", "NiÃ±o", "NiÃ±a", "Ninos", "Ninas"):
        return "NiÃ±os"
    if normalized in ("Hombre", "Mujer", "Unisex"):
        return "Adultos"
    return ""


def centry_footwear_cane_from_row(row):
    return first_non_empty(
        centry_tag_value(row, "Altura De Taco", "Altura de Taco", "Tipo de taco", "Tipo de caÃ±a"),
        row.get("Altura De Taco"),
        row.get("Tipo de taco"),
    )


def centry_tag_or_column_value(row, *labels):
    candidates = [centry_tag_value(row, *labels)]
    for label in labels:
        candidates.append(row.get(label))
    return first_non_empty(*candidates)


def centry_labeled_characteristics(row, vendor, product_type, color, gender, material, composition, footwear_cane):
    items = [
        ("Tipo De Producto", product_type),
        ("GÃ©nero", gender),
        ("Color", color),
        ("Marca", vendor),
        ("OcasiÃ³n", centry_tag_or_column_value(row, "OcasiÃ³n", "Ocasion")),
        ("ColecciÃ³n", centry_tag_or_column_value(row, "ColecciÃ³n", "Coleccion")),
        ("Material", material),
        ("ComposiciÃ³n", composition),
        ("Altura De Taco", footwear_cane),
    ]
    return " |".join(
        f"{label} : {centry_value(value).strip()}"
        for label, value in items
        if centry_value(value)
    )


def centry_looks_like_color_code(value, color_code=""):
    text = clean_value(value).strip().upper()
    code = clean_value(color_code).strip().upper()
    if not text:
        return False
    if code and text == code:
        return True
    return bool(re.fullmatch(r"[A-Z]{0,3}\d{1,4}[A-Z]{0,3}", text))


def centry_color_name_from_row(row, color_code=""):
    if row is None:
        return ""
    color = first_non_empty(
        row.get("Color Web"),
        row.get("Color"),
        row.get("COLOR"),
        row.get("Nombre Color"),
        row.get("Color Nombre"),
        row.get("Metafield: custom.color_forus [single_line_text_field]"),
        row.get("Metafield: theme.siblings_color [single_line_text_field]"),
        row.get("Metafield: custom.siblings_color [single_line_text_field]"),
        row.get("Metafield: custom.color [single_line_text_field]"),
        row.get("Option2 Value"),
        centry_tag_value(row, "Color", "Color Comercial", "Color Web"),
    )
    if centry_looks_like_color_code(color, color_code):
        return ""
    return color


def centry_gender_marketplace(gender):
    if gender == "Masculino":
        return "Hombre"
    if gender == "Femenino":
        return "Mujer"
    return centry_value(gender)


def centry_product_type_lower(product_type):
    return normalize_header(product_type)


def centry_is_accessory_type(product_type, category_record=None):
    text = centry_product_type_lower(product_type)
    category_text = centry_product_type_lower((category_record or {}).get("class")) + " " + centry_product_type_lower((category_record or {}).get("category"))
    return "accesorio" in category_text or any(
        word in text
        for word in (
            "bolso", "mochila", "billetera", "cartera", "canguro", "maleta", "gorro",
            "jockey", "sombrero", "beanie", "cuello", "botella", "calcetin", "calcetines",
            "cinturon", "banano", "lentes",
        )
    )


def centry_apparel_top_bottom_column(product_type):
    text = centry_product_type_lower(product_type)
    if any(word in text for word in ("pantalon", "short", "bermuda", "buzo", "jogger", "legging", "falda")):
        return "Tipo de pantalÃ³n largo/corto - Ropa y accesorios (Falabella GSC PerÃº)"
    if any(word in text for word in ("casaca", "chaqueta", "parka", "chaleco", "polar")):
        return "Tipo de chaqueta/chaleco - Ropa y accesorios (Falabella GSC PerÃº)"
    return "Tipo de camisa/blusa/polo/camiseta - Ropa y accesorios (Falabella GSC PerÃº)"


def centry_apply_accessory_fields(centry_row, product_type, gender, material):
    text = centry_product_type_lower(product_type)
    centry_row["GÃ©nero de vestuario - Ropa y accesorios (Falabella GSC PerÃº)"] = centry_gender_marketplace(gender)
    centry_row["Material del accesorio - Ropa y accesorios (Falabella GSC PerÃº)"] = material
    centry_row["Material principal - Material de vestuario - Ropa y accesorios (Falabella GSC PerÃº)"] = first_non_empty(material, "Otros")
    centry_row["ComposiciÃ³n - Ropa y accesorios (Falabella GSC PerÃº)"] = first_non_empty(material, "-")
    centry_row["ComposiciÃ³n (MercadoLibre PerÃº)"] = first_non_empty(material, "")
    if any(word in text for word in ("gorro", "jockey", "sombrero", "beanie")):
        centry_row["Tipo de ropa para la cabeza - Ropa y accesorios (Falabella GSC PerÃº)"] = product_type
        centry_row["Tipo de sombrero (MercadoLibre PerÃº)"] = product_type
    if "cuello" in text:
        centry_row["Tipo de ropa para el cuello - Ropa y accesorios (Falabella GSC PerÃº)"] = product_type
    if any(word in text for word in ("mochila", "bolso", "cartera", "billetera", "canguro", "banano")):
        centry_row["Tipo de bolso mochila funda - Ropa y accesorios (Falabella GSC PerÃº)"] = product_type
        centry_row["Tipo de cartera (MercadoLibre PerÃº)"] = product_type
        centry_row["Uso de la cartera/mochila/bolsa - Uso de la cartera mochila bolsa - Ropa y accesorios (Falabella GSC PerÃº)"] = "Urbano"
    if "maleta" in text:
        centry_row["Tipo de maleta - Ropa y accesorios (Falabella GSC PerÃº)"] = product_type
        centry_row["Material de la maleta - Material de la maleta - Ropa y accesorios (Falabella GSC PerÃº)"] = first_non_empty(material, "")
    centry_row["Contenido del paquete - Package content - Almacenamiento (Falabella GSC PerÃº)"] = "1"
    centry_row["Con monedero (MercadoLibre PerÃº)"] = "No"
    centry_row["Con cierre ajustable (MercadoLibre PerÃº)"] = "No"
    centry_row["Es reversible (MercadoLibre PerÃº)"] = "No"
    centry_row["Con cierre (MercadoLibre PerÃº)"] = "No"


def centry_apply_apparel_fields(centry_row, product_type, gender, material, composition):
    centry_row["GÃ©nero de vestuario - Ropa y accesorios (Falabella GSC PerÃº)"] = centry_gender_marketplace(gender)
    centry_row["Material de vestuario - Ropa y accesorios (Falabella GSC PerÃº)"] = material
    centry_row["Material principal - Material de vestuario - Ropa y accesorios (Falabella GSC PerÃº)"] = first_non_empty(material, "Otros")
    centry_row["ComposiciÃ³n - Ropa y accesorios (Falabella GSC PerÃº)"] = first_non_empty(composition, material, "-")
    target_column = centry_apparel_top_bottom_column(product_type)
    centry_row[target_column] = product_type
    centry_row["Tipo de prenda para la parte superior - Ropa y accesorios (Falabella GSC PerÃº)"] = product_type
    centry_row["DiseÃ±o - Ropa y accesorios (Falabella GSC PerÃº)"] = "-"
    centry_row["Tipo de cierre - Ropa y accesorios (Falabella GSC PerÃº)"] = ""


def filter_centry_size_rows(df, issues, size_column, key_column="Mod", output_label="Centry"):
    if df is None or df.empty or size_column not in df.columns:
        return df, issues
    result = df.copy()
    key_values = result[key_column].map(clean_value) if key_column in result.columns else pd.Series("", index=result.index)

    k_mask = result[size_column].map(is_internal_k_size)
    if k_mask.any():
        issues.append({"Mod-Col": output_label, "Problema": f"Se eliminaron {safe_int_value(k_mask.sum())} filas con talla interna K"})
        result = result[~k_mask].copy()
        key_values = key_values.loc[result.index]

    zero_block_mask = result.apply(lambda row: is_zero_size(row.get(size_column)) and centry_output_blocks_zero_size(row), axis=1)
    if zero_block_mask.any():
        issues.append({"Mod-Col": output_label, "Problema": f"Se eliminaron {safe_int_value(zero_block_mask.sum())} filas con talla 0/000 en vestuario/calzado"})
        result = result[~zero_block_mask].copy()
        key_values = key_values.loc[result.index]

    if result.empty:
        return result, issues

    drop_accessory_zero = pd.Series(False, index=result.index)
    for key, group in result.groupby(key_values, sort=False):
        if not clean_value(key):
            continue
        accessory_group = group.apply(centry_output_is_accessory, axis=1)
        if not accessory_group.any():
            continue
        accessory_rows = group.loc[accessory_group]
        has_real_size = (
            accessory_rows[size_column].map(clean_value).ne("")
            & ~accessory_rows[size_column].map(is_zero_size)
        ).any()
        if has_real_size:
            group_zero = accessory_rows[size_column].map(is_zero_size)
            drop_accessory_zero.loc[accessory_rows.index[group_zero]] = True
    if drop_accessory_zero.any():
        issues.append({"Mod-Col": output_label, "Problema": f"Se eliminaron {safe_int_value(drop_accessory_zero.sum())} filas accesorio talla 0/000 porque existe una talla real"})
        result = result[~drop_accessory_zero].copy()

    one_size_mask = result[size_column].map(is_one_size)
    if one_size_mask.any():
        result.loc[one_size_mask, size_column] = "O/S"
    return result, issues


def build_centry_from_matrixify(matrixify_df, brand_config=None, only_codes=None, arti_df=None):
    if matrixify_df is None or matrixify_df.empty:
        return pd.DataFrame(columns=CENTRY_COLUMNS), pd.DataFrame(columns=["Mod-Col", "Problema"])
    brand_config = brand_config or {}
    df = coalesce_duplicate_columns(matrixify_df).copy()
    for column in MATRIXIFY_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    product_fields = [
        "Title", "Body HTML", "Vendor", "Type", "Tags", "Status", "Published", "Image Src",
        "Metafield: custom.codigo_modelo_color [id]", "Metafield: custom.estilo [single_line_text_field]",
        "Metafield: custom.color [single_line_text_field]",
        "Metafield: custom.materialidad [single_line_text_field]",
        "Metafield: custom.tecnologia [list.single_line_text_field]",
    ]
    for column in product_fields:
        if column in df.columns:
            df[column] = df[column].replace("", pd.NA).ffill().fillna("")
    only_codes_set = {clean_value(code).upper() for code in (only_codes or []) if clean_value(code)}
    rows = []
    issues = []
    normalized_arti_df = normalize_arti_columns_for_app(arti_df) if arti_df is not None else arti_df
    if normalized_arti_df is not None and not normalized_arti_df.empty and normalized_arti_df["CodBarras"].map(clean_value).eq("").all():
        issues.append({"Mod-Col": "BigQuery/ARTI", "Problema": "La fuente maestra no trajo CodBarras/EAN/barcode reconocible"})
    arti_lookup = build_centry_arti_lookup(normalized_arti_df)
    current_mod_col = ""
    for _, row in df.iterrows():
        variant_sku = centry_value(row.get("Variant SKU"))
        if not variant_sku:
            continue
        mod_col = centry_mod_col_from_row(row)
        current_mod_col = mod_col or current_mod_col
        model, color_code = split_model_color(current_mod_col)
        if only_codes_set and current_mod_col not in only_codes_set and model not in only_codes_set:
            continue
        images = centry_split_images(row.get("Image Src"))
        title = centry_value(row.get("Title"))
        vendor = centry_value(row.get("Vendor"), brand_config.get("label", ""))
        product_type = centry_value(row.get("Type"))
        raw_size = first_non_empty(row.get("__CENTRY_RAW_SIZE"), row.get("Option1 Value"))
        gender = centry_gender(row)
        is_footwear = centry_is_footwear(row)
        fallback_package = centry_package_values(row)
        dimensions = centry_lookup_dimensions(product_type, fallback_package)
        category_record = centry_lookup_category(product_type, gender, vendor, centry_category(row))
        category_probe = pd.Series({"Clase": category_record.get("class"), "CategorÃ­a": category_record.get("category"), "Type": product_type})
        arti_item = centry_arti_item_for_row(row, arti_lookup, current_mod_col, raw_size)
        color = first_non_empty(
            centry_color_name_from_row(row, color_code),
            "" if centry_looks_like_color_code(arti_item.get("color_name"), color_code) else arti_item.get("color_name"),
        )
        size = centry_display_size(
            raw_size,
            centry_output_is_accessory(category_probe) and (is_one_size(raw_size) or is_zero_size(raw_size)),
        )
        barcode = first_non_empty(arti_item.get("barcode"), row.get("Variant Barcode"))
        tal_value = first_non_empty(arti_item.get("raw_size"), raw_size)
        variant_centry_sku = barcode or variant_sku
        class_name = category_record.get("class") or ("Calzado" if is_footwear else "Vestuario")
        material = centry_material_from_row(row)
        composition = centry_composition_from_row(row)
        age = centry_age_from_gender(gender)
        footwear_cane = centry_footwear_cane_from_row(row)
        characteristics = centry_labeled_characteristics(
            row,
            vendor,
            product_type,
            color,
            gender,
            material,
            composition,
            footwear_cane,
        )
        centry_row = {column: "" for column in CENTRY_COLUMNS}
        centry_row.update(
            {
                "Nombre del Producto": title,
                "Marca": vendor,
                "Descripcion": strip_html(row.get("Body HTML")),
                "Listado de caracterÃ­sticas": characteristics,
                "GarantÃ­a": "3 meses, GarantÃ­a del vendedor",
                "Alto del paquete": first_non_empty(dimensions.get("height"), fallback_package["height"]),
                "Ancho del paquete": first_non_empty(dimensions.get("width"), fallback_package["width"]),
                "Largo del paquete": first_non_empty(dimensions.get("length"), fallback_package["length"]),
                "Peso del paquete": first_non_empty(dimensions.get("weight"), fallback_package["weight"]),
                "CategorÃ­a": category_record.get("category") or centry_category(row),
                "Precio": centry_value(row.get("Variant Price")),
                "SKU del producto": current_mod_col,
                "SKU de la variante": variant_centry_sku,
                "CÃ³digo de barra variante (EAN/UPC/ISBN)": barcode,
                "Color": color,
                "Talla": size,
                "CondiciÃ³n del Producto": "Nuevo",
                "AÃ±o de temporada": datetime.now().year,
                "Temporada": centry_value(row.get("Temporada"), "Verano"),
                "GÃ©nero": gender,
                "Estado": "Activo",
                "cccc": current_mod_col,
                "Incluir en MercadoLibrePe / MercadoLibrePe": "SI",
                "Incluir en Falabella Global Peru / FalabellaGlobalProduction": "SI",
                "Es inflamable (MercadoLibre PerÃº)": "No",
                "ComposiciÃ³n (MercadoLibre PerÃº)": composition,
                "Material principal - Material de vestuario - Ropa y accesorios (Falabella GSC PerÃº)": first_non_empty(material, "Otros"),
                "DiseÃ±o - Ropa y accesorios (Falabella GSC PerÃº)": "-",
                "Contenido del paquete - Package content - Almacenamiento (Falabella GSC PerÃº)": "1",
                "DiseÃ±o de la tela (MercadoLibre PerÃº)": "-",
                "Con monedero (MercadoLibre PerÃº)": "No",
                "Con cierre ajustable (MercadoLibre PerÃº)": "No",
                "Es reversible (MercadoLibre PerÃº)": "No",
                "Con cierre (MercadoLibre PerÃº)": "No",
                "INFORMACIÃ“N ADICIONAL": "",
                "Unnamed: 52": "",
                "Unnamed: 53": product_type,
                "Unnamed: 54": f"{product_type}{gender}{vendor}",
                "Mod": model,
                "Col": color_code,
                "Tal": tal_value,
                "Cod Mod Col Talla": f"{current_mod_col}-{size}" if current_mod_col and size else current_mod_col,
                "Clase": class_name,
                "GuÃ­a de tallas": centry_size_guide(gender, class_name, vendor),
                "Base de categorÃ­a": category_record.get("base") or f"{product_type}{gender}{vendor}",
            }
        )
        for index, image in enumerate(images, start=1):
            centry_row["URL imagen principal" if index == 1 else f"URL imagen {index}"] = image
        if is_footwear:
            centry_row["GÃ©nero - Calzado (Falabella GSC PerÃº)"] = centry_gender_marketplace(gender)
            centry_row["Tipo - Calzado (Falabella GSC PerÃº)"] = product_type or "Zapatillas"
            centry_row["Horma - Calzado (Falabella GSC PerÃº)"] = "Normal"
            centry_row["Material principal - Calzado (Falabella GSC PerÃº)"] = material
            centry_row["Tipo de calzado (MercadoLibre PerÃº)"] = product_type or "Zapatillas"
            centry_row["Edad (MercadoLibre PerÃº)"] = age
            centry_row["Material del calzado (MercadoLibre PerÃº)"] = first_non_empty(composition, material)
            centry_row["Tipo de taco (MercadoLibre PerÃº)"] = footwear_cane
        elif centry_is_accessory_type(product_type, category_record):
            centry_apply_accessory_fields(centry_row, product_type, gender, material)
        else:
            centry_apply_apparel_fields(centry_row, product_type, gender, material, composition)
        if not title:
            issues.append({"Mod-Col": current_mod_col, "Problema": "Sin nombre de producto"})
        if not images:
            issues.append({"Mod-Col": current_mod_col, "Problema": "Sin imagen principal"})
        rows.append(centry_row)
    centry_df = pd.DataFrame(rows, columns=CENTRY_COLUMNS).fillna("")
    centry_df = centry_df.replace({"#N/D": "", "#ND": "", "#N/A": ""})
    centry_df, issues = filter_centry_size_rows(centry_df, issues, "Talla", key_column="SKU del producto", output_label="Centry")
    if not centry_df.empty:
        barcode_column = "CÃ³digo de barra variante (EAN/UPC/ISBN)"
        missing_barcode = centry_df[centry_df[barcode_column].map(clean_value) == ""].copy()
        for _, missing_row in missing_barcode.iterrows():
            issues.append(
                {
                    "Mod-Col": clean_value(missing_row.get("SKU del producto")) or "Centry",
                    "Problema": f"Sin codigo de barras en SKU {clean_value(missing_row.get('SKU de la variante'))}",
                }
            )
    return repair_mojibake_dataframe(centry_df), repair_mojibake_dataframe(
        pd.DataFrame(issues, columns=["Mod-Col", "Problema"]).drop_duplicates()
    )


def centry_package_values(row):
    is_footwear = centry_is_footwear(row)
    return {
        "weight": 900 if is_footwear else 200,
        "length": 35 if is_footwear else 29,
        "width": 21 if is_footwear else 27,
        "height": 12 if is_footwear else 1,
    }


def centry_bullets(row, vendor, product_type, color, gender):
    tags = [part.strip() for part in centry_value(row.get("Tags")).split(",") if part.strip()]
    items = [
        ("Tipo Producto", product_type),
        ("GÃ©nero", gender),
        ("Color Comercial", color),
        ("Marca", vendor),
    ]
    for tag in tags[:8]:
        if "|" in tag:
            continue
        items.append(("CaracterÃ­stica", tag))
    return ", ".join(f"{label} | {value}" for label, value in items if centry_value(value))


def build_centry_sial_from_matrixify(matrixify_df, brand_config=None):
    if matrixify_df is None or matrixify_df.empty:
        return pd.DataFrame(columns=CENTRY_SIAL_COLUMNS)
    brand_config = brand_config or {}
    df = coalesce_duplicate_columns(matrixify_df).copy()
    for column in MATRIXIFY_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    product_fields = [
        "Title", "Body HTML", "Vendor", "Type", "Tags", "Image Src",
        "Metafield: custom.codigo_modelo_color [id]", "Metafield: custom.color [single_line_text_field]",
        "Metafield: custom.materialidad [single_line_text_field]",
        "Metafield: custom.tecnologia [list.single_line_text_field]",
    ]
    for column in product_fields:
        if column not in df.columns:
            df[column] = ""
        df[column] = df[column].replace("", pd.NA).ffill().fillna("")

    rows = []
    for _, row in df.iterrows():
        sku = centry_value(row.get("Variant SKU"))
        if not sku:
            continue
        mod_col = centry_mod_col_from_row(row)
        model, color_code = split_model_color(mod_col)
        vendor = centry_value(row.get("Vendor"), brand_config.get("label", ""))
        product_type = centry_value(row.get("Type"))
        color = centry_color_name_from_row(row, color_code)
        raw_size = first_non_empty(row.get("__CENTRY_RAW_SIZE"), row.get("Option1 Value"))
        gender = centry_gender(row)
        images = centry_split_images(row.get("Image Src"))
        image = images[0] if images else ""
        fallback_package = centry_package_values(row)
        package = centry_sial_dimension_package(row, product_type, fallback_package)
        category_record = centry_lookup_category(product_type, gender, vendor, centry_category(row))
        category = category_record.get("class") or ("CALZADO" if centry_is_footwear(row) else ("ACCESORIOS" if "accesorio" in centry_category(row).lower() else "VESTUARIO"))
        category_probe = pd.Series({"Clase": category, "Categoria ": category, "Type": product_type})
        size = centry_display_size(
            raw_size,
            centry_output_is_accessory(category_probe) and (is_one_size(raw_size) or is_zero_size(raw_size)),
        )
        tal_value = first_non_empty(row.get("__CENTRY_RAW_SIZE"), raw_size)
        rows.append(
            {
                "Mod": model,
                "Col": color_code,
                "Tal": tal_value,
                "Product Name ": centry_value(row.get("Title")) or mod_col,
                "Product Bullets": centry_bullets(row, vendor, product_type, color, gender),
                "Product Description": strip_html(row.get("Body HTML")),
                "Image URL": image,
                "Product Weight": first_non_empty(package.get("weight"), fallback_package["weight"]),
                "Product Length": first_non_empty(package.get("length"), fallback_package["length"]),
                "Product Width": first_non_empty(package.get("width"), fallback_package["width"]),
                "Product Height": first_non_empty(package.get("height"), fallback_package["height"]),
                "Package Weight": first_non_empty(package.get("weight"), fallback_package["weight"]),
                "Package Length": first_non_empty(package.get("length"), fallback_package["length"]),
                "Package Width": first_non_empty(package.get("width"), fallback_package["width"]),
                "Package Height": first_non_empty(package.get("height"), fallback_package["height"]),
                "Boost ": "",
                "Talla Web ": size,
                "Color Web": color,
                "Categoria ": category,
                "Sub Categoria": product_type,
                "Genero": gender,
                "Estilo ": model,
                "Colecciones ": "",
                "Temporada ": centry_value(row.get("Temporada"), "Verano"),
                "Modelo": model,
                "Marca": vendor,
                "Tecnologias ": limit_words(first_non_empty(row.get("Metafield: custom.tecnologia [list.single_line_text_field]"), ""), 45),
                "Caracteristicas": limit_words(centry_value(row.get("Tags")).replace(",", " |"), 45),
                "Tipo de Boardshort": "",
                "Tipo de Bikini": "",
                "Iniciativas": "",
                "Tipo de Material": first_non_empty(row.get("Metafield: custom.materialidad [single_line_text_field]"), ""),
                "1": "",
                "Tipo de Prenda": product_type,
                "Adicional 2 ": "",
                "Adicional 3 ": "",
                "Adicional 4 ": "",
                "Adicional 5 ": "",
                "Adicional 6 ": "",
                "Adicional 7 ": "",
                "Adicional 8 ": "",
                "Adicional 9 ": "",
                "Adicional 10": "",
                "Mod-Col": mod_col,
                "Sku - Sial": sku,
                "Nuevo o Actualizar (Rockford.pe)": "Crear",
                "Sku - Supermall.pe": sku,
                "Porduct Id - Supermall.pe": "",
            }
        )
    sial_df = pd.DataFrame(rows, columns=CENTRY_SIAL_COLUMNS).fillna("")
    sial_df, _ = filter_centry_size_rows(sial_df, [], "Tal", key_column="Mod-Col", output_label="Carga Sial Centry")
    if not sial_df.empty and "Talla Web " in sial_df.columns:
        sial_df["Talla Web "] = sial_df["Tal"].map(centry_display_size)
    return repair_mojibake_dataframe(sial_df)


def render_centry_preview(centry_df, issues_df=None, title="Vista previa Centry"):
    if centry_df is None or centry_df.empty:
        return
    df = centry_df.copy()
    total_rows = len(df)
    total_products = df.get("SKU del producto", pd.Series(dtype=object)).map(clean_value).nunique()
    no_barcode = safe_int_value((df.get("CÃ³digo de barra variante (EAN/UPC/ISBN)", pd.Series(dtype=object)).map(clean_value) == "").sum())
    no_image = safe_int_value((df.get("URL imagen principal", pd.Series(dtype=object)).map(clean_value) == "").sum())
    no_price = safe_int_value((df.get("Precio", pd.Series(dtype=object)).map(clean_value) == "").sum())
    issue_count = 0 if issues_df is None or issues_df.empty else len(issues_df)
    render_html(
        f"""
        <div class="combo-card">
            <div class="combo-card-head">
                <div>
                    <div class="combo-title"><span class="combo-title-icon">C</span> {title}</div>
                    <p>Revision rapida de completitud antes de enviar el archivo al canal.</p>
                </div>
                <div class="combo-chip">{format_kpi_number(total_rows)} filas</div>
            </div>
            <div class="commercial-summary-grid">
                <div class="commercial-summary-tile ok"><span>Productos</span><b>&#10003;</b><strong>{format_kpi_number(total_products)}</strong></div>
                <div class="commercial-summary-tile {'ok' if no_barcode == 0 else 'bad'}"><span>EAN faltante</span><b>{'&#10003;' if no_barcode == 0 else '&#10005;'}</b><strong>{format_kpi_number(no_barcode)}</strong></div>
                <div class="commercial-summary-tile {'ok' if no_image == 0 else 'bad'}"><span>Imagen faltante</span><b>{'&#10003;' if no_image == 0 else '&#10005;'}</b><strong>{format_kpi_number(no_image)}</strong></div>
                <div class="commercial-summary-tile {'ok' if no_price == 0 else 'bad'}"><span>Precio faltante</span><b>{'&#10003;' if no_price == 0 else '&#10005;'}</b><strong>{format_kpi_number(no_price)}</strong></div>
                <div class="commercial-summary-tile {'ok' if issue_count == 0 else 'bad'}"><span>Observaciones</span><b>{'&#10003;' if issue_count == 0 else '&#10005;'}</b><strong>{format_kpi_number(issue_count)}</strong></div>
            </div>
        </div>
        """
    )
    st.dataframe(df.head(120), use_container_width=True, height=360)
    if issues_df is not None and not issues_df.empty:
        st.warning(f"Centry tiene {len(issues_df):,} observaciones.")
        st.dataframe(issues_df, use_container_width=True)


def model_codes_from_text(value):
    return [part.strip().upper() for part in re.split(r"[\n,; ]+", clean_value(value)) if part.strip()]


def model_codes_from_excel(df):
    if df is None or df.empty:
        return []
    code_column = first_existing_column(
        df,
        [
            "Mod-Col",
            "Mod Col",
            "Codigo Modelo Color",
            "Codigo modelo-color",
            "Codigo modelo color",
            "Codigo modelo",
            "Cod Mod Col",
            "Modelo Color",
            "Modelo-Color",
            "modelo_color",
            "sku",
        ],
    )
    if code_column is None:
        non_empty_columns = [col for col in df.columns if df[col].map(clean_value).ne("").any()]
        if len(non_empty_columns) == 1:
            code_column = non_empty_columns[0]
    if code_column is None:
        return []
    values = df[code_column].dropna().map(clean_value).tolist()
    codes = []
    seen = set()
    for value in values:
        for code in model_codes_from_text(value):
            if code not in seen:
                codes.append(code)
                seen.add(code)
    return codes


def centry_ean_column(df):
    if df is None or df.empty:
        return ""
    candidates = [
        "CÃ³digo de barra variante (EAN/UPC/ISBN)",
        "CÃƒÂ³digo de barra variante (EAN/UPC/ISBN)",
    ]
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
    for column in df.columns:
        normalized = normalize_header(column)
        if "codigodebarravariante" in normalized or "eanupcisbn" in normalized:
            return column
    return ""


def centry_missing_ean_count(centry_df):
    column = centry_ean_column(centry_df)
    if not column:
        return len(centry_df) if centry_df is not None else 0
    return safe_int_value(centry_df[column].map(clean_value).eq("").sum())


def build_centry_matrixify_from_master(codes, shopify_matrixify_df, arti_df, brand_config):
    codes = [clean_value(code).upper() for code in codes if clean_value(code)]
    if not codes:
        return pd.DataFrame(columns=MATRIXIFY_COLUMNS), pd.DataFrame(columns=["Mod-Col", "Problema"])

    issues = []
    code_set = set(codes)
    model_only_set = {code for code in codes if "-" not in code}
    shopify_df = coalesce_duplicate_columns(shopify_matrixify_df).copy() if shopify_matrixify_df is not None else pd.DataFrame()
    if not shopify_df.empty:
        for column in MATRIXIFY_COLUMNS:
            if column not in shopify_df.columns:
                shopify_df[column] = ""
        for column in ("Title", "Body HTML", "Vendor", "Type", "Tags", "Image Src", "Metafield: custom.codigo_modelo_color [id]"):
            if column in shopify_df.columns:
                shopify_df[column] = shopify_df[column].replace("", pd.NA).ffill().fillna("")
        shopify_df["__CENTRY_KEY"] = shopify_df.apply(centry_mod_col_from_row, axis=1)
    else:
        shopify_df["__CENTRY_KEY"] = pd.Series(dtype=object)

    product_lookup = {}
    if not shopify_df.empty and "__CENTRY_KEY" in shopify_df.columns:
        for key, group in shopify_df.groupby("__CENTRY_KEY", sort=False):
            key = clean_value(key).upper()
            if key and key not in product_lookup:
                product_lookup[key] = group.iloc[0]

    arti = normalize_arti_columns_for_app(arti_df).copy() if arti_df is not None else pd.DataFrame()
    for column in ("CODINT_MA", "COD MOD COL", "Mod-Col", "TALNUM_MA", "MARCA_MA", "Precio", "CodBarras"):
        if column not in arti.columns:
            arti[column] = ""
    issues.append({"Mod-Col": "Diagnostico EAN", "Problema": arti_barcode_diagnostics(arti)})
    if arti["CodBarras"].map(clean_value).eq("").all():
        issues.append({"Mod-Col": "Diagnostico BigQuery", "Problema": bigquery_barcode_schema_diagnostics(get_bigquery_config())})
    arti["__KEY"] = arti["Mod-Col"].where(arti["Mod-Col"].map(clean_value) != "", arti["COD MOD COL"]).map(lambda value: clean_value(value).upper())
    arti["__MODEL"] = arti["__KEY"].map(lambda value: value.rsplit("-", 1)[0] if "-" in value else value)
    allowed = {clean_value(value).upper() for value in brand_config.get("allowed_arti_brands", [])}
    if allowed and "MARCA_MA" in arti.columns:
        arti = arti[arti["MARCA_MA"].map(lambda value: clean_value(value).upper()).isin(allowed)].copy()
    arti = arti[(arti["__KEY"].isin(code_set)) | (arti["__MODEL"].isin(model_only_set))].copy()
    if arti.empty:
        return pd.DataFrame(columns=MATRIXIFY_COLUMNS), pd.DataFrame(
            [{"Mod-Col": code, "Problema": "Codigo no encontrado en BigQuery/ARTI"} for code in codes],
            columns=["Mod-Col", "Problema"],
        )

    arti["__SIZE"] = arti["TALNUM_MA"].map(normalize_master_size)
    invalid_size = arti[arti["__SIZE"].map(clean_value) == ""].copy()
    for key, group in invalid_size.groupby("__KEY"):
        issues.append({"Mod-Col": key, "Problema": f"{len(group):,} filas BigQuery/ARTI omitidas por talla vacia o no reconocida"})
    arti = arti[arti["__SIZE"].map(clean_value) != ""].copy()
    if arti.empty:
        return pd.DataFrame(columns=MATRIXIFY_COLUMNS), pd.DataFrame(issues, columns=["Mod-Col", "Problema"]).drop_duplicates()

    rows = []
    for key, variants in arti.groupby("__KEY", sort=False):
        variants = variants.copy()
        variants = variants[variants["CODINT_MA"].map(clean_value) != ""].copy()
        if variants.empty:
            issues.append({"Mod-Col": key, "Problema": "Sin SKU valido en BigQuery/ARTI"})
            continue
        kept_indexes = []
        seen_skus = set()
        duplicate_skus = []
        for variant_index, variant_row in variants.iterrows():
            sku_key = clean_value(variant_row.get("CODINT_MA")).upper()
            size_key = clean_value(display_size_for_site(variant_row.get("__SIZE"), brand_config)).upper()
            if sku_key and sku_key in seen_skus:
                duplicate_skus.append(f"{sku_key} ({size_key or 'sin talla'})")
                continue
            if sku_key:
                seen_skus.add(sku_key)
            kept_indexes.append(variant_index)
        variants = variants.loc[kept_indexes].copy()
        if duplicate_skus:
            issues.append({"Mod-Col": key, "Problema": f"Se omitieron {len(duplicate_skus):,} variantes duplicadas por SKU: {', '.join(duplicate_skus[:10])}"})
        if variants.empty:
            issues.append({"Mod-Col": key, "Problema": "Todas las variantes fueron omitidas por duplicidad de SKU"})
            continue
        variants = variants.sort_values("__SIZE", key=lambda series: series.map(master_size_sort_key))
        product_row = product_lookup.get(key)
        raw_brand = first_non_empty(variants.iloc[0].get("MARCA_MA"), product_row.get("Vendor") if product_row is not None else "", brand_config.get("label", ""))
        vendor = brand_display_name(raw_brand, brand_config.get("label", ""))
        image_config = brand_image_config(raw_brand, brand_config)
        image_urls = image_candidates(key, image_config)
        title = first_non_empty(product_row.get("Title") if product_row is not None else "", key)
        body_html = first_non_empty(product_row.get("Body HTML") if product_row is not None else "", "")
        product_type = first_non_empty(product_row.get("Type") if product_row is not None else "", "")
        tags = first_non_empty(product_row.get("Tags") if product_row is not None else "", vendor)
        materiality = first_non_empty(product_row.get("Metafield: custom.materialidad [single_line_text_field]") if product_row is not None else "", "")
        technology = first_non_empty(product_row.get("Metafield: custom.tecnologia [list.single_line_text_field]") if product_row is not None else "", "")
        color = centry_color_name_from_row(product_row, split_model_color(key)[1]) if product_row is not None else ""
        if product_row is None:
            issues.append({"Mod-Col": key, "Problema": "No existe en Shopify; se completo Centry con BigQuery/ARTI y fotos S3"})

        for position, (_, variant) in enumerate(variants.iterrows(), start=1):
            size = display_size_for_site(variant.get("__SIZE"), brand_config)
            rows.append(
                {
                    "Handle": clean_value(product_row.get("Handle")) if product_row is not None else key.lower(),
                    "Title": title,
                    "Body HTML": body_html,
                    "Vendor": vendor,
                    "Type": product_type,
                    "Tags": tags,
                    "Image Src": "; ".join(image_urls),
                    "Variant SKU": clean_value(variant.get("CODINT_MA")),
                    "__CENTRY_RAW_SIZE": clean_value(variant.get("TALNUM_MA")),
                    "Option1 Name": "Talla",
                    "Option1 Value": size,
                    "Option2 Name": "Color",
                    "Option2 Value": color,
                    "Variant Barcode": clean_value(variant.get("CodBarras")),
                    "Variant Price": clean_value(variant.get("Precio")),
                    "Metafield: custom.codigo_modelo_color [id]": key,
                    "Metafield: custom.marca [single_line_text_field]": vendor,
                    "Metafield: custom.color [single_line_text_field]": color,
                    "Metafield: custom.materialidad [single_line_text_field]": materiality,
                    "Metafield: custom.tecnologia [list.single_line_text_field]": technology,
                    "Status": clean_value(product_row.get("Status")) if product_row is not None else "",
                    "Published": clean_value(product_row.get("Published")) if product_row is not None else "",
                    "Variant Position": position,
                }
            )

    matrixify_like = pd.DataFrame(rows)
    for column in MATRIXIFY_COLUMNS:
        if column not in matrixify_like.columns:
            matrixify_like[column] = ""
    return matrixify_like, pd.DataFrame(issues, columns=["Mod-Col", "Problema"]).drop_duplicates()


def _split_tags(value):
    return [tag.strip() for tag in clean_value(value).split(",") if tag.strip()]


def _join_tags(values):
    return ", ".join(dict.fromkeys(tag for tag in values if clean_value(tag)))


def _split_semicolon_values(value):
    return [item.strip() for item in re.split(r"[;\n\r]+", clean_value(value)) if item.strip()]


def _product_lookup_from_shopify(records):
    by_key = {}
    by_handle = {}

    def add_key(value, record):
        for candidate in product_lookup_candidates(value):
            if candidate and candidate not in by_key:
                by_key[candidate] = record

    for record in records:
        key = clean_value(record.get("Mod-Col")).upper()
        handle = clean_value(record.get("Handle"))
        add_key(key, record)
        for variant in record.get("Variants") or []:
            for candidate in variant_mod_col_candidates(variant):
                add_key(candidate, record)
        if handle and handle not in by_handle:
            by_handle[handle] = record
    return by_key, by_handle


UPDATE_KEY_COLUMNS = [
    "Mod-Col",
    "Mod Col",
    "MOD COL",
    "MOD-COL",
    "COD MOD COL",
    "Cod Mod Col",
    "Cod-Mod-Col",
    "COD_MOD_COL",
    "cod_mod_col",
    "Codigo Modelo Color",
    "Código Modelo Color",
    "CÃ³digo Modelo Color",
    "codigo_modelo_color",
    "Codigo Modelo-Color",
    "Código Modelo-Color",
    "Modelo Color",
    "Modelo-Color",
    "Modelo + Color",
    "Cod Modelo Color",
    "Código Modelo Color Shopify",
    "Metafield: custom.codigo_modelo_color [id]",
    "custom.codigo_modelo_color",
]


TAG_UPDATE_COLUMNS = [
    "Tags",
    "Tag",
    "tags",
    "tag",
    "Etiquetas",
    "Etiqueta",
    "Nueva etiqueta",
    "Nuevo tag",
    "Nuevo Tag",
    "Nuevo Tags",
    "Tag nuevo",
    "Tags nuevos",
    "New Tag",
    "New Tags",
]


HANDLE_UPDATE_COLUMNS = [
    "Handle",
    "handle",
    "Product Handle",
    "Handle Shopify",
    "Shopify Handle",
    "URL",
    "Url",
    "Slug",
]


SIZE_GUIDE_UPDATE_COLUMNS = [
    "Guia de tallas",
    "Guía de tallas",
    "Guia Tallas",
    "Guía Tallas",
    "Size Guide",
    "Size guide",
    "Tabla de tallas",
    "Guia",
    "Guía",
    "Metafield: custom.guia_de_tallas [page_reference]",
    "custom.guia_de_tallas",
]


def _source_key_for_update(row):
    for column in UPDATE_KEY_COLUMNS:
        value = clean_value(row.get(column))
        if value:
            return value.upper()
    return ""


def normalize_partial_update_input(df, operation=""):
    if df is None:
        return pd.DataFrame()
    source = df.dropna(how="all").copy()
    if source.empty:
        return source

    key_col = first_existing_column(source, UPDATE_KEY_COLUMNS)
    if key_col and key_col != "Mod-Col":
        source["Mod-Col"] = source[key_col]

    if operation == "tags":
        tags_col = first_existing_column(source, TAG_UPDATE_COLUMNS)
        if tags_col and tags_col != "Tags":
            source["Tags"] = source[tags_col]

    handle_col = first_existing_column(source, HANDLE_UPDATE_COLUMNS)
    if handle_col and handle_col != "Handle":
        source["Handle"] = source[handle_col]

    if not key_col and len(source.columns) == 1:
        first_column = source.columns[0]
        if looks_like_mod_col(first_column):
            values = [first_column]
            values.extend(source[first_column].dropna().tolist())
            source = pd.DataFrame({"Mod-Col": [clean_value(value).upper() for value in values if clean_value(value)]})
        elif operation == "tags":
            source["Tags"] = source[first_column]
    return source


def normalize_photo_update_input(df):
    if df is None:
        return None
    source = df.dropna(how="all").copy()
    if source.empty:
        return source

    mod_col = first_existing_column(
        source,
        [
            "Mod-Col",
            "Mod Col",
            "COD MOD COL",
            "Codigo Modelo Color",
            "CÃ³digo Modelo Color",
            "codigo_modelo_color",
            "Modelo Color",
            "Modelo-Color",
        ],
    )
    if mod_col:
        source["Mod-Col"] = source[mod_col]
        return source

    first_column = source.columns[0]
    values = []
    if looks_like_mod_col(first_column):
        values.append(first_column)
    values.extend(source[first_column].dropna().tolist())
    values = [clean_value(value).upper() for value in values if clean_value(value)]
    return pd.DataFrame({"Mod-Col": values})


TECHNOLOGY_MAINTAINER = [
    {
        "brand": "Columbia",
        "name": "Omni-Tech",
        "keywords": ["omni-tech", "omni tech", "omnitech"],
        "logo": "logo.omni-tech-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni-Heat",
        "keywords": ["omni-heat", "omni heat", "omniheat"],
        "logo": "logo.omni-heat-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni-Heat Infinity",
        "keywords": ["omni-heat infinity", "omni heat infinity", "omniheat infinity"],
        "logo": "logo.omni-heat-infinity-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni-Shade",
        "keywords": ["omni-shade", "omni shade", "omnishade"],
        "logo": "logo.omni-shade-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni Shade Broad Spectrum",
        "keywords": ["omni shade broad spectrum", "omni-shade broad spectrum", "omni shade upf"],
        "logo": "Omni Shade Broad Spectrum",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni-Wick",
        "keywords": ["omni-wick", "omni wick", "omniwick"],
        "logo": "logo.omni-wick-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni Wick Evap",
        "keywords": ["omni wick evap", "omni-wick evap", "omniwick evap"],
        "logo": "Omni Wick Evap",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni Wind Block",
        "keywords": ["omni wind block", "omni-wind block", "omniwind block"],
        "logo": "Omni Wind Block",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni-Shield",
        "keywords": ["omni-shield", "omni shield", "omnishield"],
        "logo": "logo.omni-shield-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni-Grip",
        "keywords": ["omni-grip", "omni grip", "omnigrip"],
        "logo": "logo.omni-grip-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni Grip LT",
        "keywords": ["omni grip lt", "omni-grip lt", "omnigrip lt"],
        "logo": "Omni Grip LT",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni-Max",
        "keywords": ["omni-max", "omni max", "omnimax"],
        "logo": "logo.omni-max-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "OutDry",
        "keywords": ["outdry", "out-dry", "out dry"],
        "logo": "logo.out-dry-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Techlite",
        "keywords": ["techlite", "tech-lite", "tech lite"],
        "logo": "logo.tech-lite-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "TechLite Plus",
        "keywords": ["techlite plus", "tech-lite plus", "tech lite plus"],
        "logo": "TechLite Plus",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Thermarator",
        "keywords": ["thermarator"],
        "logo": "Thermarator",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Waterproof",
        "keywords": ["waterproof", "water proof"],
        "logo": "Waterproof",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Water Repellent",
        "keywords": ["water repellent", "water-repellent"],
        "logo": "Water Repellent",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni Freeze Zero Ice",
        "keywords": ["omni freeze zero ice", "omni-freeze zero ice", "omnifreeze zero ice"],
        "logo": "Omni Freeze Zero Ice",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Omni Heat Thermal Insulation",
        "keywords": [
            "omni heat thermal insulation",
            "omni-heat thermal insulation",
            "omni heat thermal insulaion",
            "omni-heat thermal insulaion",
        ],
        "logo": "Omni Heat Thermal Insulation",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Adapt Trax",
        "keywords": ["adapt trax", "adapt-trax", "adapttrax"],
        "logo": "logo.adapt-trax-clb",
        "active": True,
    },
    {
        "brand": "Columbia",
        "name": "Navic Fit",
        "keywords": ["navic fit", "navic-fit", "navicfit"],
        "logo": "logo.navic-fit-clb",
        "active": True,
    },
]


TECHNOLOGY_LOGO_ALIASES = {
    "omni-tech": ["Omni Tech", "Omni-Tech", "logo.omni-tech-clb"],
    "omni-heat": ["Omni Heat", "Omni-Heat", "logo.omni-heat-clb"],
    "omni-heat-infinity": ["Omni Heat Infinity", "Omni-Heat Infinity", "logo.omni-heat-infinity-clb"],
    "omni-shade": ["Omni Shade", "Omni-Shade", "logo.omni-shade-clb"],
    "omni-shade-broad-spectrum": [
        "Omni Shade Broad Spectrum",
        "Omni-Shade Broad Spectrum",
        "logo.omni-shade-broad-spectrum-clb",
    ],
    "omni-wick": ["Omni Wick", "Omni-Wick", "logo.omni-wick-clb"],
    "omni-wick-evap": ["Omni Wick Evap", "Omni-Wick Evap", "logo.omni-wick-evap-clb"],
    "omni-wind-block": ["Omni Wind Block", "Omni-Wind Block", "logo.omni-wind-block-clb"],
    "omni-shield": ["Omni Shield", "Omni-Shield", "logo.omni-shield-clb"],
    "omni-grip": ["Omni Grip", "Omni-Grip", "logo.omni-grip-clb"],
    "omni-grip-lt": ["Omni Grip LT", "Omni-Grip LT", "logo.omni-grip-lt-clb"],
    "omni-max": ["Omni Max", "Omni-Max", "logo.omni-max-clb"],
    "out-dry": ["OutDry", "Out Dry", "Out-Dry", "logo.out-dry-clb"],
    "techlite": ["Techlite", "Tech Lite", "Tech-Lite", "logo.tech-lite-clb"],
    "techlite-plus": ["TechLite Plus", "Techlite Plus", "Tech Lite Plus", "logo.tech-lite-plus-clb"],
    "adapt-trax": ["Adapt Trax", "Adapt-Trax", "logo.adapt-trax-clb"],
    "navic-fit": ["Navic Fit", "Navic-Fit", "logo.navic-fit-clb"],
    "thermarator": ["Thermarator", "logo.thermarator-clb"],
    "waterproof": ["Waterproof", "logo.waterproof-clb"],
    "water-repellent": ["Water Repellent", "Water-Repellent", "logo.water-repellent-clb"],
    "omni-freeze-zero-ice": ["Omni Freeze Zero Ice", "logo.omni-freeze-zero-ice-clb"],
    "omni-heat-thermal-insulation": [
        "Omni Heat Thermal Insulation",
        "Omni-Heat Thermal Insulation",
        "Omni Heat Thermal Insulaion",
        "logo.omni-heat-thermal-insulation-clb",
    ],
}


TECHNOLOGY_LOGO_METAOBJECT_GIDS = {
    "Amortiguación Alta": "gid://shopify/Metaobject/524803440828",
    "Amortiguación Baja": "gid://shopify/Metaobject/524803473596",
    "Light Warm": "gid://shopify/Metaobject/524803506364",
    "Maximum Warm": "gid://shopify/Metaobject/524803539132",
    "Medium Warm": "gid://shopify/Metaobject/524803571900",
    "Amortiguación Media": "gid://shopify/Metaobject/524803604668",
    "Navic Fit": "gid://shopify/Metaobject/524803637436",
    "Adapt Trax": "gid://shopify/Metaobject/524803670204",
    "Omni Heat Helix": "gid://shopify/Metaobject/524803702972",
    "Omni Heat Arctic": "gid://shopify/Metaobject/524803735740",
    "Omni Heat Reflective": "gid://shopify/Metaobject/524803768508",
    "Omni Heat Infinity": "gid://shopify/Metaobject/524803801276",
    "Omni Shield": "gid://shopify/Metaobject/524803834044",
    "Omni Tech": "gid://shopify/Metaobject/524803866812",
    "Omni Shade": "gid://shopify/Metaobject/524803899580",
    "Omni Wick": "gid://shopify/Metaobject/524803932348",
    "Omni Freeze Zero": "gid://shopify/Metaobject/524803965116",
    "Omni Freeze": "gid://shopify/Metaobject/524803997884",
    "OmniFreezeZeroIce": "gid://shopify/Metaobject/524804030652",
    "Omni Grip": "gid://shopify/Metaobject/524804063420",
    "Omni Shade Sun Deflector": "gid://shopify/Metaobject/524804096188",
    "Out Dry Extreme": "gid://shopify/Metaobject/524804128956",
    "Skin Cancer Foundation": "gid://shopify/Metaobject/524804161724",
    "RDS": "gid://shopify/Metaobject/524804194492",
    "Tech Lite": "gid://shopify/Metaobject/524804227260",
    "Thermarator": "gid://shopify/Metaobject/524804260028",
    "Omni Max": "gid://shopify/Metaobject/524804292796",
    "Fishing": "gid://shopify/Metaobject/524804325564",
    "Double Wall": "gid://shopify/Metaobject/524804358332",
    "Hiking": "gid://shopify/Metaobject/524804391100",
    "Heat Seal": "gid://shopify/Metaobject/524804423868",
    "Everyday": "gid://shopify/Metaobject/524804456636",
    "Omni Shield Release": "gid://shopify/Metaobject/524804489404",
    "Omni Shade Broad Spectrum AF": "gid://shopify/Metaobject/524804522172",
    "Omni Shade Broad Spectrum UPF50": "gid://shopify/Metaobject/524804554940",
    "Omni Shield Blood N Guts": "gid://shopify/Metaobject/524804587708",
    "Tech Lite Plush": "gid://shopify/Metaobject/524804620476",
    "Out Dry": "gid://shopify/Metaobject/524804653244",
    "Omni Heat Black Dot": "gid://shopify/Metaobject/524804686012",
    "Omni Heat Synthetic Down": "gid://shopify/Metaobject/524804718780",
    "Ski Snow": "gid://shopify/Metaobject/524804751548",
    "Trail Running": "gid://shopify/Metaobject/524804784316",
    "Turbo Down": "gid://shopify/Metaobject/524804817084",
    "Water": "gid://shopify/Metaobject/524804849852",
    "Omni Grip LT": "gid://shopify/Metaobject/524804882620",
    "Omni Heat Thermal Insulation": "gid://shopify/Metaobject/524804915388",
    "Omni Freeze Zero Ice": "gid://shopify/Metaobject/524804948156",
    "Water Repellent": "gid://shopify/Metaobject/524804980924",
    "Waterproof": "gid://shopify/Metaobject/524805013692",
    "Omni Wind Block": "gid://shopify/Metaobject/524805046460",
    "Omni Wick Evap": "gid://shopify/Metaobject/524805079228",
    "Omni Shade Broad Spectrum": "gid://shopify/Metaobject/524805111996",
    "TechLite Plus": "gid://shopify/Metaobject/524805144764",
}


TECHNOLOGY_SOURCE_COLUMNS = [
    "Metafield: custom.tecnologia [list.single_line_text_field]",
    "Metafield: custom.tecnologia [single_line_text_field]",
    "Tecnologias ",
    "Tecnologias",
    "Tecnología",
    "Tecnologia",
    "METAFIELD TECNOLOGÍAS",
    "METAFIELD TECNOLOGIAS",
    "Title",
    "Titulo",
    "Nombre",
    "Body HTML",
    "Description",
    "Descripcion",
    "Descripción",
    "Caracteristicas",
    "Características",
    "Tags",
]


MATERIAL_RECOVERY_COLUMNS = [
    "Metafield: custom.materialidad [single_line_text_field]",
    "Material",
    "Tipo de Material",
    "Materialidad",
    "MATERIALIDAD",
    "Composición",
    "Composicion",
    "COMPOSICION",
    "ComposiciÃ³n",
    "Material principal",
    "MATERIAL",
]


def material_recovery_by_key(arti_df):
    if arti_df is None or not isinstance(arti_df, pd.DataFrame) or arti_df.empty:
        return {}
    key_col = first_existing_column(
        arti_df,
        [
            "Mod-Col",
            "COD MOD COL",
            "COD_MOD_COL",
            "Codigo Modelo Color",
            "Código Modelo Color",
            "codigo_modelo_color",
            "CONCA",
            "conca",
        ],
    )
    if not key_col:
        return {}
    value_cols = [column for column in MATERIAL_RECOVERY_COLUMNS if column in arti_df.columns]
    if not value_cols:
        return {}
    recovered = {}
    for _, item in arti_df.iterrows():
        key = clean_value(item.get(key_col)).upper()
        if not key or key in recovered:
            continue
        for column in value_cols:
            value = clean_value(item.get(column))
            if value:
                try:
                    from generate_columbia_matrixify import valid_body_section_text

                    value = valid_body_section_text(value)
                except Exception:
                    pass
            if value:
                recovered[key] = value
                break
    return recovered


def _technology_text_from_records(*records):
    pieces = []
    for record in records:
        if record is None:
            continue
        for column in TECHNOLOGY_SOURCE_COLUMNS:
            try:
                value = record.get(column)
            except Exception:
                value = ""
            value = clean_value(strip_html(value))
            if value:
                pieces.append(value)
    return " | ".join(pieces)


def _explicit_technology_items(*records):
    from generate_columbia_matrixify import split_technology_items

    explicit_columns = TECHNOLOGY_SOURCE_COLUMNS[:8]
    items = []
    seen = set()
    for record in records:
        if record is None:
            continue
        for column in explicit_columns:
            try:
                value = record.get(column)
            except Exception:
                value = ""
            for item in split_technology_items(value):
                key = clean_value(item).lower()
                if key and key not in seen:
                    items.append(item)
                    seen.add(key)
    return items


def detect_product_technologies(source_row=None, shopify_product=None, brand_config=None):
    from generate_columbia_matrixify import format_technology_logos, split_technology_items

    site_brand = clean_value((brand_config or {}).get("brand_name") or (brand_config or {}).get("site_label")).lower()
    text = _technology_text_from_records(source_row, shopify_product).lower()
    explicit_items = _explicit_technology_items(source_row, shopify_product)
    detected_names = []
    detected_logos = []
    seen_names = set()
    seen_logos = set()

    def technology_identity_key(value):
        text_value = clean_value(value).lower()
        text_value = text_value.replace("tech-lite", "techlite").replace("tech lite", "techlite")
        text_value = text_value.replace("out-dry", "outdry").replace("out dry", "outdry")
        return re.sub(r"[^a-z0-9]+", "", text_value)

    def logo_aliases_for_value(value):
        text_value = clean_value(value)
        normalized = re.sub(r"[^a-z0-9]+", "-", text_value.lower()).strip("-")
        if normalized.startswith("logo-"):
            normalized = normalized[5:]
        if normalized.endswith("-clb"):
            normalized = normalized[:-4]
        aliases = [text_value]
        aliases.extend(TECHNOLOGY_LOGO_ALIASES.get(normalized, []))
        return [clean_value(item) for item in aliases if clean_value(item)]

    def add_detected(name, logo):
        name = clean_value(name)
        logo = clean_value(logo)
        name_key = technology_identity_key(name)
        if any(
            existing.startswith(f"{name_key} ") or existing.startswith(f"{name_key}-")
            for existing in seen_names
        ):
            return
        if name and name_key not in seen_names:
            detected_names.append(name)
            seen_names.add(name_key)
        logo_aliases = logo_aliases_for_value(logo or name)
        preferred_logo = next((item for item in logo_aliases if clean_value(item).lower().startswith("logo.")), "")
        preferred_logo = preferred_logo or (logo_aliases[0] if logo_aliases else "")
        logo_key = _static_logo_metaobject_gid(preferred_logo) or technology_identity_key(preferred_logo)
        if preferred_logo and logo_key not in seen_logos:
            detected_logos.append(preferred_logo)
            seen_logos.add(logo_key)

    for tech in sorted(
        TECHNOLOGY_MAINTAINER,
        key=lambda item: max([len(clean_value(keyword)) for keyword in item.get("keywords", [])] or [0]),
        reverse=True,
    ):
        if not tech.get("active", True):
            continue
        tech_brand = clean_value(tech.get("brand")).lower()
        if tech_brand and site_brand and tech_brand not in site_brand and site_brand not in tech_brand:
            continue
        keywords = [clean_value(item).lower() for item in tech.get("keywords", []) if clean_value(item)]
        explicit_match = any(
            clean_value(item).lower() == clean_value(tech.get("name")).lower()
            or clean_value(tech.get("name")).lower() in clean_value(item).lower()
            for item in explicit_items
        )
        if explicit_match or any(keyword and keyword in text for keyword in keywords):
            add_detected(tech.get("name"), tech.get("logo"))

    for item in explicit_items:
        logos = format_technology_logos(item)
        for parsed in split_technology_items(item):
            if parsed:
                add_detected(parsed, "")
        for logo in _split_tags(logos):
            logo_key = _static_logo_metaobject_gid(logo) or technology_identity_key(logo)
            if logo and logo_key not in seen_logos:
                detected_logos.append(logo)
                seen_logos.add(logo_key)

    return detected_names, detected_logos


def partial_preview_summary(preview_df, issues_df=None):
    preview = preview_df if isinstance(preview_df, pd.DataFrame) else pd.DataFrame()
    issues = issues_df if isinstance(issues_df, pd.DataFrame) else pd.DataFrame()
    if preview.empty:
        return pd.DataFrame(
            [
                {"Indicador": "Productos procesados", "Valor": 0},
                {"Indicador": "Errores u observaciones", "Valor": len(issues)},
            ]
        )
    operation = preview["Operacion"].map(clean_value).str.lower() if "Operacion" in preview.columns else pd.Series([], dtype=object)
    html_count = int((operation == "body").sum()) if not operation.empty else 0
    tech_count = int((operation == "technologies").sum()) if not operation.empty else 0
    logo_count = 0
    if "Valor nuevo logos" in preview.columns:
        logo_count = int(preview["Valor nuevo logos"].map(lambda value: len(_split_tags(value))).sum())
    return pd.DataFrame(
        [
            {"Indicador": "Productos procesados", "Valor": int(preview["Handle"].map(clean_value).nunique() if "Handle" in preview.columns else len(preview))},
            {"Indicador": "HTML normalizado", "Valor": html_count},
            {"Indicador": "Tecnologías detectadas", "Valor": tech_count},
            {"Indicador": "Logos inyectados", "Valor": logo_count},
            {"Indicador": "Errores u observaciones", "Valor": len(issues)},
        ]
    )


PARTIAL_REQUIRED_HTML_SECTIONS = ("Caracteristicas", "Materiales", "Cuidados")
PARTIAL_INVALID_VALUES = {"", "-", "--", "0", "00", "000", "n/a", "na", "null", "none", "sin info", "sin informacion"}
PARTIAL_HTML_DANGER_TOKENS = ("<script", "<style", " onerror=", " onclick=", "javascript:", "<iframe", "<object", "<embed")


def read_uploaded_excel_sheet_or_first(uploaded_file, state_prefix, preferred_sheet="Products"):
    """Read the preferred worksheet when present; otherwise fall back to the first sheet."""
    if uploaded_file is None:
        return pd.DataFrame()
    try:
        return read_uploaded_excel_cached(uploaded_file, state_prefix, sheet_name=preferred_sheet)
    except ValueError as exc:
        if "Worksheet named" not in clean_value(exc):
            raise
        return read_uploaded_excel_cached(uploaded_file, f"{state_prefix}_first_sheet", sheet_name=0)


def _partial_status_series(df):
    if df is None or df.empty or "Estado validacion" not in df.columns:
        return pd.Series([], dtype=object)
    return df["Estado validacion"].map(lambda value: normalize_header(clean_value(value)))


def _partial_problem_series(df):
    if df is None or df.empty or "Problema" not in df.columns:
        return pd.Series([], dtype=object)
    return df["Problema"].map(lambda value: normalize_header(clean_value(value)))


def _partial_bad_value(value):
    text = normalize_header(value)
    return text in PARTIAL_INVALID_VALUES or text in {"sincontenido", "sininformacion", "pendiente", "porcompletar", "tbd"}


def _html_tag_balance_warning(html):
    lowered = clean_value(html).lower()
    for tag in ("section", "div", "ul", "ol", "li", "p", "strong"):
        opening = len(re.findall(rf"<{tag}(\s|>|/)", lowered))
        closing = lowered.count(f"</{tag}>")
        if opening != closing:
            return f"HTML posiblemente mal cerrado en etiqueta {tag}."
    return ""


def _partial_issue_text(row):
    pieces = []
    for column in ("Problema", "Detalle", "Observacion"):
        value = clean_value(row.get(column))
        if value:
            pieces.append(value)
    return " | ".join(dict.fromkeys(pieces))


def _partial_row_base(row, operation="", status="Listo", problem="", action="Actualizar campo seleccionado"):
    return {
        "Mod-Col": clean_value(row.get("Mod-Col") or row.get("Código modelo color") or row.get("Codigo modelo color")),
        "Handle": clean_value(row.get("Handle")),
        "Producto": clean_value(row.get("Title") or row.get("Nombre") or row.get("Product Title")),
        "Marca": clean_value(row.get("Vendor") or row.get("Marca") or row.get("Sitio")),
        "Categoria": clean_value(row.get("Categoria") or row.get("Categoría") or row.get("Metafield: custom.categoria [single_line_text_field]")),
        "Tipo": clean_value(row.get("Type") or row.get("Tipo") or row.get("Tipo de prenda") or row.get("Metafield: custom.tipo [single_line_text_field]")),
        "Genero": clean_value(row.get("Genero") or row.get("Género") or row.get("Metafield: custom.genero [single_line_text_field]")),
        "Operacion": clean_value(operation or row.get("Operacion")),
        "Estado validacion": status,
        "Problema": clean_value(problem),
        "Accion sugerida": clean_value(action),
    }


def validate_partial_body_html(value, title="", mod_col=""):
    html = clean_value(value)
    text = strip_html(html)
    normalized_text = normalize_header(text)
    if _partial_bad_value(html) or _partial_bad_value(text):
        return "Bloqueado", "Body HTML vacío o placeholder inválido."
    if len(text) < 45:
        return "Bloqueado", "Body HTML demasiado corto para publicación."
    if normalize_header(title) and normalized_text == normalize_header(title):
        return "Bloqueado", "Body HTML solo contiene el título del producto."
    if normalize_header(mod_col) and normalized_text == normalize_header(mod_col):
        return "Bloqueado", "Body HTML solo contiene el código modelo-color."
    lowered = html.lower()
    if any(token in lowered for token in PARTIAL_HTML_DANGER_TOKENS):
        return "Bloqueado", "HTML contiene scripts, estilos o atributos no permitidos."
    visible_tags = ("&lt;" in lowered and "&gt;" in lowered) or ("<p>" in text.lower() and "</p>" in text.lower())
    if visible_tags:
        return "Observación", "El texto parece contener etiquetas visibles; revisar antes de aplicar."
    missing_sections = [
        label
        for label in PARTIAL_REQUIRED_HTML_SECTIONS
        if normalize_header(label) not in normalize_header(text)
    ]
    if missing_sections:
        return "Observación", f"Faltan secciones esperadas: {', '.join(missing_sections)}."
    balance_warning = _html_tag_balance_warning(html)
    if balance_warning:
        return "Observación", balance_warning
    return "Listo", "HTML listo para actualizar solo Body HTML."


def validate_partial_image_urls(value):
    urls = [clean_value(item) for item in re.split(r"[;|\n]+", clean_value(value)) if clean_value(item)]
    if not urls:
        return "Bloqueado", "No hay URLs de fotos para cargar.", 0, 0
    duplicates = len(urls) - len(dict.fromkeys(urls))
    invalid = [
        url for url in urls
        if not re.match(r"^https?://", url, flags=re.IGNORECASE)
        or " " in url
    ]
    if invalid:
        return "Bloqueado", f"{len(invalid)} URL(s) inválida(s).", len(urls), duplicates
    if duplicates:
        return "Observación", f"{duplicates} URL(s) duplicada(s); se recomienda revisar.", len(urls), duplicates
    return "Listo", "Fotos listas para actualizar solo media.", len(urls), 0


def partial_business_entity_count(df):
    if df is None or df.empty:
        return 0
    for column in ("Mod-Col", "Handle", "Product ID"):
        if column in df.columns:
            values = df[column].map(clean_value)
            values = values[values != ""]
            if not values.empty:
                return int(values.nunique())
    return int(len(df))


def _partial_proposed_value(row, operation):
    operation = clean_value(operation).lower()
    if operation == "body":
        return (
            row.get("Valor nuevo")
            or row.get("Body HTML")
            or row.get("Description HTML")
            or row.get("Descripcion HTML")
            or row.get("Descripción HTML")
        )
    if operation == "photos":
        return row.get("Valor nuevo") or row.get("Image Src") or row.get("Imagen") or row.get("URL Imagen")
    if operation == "size_guides":
        return (
            row.get("Valor nuevo")
            or row.get("Guía propuesta")
            or row.get("Guia propuesta")
            or row.get("Metafield: custom.guia_de_tallas [page_reference]")
            or row.get("Guía de tallas")
            or row.get("Guia de tallas")
        )
    return row.get("Valor nuevo")


def _partial_current_value(row, operation):
    operation = clean_value(operation).lower()
    if clean_value(row.get("Valor actual")):
        return row.get("Valor actual")
    if operation == "body":
        return row.get("Body HTML actual") or row.get("Current Body HTML") or row.get("Body HTML")
    if operation == "photos":
        return row.get("Fotos actuales") or row.get("Current Image Src") or row.get("Image Src actual")
    if operation == "size_guides":
        return row.get("Guía actual") or row.get("Guia actual")
    return ""


def _size_guide_decision_for_row(row):
    current = clean_value(_partial_proposed_value(row, "size_guides") or _partial_current_value(row, "size_guides"))
    return resolve_size_guide(
        brand=clean_value(row.get("Marca") or row.get("Vendor") or row.get("Sitio")),
        category=clean_value(row.get("Categoria") or row.get("Categoría")),
        product_type=clean_value(row.get("Tipo") or row.get("Type") or row.get("Tipo de prenda")),
        gender=clean_value(row.get("Genero") or row.get("Género")),
        age_group=clean_value(row.get("Edad") or row.get("Age Group")),
        current_guide=current,
    )


def filter_preview_by_diagnostic_ready(preview_df, diagnostic_df):
    preview = preview_df if isinstance(preview_df, pd.DataFrame) else pd.DataFrame()
    diagnostic = diagnostic_df if isinstance(diagnostic_df, pd.DataFrame) else pd.DataFrame()
    if preview.empty or diagnostic.empty or "Estado validacion" not in diagnostic.columns:
        return preview
    ready = diagnostic[diagnostic["Estado validacion"].map(clean_value).str.lower() == "listo"].copy()
    if ready.empty:
        return preview.iloc[0:0].copy()
    filtered = preview.copy()
    masks = []
    if "Handle" in ready.columns and "Handle" in filtered.columns:
        ready_handles = {clean_value(value) for value in ready["Handle"] if clean_value(value)}
        if ready_handles:
            masks.append(filtered["Handle"].map(clean_value).isin(ready_handles))
    if "Mod-Col" in ready.columns and "Mod-Col" in filtered.columns:
        ready_keys = {clean_value(value).upper() for value in ready["Mod-Col"] if clean_value(value)}
        if ready_keys:
            masks.append(filtered["Mod-Col"].map(lambda value: clean_value(value).upper()).isin(ready_keys))
    if not masks:
        return filtered.iloc[0:0].copy()
    mask = masks[0].copy()
    for extra_mask in masks[1:]:
        mask = mask | extra_mask
    return filtered[mask].copy()


def build_partial_diagnostic_table(preview_df, issues_df=None, operation=""):
    preview = preview_df if isinstance(preview_df, pd.DataFrame) else pd.DataFrame()
    issues = issues_df if isinstance(issues_df, pd.DataFrame) else pd.DataFrame()
    rows = []
    operation = clean_value(operation).lower()

    for _, row in preview.iterrows():
        base = _partial_row_base(row, operation=operation, status="Listo")
        base["Campo Shopify"] = clean_value(row.get("Campo"))
        current_value = _partial_current_value(row, operation)
        proposed_value = _partial_proposed_value(row, operation)
        base["Valor actual"] = clean_value(current_value)[:1200]
        base["Valor propuesto"] = clean_value(proposed_value)[:1200]
        if operation == "body":
            status, problem = validate_partial_body_html(proposed_value, row.get("Producto") or row.get("Title"), row.get("Mod-Col"))
            current_status, current_problem = validate_partial_body_html(current_value, row.get("Producto") or row.get("Title"), row.get("Mod-Col"))
            base.update(
                {
                    "Estado validacion": status,
                    "Problema": "" if status == "Listo" else problem,
                    "Accion sugerida": "Actualizar solo Body HTML" if status == "Listo" else "Revisar contenido antes de aplicar",
                    "Preview texto": strip_html(proposed_value)[:600],
                    "HTML actual valido": "Sí" if current_status == "Listo" else "No",
                    "Problema HTML actual": "" if current_status == "Listo" else current_problem,
                    "Campo afectado": "Body HTML",
                }
            )
        elif operation == "photos":
            status, problem, photo_count, duplicate_count = validate_partial_image_urls(proposed_value)
            current_urls = _split_semicolon_values(current_value)
            base.update(
                {
                    "Estado validacion": status,
                    "Problema": "" if status == "Listo" else problem,
                    "Accion sugerida": "Actualizar solo fotos/media" if status == "Listo" else "Corregir URLs de fotos",
                    "Cantidad fotos": photo_count,
                    "Duplicadas": duplicate_count,
                    "Tiene foto actual": "Sí" if current_urls or clean_value(row.get("Valor actual")) else "No",
                    "Tiene foto principal propuesta": "Sí" if photo_count > 0 else "No",
                    "Campo afectado": "Media/Fotos",
                }
            )
        elif operation == "size_guides":
            decision = _size_guide_decision_for_row(row)
            status = "Listo" if clean_value(decision.get("status")) in ("approved", "") else "Bloqueado" if clean_value(decision.get("status")) == "blocked" else "Observación"
            base.update(
                {
                    "Estado validacion": status,
                    "Problema": clean_value(decision.get("warning")) if status != "Listo" else "",
                    "Accion sugerida": "Actualizar solo guía de tallas" if status == "Listo" else "Revisar regla categoría/tipo/género",
                    "Guía actual": clean_value(row.get("Valor actual")),
                    "Guía propuesta": clean_value(row.get("Valor nuevo")),
                    "Regla aplicada": clean_value(decision.get("rule")),
                    "Campo afectado": "Metafield: custom.guia_de_tallas",
                }
            )
        rows.append(base)

    for _, issue in issues.iterrows():
        base = _partial_row_base(issue, operation=operation, status="Bloqueado", problem=_partial_issue_text(issue), action="No actualizar hasta corregir")
        base["Campo Shopify"] = ""
        base["Valor actual"] = ""
        base["Valor propuesto"] = ""
        rows.append(base)

    return pd.DataFrame(rows)


def partial_diagnostic_summary(diagnostic_df, operation=""):
    df = diagnostic_df if isinstance(diagnostic_df, pd.DataFrame) else pd.DataFrame()
    operation = clean_value(operation).lower()
    if df.empty:
        return pd.DataFrame(
            [
                {"Indicador": "Total analizado", "Valor": 0},
                {"Indicador": "Listos para actualizar", "Valor": 0},
                {"Indicador": "Bloqueados", "Valor": 0},
                {"Indicador": "Con observación", "Valor": 0},
            ]
        )
    status = _partial_status_series(df)
    problems = _partial_problem_series(df)
    total = partial_business_entity_count(df)
    ready = int((status == "listo").sum())
    blocked = int((status == "bloqueado").sum())
    observed = int((status == "observacion").sum())
    rows = []
    if operation == "body":
        rows = [
            {"Indicador": "Total analizado", "Valor": total},
            {"Indicador": "Body HTML válido", "Valor": ready},
            {"Indicador": "Sin descripción", "Valor": int(problems.str.contains("vacio|placeholder|nocontenido|nohaycontenido", na=False).sum())},
            {"Indicador": "Contenido insuficiente", "Valor": int(problems.str.contains("demasiadocorto|solocontiene", na=False).sum())},
            {"Indicador": "Valores sin sentido", "Valor": int(problems.str.contains("placeholder|codigo|titulo", na=False).sum())},
            {"Indicador": "HTML mal formado", "Valor": int(problems.str.contains("malcerrado|etiquetasvisibles|scripts|estilos|atributosnopermitidos", na=False).sum())},
            {"Indicador": "Listos para actualizar", "Valor": ready},
            {"Indicador": "Bloqueados", "Valor": blocked},
        ]
    elif operation == "photos":
        photo_counts = pd.to_numeric(df.get("Cantidad fotos", pd.Series(0, index=df.index)), errors="coerce").fillna(0)
        duplicates = pd.to_numeric(df.get("Duplicadas", pd.Series(0, index=df.index)), errors="coerce").fillna(0)
        current_photo = df.get("Tiene foto actual", pd.Series("", index=df.index)).map(clean_value).str.lower()
        rows = [
            {"Indicador": "Modelos analizados", "Valor": total},
            {"Indicador": "Con fotografías", "Valor": int((photo_counts > 0).sum())},
            {"Indicador": "Sin fotografías", "Valor": int((photo_counts <= 0).sum())},
            {"Indicador": "Sin imagen principal", "Valor": int((current_photo != "sí").sum() if not current_photo.empty else 0)},
            {"Indicador": "URLs inválidas/vacías", "Valor": int(problems.str.contains("url|nohayurls|invalida|inválida", na=False).sum())},
            {"Indicador": "Imágenes duplicadas", "Valor": int((duplicates > 0).sum())},
            {"Indicador": "Listos para actualizar", "Valor": ready},
            {"Indicador": "Bloqueados", "Valor": blocked},
        ]
    elif operation == "size_guides":
        rows = [
            {"Indicador": "Total analizado", "Valor": total},
            {"Indicador": "Guía válida", "Valor": ready},
            {"Indicador": "Sin guía", "Valor": int(problems.str.contains("vacia|vacía|noguia|noexiste", na=False).sum())},
            {"Indicador": "Posible error", "Valor": observed},
            {"Indicador": "Categoría/guía incompatible", "Valor": int(problems.str.contains("incompatible|calzado|vestuario|categoria|categoría", na=False).sum())},
            {"Indicador": "Listos para actualizar", "Valor": ready},
            {"Indicador": "Bloqueados", "Valor": blocked},
        ]
    else:
        rows = [
            {"Indicador": "Total analizado", "Valor": total},
            {"Indicador": "Listos para actualizar", "Valor": ready},
            {"Indicador": "Bloqueados", "Valor": blocked},
            {"Indicador": "Con observación", "Valor": observed},
        ]
    return pd.DataFrame(rows)


def render_partial_diagnostic_panel(diagnostic_df, operation=""):
    if diagnostic_df is None or diagnostic_df.empty:
        st.warning("No hay registros válidos para diagnosticar.")
        return diagnostic_df
    summary_df = partial_diagnostic_summary(diagnostic_df, operation)
    st.markdown("#### Diagnóstico antes de actualizar Shopify")
    st.caption("El análisis es una simulación: solo las filas con estado Listo pasan a la actualización. Los bloqueados quedan fuera.")
    cards_html = ['<div class="partial-kpi-grid">']
    for _, row in summary_df.iterrows():
        cards_html.append(
            '<div class="partial-kpi-card">'
            f'<span>{clean_value(row.get("Indicador"))}</span>'
            f'<strong>{format_kpi_number(safe_int_value(row.get("Valor")))}</strong>'
            '</div>'
        )
    cards_html.append("</div>")
    st.markdown(
        """
        <style>
        .partial-kpi-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin:10px 0 18px}
        .partial-kpi-card{background:#fff;border:1px solid #dbe5f2;border-radius:14px;padding:14px 16px;min-height:84px;box-shadow:0 8px 24px rgba(15,40,80,.05)}
        .partial-kpi-card span{display:block;color:#53657c;font-weight:800;font-size:12px;line-height:1.25}
        .partial-kpi-card strong{display:block;color:#001b44;font-size:26px;line-height:1.05;margin-top:8px}
        @media(max-width:1100px){.partial-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}}
        </style>
        """
        + "".join(cards_html),
        unsafe_allow_html=True,
    )
    filter_cols = st.columns([1.2, 1.2, 1.2, 2])
    brands = ["Todas"] + sorted([value for value in diagnostic_df.get("Marca", pd.Series(dtype=object)).map(clean_value).unique() if value])
    categories = ["Todas"] + sorted([value for value in diagnostic_df.get("Categoria", pd.Series(dtype=object)).map(clean_value).unique() if value])
    statuses = ["Todos"] + sorted([value for value in diagnostic_df.get("Estado validacion", pd.Series(dtype=object)).map(clean_value).unique() if value])
    brand = filter_cols[0].selectbox("Marca", brands, key=f"partial_diag_brand_{operation}")
    category = filter_cols[1].selectbox("Categoría", categories, key=f"partial_diag_category_{operation}")
    status = filter_cols[2].selectbox("Estado", statuses, key=f"partial_diag_status_{operation}")
    search = filter_cols[3].text_input("Buscar Mod-Col, handle o producto", key=f"partial_diag_search_{operation}")

    filtered = diagnostic_df.copy()
    if brand != "Todas" and "Marca" in filtered.columns:
        filtered = filtered[filtered["Marca"].map(clean_value) == brand]
    if category != "Todas" and "Categoria" in filtered.columns:
        filtered = filtered[filtered["Categoria"].map(clean_value) == category]
    if status != "Todos" and "Estado validacion" in filtered.columns:
        filtered = filtered[filtered["Estado validacion"].map(clean_value) == status]
    if clean_value(search):
        needle = normalize_header(search)
        searchable = filtered.apply(lambda row: " ".join(clean_value(value) for value in row.values), axis=1).map(normalize_header)
        filtered = filtered[searchable.str.contains(re.escape(needle), na=False)]
    st.dataframe(filtered.head(500), use_container_width=True, height=360)
    return filtered


def build_shopify_update_preview(
    shopify_products,
    update_input_df,
    operation,
    brand_config,
    shopify_config=None,
    arti_df=None,
    tag_mode="merge",
    image_mode="replace",
    only_missing_images=True,
    body_mode="from_input",
):
    by_key, by_handle = _product_lookup_from_shopify(shopify_products)
    rows = []
    issues = []
    operation = clean_value(operation)
    arti_material_lookup = material_recovery_by_key(arti_df) if operation == "body" else {}

    if operation == "siblings":
        products_df = pd.DataFrame(shopify_products)
        if products_df.empty:
            return pd.DataFrame(), pd.DataFrame([{"Problema": "Shopify no devolvio productos"}]), pd.DataFrame()
        products_df["__MODEL"] = products_df["Mod-Col"].map(lambda value: clean_value(value).upper().rsplit("-", 1)[0])
        siblings_by_model = (
            products_df[products_df["__MODEL"] != ""]
            .groupby("__MODEL")["Handle"]
            .apply(lambda values: ", ".join(dict.fromkeys(clean_value(value) for value in values if clean_value(value))))
            .to_dict()
        )
        custom_siblings_by_model = (
            products_df[products_df["__MODEL"] != ""]
            .groupby("__MODEL")["Product ID"]
            .apply(lambda values: json.dumps(list(dict.fromkeys(clean_value(value) for value in values if clean_value(value)))))
            .to_dict()
        )
        for _, product in products_df.iterrows():
            new_value = siblings_by_model.get(product["__MODEL"], "")
            custom_new_value = custom_siblings_by_model.get(product["__MODEL"], "[]")
            current_theme = clean_value(product.get("Siblings"))
            current_custom = clean_value(product.get("Custom Siblings"))
            if not new_value or (current_theme == new_value and current_custom == custom_new_value):
                continue
            rows.append(
                {
                    "Accion": "Actualizar",
                    "Sitio": brand_config["site_label"],
                    "Operacion": "siblings",
                    "Mod-Col": product.get("Mod-Col"),
                    "Product ID": product.get("Product ID"),
                    "Handle": product.get("Handle"),
                    "Campo": "Metafield: theme.siblings + Metafield: custom.siblings",
                    "Valor actual": f"theme: {current_theme} | custom: {current_custom}",
                    "Valor nuevo": new_value,
                    "Valor nuevo custom": custom_new_value,
                    "Metafield: theme.siblings [single_line_text_field]": new_value,
                    "Metafield: custom.siblings [list.product_reference]": custom_new_value,
                    "Estado": "OK",
                    "Observacion": f"{len(_split_tags(new_value))} handles del mismo modelo",
                }
            )
        return pd.DataFrame(rows), pd.DataFrame(issues), pd.DataFrame()

    if operation == "photos":
        update_input_df = normalize_photo_update_input(update_input_df)
    source_df = (
        normalize_partial_update_input(update_input_df, operation)
        if operation != "photos"
        else (update_input_df.dropna(how="all").copy() if update_input_df is not None else pd.DataFrame())
    )
    if operation in ("photos", "technologies") and source_df.empty:
        source_df = pd.DataFrame(shopify_products)
    if operation == "body" and body_mode == "fix_catalog" and source_df.empty:
        source_df = pd.DataFrame(shopify_products)
    if operation == "size_guides" and source_df.empty:
        source_df = pd.DataFrame(shopify_products)

    matrixify_rows = []
    for input_index, row in source_df.iterrows():
        key = _source_key_for_update(row)
        handle = clean_value(row.get("Handle"))
        product = next((by_key.get(candidate) for candidate in product_lookup_candidates(key) if by_key.get(candidate)), None) or by_handle.get(handle)
        if not product:
            raw_key_values = [
                clean_value(row.get(column))
                for column in UPDATE_KEY_COLUMNS
                if clean_value(row.get(column))
            ]
            issues.append(
                {
                    "Mod-Col": key or "Sin Mod-Col detectado",
                    "Handle": handle,
                    "Problema": "No se encontro producto en Shopify",
                    "Detalle": f"Columnas leidas: {', '.join(raw_key_values[:3])}" if raw_key_values else "No se detecto columna Mod-Col/Handle reconocible",
                    "Fila": input_index + 2,
                }
            )
            continue

        product_id = product.get("Product ID")
        product_key = key or clean_value(product.get("Mod-Col")).upper()
        if operation == "tags":
            tags_col = first_existing_column(source_df, TAG_UPDATE_COLUMNS)
            if not tags_col:
                issues.append(
                    {
                        "Mod-Col": product_key,
                        "Handle": product.get("Handle"),
                        "Problema": "No se encontro columna Tags",
                        "Detalle": f"Columnas recibidas: {', '.join(clean_value(column) for column in source_df.columns[:12])}",
                    }
                )
                continue
            current_tags = _split_tags(product.get("Tags"))
            incoming_tags = _split_tags(row.get(tags_col))
            new_tags = _join_tags(incoming_tags if tag_mode == "replace" else current_tags + incoming_tags)
            rows.append(
                {
                    "Accion": "Actualizar",
                    "Sitio": brand_config["site_label"],
                    "Operacion": "tags",
                    "Mod-Col": product_key,
                    "Product ID": product_id,
                    "Handle": product.get("Handle"),
                    "Campo": "Tags",
                    "Valor actual": product.get("Tags"),
                    "Valor nuevo": new_tags,
                    "Estado": "OK",
                    "Observacion": "REPLACE seguro: se envia la lista final completa",
                }
            )
        elif operation == "title":
            title_col = first_existing_column(
                source_df,
                [
                    "Title",
                    "Product Title",
                    "Product Name",
                    "Nombre del Producto",
                    "Nombre Producto",
                    "Nombre Web",
                    "Titulo",
                    "TÃ­tulo",
                    "Título",
                    "Nombre",
                ],
            )
            if not title_col:
                issues.append({"Mod-Col": product_key, "Handle": product.get("Handle"), "Problema": "No se encontro columna Title"})
                continue
            new_title = clean_value(row.get(title_col))
            rows.append(
                {
                    "Accion": "Actualizar",
                    "Sitio": brand_config["site_label"],
                    "Operacion": "title",
                    "Mod-Col": product_key,
                    "Product ID": product_id,
                    "Handle": product.get("Handle"),
                    "Campo": "Title",
                    "Valor actual": product.get("Title"),
                    "Valor nuevo": new_title,
                    "Estado": "OK",
                    "Observacion": "",
                }
            )
        elif operation == "body":
            if body_mode == "from_input":
                from generate_columbia_matrixify import build_body_html

                new_body = build_body_html(row)
                if not new_body:
                    issues.append({"Mod-Col": product_key, "Handle": product.get("Handle"), "Problema": "No hay contenido para Body HTML"})
                    continue
            else:
                from generate_columbia_matrixify import _body_needs_material_care_fix, _split_labeled_body_text, build_body_html, valid_body_section_text

                current_body = clean_value(product.get("Body HTML"))
                if not _body_needs_material_care_fix(current_body):
                    continue
                features, material, care = _split_labeled_body_text(current_body)
                body_material = valid_body_section_text(material)
                shopify_material = valid_body_section_text(product.get("Metafield: custom.materialidad [single_line_text_field]"))
                arti_material = valid_body_section_text(
                    arti_material_lookup.get(product_key)
                    or arti_material_lookup.get(product_lookup_key(product_key))
                    or arti_material_lookup.get(clean_value(product.get("Mod-Col")).upper())
                )
                material = body_material or shopify_material or arti_material
                if not material and "nweb__materiales" in clean_value(current_body).lower():
                    issues.append(
                        {
                            "Mod-Col": product_key,
                            "Handle": product.get("Handle"),
                            "Problema": "Material invalido en Body HTML y no encontre respaldo en Shopify/ARTI. No se sobrescribe.",
                        }
                    )
                    continue
                new_body = build_body_html(
                    {
                        "Body HTML": "",
                        "Caracteristicas": features,
                        "Material": material,
                        "Cuidado": care,
                    }
                )
                if not new_body:
                    issues.append({"Mod-Col": product_key, "Handle": product.get("Handle"), "Problema": "No pude normalizar Body HTML"})
                    continue
            rows.append(
                {
                    "Accion": "Actualizar",
                    "Sitio": brand_config["site_label"],
                    "Operacion": "body",
                    "Mod-Col": product_key,
                    "Product ID": product_id,
                    "Handle": product.get("Handle"),
                    "Campo": "Body HTML",
                    "Valor actual": product.get("Body HTML"),
                    "Valor nuevo": new_body,
                    "Estado": "OK",
                    "Observacion": "HTML normalizado con Caracteristicas, Material y Cuidado separados",
                }
            )
        elif operation == "size_guides":
            guide_col = first_existing_column(source_df, SIZE_GUIDE_UPDATE_COLUMNS)
            current_guide = clean_value(
                product.get("Metafield: custom.guia_de_tallas [page_reference]")
                or product.get("Guia de tallas")
                or product.get("Guía de tallas")
            )
            input_guide = clean_value(row.get(guide_col)) if guide_col else ""
            proposed_seed = input_guide or current_guide
            category = clean_value(
                row.get("Categoria")
                or row.get("Categoría")
                or product.get("Metafield: custom.categoria [single_line_text_field]")
                or product.get("Type")
            )
            product_type = clean_value(
                row.get("Tipo de prenda")
                or row.get("Tipo")
                or product.get("Metafield: custom.tipo [single_line_text_field]")
                or product.get("Type")
            )
            gender = clean_value(
                row.get("Genero")
                or row.get("Género")
                or product.get("Metafield: custom.genero [single_line_text_field]")
            )
            decision = resolve_size_guide(
                brand=clean_value(row.get("Marca") or product.get("Vendor") or brand_config.get("label")),
                category=category,
                product_type=product_type,
                gender=gender,
                age_group=clean_value(row.get("Edad") or row.get("Age Group")),
                current_guide=proposed_seed,
            )
            if clean_value(decision.get("status")) == "blocked":
                issues.append(
                    {
                        "Mod-Col": product_key,
                        "Handle": product.get("Handle"),
                        "Problema": "Guía de talla incompatible",
                        "Detalle": clean_value(decision.get("warning")),
                        "Fila": input_index + 2,
                    }
                )
                continue
            proposed_guide = clean_value(input_guide or decision.get("guide"))
            if normalize_header(proposed_guide) in PARTIAL_INVALID_VALUES:
                issues.append(
                    {
                        "Mod-Col": product_key,
                        "Handle": product.get("Handle"),
                        "Problema": "Guía de talla vacía o inválida",
                        "Detalle": clean_value(decision.get("warning")) or "No se detectó guía de talla válida.",
                        "Fila": input_index + 2,
                    }
                )
                continue
            if current_guide and clean_value(current_guide) == proposed_guide:
                continue
            rows.append(
                {
                    "Accion": "Actualizar",
                    "Sitio": brand_config["site_label"],
                    "Operacion": "size_guides",
                    "Mod-Col": product_key,
                    "Product ID": product_id,
                    "Handle": product.get("Handle"),
                    "Campo": "Metafield: custom.guia_de_tallas",
                    "Valor actual": current_guide,
                    "Valor nuevo": proposed_guide,
                    "Metafield: custom.guia_de_tallas [page_reference]": proposed_guide,
                    "Estado": "OK" if clean_value(decision.get("status")) == "approved" else "PARCIAL",
                    "Observacion": clean_value(decision.get("warning")) or clean_value(decision.get("rule")),
                    "Marca": clean_value(row.get("Marca") or product.get("Vendor")),
                    "Categoria": category,
                    "Tipo": product_type,
                    "Genero": gender,
                }
            )
        elif operation == "technologies":
            from generate_columbia_matrixify import (
                format_technology_for_site,
                site_uses_technology_logo_metaobjects,
                technology_metafield_column,
            )

            technology_names, logo_refs = detect_product_technologies(row, product, brand_config)
            if not technology_names:
                issues.append(
                    {
                        "Mod-Col": product_key,
                        "Handle": product.get("Handle"),
                        "Problema": "No se detectaron tecnologias en input, titulo, descripcion, tags o metafields",
                        "Fila": input_index + 2,
                    }
                )
                continue
            technology_column = technology_metafield_column(brand_config)
            technology_type = "list.single_line_text_field" if "[list.single_line_text_field]" in technology_column else "single_line_text_field"
            uses_logo_metaobjects = site_uses_technology_logo_metaobjects(brand_config)
            technology_value = format_technology_for_site(" | ".join(technology_names), brand_config)
            logo_value = (
                ", ".join(dict.fromkeys(clean_value(value) for value in logo_refs if clean_value(value)))
                if uses_logo_metaobjects
                else ""
            )
            current_technology = clean_value(
                product.get(technology_column)
                or product.get("Metafield: custom.tecnologia [list.single_line_text_field]")
                or product.get("Metafield: custom.tecnologia [single_line_text_field]")
            )
            current_logo = clean_value(product.get("Metafield: custom.logo [list.metaobject_reference]")) if uses_logo_metaobjects else ""
            logo_resolved, logo_missing = (
                _validate_logo_metaobject_refs(shopify_config, logo_value)
                if uses_logo_metaobjects
                else ([], [])
            )
            status = "OK"
            observation = f"{len(technology_names)} tecnologia(s)"
            if uses_logo_metaobjects:
                observation += f", {len(_split_tags(logo_value))} logo(s)"
            if logo_missing:
                status = "PARCIAL"
                observation = f"{observation}. Faltan logo/metaobject: {', '.join(logo_missing)}"
                issues.append(
                    {
                        "Mod-Col": product_key,
                        "Handle": product.get("Handle"),
                        "Problema": "Tecnologia detectada, pero faltan logos/metaobjects en Shopify",
                        "Detalle": ", ".join(logo_missing),
                        "Fila": input_index + 2,
                    }
                )
            elif logo_value and shopify_config:
                observation = f"{observation}. Logos validados: {len(logo_resolved)}"
            preview_row = {
                    "Accion": "Actualizar",
                    "Sitio": brand_config["site_label"],
                    "Operacion": "technologies",
                    "Mod-Col": product_key,
                    "Product ID": product_id,
                    "Handle": product.get("Handle"),
                    "Campo": "Metafield: custom.tecnologia + Metafield: custom.logo" if uses_logo_metaobjects else "Metafield: custom.tecnologia",
                    "Valor actual": f"tecnologia: {current_technology} | logo: {current_logo}" if uses_logo_metaobjects else current_technology,
                    "Valor nuevo": technology_value,
                    "Tipo tecnologia": technology_type,
                    "Estado": status,
                    "Observacion": observation,
                }
            preview_row[technology_column] = technology_value
            if uses_logo_metaobjects:
                preview_row.update(
                    {
                        "Valor nuevo logos": logo_value,
                        "Logos validados": ", ".join(logo_resolved),
                        "Logos faltantes": ", ".join(logo_missing),
                        "Metafield: custom.logo [list.metaobject_reference]": logo_value,
                    }
                )
            rows.append(preview_row)
        elif operation == "photos":
            current_images = clean_value(product.get("Image Src"))
            if only_missing_images and current_images:
                continue
            from generate_columbia_matrixify import brand_image_config, image_candidates

            row_brand_config = brand_image_config(row.get("Marca") or product.get("Vendor"), brand_config)
            urls = image_candidates(product_key, row_brand_config)
            urls_text = "; ".join(urls)
            rows.append(
                {
                    "Accion": "Sincronizar Shopify" if image_mode == "replace" else "Agregar fotos Shopify",
                    "Sitio": brand_config["site_label"],
                    "Operacion": "photos",
                    "Mod-Col": product_key,
                    "Product ID": product_id,
                    "Handle": product.get("Handle"),
                    "Campo": "Fotos",
                    "Valor actual": current_images,
                    "Valor nuevo": urls_text,
                    "Modo fotos": image_mode,
                    "Media IDs": product.get("Media IDs"),
                    "Estado": "OK",
                    "Observacion": "REPLACE elimina fotos actuales antes de subir las 10 URLs nuevas; MERGE agrega las URLs nuevas.",
                }
            )
            matrixify_rows.append(
                {
                    "ID": product.get("Legacy ID"),
                    "Handle": product.get("Handle"),
                    "Command": "MERGE",
                    "Image Src": urls_text,
                    "Image Command": "REPLACE" if image_mode == "replace" else "MERGE",
                    "Image Position": "",
                    "Image Alt Text": product.get("Title"),
                }
            )

    return pd.DataFrame(rows), pd.DataFrame(issues), pd.DataFrame(matrixify_rows)


def _sync_result_summary(result_df):
    if result_df is None or result_df.empty or "Resultado" not in result_df.columns:
        return {"total": 0, "ok": 0, "partial": 0, "errors": 0, "skipped": 0}
    result = result_df["Resultado"].map(lambda value: clean_value(value).upper())
    return {
        "total": int(len(result_df)),
        "ok": int((result == "OK").sum()),
        "partial": int((result == "PARCIAL").sum()),
        "errors": int((result == "ERROR").sum()),
        "skipped": int((result == "OMITIDO").sum()),
    }


def render_sync_result_summary(result_df, label="sincronizacion"):
    summary = _sync_result_summary(result_df)
    if not summary["total"]:
        st.warning(f"No se genero resultado de {label}.")
        return
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Procesados", f"{summary['total']:,}")
    c2.metric("Sin observaciones", f"{summary['ok']:,}")
    c3.metric("Creados con observacion", f"{summary['partial']:,}")
    c4.metric("Errores", f"{summary['errors']:,}")
    if summary["errors"]:
        st.error(f"{summary['errors']:,} filas terminaron con error. Revisa el reporte de sincronizacion.")
    elif summary["partial"] or summary["skipped"]:
        st.info(
            "La sincronizacion termino con observaciones: los productos se crearon/actualizaron, "
            "pero hay datos por revisar en el reporte."
        )
    else:
        st.success("Sincronizacion finalizada correctamente.")


def sync_error_result(stage, exc):
    return pd.DataFrame(
        [
            {
                "Handle": "",
                "ID": "",
                "Resultado": "ERROR",
                "Etapa": stage,
                "Mensaje": clean_value(exc),
            }
        ]
    )


def make_sync_progress_callback(label="Sincronizacion"):
    started_at = time.perf_counter()
    progress_bar = st.progress(0)
    status_box = st.empty()
    log_box = st.empty()
    events = []

    def progress_callback(current, total, handle, stage, message=""):
        total = max(int(total or 0), 1)
        current = max(0, min(int(current or 0), total))
        percent = min(1.0, current / total)
        elapsed = round(time.perf_counter() - started_at, 1)
        progress_bar.progress(percent)
        status_box.info(
            f"{label}: {current:,}/{total:,} | {stage}"
            + (f" | {handle}" if clean_value(handle) else "")
            + f" | {elapsed}s"
        )
        events.append(
            {
                "N": len(events) + 1,
                "Producto": clean_value(handle),
                "Etapa": clean_value(stage),
                "Detalle": clean_value(message)[:260],
                "Segundos": elapsed,
            }
        )
        log_box.dataframe(pd.DataFrame(events[-12:]), use_container_width=True, height=220)

    return progress_callback


def apply_shopify_preview(shopify_config, preview_df, progress_callback=None):
    results = []
    total_rows = len(preview_df) if preview_df is not None else 0
    for position, (_, row) in enumerate(preview_df.iterrows(), start=1):
        status = "OK"
        message = ""
        handle = clean_value(row.get("Handle"))
        operation = clean_value(row.get("Operacion"))
        if progress_callback:
            progress_callback(position, total_rows, handle, f"Aplicando {operation or 'cambio'}")
        try:
            product_id = clean_value(row.get("Product ID"))
            if operation == "tags":
                product_update(shopify_config, product_id, tags=_split_tags(row.get("Valor nuevo")))
            elif operation == "title":
                product_update(shopify_config, product_id, title=clean_value(row.get("Valor nuevo")))
            elif operation == "body":
                product_update(shopify_config, product_id, body_html=clean_value(row.get("Valor nuevo")))
            elif operation == "size_guides":
                status = "OMITIDO"
                message = (
                    "custom.guia_de_tallas es page_reference; la API requiere gid://shopify/Page/... "
                    "y se mantiene para descarga Matrixify/validación segura."
                )
            elif operation == "photos":
                image_urls = _split_semicolon_values(row.get("Valor nuevo"))
                media_ids = _split_semicolon_values(row.get("Media IDs"))
                image_mode = clean_value(row.get("Modo fotos")).lower() or "replace"
                message = _sync_product_photos_direct(
                    shopify_config,
                    product_id,
                    image_urls,
                    existing_media_ids=media_ids,
                    image_mode=image_mode,
                    alt_text=clean_value(row.get("Handle")) or clean_value(row.get("Mod-Col")),
                )
            elif operation == "technologies":
                from generate_columbia_matrixify import (
                    site_uses_technology_logo_metaobjects,
                    technology_metafield_column,
                )

                technology_column = technology_metafield_column(brand_config)
                technology_type = clean_value(row.get("Tipo tecnologia")) or (
                    "list.single_line_text_field"
                    if "[list.single_line_text_field]" in technology_column
                    else "single_line_text_field"
                )
                technology_value = clean_value(
                    row.get("Valor nuevo")
                    or row.get(technology_column)
                    or row.get("Metafield: custom.tecnologia [list.single_line_text_field]")
                    or row.get("Metafield: custom.tecnologia [single_line_text_field]")
                )
                uses_logo_metaobjects = site_uses_technology_logo_metaobjects(brand_config)
                logo_value = (
                    clean_value(
                        row.get("Valor nuevo logos")
                        or row.get("Metafield: custom.logo [list.metaobject_reference]")
                    )
                    if uses_logo_metaobjects
                    else ""
                )
                updated_parts = []
                failed_parts = []
                if technology_value:
                    try:
                        metafields_set(
                            shopify_config,
                            [
                                {
                                    "ownerId": product_id,
                                    "namespace": "custom",
                                    "key": "tecnologia",
                                    "type": technology_type,
                                    "value": (
                                        _list_text_metafield_value(technology_value)
                                        if technology_type == "list.single_line_text_field"
                                        else technology_value
                                    ),
                                }
                            ],
                        )
                        updated_parts.append("tecnologia")
                    except Exception as exc:
                        failed_parts.append(f"tecnologia: {exc}")
                if logo_value:
                    try:
                        metafields_set(
                            shopify_config,
                            [
                                {
                                    "ownerId": product_id,
                                    "namespace": "custom",
                                    "key": "logo",
                                    "type": "list.metaobject_reference",
                                    "value": _metafield_value_for_api(
                                        "Metafield: custom.logo [list.metaobject_reference]",
                                        logo_value,
                                        shopify_config,
                                    ),
                                }
                            ],
                        )
                        updated_parts.append("logo")
                    except Exception as exc:
                        failed_parts.append(f"logo: {exc}")
                if failed_parts:
                    status = "PARCIAL" if updated_parts else "ERROR"
                    message = " | ".join(failed_parts)
                else:
                    message = f"Actualizado: {', '.join(updated_parts)}" if updated_parts else "Sin tecnologias/logos para actualizar"
            elif operation == "siblings":
                sibling_value = clean_value(row.get("Valor nuevo"))
                custom_sibling_value = clean_value(
                    row.get("Valor nuevo custom") or row.get("Metafield: custom.siblings [list.product_reference]")
                )
                if not custom_sibling_value:
                    custom_sibling_value = "[]"
                metafields_set(
                    shopify_config,
                    [
                        {
                            "ownerId": product_id,
                            "namespace": "theme",
                            "key": "siblings",
                            "type": "list.single_line_text_field",
                            "value": _list_text_metafield_value(sibling_value),
                        },
                        {
                            "ownerId": product_id,
                            "namespace": "custom",
                            "key": "siblings",
                            "type": "list.product_reference",
                            "value": custom_sibling_value,
                        }
                    ],
                )
            else:
                status = "OMITIDO"
                message = "Operacion no habilitada para escritura directa"
        except Exception as exc:
            status = "ERROR"
            message = str(exc)
        results.append(
            {
                "Mod-Col": row.get("Mod-Col"),
                "Handle": row.get("Handle"),
                "Operacion": row.get("Operacion"),
                "Campo": row.get("Campo"),
                "Resultado": status,
                "Mensaje": message,
            }
        )
        if progress_callback:
            progress_callback(position, total_rows, handle, f"Resultado {status}", message)
    return pd.DataFrame(results)


def shopify_products_to_matrixify_df(shopify_products):
    default_columns = [
        "ID",
        "Handle",
        "Command",
        "Title",
        "Body HTML",
        "Vendor",
        "Type",
        "Tags",
        "Image Src",
        "Variant SKU",
        "Variant Barcode",
        "Variant Inventory Item ID",
        "Variant ID",
        "Variant Image",
        "Metafield: custom.codigo_modelo_color [id]",
        "Metafield: theme.siblings [single_line_text_field]",
        "Metafield: theme.siblings_color [single_line_text_field]",
        "Metafield: custom.siblings [single_line_text_field]",
        "Metafield: custom.siblings_color [single_line_text_field]",
    ]
    try:
        if Path(DEFAULT_MATRIXIFY_PATH).exists():
            default_columns = list(pd.read_excel(DEFAULT_MATRIXIFY_PATH, sheet_name="Products", nrows=0).columns)
    except Exception:
        pass
    for column in (
        "Metafield: custom.marca [single_line_text_field]",
        "Metafield: custom.materialidad [single_line_text_field]",
        "Metafield: custom.tecnologia [list.single_line_text_field]",
        "Metafield: custom.logo [list.metaobject_reference]",
        "Metafield: custom.color_forus [single_line_text_field]",
        "Metafield: custom.grupo_color [single_line_text_field]",
        "Metafield: custom.genero [single_line_text_field]",
        "Metafield: custom.tipo [single_line_text_field]",
        "Metafield: custom.categoria [single_line_text_field]",
        "Metafield: custom.sub_categoria [single_line_text_field]",
        "Metafield: custom.guia_de_tallas [page_reference]",
        "Metafield: custom.nombre_corto [single_line_text_field]",
        "Metafield: custom.descripcion_corta [single_line_text_field]",
        "Metafield: custom.pais_de_fabricacion [single_line_text_field]",
        "Metafield: theme.siblings [single_line_text_field]",
        "Metafield: theme.siblings_color [single_line_text_field]",
        "Metafield: custom.siblings [single_line_text_field]",
        "Metafield: custom.siblings_color [single_line_text_field]",
    ):
        if column not in default_columns:
            default_columns.append(column)

    rows = []
    for product in shopify_products:
        variants = product.get("Variants") or [{}]
        for index, variant in enumerate(variants):
            row = {column: "" for column in default_columns}
            row.update(
                {
                    "ID": product.get("Legacy ID"),
                    "Handle": product.get("Handle"),
                    "Command": "MERGE",
                    "Title": product.get("Title") if index == 0 else "",
                    "Body HTML": product.get("Body HTML") if index == 0 else "",
                    "Vendor": product.get("Vendor") if index == 0 else "",
                    "Type": product.get("Type") if index == 0 else "",
                    "Tags": product.get("Tags") if index == 0 else "",
                    "Image Src": product.get("Image Src") if index == 0 else "",
                    "Variant SKU": variant.get("Variant SKU", ""),
                    "Variant Barcode": variant.get("Variant Barcode", ""),
                    "Variant Price": variant.get("Variant Price", ""),
                    "Variant Inventory Item ID": variant.get("Variant Inventory Item ID", ""),
                    "Variant ID": variant.get("Variant ID", ""),
                    "Variant Image": variant.get("Variant Image", ""),
                    "Option1 Name": variant.get("Option1 Name", "") or "Talla",
                    "Option1 Value": variant.get("Option1 Value", ""),
                    "Option2 Name": variant.get("Option2 Name", ""),
                    "Option2 Value": variant.get("Option2 Value", ""),
                    "Metafield: custom.codigo_modelo_color [id]": product.get("Mod-Col") if index == 0 else "",
                    "Metafield: custom.materialidad [single_line_text_field]": (
                        product.get("Metafield: custom.materialidad [single_line_text_field]")
                    )
                    if index == 0
                    else "",
                    "Metafield: custom.tecnologia [list.single_line_text_field]": (
                        product.get("Metafield: custom.tecnologia [list.single_line_text_field]")
                    )
                    if index == 0
                    else "",
                    "Metafield: custom.logo [list.metaobject_reference]": (
                        product.get("Metafield: custom.logo [list.metaobject_reference]")
                    )
                    if index == 0
                    else "",
                    "Metafield: custom.guia_de_tallas [page_reference]": (
                        product.get("Metafield: custom.guia_de_tallas [page_reference]")
                        or product.get("Guia de tallas")
                        or product.get("Guía de tallas")
                    )
                    if index == 0
                    else "",
                    "Metafield: theme.siblings [single_line_text_field]": product.get("Siblings") if index == 0 else "",
                    "Metafield: theme.siblings_color [single_line_text_field]": product.get("Siblings Color") if index == 0 else "",
                    "Metafield: custom.siblings [single_line_text_field]": (
                        product.get("Custom Siblings") or product.get("Siblings")
                    )
                    if index == 0
                    else "",
                    "Metafield: custom.siblings_color [single_line_text_field]": (
                        product.get("Custom Siblings Color") or product.get("Siblings Color")
                    )
                    if index == 0
                    else "",
                }
            )
            rows.append(row)
    return pd.DataFrame(rows, columns=default_columns)


STOCK_QUERY_DEFAULT = """
WITH stock_base AS (
  SELECT
    fecha_corte,
    id_producto,
    conca || '-' || talla AS key_producto,
    codigo_tienda,
    CONCAT_TIENDA,
    stock_tiendas,
    stock_bodega
  FROM `forus-analitica-prod-datalake.bronze.stg_pe_central_stock_bi`
  WHERE fecha_corte = (
    SELECT MAX(fecha_corte)
    FROM `forus-analitica-prod-datalake.bronze.stg_pe_central_stock_bi`
  )
)
SELECT
  fecha_corte,
  id_producto,
  UPPER(TRIM(key_producto)) AS key_producto,
  codigo_tienda,
  CONCAT_TIENDA,
  COALESCE(stock_tiendas, 0) AS stock_tiendas,
  COALESCE(stock_bodega, 0) AS stock_bodega,
  COALESCE(stock_tiendas, 0) + COALESCE(stock_bodega, 0) AS stock_total
FROM stock_base
"""

STOCK_QUERY_SAFE = """
WITH stock_base AS (
  SELECT
    fecha_corte,
    id_producto,
    conca || '-' || talla AS key_producto,
    codigo_tienda,
    stock_tiendas,
    stock_bodega
  FROM `forus-analitica-prod-datalake.bronze.stg_pe_central_stock_bi`
  WHERE fecha_corte = (
    SELECT MAX(fecha_corte)
    FROM `forus-analitica-prod-datalake.bronze.stg_pe_central_stock_bi`
  )
)
SELECT
  fecha_corte,
  id_producto,
  UPPER(TRIM(key_producto)) AS key_producto,
  codigo_tienda,
  COALESCE(stock_tiendas, 0) AS stock_tiendas,
  COALESCE(stock_bodega, 0) AS stock_bodega,
  COALESCE(stock_tiendas, 0) + COALESCE(stock_bodega, 0) AS stock_total
FROM stock_base
"""


def read_current_stock_from_bigquery(bigquery_config):
    try:
        from google.cloud import bigquery
        from google.oauth2 import service_account
    except ImportError as exc:
        raise RuntimeError("Faltan dependencias de BigQuery para leer stock.") from exc

    config = dict(bigquery_config or {})
    credentials_info = config.get("service_account_info")
    credentials = None
    project_id = clean_value(config.get("project_id"))
    if credentials_info:
        credentials = service_account.Credentials.from_service_account_info(dict(credentials_info))
        project_id = project_id or credentials.project_id

    job_project_id = clean_value(config.get("job_project_id")) or project_id
    client = bigquery.Client(project=job_project_id or None, credentials=credentials)
    job_config = bigquery.QueryJobConfig(use_legacy_sql=False)
    location = clean_value(config.get("location")) or None
    configured_query = clean_value(config.get("stock_query"))
    query = configured_query or STOCK_QUERY_DEFAULT

    def run_query(query_text):
        return client.query(query_text, job_config=job_config, location=location).to_dataframe()

    try:
        df = run_query(query)
    except Exception:
        if query.strip() == STOCK_QUERY_SAFE.strip():
            raise
        df = run_query(STOCK_QUERY_SAFE)
    if df.empty and configured_query:
        try:
            df = run_query(STOCK_QUERY_DEFAULT)
        except Exception:
            df = run_query(STOCK_QUERY_SAFE)
    df = standardize_stock_columns(df)
    for column in ("fecha_corte", "id_producto", "key_producto", "codigo_tienda", "CONCAT_TIENDA", "stock_tiendas", "stock_bodega", "stock_total"):
        if column not in df.columns:
            df[column] = 0 if column.startswith("stock_") else ""
    df = filter_latest_stock_cutoff(df)
    df["key_producto"] = df["key_producto"].map(lambda value: clean_value(value).upper())
    df["codigo_tienda"] = df["codigo_tienda"].map(normalize_warehouse_code)
    for column in ("stock_tiendas", "stock_bodega", "stock_total"):
        df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0)
    return df


def normalize_site_for_stock(value):
    return re.sub(r"[^a-z0-9]+", "", clean_value(value).lower())


def normalize_warehouse_code(value):
    text = clean_value(value)
    if not text:
        return ""
    try:
        numeric = float(text)
        if numeric.is_integer():
            code = int(numeric)
            return str(code) if 1 <= code <= 400 else ""
    except ValueError:
        pass
    tokens = re.findall(r"\d+", text)
    valid_codes = [int(token) for token in tokens if token.isdigit() and 1 <= int(token) <= 400]
    return str(valid_codes[0]) if valid_codes else ""


def store_code_from_concat_tienda(value):
    text = clean_value(value)
    if not text:
        return ""
    numbers = [int(token) for token in re.findall(r"\d+", text)]
    if len(numbers) >= 2:
        store = numbers[-1]
        return str(store) if 1 <= store <= 400 else ""
    return normalize_warehouse_code(text)


def stock_units_from_concat_tienda(value, store_code):
    text = clean_value(value)
    store = normalize_warehouse_code(store_code)
    if not text:
        return 0
    numbers = [int(token) for token in re.findall(r"\d+", text)]
    if not numbers:
        return 0
    if len(numbers) >= 2:
        if store and str(numbers[-1]) == store:
            return numbers[0]
        if store and str(numbers[0]) == store:
            return numbers[-1]
        if not store and 1 <= numbers[-1] <= 400:
            return numbers[0]
    if not store:
        return 0
    candidates = [number for number in numbers if str(number) != store]
    if not candidates:
        return 0
    return max(candidates)


def ecomm_stock_units_series(stock, store_column):
    if stock.empty:
        return pd.Series(dtype=float)
    store_codes = stock.get(store_column, pd.Series("", index=stock.index)).map(normalize_warehouse_code)
    stock_tiendas = numeric_stock_series(stock, "stock_tiendas")
    stock_bodega = numeric_stock_series(stock, "stock_bodega")
    stock_total = numeric_stock_series(stock, "stock_total")
    units = stock_tiendas.where(store_codes.ne("320"), stock_tiendas + stock_bodega)
    units = units.where(~((units <= 0) & (stock_total > 0)), stock_total)
    if "CONCAT_TIENDA" in stock.columns:
        fallback_mask = units <= 0
        if fallback_mask.any():
            fallback = stock.loc[fallback_mask].apply(
                lambda row: stock_units_from_concat_tienda(row.get("CONCAT_TIENDA"), normalize_warehouse_code(row.get(store_column))),
                axis=1,
            )
            units.loc[fallback_mask] = fallback
    return units.clip(lower=0)


def numeric_stock_series(df, column, default=0):
    values = df[column] if column in df.columns else pd.Series(default, index=df.index)
    return pd.to_numeric(values, errors="coerce").fillna(default)


def filter_latest_stock_cutoff(df):
    result = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    if result.empty or "fecha_corte" not in result.columns:
        return result
    before_rows = len(result)
    cutoff_dates = pd.to_datetime(result["fecha_corte"], errors="coerce")
    if cutoff_dates.notna().any():
        latest = cutoff_dates.max()
        filtered = result.loc[cutoff_dates.eq(latest)].copy()
        filtered.attrs["stock_rows_before_cutoff"] = before_rows
        filtered.attrs["stock_latest_cutoff"] = latest.isoformat()
        return filtered
    cutoff_text = result["fecha_corte"].map(clean_value)
    cutoff_text = cutoff_text[cutoff_text != ""]
    if cutoff_text.empty:
        return result
    latest_text = cutoff_text.max()
    filtered = result.loc[result["fecha_corte"].map(clean_value).eq(latest_text)].copy()
    filtered.attrs["stock_rows_before_cutoff"] = before_rows
    filtered.attrs["stock_latest_cutoff"] = latest_text
    return filtered


def standardize_stock_columns(df):
    result = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    rename_map = {}
    candidates_by_target = {
        "fecha_corte": ["fecha_corte", "fecha corte", "fecha", "cutoff", "fecha_stock", "fecha stock"],
        "id_producto": ["id_producto", "id producto", "producto_id", "product_id"],
        "key_producto": ["key_producto", "key producto"],
        "codigo_tienda": ["codigo_tienda", "codigo tienda", "cod tda", "cod_tda", "codtda", "codigo_tda", "codigo tda"],
        "CONCAT_TIENDA": ["CONCAT_TIENDA", "concat_tienda", "concat tienda", "CONCAT TIENDA"],
        "stock_tiendas": ["stock_tiendas", "stock tiendas"],
        "stock_bodega": ["stock_bodega", "stock bodega"],
        "stock_total": ["stock_total", "stock total"],
    }
    for target, candidates in candidates_by_target.items():
        found = first_existing_column(result, candidates)
        if found and found != target:
            rename_map[found] = target
    if rename_map:
        result = result.rename(columns=rename_map)
    if "codigo_tienda" in result.columns:
        result["codigo_tienda"] = result["codigo_tienda"].map(normalize_warehouse_code)
    if "CONCAT_TIENDA" in result.columns:
        if "codigo_tienda" not in result.columns:
            result["codigo_tienda"] = ""
        missing_store = result["codigo_tienda"].map(clean_value) == ""
        result.loc[missing_store, "codigo_tienda"] = result.loc[missing_store, "CONCAT_TIENDA"].map(
            store_code_from_concat_tienda
        )
    return result


def default_ecomm_stock_rules():
    rows = []
    for site_norm, codes in DEFAULT_ECOMM_SITE_WAREHOUSES.items():
        for code in codes:
            rows.append(
                {
                    "bodega_code": code,
                    "site_norm": site_norm,
                    "stock_seguridad": DEFAULT_ECOMM_STOCK_SECURITY.get(code, 0),
                    "stock_activo": 1,
                    "bodega_nombre": f"Bodega {code}",
                }
            )
    return pd.DataFrame(rows)


@st.cache_data(show_spinner=False)
def load_ecomm_stock_rules(path_text):
    path = Path(path_text)
    if not path.exists():
        return default_ecomm_stock_rules()
    try:
        assigned = pd.read_excel(path, sheet_name="Bodegas Asginadas", dtype=object).dropna(how="all")
        warehouses = pd.read_excel(path, sheet_name="Bodegas", dtype=object).dropna(how="all")
    except Exception:
        return default_ecomm_stock_rules()
    assigned.columns = [clean_value(column).lower() for column in assigned.columns]
    warehouses.columns = [clean_value(column).lower() for column in warehouses.columns]
    assigned["bodega_code"] = assigned.get("bodega", pd.Series(dtype=object)).map(normalize_warehouse_code)
    assigned["site_norm"] = assigned.get("sitio", pd.Series(dtype=object)).map(normalize_site_for_stock)
    assigned["stock_activo"] = pd.to_numeric(assigned.get("on/off", 0), errors="coerce").fillna(0).astype(int)
    warehouses["bodega_code"] = warehouses.get("numbodega", pd.Series(dtype=object)).map(normalize_warehouse_code)
    warehouses["stock_seguridad"] = pd.to_numeric(
        warehouses.get("stock_seguridad", 0), errors="coerce"
    ).fillna(0)
    warehouses["warehouse_estado"] = pd.to_numeric(warehouses.get("estado", 1), errors="coerce").fillna(1).astype(int)
    warehouses["bodega_nombre"] = warehouses.get("nombrebodega", "")
    rules = assigned.merge(
        warehouses[["bodega_code", "stock_seguridad", "warehouse_estado", "bodega_nombre"]],
        how="left",
        on="bodega_code",
    )
    rules["stock_seguridad"] = pd.to_numeric(rules["stock_seguridad"], errors="coerce").fillna(0)
    rules["warehouse_estado"] = pd.to_numeric(rules["warehouse_estado"], errors="coerce").fillna(1).astype(int)
    return rules[
        (rules["bodega_code"] != "")
        & (rules["site_norm"] != "")
        & (rules["stock_activo"] == 1)
        & (rules["warehouse_estado"] == 1)
    ].copy()


def ecomm_stock_rule_codes_for_site(brand_config):
    rules = load_ecomm_stock_rules(str(DEFAULT_ECOMM_WAREHOUSES_PATH))
    site_norm = normalize_site_for_stock(brand_config.get("site_label"))
    if rules.empty:
        return []
    site_rules = rules[rules["site_norm"] == site_norm].copy()
    return sorted(site_rules["bodega_code"].map(clean_value).loc[lambda series: series.ne("")].unique().tolist(), key=safe_int_value)


def apply_ecomm_stock_rules(stock_df, brand_config):
    empty_filtered = pd.DataFrame(
        columns=["fecha_corte", "id_producto", "key_producto", "stock_tiendas", "stock_bodega", "stock_total"]
    )
    stock = standardize_stock_columns(stock_df)
    if stock.empty:
        return stock
    for column in ("fecha_corte", "id_producto", "key_producto", "stock_tiendas", "stock_bodega", "stock_total"):
        if column not in stock.columns:
            stock[column] = 0 if column.startswith("stock_") else ""
    if "codigo_tienda" not in stock.columns or not stock["codigo_tienda"].map(clean_value).any():
        return empty_filtered

    rules = load_ecomm_stock_rules(str(DEFAULT_ECOMM_WAREHOUSES_PATH))
    site_norm = normalize_site_for_stock(brand_config.get("site_label"))
    site_rules = rules[rules["site_norm"] == site_norm].copy() if not rules.empty else pd.DataFrame()
    if site_rules.empty:
        return empty_filtered

    stock["codigo_tienda_norm"] = stock["codigo_tienda"].map(normalize_warehouse_code)
    stock = stock.merge(
        site_rules[["bodega_code", "stock_seguridad"]],
        how="inner",
        left_on="codigo_tienda_norm",
        right_on="bodega_code",
    )
    if stock.empty:
        return empty_filtered

    for column in ("stock_tiendas", "stock_bodega", "stock_total", "stock_seguridad"):
        stock[column] = numeric_stock_series(stock, column)
    stock["stock_bruto_ecomm"] = ecomm_stock_units_series(stock, "codigo_tienda_norm")
    stock["stock_total"] = (stock["stock_bruto_ecomm"] - stock["stock_seguridad"]).clip(lower=0)
    stock["stock_seguridad_aplicado"] = stock["stock_bruto_ecomm"] - stock["stock_total"]

    grouped = (
        stock.groupby(["fecha_corte", "key_producto"], as_index=False)
        .agg(
            id_producto=("id_producto", "first"),
            stock_tiendas=("stock_tiendas", "sum"),
            stock_bodega=("stock_bodega", "sum"),
            stock_total=("stock_total", "sum"),
            stock_bruto_ecomm=("stock_bruto_ecomm", "sum"),
            stock_seguridad_aplicado=("stock_seguridad_aplicado", "sum"),
            bodegas_ecomm=("codigo_tienda_norm", "nunique"),
        )
    )
    return grouped


def build_ecomm_stock_match_summary(stock_df, brand_config):
    rules = load_ecomm_stock_rules(str(DEFAULT_ECOMM_WAREHOUSES_PATH))
    site_norm = normalize_site_for_stock(brand_config.get("site_label"))
    site_rules = rules[rules["site_norm"] == site_norm].copy() if not rules.empty else pd.DataFrame()
    if site_rules.empty:
        return pd.DataFrame(
            [
                {
                    "Bodega": "Sin reglas",
                    "Nombre": f"Archivo existe: {'Si' if DEFAULT_ECOMM_WAREHOUSES_PATH.exists() else 'No'} | sitio: {brand_config.get('site_label')}",
                    "Stock seguridad": 0,
                    "Aparece en query": "No",
                    "Filas query": 0,
                    "Stock bruto": 0,
                    "Stock efectivo": 0,
                }
            ]
        )

    stock = standardize_stock_columns(stock_df)
    if stock.empty or "codigo_tienda" not in stock.columns:
        stock_summary = pd.DataFrame(columns=["bodega_code", "Filas query", "Stock bruto", "Stock efectivo"])
    else:
        stock["bodega_code"] = stock["codigo_tienda"].map(normalize_warehouse_code)
        if "key_producto" not in stock.columns:
            stock["key_producto"] = ""
        for column in ("stock_tiendas", "stock_bodega", "stock_total"):
            stock[column] = numeric_stock_series(stock, column)
        stock["stock_bruto"] = ecomm_stock_units_series(stock, "bodega_code")
        stock = stock.merge(
            site_rules[["bodega_code", "stock_seguridad"]],
            how="left",
            on="bodega_code",
        )
        stock["stock_seguridad"] = pd.to_numeric(stock["stock_seguridad"], errors="coerce").fillna(0)
        stock["stock_efectivo"] = (stock["stock_bruto"] - stock["stock_seguridad"]).clip(lower=0)
        stock_summary = (
            stock.groupby("bodega_code", as_index=False)
            .agg(
                **{
                    "Filas query": ("key_producto", "count"),
                    "Stock bruto": ("stock_bruto", "sum"),
                    "Stock efectivo": ("stock_efectivo", "sum"),
                }
            )
        )

    summary = site_rules[["bodega_code", "bodega_nombre", "stock_seguridad"]].drop_duplicates("bodega_code").merge(
        stock_summary,
        how="left",
        on="bodega_code",
    )
    summary["Filas query"] = pd.to_numeric(summary["Filas query"], errors="coerce").fillna(0).astype(int)
    summary["Stock bruto"] = pd.to_numeric(summary["Stock bruto"], errors="coerce").fillna(0)
    summary["Stock efectivo"] = pd.to_numeric(summary["Stock efectivo"], errors="coerce").fillna(0)
    summary["Aparece en query"] = summary["Filas query"].map(lambda value: "Si" if safe_int_value(value) > 0 else "No")
    return summary.rename(
        columns={
            "bodega_code": "Bodega",
            "bodega_nombre": "Nombre",
            "stock_seguridad": "Stock seguridad",
        }
    )[["Bodega", "Nombre", "Stock seguridad", "Aparece en query", "Filas query", "Stock bruto", "Stock efectivo"]]


def stock_key_from_parts(mod_col, size):
    mod_col = clean_value(mod_col).upper()
    size = clean_value(normalize_size(size)).upper()
    return f"{mod_col}-{size}" if mod_col and size else ""


def filter_visible_kpi_sizes(df):
    if df is None or df.empty or "Mod-Col KPI" not in df.columns or "Talla KPI" not in df.columns:
        return df
    result = df.copy()
    result = result[~result["Talla KPI"].map(is_internal_k_size)].copy()
    keep_parts = []
    for _, group in result.groupby("Mod-Col KPI", dropna=False):
        has_one_size = group["Talla KPI"].map(is_one_size).any()
        if has_one_size:
            group = group[~group["Talla KPI"].map(is_zero_size)].copy()
        keep_parts.append(group)
    if not keep_parts:
        return result.iloc[0:0].copy()
    return pd.concat(keep_parts, ignore_index=True)


def valid_kpi_price(value):
    text = clean_value(value)
    if not text:
        return False
    try:
        return float(text.replace(",", ".")) > 0
    except ValueError:
        return False


def numeric_kpi_value(value):
    text = clean_value(value).replace(",", ".")
    if not text:
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def row_first_value(row, columns):
    normalized_lookup = {}
    try:
        normalized_lookup = {normalize_header(column): column for column in row.index}
    except Exception:
        normalized_lookup = {}
    for column in columns:
        source_column = column if column in row.index else normalized_lookup.get(normalize_header(column))
        if source_column is None:
            continue
        value = clean_value(row.get(source_column))
        if value:
            return repair_mojibake_text(value)
    return ""


def split_mod_col_code(mod_col):
    text = clean_value(mod_col).upper()
    if "-" not in text:
        return text, ""
    model, color = text.rsplit("-", 1)
    return model, color


def suggested_handle(title, mod_col, brand_label="", product_type="", gender=""):
    technical_handle = build_catalog_handle(product_type=product_type, gender=gender, brand=brand_label, mod_col=mod_col)
    if technical_handle:
        return technical_handle
    base = first_non_empty(title, mod_col)
    text = f"{base} {brand_label} {mod_col}".strip()
    text = unicodedata.normalize("NFKD", clean_value(text)).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").lower()
    return text or clean_value(mod_col).lower()


def pluralize_spanish_label(value):
    text = clean_value(value)
    if not text:
        return ""
    words = text.split()
    last = words[-1]
    lower = last.lower()
    if lower.endswith("s"):
        return text
    if lower.endswith("z"):
        words[-1] = last[:-1] + ("ces" if last.islower() else "CES")
    elif lower[-1:] in "aeiouáéíóú":
        words[-1] = last + "s"
    else:
        words[-1] = last + "es"
    return " ".join(words)


def compact_join_values(values, separator=", "):
    cleaned = [clean_value(value) for value in values if clean_value(value)]
    return separator.join(dict.fromkeys(cleaned))


def build_missing_models_input_export(expected, missing_models, brand_config):
    if expected is None or expected.empty or missing_models is None or missing_models.empty:
        return (
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame(columns=["Campo", "Responsable", "Descripcion"]),
        )

    missing_keys = {
        clean_value(value).upper()
        for value in missing_models.get("Mod-Col KPI", pd.Series(dtype=object))
        if clean_value(value)
    }
    source = expected[expected["Mod-Col KPI"].map(lambda value: clean_value(value).upper() in missing_keys)].copy()
    if source.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    product_rows = []
    variant_rows = []
    missing_field_rows = []
    brand_label = clean_value(brand_config.get("label"))

    for mod_col, group in source.groupby("Mod-Col KPI", sort=False):
        group = group.copy()
        group = group[group["SKU"].map(clean_value) != ""].copy() if "SKU" in group.columns else group
        if group.empty:
            continue
        group = group.drop_duplicates(subset=["Talla KPI", "SKU"], keep="first")
        first_row = group.iloc[0]
        model_code, color_code = split_mod_col_code(mod_col)
        vendor = brand_display_name(
            first_non_empty(row_first_value(first_row, ["MARCA_MA", "Marca", "Vendor"]), brand_label),
            brand_label,
        )
        title = row_first_value(
            first_row,
            [
                "Title",
                "NombreModelo",
                "Nombre del Producto",
                "Nombre Producto",
                "NOMBRE_PRODUCTO",
                "Descripcion Producto",
                "Descripcion Modelo",
                "DESCRIPCION_MA",
                "Descripcion",
                "MODELO",
            ],
        )
        product_type = row_first_value(
            first_row,
            ["TipoProducto", "Type", "Tipo De Producto", "Tipo de Producto", "TIPO", "TIPO_MA", "Tipo", "Categoria", "CATEGORIA"],
        )
        product_type_plural = pluralize_spanish_label(product_type)
        category = row_first_value(first_row, ["Categoria", "Categoría", "CATEGORIA", "Familia", "FAMILIA"])
        subcategory = row_first_value(first_row, ["SubCategoria", "Sub Categoria", "Sub Categoría", "SUBCATEGORIA", "SUB CATEGORIA"])
        gender = row_first_value(first_row, ["Genero", "Género", "GENERO", "Sexo", "SEXO"])
        season = row_first_value(first_row, ["Temporada", "TEMPORADA", "Season"])
        collection = row_first_value(first_row, ["Coleccion", "Colección", "COLECCION", "Collection"])
        occasion = row_first_value(first_row, ["Ocasion", "Ocasión", "OCASION", "Occasion"])
        sport = row_first_value(first_row, ["Deporte", "DEPORTE", "Sport", "Activity"])
        color_name = row_first_value(
            first_row,
            ["ColorNombre", "Color Forus", "Color Web", "COLOR_WEB", "Color", "COLOR", "DESC_COLOR", "COLOR_MA"],
        )
        features = row_first_value(first_row, ["Caracteristicas", "Características", "Features", "Beneficios"])
        composition = row_first_value(
            first_row,
            ["Material", "Composicion", "Composición", "COMPOSICION", "MATERIAL", "Materialidad"],
        )
        technology = row_first_value(
            first_row,
            ["Tecnologia", "Tecnología", "TECNOLOGIA", "TECNOLOGÍA", "Technology", "TECHNOLOGY"],
        )
        care = row_first_value(first_row, ["Cuidado", "CUIDADO", "Cuidados", "Care"])
        body_html = row_first_value(first_row, ["Body HTML", "DescripcionWeb", "Descripcion Web", "Descripción Web", "DESCRIPCION_WEB"])
        if not body_html:
            try:
                body_html = build_matrixify_body_html(first_row)
            except Exception:
                body_html = ""
        image_src = row_first_value(first_row, ["Imagen", "Image Src", "IMAGEN", "Foto", "FOTO"])
        image_folder = ""
        if not image_src:
            try:
                image_config = brand_image_config(vendor, brand_config)
                image_urls = image_candidates(mod_col, image_config)
                image_src = "; ".join(image_urls)
                image_folder = clean_value((image_config or {}).get("folder"))
            except Exception:
                image_src = ""
        title_suggested = first_non_empty(
            title,
            compact_join_values([vendor, product_type, gender, color_name], " "),
            mod_col,
        )
        tags_parts = [
            vendor,
            brand_label,
            category,
            subcategory,
            gender,
            product_type_plural or product_type,
            color_name,
            season,
            collection,
            occasion,
            sport,
            technology,
            mod_col,
        ]
        tags = compact_join_values(tags_parts)
        valid_sizes = list(dict.fromkeys(group.get("Talla KPI", pd.Series(dtype=object)).map(clean_value).tolist()))
        skus = list(dict.fromkeys(group.get("SKU", pd.Series(dtype=object)).map(clean_value).tolist()))
        stock_total = safe_int_value(pd.to_numeric(group.get("stock_total", 0), errors="coerce").fillna(0).sum())
        price = row_first_value(first_row, ["Precio", "PRECIO", "Variant Price", "Price"])
        compare_at = row_first_value(first_row, ["Compare At Price", "Precio Compare At", "PRECIO_ANTES"])

        missing_notes = []
        if not title:
            missing_notes.append("Completar titulo comercial")
        if not body_html:
            missing_notes.append("Completar descripcion/body HTML")
        if not image_src:
            missing_notes.append("Validar fotos")
        if not price:
            missing_notes.append("Validar precio")
        if not composition:
            missing_notes.append("Completar composicion/material")
        if not care:
            missing_notes.append("Completar cuidados")
        validation_row = {
            "Mod-Col": mod_col,
            "Marca": vendor,
            "Genero": gender,
            "Categoria": category,
            "Sub Categoria": subcategory,
            "Tipo de prenda": product_type_plural or product_type,
            "Color web": color_name or color_code,
            "Title": title_suggested,
            "Body HTML": body_html,
            "Talla": first_non_empty(*(valid_sizes or [""])),
            "SKU": first_non_empty(*(skus or [""])),
            "Precio": price,
            "Guia de tallas": row_first_value(first_row, ["Guia de tallas", "Guía de tallas", "Size Guide"]),
        }
        validation_result = validate_catalog_row(validation_row)
        size_decision = validation_result.get("size_guide_decision") or {}
        validation_issues = validation_result.get("issues") or []
        if validation_issues:
            missing_notes.extend(
                [
                    f"{clean_value(issue.get('level')).upper()}: {clean_value(issue.get('field'))} - {clean_value(issue.get('message'))}"
                    for issue in validation_issues
                    if clean_value(issue.get("message"))
                ]
            )

        product_rows.append(
            {
                "Mod-Col": mod_col,
                "Codigo modelo color": mod_col,
                "Codigo modelo": model_code,
                "Codigo color": color_code,
                "Nombre modelo ARTI": title,
                "Nombre web sugerido": title_suggested,
                "Handle sugerido": suggested_handle(
                    title_suggested,
                    mod_col,
                    vendor,
                    product_type_plural or product_type,
                    gender,
                ),
                "Title": title_suggested,
                "Body HTML": body_html,
                "Vendor": vendor,
                "Type": product_type_plural or product_type,
                "Tags sugeridos": tags,
                "Status recomendado": "ACTIVE",
                "Published recomendado": "TRUE",
                "Image Src": image_src,
                "Ruta fotos esperada": image_folder,
                "Marca": vendor,
                "Genero": gender,
                "Tipo de prenda": product_type_plural or product_type,
                "Categoria": category,
                "Sub Categoria": subcategory,
                "Color web": color_name or color_code,
                "Color": color_name or color_code,
                "Temporada": season,
                "Coleccion": collection,
                "Ocasion": occasion,
                "Deporte": sport,
                "Tallas validas": ", ".join([value for value in valid_sizes if value]),
                "Variantes a crear": len([value for value in valid_sizes if value]),
                "SKUs": ", ".join([value for value in skus if value]),
                "Precio": price,
                "Compare At Price": compare_at,
                "Stock disponible": stock_total,
                "Caracteristicas": features,
                "Composicion": composition,
                "Tecnologia": technology,
                "Cuidado": care,
                "Metafield: custom.codigo_modelo_color [id]": mod_col,
                "Metafield: custom.marca [single_line_text_field]": vendor,
                "Metafield: custom.tecnologia [list.single_line_text_field]": technology,
                "Metafield: custom.materialidad [single_line_text_field]": composition,
                "Metafield: custom.tipo [single_line_text_field]": product_type_plural or product_type,
                "Metafield: custom.genero [single_line_text_field]": gender,
                "Metafield: custom.color_forus [single_line_text_field]": color_name or color_code,
                "Guia de talla sugerida": clean_value(size_decision.get("guide")),
                "Regla guia talla": clean_value(size_decision.get("rule")),
                "Estado validacion": (
                    "BLOQUEADO"
                    if any(clean_value(issue.get("level")).lower() == "bloqueo" for issue in validation_issues)
                    else "ADVERTENCIA"
                    if validation_issues or clean_value(size_decision.get("status")).lower() == "warning"
                    else "APROBADO"
                ),
                "Campos que debe completar marca": "; ".join(missing_notes) if missing_notes else "Revisar y aprobar",
                "Observaciones": "Producto no creado en Shopify. Input sugerido desde ARTI/BigQuery.",
            }
        )
        for _, variant in group.iterrows():
            variant_rows.append(
                {
                    "Mod-Col": mod_col,
                    "Codigo modelo": model_code,
                    "Color": color_name or color_code,
                    "Talla": clean_value(variant.get("Talla KPI")),
                    "SKU": clean_value(variant.get("SKU")),
                    "EAN / Barcode": row_first_value(variant, ["CodBarras", "EAN", "Barcode", "CODBARRAS", "COD_BARRAS"]),
                    "Precio": row_first_value(variant, ["Precio", "PRECIO", "Variant Price", "Price"]),
                    "Stock disponible": safe_int_value(variant.get("stock_total")),
                    "Stock Key": clean_value(variant.get("Stock Key")),
                    "Accion": "Crear variante",
                }
            )
        for field, value in (
            ("Title", title),
            ("Body HTML", body_html),
            ("Image Src", image_src),
            ("Precio", price),
            ("Composicion", composition),
            ("Cuidado", care),
        ):
            if not clean_value(value):
                missing_field_rows.append(
                    {
                        "Mod-Col": mod_col,
                        "Campo": field,
                        "Responsable": "Marca / Brand Manager",
                        "Descripcion": "Completar antes de carga final",
                    }
                )

    required_fields = pd.DataFrame(
        [
            {"Campo": "Mod-Col", "Responsable": "Sistema", "Descripcion": "Codigo modelo-color fuente de verdad."},
            {"Campo": "SKU / Talla", "Responsable": "ARTI/BigQuery", "Descripcion": "Variantes validas; no se inventan tallas."},
            {"Campo": "Title / Body HTML / Fotos", "Responsable": "Marca", "Descripcion": "Campos comerciales a revisar."},
            {"Campo": "Metafields", "Responsable": "Sistema + Marca", "Descripcion": "Se precargan si existen datos fuente."},
        ]
        + missing_field_rows
    )
    return pd.DataFrame(product_rows), pd.DataFrame(variant_rows), required_fields


def flatten_shopify_for_kpis(shopify_products):
    product_rows = []
    variant_rows = []
    for product in shopify_products or []:
        mod_col = clean_value(product.get("Mod-Col")).upper()
        status = clean_value(product.get("Status")).upper()
        online_url = clean_value(product.get("Online Store URL"))
        published_field = clean_value(product.get("Published Online Store")).upper()
        if published_field:
            published_online = published_field in ("SI", "YES", "TRUE", "1", "PUBLISHED")
            published_source = "publishedOnPublication"
        else:
            published_online = bool(online_url)
            published_source = "onlineStoreUrl"
        visible_online = status == "ACTIVE" and published_online
        variants = product.get("Variants") or []
        has_price = any(valid_kpi_price(variant.get("Variant Price")) for variant in variants)
        product_rows.append(
            {
                "Mod-Col": mod_col,
                "Handle": clean_value(product.get("Handle")),
                "Title": clean_value(product.get("Title")),
                "Status": status,
                "Publicado": "SI" if published_online else "NO",
                "Publicado fuente": published_source,
                "Online Store URL": online_url,
                "Visible": visible_online,
                "Tiene precio": has_price,
                "Fotos": len([url for url in clean_value(product.get("Image Src")).split(";") if clean_value(url)]),
            }
        )
        for variant in variants:
            variant_rows.append(
                {
                    "Mod-Col": mod_col,
                    "Handle": clean_value(product.get("Handle")),
                    "Status": status,
                    "Publicado": "SI" if published_online else "NO",
                    "Publicado fuente": published_source,
                    "Online Store URL": online_url,
                    "Visible": visible_online,
                    "Variant SKU": clean_value(variant.get("Variant SKU")),
                    "Variant Price": clean_value(variant.get("Variant Price")),
                    "Variant Inventory Qty": numeric_kpi_value(variant.get("Variant Inventory Qty")),
                    "Variant Inventory Item ID": clean_value(variant.get("Variant Inventory Item ID")),
                    "Variant Inventory Item GID": clean_value(variant.get("Variant Inventory Item GID")),
                    "Tiene precio": valid_kpi_price(variant.get("Variant Price")),
                }
            )
    return pd.DataFrame(product_rows), pd.DataFrame(variant_rows)


def build_catalog_kpis(arti_df, stock_df, shopify_products, brand_config):
    arti = arti_df.copy() if isinstance(arti_df, pd.DataFrame) else pd.DataFrame()
    arti = normalize_arti_columns_for_app(arti)
    stock = stock_df.copy() if isinstance(stock_df, pd.DataFrame) else pd.DataFrame()
    allowed = set(brand_config.get("allowed_arti_brands") or [])
    if "MARCA_MA" in arti.columns and allowed:
        arti = arti[arti["MARCA_MA"].map(lambda value: clean_value(value).upper()).isin(allowed)].copy()
    if "Mod-Col" not in arti.columns:
        arti["Mod-Col"] = ""
    if "COD MOD COL" not in arti.columns:
        arti["COD MOD COL"] = ""
    for column in ("CODINT_MA", "TALNUM_MA", "MARCA_MA"):
        if column not in arti.columns:
            arti[column] = ""

    arti["Mod-Col KPI"] = arti["Mod-Col"].where(arti["Mod-Col"].map(clean_value) != "", arti["COD MOD COL"])
    arti["Mod-Col KPI"] = arti["Mod-Col KPI"].map(lambda value: clean_value(value).upper())
    arti["Talla KPI"] = arti["TALNUM_MA"].map(normalize_size)
    arti["Stock Key"] = arti.apply(lambda row: stock_key_from_parts(row.get("Mod-Col KPI"), row.get("Talla KPI")), axis=1)
    expected = arti[(arti["Mod-Col KPI"] != "") & (arti["Stock Key"] != "")].copy()
    expected = filter_visible_kpi_sizes(expected)

    if stock.empty:
        stock = pd.DataFrame(columns=["key_producto", "stock_tiendas", "stock_bodega", "stock_total", "fecha_corte"])
    stock["key_producto"] = stock["key_producto"].map(lambda value: clean_value(value).upper())
    ecomm_stock_match = build_ecomm_stock_match_summary(stock, brand_config)
    stock = apply_ecomm_stock_rules(stock, brand_config)
    stock_ecomm_rows = safe_int_value((pd.to_numeric(stock.get("stock_total", 0), errors="coerce").fillna(0) > 0).sum()) if not stock.empty else 0
    stock_ecomm_units = safe_int_value(pd.to_numeric(stock.get("stock_total", 0), errors="coerce").fillna(0).sum()) if not stock.empty else 0
    stock_ecomm_models = (
        stock.loc[pd.to_numeric(stock.get("stock_total", 0), errors="coerce").fillna(0) > 0, "key_producto"]
        .map(lambda value: clean_value(value).rsplit("-", 1)[0])
        .nunique()
        if not stock.empty and "key_producto" in stock.columns
        else 0
    )
    if stock.empty:
        stock = pd.DataFrame(columns=["key_producto", "stock_tiendas", "stock_bodega", "stock_total", "fecha_corte"])
    stock["key_producto"] = stock["key_producto"].map(lambda value: clean_value(value).upper())
    expected = expected.merge(
        stock[["key_producto", "stock_tiendas", "stock_bodega", "stock_total", "fecha_corte"]],
        how="left",
        left_on="Stock Key",
        right_on="key_producto",
    )
    for column in ("stock_tiendas", "stock_bodega", "stock_total"):
        expected[column] = pd.to_numeric(expected[column], errors="coerce").fillna(0)

    products_df, variants_df = flatten_shopify_for_kpis(shopify_products)
    if not products_df.empty and "Fotos" in products_df.columns:
        shopify_no_photo_all = products_df[pd.to_numeric(products_df["Fotos"], errors="coerce").fillna(0) <= 0].copy()
    else:
        shopify_no_photo_all = pd.DataFrame(columns=list(products_df.columns) if isinstance(products_df, pd.DataFrame) else [])
    shopify_model_keys = {clean_value(value).upper() for value in products_df.get("Mod-Col", pd.Series(dtype=object)) if clean_value(value)}
    shopify_variant_skus = {clean_value(value) for value in variants_df.get("Variant SKU", pd.Series(dtype=object)) if clean_value(value)}
    product_status_by_key = products_df.drop_duplicates("Mod-Col").set_index("Mod-Col").to_dict("index") if not products_df.empty and "Mod-Col" in products_df.columns else {}

    expected["SKU"] = expected["CODINT_MA"].map(clean_value)
    expected["Con stock"] = expected["stock_total"] > 0
    expected["Producto creado Shopify"] = expected["Mod-Col KPI"].map(lambda value: value in shopify_model_keys)
    expected["Variante creada Shopify"] = expected["SKU"].map(lambda value: value in shopify_variant_skus)
    expected["Status Shopify"] = expected["Mod-Col KPI"].map(lambda value: clean_value(product_status_by_key.get(value, {}).get("Status")))
    expected["Publicado Shopify"] = expected["Mod-Col KPI"].map(lambda value: clean_value(product_status_by_key.get(value, {}).get("Publicado")))
    expected["Publicado fuente Shopify"] = expected["Mod-Col KPI"].map(lambda value: clean_value(product_status_by_key.get(value, {}).get("Publicado fuente")))
    expected["URL Shopify"] = expected["Mod-Col KPI"].map(lambda value: clean_value(product_status_by_key.get(value, {}).get("Online Store URL")))
    expected["Visible Shopify"] = expected["Mod-Col KPI"].map(lambda value: bool(product_status_by_key.get(value, {}).get("Visible")))
    expected["Fotos Shopify"] = expected["Mod-Col KPI"].map(lambda value: int(product_status_by_key.get(value, {}).get("Fotos") or 0))
    variant_status_by_sku = (
        variants_df.drop_duplicates("Variant SKU").set_index("Variant SKU").to_dict("index")
        if not variants_df.empty and "Variant SKU" in variants_df.columns
        else {}
    )
    expected["Stock Shopify Variante"] = expected["SKU"].map(
        lambda value: numeric_kpi_value(variant_status_by_sku.get(clean_value(value), {}).get("Variant Inventory Qty"))
    )
    expected["Inventory Item GID"] = expected["SKU"].map(
        lambda value: clean_value(variant_status_by_sku.get(clean_value(value), {}).get("Variant Inventory Item GID"))
    )
    expected["Inventory Item ID"] = expected["SKU"].map(
        lambda value: clean_value(variant_status_by_sku.get(clean_value(value), {}).get("Variant Inventory Item ID"))
    )

    model_stock = (
        expected.groupby("Mod-Col KPI", as_index=False)
        .agg(
            Marca=("MARCA_MA", "first"),
            Stock_total=("stock_total", "sum"),
            Tallas_BigQuery=("Talla KPI", "nunique"),
            Tallas_con_stock=("Con stock", "sum"),
            Producto_creado=("Producto creado Shopify", "max"),
            Visible_Shopify=("Visible Shopify", "max"),
            Status_Shopify=("Status Shopify", "first"),
            Publicado_Shopify=("Publicado Shopify", "first"),
            Publicado_Fuente_Shopify=("Publicado fuente Shopify", "first"),
            URL_Shopify=("URL Shopify", "first"),
            Fotos_Shopify=("Fotos Shopify", "max"),
        )
    )
    model_stock["Debe estar visible"] = model_stock["Stock_total"] > 0
    model_stock["Estado"] = model_stock.apply(
        lambda row: (
            "OK visible con stock"
            if row["Debe estar visible"] and row["Visible_Shopify"]
            else "Con stock no visible"
            if row["Debe estar visible"] and not row["Visible_Shopify"]
            else "Sin stock visible"
            if not row["Debe estar visible"] and row["Visible_Shopify"]
            else "OK apagado sin stock"
        ),
        axis=1,
    )

    missing_models = model_stock[(model_stock["Debe estar visible"]) & (~model_stock["Producto_creado"])].copy()
    stock_not_visible = model_stock[
        (model_stock["Debe estar visible"]) & (model_stock["Producto_creado"]) & (~model_stock["Visible_Shopify"])
    ].copy()
    no_stock_visible = model_stock[(~model_stock["Debe estar visible"]) & (model_stock["Visible_Shopify"])].copy()
    missing_stock_variants = expected[(expected["Con stock"]) & (~expected["Variante creada Shopify"])].copy()
    models_with_missing_stock_variants = {
        clean_value(value)
        for value in missing_stock_variants.get("Mod-Col KPI", pd.Series(dtype=object)).dropna()
    }
    model_stock["Variantes_stock_incompletas"] = model_stock["Mod-Col KPI"].map(
        lambda value: clean_value(value) in models_with_missing_stock_variants
    )

    price_by_model = (
        variants_df.groupby("Mod-Col", as_index=False)["Tiene precio"].max()
        if not variants_df.empty and "Mod-Col" in variants_df.columns
        else pd.DataFrame(columns=["Mod-Col", "Tiene precio"])
    )
    shopify_stock_by_model = (
        variants_df.groupby("Mod-Col", as_index=False)["Variant Inventory Qty"].sum()
        if not variants_df.empty and "Mod-Col" in variants_df.columns and "Variant Inventory Qty" in variants_df.columns
        else pd.DataFrame(columns=["Mod-Col", "Variant Inventory Qty"])
    )
    model_stock = model_stock.merge(
        shopify_stock_by_model.rename(columns={"Variant Inventory Qty": "Stock_Shopify"}),
        how="left",
        left_on="Mod-Col KPI",
        right_on="Mod-Col",
    )
    model_stock["Stock_Shopify"] = pd.to_numeric(model_stock["Stock_Shopify"], errors="coerce").fillna(0)
    if "Mod-Col" in model_stock.columns:
        model_stock = model_stock.drop(columns=["Mod-Col"])

    no_price_models = model_stock[model_stock["Producto_creado"] & model_stock["Debe estar visible"]].merge(
        price_by_model,
        how="left",
        left_on="Mod-Col KPI",
        right_on="Mod-Col",
    )
    no_price_models["Tiene precio"] = no_price_models["Tiene precio"].fillna(False)
    no_price_models = no_price_models[~no_price_models["Tiene precio"]].copy()
    no_photo_models = model_stock[
        model_stock["Debe estar visible"] & model_stock["Producto_creado"] & (model_stock["Fotos_Shopify"] <= 0)
    ].copy()
    no_shopify_stock_models = model_stock[
        model_stock["Debe estar visible"] & model_stock["Producto_creado"] & (model_stock["Stock_Shopify"] <= 0)
    ].copy()
    stock_location_activation_audit = expected[
        (expected["Con stock"])
        & (expected["Producto creado Shopify"])
        & (expected["Variante creada Shopify"])
        & (expected["Stock Shopify Variante"] <= 0)
    ].copy()
    if not stock_location_activation_audit.empty:
        stock_location_activation_audit = stock_location_activation_audit[
            [
                "Mod-Col KPI",
                "MARCA_MA",
                "Talla KPI",
                "SKU",
                "stock_total",
                "Stock Shopify Variante",
                "Inventory Item GID",
                "Inventory Item ID",
                "Status Shopify",
                "Publicado Shopify",
                "URL Shopify",
            ]
        ].rename(
            columns={
                "Mod-Col KPI": "Mod-Col",
                "MARCA_MA": "Marca",
                "Talla KPI": "Talla",
                "stock_total": "Stock eComm BigQuery",
                "Stock Shopify Variante": "Stock Shopify",
            }
        )
        stock_location_activation_audit["Diagnostico"] = (
            "SKU existe en Shopify y tiene stock eComm, pero Shopify no refleja stock en la variante."
        )
        stock_location_activation_audit["Accion sugerida"] = (
            "Revisar/activar sucursales del inventory item y luego sincronizar stock."
        )
    else:
        stock_location_activation_audit = pd.DataFrame(
            columns=[
                "Mod-Col",
                "Marca",
                "Talla",
                "SKU",
                "Stock eComm BigQuery",
                "Stock Shopify",
                "Inventory Item GID",
                "Inventory Item ID",
                "Status Shopify",
                "Publicado Shopify",
                "URL Shopify",
                "Diagnostico",
                "Accion sugerida",
            ]
        )
    no_price_keys = {clean_value(value) for value in no_price_models.get("Mod-Col KPI", pd.Series(dtype=object))}
    model_stock["Sin_precio_shopify"] = model_stock["Mod-Col KPI"].map(lambda value: clean_value(value) in no_price_keys)
    no_photo_keys = {clean_value(value) for value in no_photo_models.get("Mod-Col KPI", pd.Series(dtype=object))}
    model_stock["Sin_foto_shopify"] = model_stock["Mod-Col KPI"].map(lambda value: clean_value(value) in no_photo_keys)
    no_shopify_stock_keys = {
        clean_value(value) for value in no_shopify_stock_models.get("Mod-Col KPI", pd.Series(dtype=object))
    }
    model_stock["Sin_stock_shopify"] = model_stock["Mod-Col KPI"].map(
        lambda value: clean_value(value) in no_shopify_stock_keys
    )
    model_stock["Con_foto_shopify"] = model_stock["Fotos_Shopify"] > 0
    model_stock["Con_stock_shopify"] = model_stock["Stock_Shopify"] > 0
    model_stock["Listo_venta"] = (
        model_stock["Debe estar visible"]
        & model_stock["Producto_creado"]
        & model_stock["Visible_Shopify"]
        & model_stock["Con_stock_shopify"]
        & ~model_stock["Sin_precio_shopify"]
        & ~model_stock["Sin_foto_shopify"]
    )

    created_with_stock = int((model_stock["Debe estar visible"] & model_stock["Producto_creado"]).sum())
    created_without_stock = int((model_stock["Producto_creado"] & ~model_stock["Debe estar visible"]).sum())
    web_visible = int(model_stock["Listo_venta"].sum())
    non_visible_web = model_stock[
        model_stock["Debe estar visible"] & model_stock["Producto_creado"] & (~model_stock["Listo_venta"])
    ].copy()

    def non_visible_reason(row):
        if row.get("Sin_stock_shopify"):
            return "Sin stock Shopify"
        if row.get("Sin_foto_shopify"):
            return "Sin foto"
        if row.get("Sin_precio_shopify"):
            return "Sin precio"
        if clean_value(row.get("Status_Shopify")).upper() != "ACTIVE":
            return "No activo Shopify"
        if clean_value(row.get("Publicado_Shopify")).upper() != "SI":
            return "No publicado Online Store"
        return "Otros por revisar"

    def non_visible_blockers(row):
        blockers = []
        if row.get("Sin_stock_shopify"):
            blockers.append("Sin stock Shopify")
        if row.get("Sin_foto_shopify"):
            blockers.append("Sin foto")
        if row.get("Sin_precio_shopify"):
            blockers.append("Sin precio")
        if clean_value(row.get("Status_Shopify")).upper() != "ACTIVE":
            blockers.append("No activo Shopify")
        if clean_value(row.get("Publicado_Shopify")).upper() != "SI":
            blockers.append("No publicado Online Store")
        return " + ".join(blockers) if blockers else "Otros por revisar"

    def non_visible_state(row):
        pieces = [
            "Stock Shopify OK" if row.get("Con_stock_shopify") else "Sin stock Shopify",
            "Con foto" if row.get("Con_foto_shopify") else "Sin foto",
            "Con precio" if not row.get("Sin_precio_shopify") else "Sin precio",
            clean_value(row.get("Status_Shopify")) or "Sin status",
            "Publicado Online Store" if clean_value(row.get("Publicado_Shopify")).upper() == "SI" else "No publicado Online Store",
        ]
        return " | ".join(pieces)

    if not non_visible_web.empty:
        non_visible_web["Motivo principal"] = non_visible_web.apply(non_visible_reason, axis=1)
        non_visible_web["Bloqueos"] = non_visible_web.apply(non_visible_blockers, axis=1)
        non_visible_web["Estado operativo"] = non_visible_web.apply(non_visible_state, axis=1)
        non_visible_web["ModCol_BQ"] = 1
        non_visible_web["ModCol_stock_shopify"] = non_visible_web["Con_stock_shopify"].map(lambda value: 1 if value else 0)
        non_visible_counts = non_visible_web["Motivo principal"].value_counts().to_dict()
        non_visible_combo_summary = (
            non_visible_web.groupby(["Bloqueos", "Estado operativo"], as_index=False)
            .agg(
                Modelos=("Mod-Col KPI", "nunique"),
                Stock_BigQuery=("ModCol_BQ", "sum"),
                Stock_Shopify=("ModCol_stock_shopify", "sum"),
            )
            .sort_values(["Modelos", "Stock_BigQuery"], ascending=[False, False])
        )
    else:
        non_visible_web["Motivo principal"] = ""
        non_visible_web["Bloqueos"] = ""
        non_visible_web["Estado operativo"] = ""
        non_visible_web["ModCol_BQ"] = 0
        non_visible_web["ModCol_stock_shopify"] = 0
        non_visible_counts = {}
        non_visible_combo_summary = pd.DataFrame(
            columns=["Bloqueos", "Estado operativo", "Modelos", "Stock_BigQuery", "Stock_Shopify"]
        )

    kpis = {
        "modelos_con_stock": int(model_stock["Debe estar visible"].sum()),
        "modelos_creados_shopify": int(model_stock["Producto_creado"].sum()),
        "modelos_creados_con_stock": created_with_stock,
        "cobertura_shopify": float(created_with_stock / model_stock["Debe estar visible"].sum()) if model_stock["Debe estar visible"].sum() else 0,
        "modelos_pendientes": int(len(missing_models)),
        "con_stock_no_visibles": int(len(stock_not_visible)),
        "sin_stock_visibles": int(len(no_stock_visible)),
        "modelos_variantes_incompletas": int(model_stock["Variantes_stock_incompletas"].sum()),
        "productos_creados_sin_stock": created_without_stock,
        "productos_visibles": int((model_stock["Debe estar visible"] & model_stock["Visible_Shopify"]).sum()),
        "modelos_visibles_web": web_visible,
        "modelos_no_visibles_web": max(created_with_stock - web_visible, 0),
        "no_visible_sin_stock_shopify": int(non_visible_counts.get("Sin stock Shopify", 0)),
        "no_visible_sin_foto": int(non_visible_counts.get("Sin foto", 0)),
        "no_visible_sin_precio": int(non_visible_counts.get("Sin precio", 0)),
        "no_visible_no_activo": int(non_visible_counts.get("No activo Shopify", 0)),
        "no_visible_no_publicado": int(non_visible_counts.get("No publicado Online Store", 0)),
        "no_visible_otros": int(non_visible_counts.get("Otros por revisar", 0)),
        "modelos_listos_tienda": web_visible,
        "modelos_con_stock_con_foto": int(
            (model_stock["Debe estar visible"] & model_stock["Producto_creado"] & model_stock["Con_foto_shopify"]).sum()
        ),
        "modelos_visibles_con_foto": int(
            (model_stock["Debe estar visible"] & model_stock["Visible_Shopify"] & model_stock["Con_foto_shopify"]).sum()
        ),
        "modelos_con_stock_shopify": int(
            (model_stock["Debe estar visible"] & model_stock["Producto_creado"] & model_stock["Con_stock_shopify"]).sum()
        ),
        "modelos_sin_stock_shopify": int(len(no_shopify_stock_models)),
        "variantes_stock_ecomm_sin_stock_shopify": int(len(stock_location_activation_audit)),
        "sincronizacion_stock_shopify": float(
            (
                model_stock["Debe estar visible"]
                & model_stock["Producto_creado"]
                & model_stock["Con_stock_shopify"]
            ).sum()
            / created_with_stock
        ) if created_with_stock else 0,
        "modelos_listos_venta": web_visible,
        "modelos_sin_precio": int(len(no_price_models)),
        "modelos_sin_foto": int(len(no_photo_models)),
        "productos_shopify_sin_foto_total": int(len(shopify_no_photo_all)),
        "stock_ecomm_rows": int(stock_ecomm_rows),
        "stock_ecomm_units": int(stock_ecomm_units),
        "stock_ecomm_models": int(stock_ecomm_models),
        "modelos_total_auditoria": int(model_stock["Mod-Col KPI"].nunique()),
        "modelos_no_creados_shopify": int(len(missing_models)),
        "modelos_creados_no_visibles": int(len(non_visible_web)),
        "modelos_visibles_reales_web": int(web_visible),
    }
    kpi_audit = pd.DataFrame(
        [
            {"Indicador": "Total modelos fuente ARTI/BigQuery", "Valor": kpis["modelos_total_auditoria"], "Lectura": "Modelo-color detectados con tallas validas."},
            {"Indicador": "Modelos con stock eComm", "Valor": kpis["modelos_con_stock"], "Lectura": "Modelo-color que deberian venderse por stock eComm."},
            {"Indicador": "Modelos ya creados en Shopify", "Valor": kpis["modelos_creados_shopify"], "Lectura": "Existe producto Shopify con codigo modelo-color."},
            {"Indicador": "Modelos no creados en Shopify", "Valor": kpis["modelos_no_creados_shopify"], "Lectura": "Existe en fuente pero no en Shopify."},
            {"Indicador": "Modelos creados pero no visibles", "Valor": kpis["modelos_creados_no_visibles"], "Lectura": "Creado con stock eComm, pero no cumple stock/precio/foto/activo/publicado."},
            {"Indicador": "Modelos visibles reales web", "Valor": kpis["modelos_visibles_reales_web"], "Lectura": "Activo, publicado Online Store, con stock Shopify, precio y foto."},
            {"Indicador": "Modelos con stock eComm sin foto", "Valor": kpis["modelos_sin_foto"], "Lectura": "Modelo-color creado en Shopify, con stock eComm, pero sin fotos Shopify. Esta es la base operativa del KPI."},
            {"Indicador": "Productos Shopify sin foto total", "Valor": kpis["productos_shopify_sin_foto_total"], "Lectura": "Total bruto de productos Shopify sin fotos, sin filtrar por stock eComm ni por necesidad de venta web."},
        ]
    )
    model_stock["Creado_con_stock"] = model_stock["Debe estar visible"] & model_stock["Producto_creado"]
    brand_summary = (
        model_stock.groupby("Marca", as_index=False)
        .agg(
            Modelos_con_stock=("Debe estar visible", "sum"),
            Creados_Shopify=("Creado_con_stock", "sum"),
            Pendientes_creacion=("Producto_creado", lambda values: 0),
            Stock_total=("Stock_total", "sum"),
        )
    )
    if not brand_summary.empty:
        brand_summary["Pendientes_creacion"] = brand_summary["Modelos_con_stock"] - brand_summary["Creados_Shopify"]
        brand_summary["Cobertura"] = brand_summary.apply(
            lambda row: row["Creados_Shopify"] / row["Modelos_con_stock"] if row["Modelos_con_stock"] else 0,
            axis=1,
        )

    action_rows = []
    for _, row in missing_models.iterrows():
        action_rows.append({"Mod-Col": row["Mod-Col KPI"], "Marca": row["Marca"], "Problema": "Modelo con stock no creado", "Acción sugerida": "Pedir input al Brand Manager", "Stock total": row["Stock_total"]})
    for _, row in no_stock_visible.iterrows():
        action_rows.append({"Mod-Col": row["Mod-Col KPI"], "Marca": row["Marca"], "Problema": "Sin stock visible", "Acción sugerida": "Apagar producto en Shopify", "Stock total": row["Stock_total"]})
    for _, row in no_price_models.iterrows():
        action_rows.append({"Mod-Col": row["Mod-Col KPI"], "Marca": row["Marca"], "Problema": "Creado con stock sin precio", "Acción sugerida": "Cargar precio en Shopify", "Stock total": row["Stock_total"]})
    for _, row in no_photo_models.iterrows():
        action_rows.append({"Mod-Col": row["Mod-Col KPI"], "Marca": row["Marca"], "Problema": "Modelo con stock sin foto", "Acción sugerida": "Solicitar fotos al Brand Manager", "Stock total": row["Stock_total"]})
    for _, row in no_shopify_stock_models.iterrows():
        action_rows.append({"Mod-Col": row["Mod-Col KPI"], "Marca": row["Marca"], "Problema": "Modelo con stock eComm sin stock Shopify", "Acción sugerida": "Revisar sincronización de stock hacia Shopify", "Stock total": row["Stock_total"]})

    actions_df = pd.DataFrame(action_rows)
    missing_model_keys = {clean_value(value) for value in missing_models.get("Mod-Col KPI", pd.Series(dtype=object))}
    stock_web_blocker_keys = missing_model_keys | no_shopify_stock_keys
    missing_stock_variants_export = (
        missing_stock_variants[
            missing_stock_variants["Mod-Col KPI"].map(lambda value: clean_value(value) in stock_web_blocker_keys)
        ][["Mod-Col KPI", "MARCA_MA", "Talla KPI", "SKU", "stock_total"]]
        .rename(columns={"Mod-Col KPI": "Mod-Col", "Talla KPI": "Talla", "stock_total": "Stock total"})
        if not missing_stock_variants.empty
        else pd.DataFrame(columns=["Mod-Col", "MARCA_MA", "Talla", "SKU", "Stock total"])
    )
    if not missing_stock_variants_export.empty:
        missing_stock_variants_export["Motivo web"] = missing_stock_variants_export["Mod-Col"].map(
            lambda value: (
                "Modelo con stock no creado"
                if clean_value(value) in missing_model_keys
                else "Modelo con stock eComm sin stock Shopify"
                if clean_value(value) in no_shopify_stock_keys
                else "Revisar stock web"
            )
        )
    else:
        missing_stock_variants_export = pd.DataFrame(
            columns=["Mod-Col", "MARCA_MA", "Talla", "SKU", "Stock total", "Motivo web"]
        )
    missing_models_input, missing_models_variants, missing_models_fields = build_missing_models_input_export(
        expected,
        missing_models,
        brand_config,
    )
    return {
        "kpis": kpis,
        "kpi_audit": kpi_audit,
        "model_stock": model_stock,
        "brand_summary": brand_summary,
        "actions": actions_df,
        "missing_stock_variants": missing_stock_variants_export,
        "missing_models_input": missing_models_input,
        "missing_models_variants": missing_models_variants,
        "missing_models_fields": missing_models_fields,
        "no_price_models": no_price_models,
        "no_photo_models": no_photo_models,
        "shopify_no_photo_all": shopify_no_photo_all,
        "no_shopify_stock_models": no_shopify_stock_models,
        "stock_location_activation_audit": stock_location_activation_audit,
        "non_visible_web": non_visible_web,
        "non_visible_combo_summary": non_visible_combo_summary,
        "ecomm_stock_match": ecomm_stock_match,
        "stock_not_visible": stock_not_visible,
        "no_stock_visible": no_stock_visible,
    }


def load_catalog_kpi_result(brand_config, shopify_config):
    arti_df, arti_source = read_arti_for_app(brand_config)
    stock_df = read_current_stock_from_bigquery(get_bigquery_config())
    if stock_df.empty:
        raise RuntimeError(
            "BigQuery devolvio 0 filas de stock. No se actualizo el dashboard para evitar pisar KPIs validos con ceros."
        )
    shopify_products = fetch_products(shopify_config)
    result = build_catalog_kpis(arti_df, stock_df, shopify_products, brand_config)
    allowed_ecomm_codes = ecomm_stock_rule_codes_for_site(brand_config)
    result["meta"] = {
        "cache_version": KPI_CACHE_VERSION,
        "arti_source": arti_source,
        "stock_raw_rows": safe_int_value(stock_df.attrs.get("stock_rows_before_cutoff", len(stock_df))),
        "stock_cutoff_rows": len(stock_df),
        "stock_filtered_rows": result.get("kpis", {}).get("stock_ecomm_rows", 0),
        "ecomm_bodegas_usadas": ", ".join(allowed_ecomm_codes),
        "ecomm_bodegas_count": len(allowed_ecomm_codes),
        "shopify_products": len(shopify_products),
        "fecha_corte": clean_value(stock_df.attrs.get("stock_latest_cutoff"))
        or (clean_value(stock_df["fecha_corte"].max()) if not stock_df.empty and "fecha_corte" in stock_df.columns else ""),
        "refreshed_at": datetime.now(timezone.utc).isoformat(),
    }
    return result


def kpi_cache_path(site_key):
    safe_site = re.sub(r"[^a-zA-Z0-9_-]+", "_", clean_value(site_key) or "site")
    return KPI_CACHE_DIR / f"{safe_site}.pkl"


def load_cached_catalog_kpi_result(site_key):
    path = kpi_cache_path(site_key)
    if not path.exists():
        return None
    try:
        with path.open("rb") as cache_file:
            result = pickle.load(cache_file)
        if not isinstance(result, dict) or "kpis" not in result:
            return None
        meta = result.get("meta", {}) if isinstance(result.get("meta"), dict) else {}
        if meta.get("cache_version") != KPI_CACHE_VERSION:
            return None
        if safe_int_value(meta.get("stock_cutoff_rows")) <= 0 or not clean_value(meta.get("fecha_corte")):
            return None
        return result
    except Exception:
        return None


def is_current_kpi_result(result):
    if not isinstance(result, dict) or "kpis" not in result:
        return False
    meta = result.get("meta", {}) if isinstance(result.get("meta"), dict) else {}
    return (
        meta.get("cache_version") == KPI_CACHE_VERSION
        and safe_int_value(meta.get("stock_cutoff_rows")) > 0
        and clean_value(meta.get("fecha_corte")) != ""
    )


def is_stale_kpi_result(result, max_age_seconds=KPI_AUTO_REFRESH_SECONDS):
    if not is_current_kpi_result(result):
        return True
    meta = result.get("meta", {}) if isinstance(result.get("meta"), dict) else {}
    refreshed_at = parse_iso_datetime(meta.get("refreshed_at"))
    if refreshed_at is None:
        return True
    return datetime.now(timezone.utc) - refreshed_at.astimezone(timezone.utc) >= timedelta(seconds=max_age_seconds)


def save_cached_catalog_kpi_result(site_key, result):
    if not isinstance(result, dict) or "kpis" not in result:
        return
    KPI_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    with kpi_cache_path(site_key).open("wb") as cache_file:
        pickle.dump(result, cache_file)


def render_dashboard_refresh_error(exc, cached_result=None):
    error_text = clean_value(exc)
    st.error(f"No se pudo actualizar el dashboard: {error_text}")
    if isinstance(cached_result, dict) and cached_result.get("kpis"):
        refreshed_label = format_datetime_lima(cached_result.get("meta", {}).get("refreshed_at"))
        suffix = f" del {refreshed_label}" if refreshed_label else ""
        st.warning(f"Se mantienen en pantalla los ultimos KPIs validos{suffix}. No se reemplazaron por ceros.")
    else:
        st.warning(
            "No hay KPIs validos guardados para mostrar. Reintenta en unos minutos o revisa que la tabla de stock "
            "BigQuery tenga filas para el ultimo corte."
        )
    with st.expander("Diagnostico rapido", expanded=False):
        st.write(
            "- La app evita pisar el dashboard cuando BigQuery devuelve 0 filas de stock.\n"
            "- Esto suele pasar por una ventana de actualizacion de la tabla, permisos, o una consulta `stock_query` "
            "custom que quedo sin datos.\n"
            "- Si usas `stock_query` en Secrets, valida que no este amarrada a una fecha sin carga."
        )


def siblings_by_model_from_shopify(shopify_products):
    products_df = pd.DataFrame(shopify_products)
    if products_df.empty or "Mod-Col" not in products_df.columns or "Handle" not in products_df.columns:
        return {}
    products_df["__MODEL"] = products_df["Mod-Col"].map(lambda value: clean_value(value).upper().rsplit("-", 1)[0])
    return (
        products_df[products_df["__MODEL"] != ""]
        .groupby("__MODEL")["Handle"]
        .apply(lambda values: ", ".join(dict.fromkeys(clean_value(value) for value in values if clean_value(value))))
        .to_dict()
    )


def apply_shopify_siblings_to_matrixify(matrixify_df, shopify_products):
    siblings_map = siblings_by_model_from_shopify(shopify_products)
    if matrixify_df is None or matrixify_df.empty or not siblings_map:
        return matrixify_df
    df = matrixify_df.copy()
    key_column = "Metafield: custom.codigo_modelo_color [id]"
    siblings_column = "Metafield: theme.siblings [single_line_text_field]"
    custom_siblings_column = "Metafield: custom.siblings [single_line_text_field]"
    if key_column not in df.columns:
        return df
    for column in (siblings_column, custom_siblings_column):
        if column not in df.columns:
            df[column] = ""

    def sibling_value(row):
        key = clean_value(row.get(key_column)).upper()
        if not key:
            return row.get(siblings_column)
        model = key.rsplit("-", 1)[0]
        return siblings_map.get(model, row.get(siblings_column) or row.get(custom_siblings_column))

    top_rows = df["Handle"].map(clean_value) != ""
    values = df.loc[top_rows].apply(sibling_value, axis=1)
    df.loc[top_rows, siblings_column] = values
    df.loc[top_rows, custom_siblings_column] = values
    return df


def _product_gid(value):
    text = clean_value(value)
    if not text:
        return ""
    if text.startswith("gid://"):
        return text
    return f"gid://shopify/Product/{text}"


def _metafield_type_from_column(column):
    match = re.search(r"\[(.+?)\]$", clean_value(column))
    if not match:
        return "single_line_text_field"
    return match.group(1)


def _metafield_can_write_direct(column):
    namespace, key = _metafield_namespace_key(column)
    field_type = _metafield_type_from_column(column)
    if field_type in ("page_reference", "list.page_reference"):
        return False, f"{namespace}.{key} requiere IDs internos de Shopify; se mantiene para Matrixify"
    return True, ""


def _logo_lookup_keys(record):
    keys = set()
    candidates = [
        record.get("handle"),
        record.get("displayName"),
    ]
    for field in record.get("fields") or []:
        candidates.append(field.get("value"))
        reference = field.get("reference") or {}
        image = reference.get("image") or {}
        candidates.append(reference.get("url"))
        candidates.append(image.get("url"))
        for referenced_node in ((field.get("references") or {}).get("nodes")) or []:
            referenced_image = referenced_node.get("image") or {}
            candidates.append(referenced_node.get("url"))
            candidates.append(referenced_image.get("url"))

    expanded_candidates = []
    for candidate in candidates:
        expanded_candidates.append(candidate)
        text = clean_value(candidate).lower()
        if not text:
            continue
        parsed_path = unquote(urlparse(text).path or "")
        if parsed_path:
            filename = Path(parsed_path).stem
            if filename:
                expanded_candidates.append(filename)

    for candidate in expanded_candidates:
        text = clean_value(candidate).lower()
        if not text:
            continue
        normalized = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
        compact = re.sub(r"[^a-z0-9]+", "", text)
        for key in (text, normalized, compact):
            if key:
                keys.add(key)
                keys.add(f"logo.{key}")
                keys.add(f"{key}-clb")
                keys.add(f"logo.{key}-clb")
        if normalized.endswith("-clb"):
            base = normalized[:-4]
            keys.update({base, f"logo.{base}"})
        if text.startswith("logo."):
            bare = text.split(".", 1)[1]
            keys.add(bare)
            if bare.endswith("-clb"):
                keys.add(bare[:-4])
    return keys


def _logo_reference_candidates(reference, handle=""):
    values = [reference, handle]
    candidates = set()
    for value in values:
        text = clean_value(value).lower()
        if not text:
            continue
        if text.startswith("logo."):
            text = text.split(".", 1)[1]
        base_values = {text}
        if text.endswith("-clb"):
            base_values.add(text[:-4])
        for base in list(base_values):
            normalized = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
            compact = re.sub(r"[^a-z0-9]+", "", base)
            spaced = normalized.replace("-", " ")
            alias_values = set()
            if normalized.startswith("logo-"):
                normalized = normalized[5:]
            alias_values.update(TECHNOLOGY_LOGO_ALIASES.get(normalized, []))
            for item in (base, normalized, compact, spaced):
                if item:
                    candidates.add(item)
                    candidates.add(f"logo.{item}")
                    candidates.add(f"{item}-clb")
                    candidates.add(f"logo.{item}-clb")
            for alias in alias_values:
                alias_text = clean_value(alias).lower()
                alias_normalized = re.sub(r"[^a-z0-9]+", "-", alias_text).strip("-")
                alias_compact = re.sub(r"[^a-z0-9]+", "", alias_text)
                for item in (alias_text, alias_normalized, alias_compact):
                    if item:
                        candidates.add(item)
                        candidates.add(f"logo.{item}")
                        candidates.add(f"{item}-clb")
                        candidates.add(f"logo.{item}-clb")
    return {candidate.lower() for candidate in candidates if candidate}


def _static_logo_metaobject_gid(reference, handle=""):
    lookup = {}
    for label, gid in TECHNOLOGY_LOGO_METAOBJECT_GIDS.items():
        for candidate in _logo_reference_candidates(label):
            lookup.setdefault(candidate, gid)
    for normalized, aliases in TECHNOLOGY_LOGO_ALIASES.items():
        gid = ""
        for alias in aliases:
            gid = lookup.get(clean_value(alias).lower()) or next(
                (lookup.get(candidate) for candidate in _logo_reference_candidates(alias) if lookup.get(candidate)),
                "",
            )
            if gid:
                break
        if gid:
            for alias in [normalized, *aliases]:
                for candidate in _logo_reference_candidates(alias):
                    lookup.setdefault(candidate, gid)

    for candidate in _logo_reference_candidates(reference, handle):
        gid = lookup.get(candidate)
        if gid:
            return gid
    return ""


def _metaobject_gid_lookup(shopify_config, metaobject_type):
    cache_key = f"metaobject_lookup_v2_{clean_value(metaobject_type)}"
    if cache_key not in st.session_state:
        records = fetch_metaobjects(shopify_config, metaobject_type)
        lookup = {}
        for record in records:
            gid = clean_value(record.get("id"))
            if not gid:
                continue
            for key in _logo_lookup_keys(record):
                lookup[key.lower()] = gid
        st.session_state[cache_key] = lookup
    return st.session_state[cache_key]


def _all_metaobject_gid_lookup(shopify_config):
    cache_key = "metaobject_lookup_all_v2"
    if cache_key in st.session_state:
        return st.session_state[cache_key]

    by_handle = {}
    by_reference = {}
    definitions = fetch_metaobject_definitions(shopify_config)
    for definition in definitions:
        metaobject_type = clean_value(definition.get("type"))
        if not metaobject_type:
            continue
        try:
            lookup = _metaobject_gid_lookup(shopify_config, metaobject_type)
        except Exception:
            continue
        for handle, gid in lookup.items():
            by_handle.setdefault(handle.lower(), gid)
            by_reference[f"{metaobject_type}.{handle}".lower()] = gid

    st.session_state[cache_key] = {"by_handle": by_handle, "by_reference": by_reference}
    return st.session_state[cache_key]


def _metaobject_definition_ids_from_metafield(shopify_config, namespace, key):
    cache_key = f"metafield_definition_metaobjects_{namespace}_{key}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]

    definition_ids = []
    if fetch_metafield_definition is None:
        st.session_state[cache_key] = definition_ids
        return definition_ids
    try:
        definition = fetch_metafield_definition(shopify_config, "PRODUCT", namespace, key)
    except Exception:
        definition = {}
    for validation in definition.get("validations") or []:
        name = clean_value(validation.get("name"))
        value = clean_value(validation.get("value"))
        if name not in ("metaobject_definition_id", "metaobject_definition_ids") or not value:
            continue
        try:
            parsed_value = json.loads(value)
            if isinstance(parsed_value, list):
                definition_ids.extend(clean_value(item) for item in parsed_value if clean_value(item))
            elif clean_value(parsed_value):
                definition_ids.append(clean_value(parsed_value))
        except Exception:
            definition_ids.extend(
                clean_value(item)
                for item in re.split(r"[,|\s]+", value)
                if clean_value(item).startswith("gid://shopify/MetaobjectDefinition/")
            )
    st.session_state[cache_key] = list(dict.fromkeys(definition_ids))
    return st.session_state[cache_key]


def _metaobject_gid_lookup_for_metafield(shopify_config, namespace, key):
    cache_key = f"metaobject_lookup_metafield_v2_{namespace}_{key}"
    if cache_key in st.session_state:
        return st.session_state[cache_key]

    lookup = {}
    if fetch_metaobjects_for_definition is None:
        st.session_state[cache_key] = lookup
        return lookup
    for definition_id in _metaobject_definition_ids_from_metafield(shopify_config, namespace, key):
        try:
            records = fetch_metaobjects_for_definition(shopify_config, definition_id)
        except Exception:
            continue
        for record in records:
            gid = clean_value(record.get("id"))
            if not gid:
                continue
            for lookup_key in _logo_lookup_keys(record):
                lookup[lookup_key.lower()] = gid
    st.session_state[cache_key] = lookup
    return lookup


def _resolve_metaobject_reference_value(shopify_config, column, value):
    text = clean_value(value)
    field_type = _metafield_type_from_column(column)
    if field_type not in ("metaobject_reference", "list.metaobject_reference") or not text:
        return text

    references = [item.strip() for item in text.split(",") if item.strip()]
    gids = []
    missing = []
    for reference in references:
        if reference.startswith("gid://shopify/Metaobject/"):
            gids.append(reference)
            continue
        metaobject_type = ""
        handle = reference
        lookup = {}
        if "." in reference:
            metaobject_type, handle = reference.split(".", 1)
            try:
                lookup = _metaobject_gid_lookup(shopify_config, metaobject_type)
            except Exception:
                lookup = {}
        gid = ""
        for candidate in _logo_reference_candidates(reference, handle):
            gid = lookup.get(candidate)
            if gid:
                break
        if not gid:
            namespace, key = _metafield_namespace_key(column)
            metafield_lookup = _metaobject_gid_lookup_for_metafield(shopify_config, namespace, key)
            for candidate in _logo_reference_candidates(reference, handle):
                gid = metafield_lookup.get(candidate)
                if gid:
                    break
        if not gid:
            try:
                fallback_lookup = _all_metaobject_gid_lookup(shopify_config)
            except Exception:
                fallback_lookup = {"by_reference": {}, "by_handle": {}}
            for candidate in _logo_reference_candidates(reference, handle):
                gid = fallback_lookup["by_reference"].get(candidate) or fallback_lookup["by_handle"].get(candidate)
                if gid:
                    break
        if not gid:
            gid = _static_logo_metaobject_gid(reference, handle)
        if gid:
            if gid not in gids:
                gids.append(gid)
        else:
            missing.append(reference)

    if missing:
        raise ValueError(f"No encontre metaobjects para: {', '.join(missing)}")
    if field_type == "list.metaobject_reference":
        return json.dumps(gids, ensure_ascii=False)
    return gids[0] if gids else ""


def _validate_logo_metaobject_refs(shopify_config, value):
    refs = _split_tags(value)
    if not refs:
        return [], []
    if not shopify_config:
        return [], refs
    resolved = []
    missing = []
    for ref in refs:
        try:
            _resolve_metaobject_reference_value(
                shopify_config,
                "Metafield: custom.logo [list.metaobject_reference]",
                ref,
            )
            resolved.append(ref)
        except Exception:
            missing.append(ref)
    return resolved, missing


def _shopify_image_url(value):
    url = _normalize_legacy_image_url(value)
    prefix = "https://ecom-imagenes.forus-digital.xyz.peru.s3.amazonaws.com/"
    if url.startswith(prefix):
        return "https://s3.amazonaws.com/ecom-imagenes.forus-digital.xyz.peru/" + url[len(prefix):]
    return url


def _normalize_legacy_image_url(value):
    url = clean_value(value)
    replacements = {
        "COLUMBIA%20SHOPIFY": "COLUMBIA",
        "ROCKFORD%20SHOPIFY": "ROCKFORD",
        "HUSH%20PUPPIES%20SHOPIFY": "HUSHPUPPIES",
        "VANS%20SHOPIFY": "VANS",
        "KEDS%20SHOPIFY": "KEDS",
        "PATAGONIA%20SHOPIFY": "PATAGONIA",
        "SOREL%20SHOPIFY": "SOREL",
        "MOUNTAIN%20HARDWEAR%20SHOPIFY": "MOUNTAINHARDWEAR",
        "COLUMBIA SHOPIFY": "COLUMBIA",
        "ROCKFORD SHOPIFY": "ROCKFORD",
        "HUSH PUPPIES SHOPIFY": "HUSHPUPPIES",
        "HUSH%20PUPPIES": "HUSHPUPPIES",
        "HUSH PUPPIES": "HUSHPUPPIES",
        "VANS SHOPIFY": "VANS",
        "KEDS SHOPIFY": "KEDS",
        "PATAGONIA SHOPIFY": "PATAGONIA",
        "SOREL SHOPIFY": "SOREL",
        "MOUNTAIN HARDWEAR SHOPIFY": "MOUNTAINHARDWEAR",
    }
    for old, new in replacements.items():
        url = url.replace(f"/{old}/", f"/{new}/")
    return url


def _image_url_candidates(value):
    original = clean_value(value)
    normalized = _normalize_legacy_image_url(original)
    converted = _shopify_image_url(original)
    return list(dict.fromkeys([url for url in (converted, normalized, original) if url]))


def _url_is_reachable_image(url, timeout=8):
    headers = {"User-Agent": "Mozilla/5.0", "Range": "bytes=0-512"}
    for method in ("HEAD", "GET"):
        request = Request(url, method=method, headers=headers)
        try:
            with urlopen(request, timeout=timeout) as response:
                content_type = clean_value(response.headers.get("Content-Type")).lower()
                return response.status < 400 and content_type.startswith("image/")
        except HTTPError as exc:
            if exc.code in (403, 405) and method == "HEAD":
                continue
            return False
        except (URLError, TimeoutError, OSError):
            return False
    return False


def _download_image_bytes(value):
    last_error = ""
    headers = {"User-Agent": "Mozilla/5.0"}
    for url in _image_url_candidates(value):
        try:
            request = Request(url, headers=headers)
            with urlopen(request, timeout=30) as response:
                content_type = clean_value(response.headers.get("Content-Type")).lower()
                if response.status >= 400 or not content_type.startswith("image/"):
                    last_error = f"{url}: no es imagen valida"
                    continue
                data = response.read()
                if not data:
                    last_error = f"{url}: imagen vacia"
                    continue
                filename = Path(unquote(urlparse(url).path or "")).name or "product_image.jpg"
                return data, content_type.split(";")[0], filename, url
        except Exception as exc:
            last_error = f"{url}: {exc}"
    raise ShopifyApiError(last_error or "No se pudo descargar la imagen")


def _file_status_summary(file_statuses):
    ready = []
    failed = []
    pending = []
    for file_node in file_statuses or []:
        status = clean_value(file_node.get("fileStatus")).upper()
        if status == "READY":
            ready.append(file_node)
        elif status == "FAILED":
            failed.append(file_node)
        else:
            pending.append(file_node)
    return ready, failed, pending


def _file_cdn_url(file_node):
    image = file_node.get("image") or {}
    preview = file_node.get("preview") or {}
    preview_image = preview.get("image") or {}
    return clean_value(image.get("url")) or clean_value(preview_image.get("url"))


def _product_set_files_with_fallback(shopify_config, product_gid, product_files):
    try:
        product_set_files(shopify_config, product_gid, product_files)
        return "productSet staged"
    except Exception as product_set_exc:
        cdn_files = []
        file_errors = []
        for product_file in product_files:
            try:
                created = file_create(
                    shopify_config,
                    product_file.get("originalSource"),
                    alt=product_file.get("alt"),
                    content_type=product_file.get("contentType") or "IMAGE",
                )
                file_ids = [clean_value(file_node.get("id")) for file_node in created if clean_value(file_node.get("id"))]
                statuses = wait_file_statuses(shopify_config, file_ids) if file_ids else created
                ready_files, failed_files, pending_files = _file_status_summary(statuses)
                if failed_files:
                    file_errors.append(f"{product_file.get('filename')}: archivo fallido")
                    continue
                if pending_files and not ready_files:
                    file_errors.append(f"{product_file.get('filename')}: archivo en procesamiento")
                    continue
                source_node = (ready_files or statuses or created or [{}])[0]
                cdn_url = _file_cdn_url(source_node)
                if not cdn_url:
                    file_errors.append(f"{product_file.get('filename')}: sin URL CDN")
                    continue
                cdn_files.append(
                    {
                        "originalSource": cdn_url,
                        "alt": product_file.get("alt"),
                        "filename": product_file.get("filename"),
                        "contentType": "IMAGE",
                        "duplicateResolutionMode": "APPEND_UUID",
                    }
                )
            except Exception as exc:
                file_errors.append(f"{product_file.get('filename')}: {exc}")
        if cdn_files:
            product_set_files(shopify_config, product_gid, cdn_files)
            if file_errors:
                raise ShopifyApiError(
                    f"productSet staged fallo ({product_set_exc}); fallback CDN parcial: {' | '.join(file_errors[:3])}"
                )
            return "fileCreate CDN"
        raise ShopifyApiError(f"productSet staged fallo: {product_set_exc}; fallback sin archivos: {' | '.join(file_errors[:3])}")


def _sync_product_photos_direct(shopify_config, product_gid, image_urls, existing_media_ids=None, image_mode="replace", alt_text=""):
    existing_media_ids = existing_media_ids or []
    image_urls = [url for url in image_urls[:10] if clean_value(url)]
    image_mode = clean_value(image_mode).lower() or "replace"
    product_files = []
    image_errors = []
    uploaded_count = 0

    for image_index, raw_image_url in enumerate(image_urls, start=1):
        try:
            image_bytes, mime_type, filename, _ = _download_image_bytes(raw_image_url)
            resource_url = staged_upload_image(shopify_config, filename, mime_type, image_bytes)
            product_files.append(
                {
                    "originalSource": resource_url,
                    "alt": clean_value(alt_text),
                    "filename": filename,
                    "contentType": "IMAGE",
                    "duplicateResolutionMode": "APPEND_UUID",
                }
            )
        except Exception as exc:
            image_errors.append(f"foto {image_index}: {exc}")

    if not product_files:
        direct_media = product_create_media(shopify_config, product_gid, image_urls)
        if direct_media:
            route = "URL directa"
            uploaded_count = len(direct_media)
        else:
            raise ShopifyApiError("No se pudo subir ninguna foto nueva. " + " | ".join(image_errors[:3]))
    else:
        try:
            route = _product_set_files_with_fallback(shopify_config, product_gid, product_files)
            uploaded_count = len(product_files)
        except Exception as upload_exc:
            direct_media = product_create_media(shopify_config, product_gid, image_urls)
            if direct_media:
                route = f"URL directa tras fallback staged ({upload_exc})"
                uploaded_count = len(direct_media)
            else:
                raise
    deleted_count = 0
    delete_note = ""
    if image_mode == "replace" and existing_media_ids:
        try:
            deleted = product_delete_media(shopify_config, product_gid, existing_media_ids)
            deleted_count = len(deleted)
        except Exception as exc:
            detail = clean_value(exc).lower()
            if "do not exist" in detail or "does not exist" in detail:
                deleted_count = len(existing_media_ids)
                delete_note = " Las fotos anteriores ya no existian despues del reemplazo."
            else:
                raise

    message = (
        f"{deleted_count} fotos anteriores eliminadas. "
        f"{uploaded_count} fotos nuevas inyectadas por API ({route})."
        if image_mode == "replace"
        else f"{uploaded_count} fotos nuevas agregadas por API ({route})."
    )
    if delete_note:
        message += delete_note
    if image_errors:
        message += f" No se cargaron {len(image_errors)} de {len(image_urls)} URLs: {' | '.join(image_errors[:3])}"
    return message


def _metafield_value_for_api(column, value, shopify_config=None):
    text = clean_value(value)
    field_type = _metafield_type_from_column(column)
    if field_type in ("metaobject_reference", "list.metaobject_reference"):
        if shopify_config is None:
            return text
        return _resolve_metaobject_reference_value(shopify_config, column, text)
    if field_type.startswith("list.") and text and not text.startswith("["):
        items = [item.strip() for item in text.split(",") if item.strip()]
        return json.dumps(items, ensure_ascii=False)
    if field_type == "boolean":
        lowered = text.lower()
        if lowered in ("true", "1", "yes", "si", "sÃ­"):
            return "true"
        if lowered in ("false", "0", "no"):
            return "false"
    return text


def _list_text_metafield_value(value):
    text = clean_value(value)
    if not text:
        return "[]"
    if text.startswith("["):
        return text
    items = [item.strip() for item in text.split(",") if item.strip()]
    return json.dumps(items, ensure_ascii=False)


def _metafield_namespace_key(column):
    text = clean_value(column)
    if not text.startswith("Metafield: "):
        return "", ""
    name = re.sub(r"\s*\[.+?\]\s*$", "", text.replace("Metafield: ", "", 1)).strip()
    if "." not in name:
        return "", ""
    namespace, key = name.split(".", 1)
    return namespace.strip(), key.strip()


def _top_product_rows(matrixify_df):
    if matrixify_df.empty or "Handle" not in matrixify_df.columns:
        return pd.DataFrame()
    df = matrixify_df.copy()
    df["__HANDLE"] = df["Handle"].map(clean_value)
    df = df[df["__HANDLE"] != ""].copy()
    return df.drop_duplicates(subset=["__HANDLE"], keep="first").copy()


def _variant_rows_for_handle(matrixify_df, handle):
    if matrixify_df.empty or "Handle" not in matrixify_df.columns:
        return pd.DataFrame()
    current_handle = ""
    matching_indexes = []
    target_handle = clean_value(handle)
    for index, row in matrixify_df.iterrows():
        row_handle = clean_value(row.get("Handle"))
        if row_handle:
            current_handle = row_handle
        if current_handle == target_handle:
            matching_indexes.append(index)
    return matrixify_df.loc[matching_indexes].copy() if matching_indexes else pd.DataFrame()


def _product_metafield_value(product_row, product_variant_rows, column):
    value = clean_value(product_row.get(column))
    if value:
        return value
    if product_variant_rows is None or product_variant_rows.empty or column not in product_variant_rows.columns:
        return ""
    for item in product_variant_rows[column].tolist():
        value = clean_value(item)
        if value:
            return value
    return ""


def _valid_price(value):
    text = clean_value(value)
    if not text:
        return ""
    try:
        if float(text.replace(",", ".")) <= 0:
            return ""
    except Exception:
        pass
    return text


def _variant_create_price(value):
    text = clean_value(value)
    if not text:
        return ""
    try:
        number = float(text.replace(",", "."))
    except Exception:
        return text
    if number < 0:
        return ""
    if number == 0:
        return "0"
    return text


def _variant_bulk_input_from_row(
    row,
    option_id=None,
    option_name=None,
    fallback_price=None,
    fallback_compare_at_price=None,
    force_option_name=False,
):
    size = clean_value(row.get("Option1 Value"))
    sku = clean_value(row.get("Variant SKU"))
    price = _variant_create_price(row.get("Variant Price")) or _variant_create_price(fallback_price)
    if not size or not sku:
        return None

    option_name = clean_value(option_name) or clean_value(row.get("Option1 Name")) or "Talla"
    option_value = {"name": size}
    if clean_value(option_id) and not force_option_name:
        option_value["optionId"] = clean_value(option_id)
    else:
        option_value["optionName"] = option_name
    variant = {
        "optionValues": [option_value]
    }
    compare_at_price = _valid_price(row.get("Variant Compare At Price")) or _valid_price(fallback_compare_at_price)
    barcode = clean_value(row.get("Variant Barcode"))
    if price:
        variant["price"] = price
    if compare_at_price:
        variant["compareAtPrice"] = compare_at_price
    if barcode:
        variant["barcode"] = barcode
    variant["inventoryItem"] = {"sku": sku, "tracked": True}
    return variant


def _variant_validation_issues(product_variant_rows):
    issues = []
    seen_skus = set()
    seen_sizes = set()
    for _, row in product_variant_rows.iterrows():
        sku = clean_value(row.get("Variant SKU"))
        size = clean_value(row.get("Option1 Value"))
        price = _valid_price(row.get("Variant Price"))
        barcode = clean_value(row.get("Variant Barcode"))
        row_issues = []
        if not sku:
            row_issues.append("sin SKU")
        elif sku.upper() in seen_skus:
            row_issues.append("SKU duplicado")
        if sku:
            seen_skus.add(sku.upper())
        if not size:
            row_issues.append("sin talla")
        elif size.upper() in seen_sizes:
            row_issues.append("talla duplicada")
        if size:
            seen_sizes.add(size.upper())
        if not price:
            row_issues.append("sin precio")
        if not barcode:
            row_issues.append("sin barcode")
        if row_issues:
            issues.append(f"{sku or '(sin SKU)'}: {', '.join(row_issues)}")
    return issues


def _dedupe_product_variant_rows(product_variant_rows):
    if product_variant_rows is None or product_variant_rows.empty:
        return product_variant_rows, []
    kept_indexes = []
    seen_skus = set()
    duplicate_skus = []
    for index, row in product_variant_rows.iterrows():
        sku = clean_value(row.get("Variant SKU")).upper()
        size = clean_value(row.get("Option1 Value")).upper()
        if sku and sku in seen_skus:
            duplicate_skus.append(f"{sku} ({size or 'sin talla'})")
            continue
        if sku:
            seen_skus.add(sku)
        kept_indexes.append(index)
    messages = []
    if duplicate_skus:
        messages.append(
            f"Se omitieron {len(duplicate_skus)} variantes duplicadas por SKU: "
            + ", ".join(duplicate_skus[:10])
        )
    return product_variant_rows.loc[kept_indexes].copy(), messages


def _size_option_from_product_data(product_data, fallback_name="Talla"):
    options = product_data.get("options") or []
    fallback = clean_value(fallback_name).lower()
    for option in options:
        if clean_value(option.get("name")).lower() == fallback:
            return option
    for option in options:
        if clean_value(option.get("name")).lower() in ("talla", "size", "title"):
            return option
    return options[0] if options else {}


def _price_fallback_from_product_data(product_data):
    for variant in ((product_data.get("variants") or {}).get("nodes")) or []:
        price = _valid_price(variant.get("price"))
        if price:
            return price, _valid_price(variant.get("compareAtPrice"))
    return "", ""


def _price_fallback_from_rows(product_variant_rows):
    for _, row in product_variant_rows.iterrows():
        price = _valid_price(row.get("Variant Price"))
        if price:
            return price, _valid_price(row.get("Variant Compare At Price"))
    return "", ""


def _missing_variant_inputs_from_shopify(product_variant_rows, product_data, force_option_name=False):
    if product_variant_rows.empty:
        return []
    requested_option_name = clean_value(product_variant_rows.iloc[0].get("Option1 Name")) or "Talla"
    size_option = _size_option_from_product_data(product_data, requested_option_name)
    option_id = clean_value(size_option.get("id"))
    option_name = clean_value(size_option.get("name")) or requested_option_name
    existing_skus = {
        clean_value(variant.get("sku") or (variant.get("inventoryItem") or {}).get("sku")).upper()
        for variant in ((product_data.get("variants") or {}).get("nodes")) or []
        if clean_value(variant.get("sku") or (variant.get("inventoryItem") or {}).get("sku"))
    }
    fallback_price, fallback_compare_at_price = _price_fallback_from_product_data(product_data)
    if not fallback_price:
        fallback_price, fallback_compare_at_price = _price_fallback_from_rows(product_variant_rows)

    variants = []
    seen_sizes = set()
    seen_skus = set()
    for _, variant_row in product_variant_rows.iterrows():
        sku = clean_value(variant_row.get("Variant SKU")).upper()
        size = clean_value(variant_row.get("Option1 Value"))
        if not size or not sku:
            continue
        size_keys = _size_lookup_keys(size)
        if sku and (sku in existing_skus or sku in seen_skus):
            continue
        if not sku and any(size_key in seen_sizes for size_key in size_keys):
            continue
        payload = _variant_bulk_input_from_row(
            variant_row,
            option_id=option_id,
            option_name=option_name,
            fallback_price=fallback_price,
            fallback_compare_at_price=fallback_compare_at_price,
            force_option_name=force_option_name,
        )
        if payload:
            variants.append(payload)
            seen_sizes.update(size_keys)
            if sku:
                seen_skus.add(sku)
    return variants


def _expected_variant_skus(product_variant_rows):
    if product_variant_rows is None or product_variant_rows.empty:
        return []
    skus = []
    seen = set()
    for _, row in product_variant_rows.iterrows():
        sku = clean_value(row.get("Variant SKU")).upper()
        size = clean_value(row.get("Option1 Value"))
        if not sku or not size or sku in seen:
            continue
        seen.add(sku)
        skus.append(sku)
    return skus


def _actual_variant_skus(product_data):
    skus = set()
    for variant in ((product_data.get("variants") or {}).get("nodes")) or []:
        inventory_item = variant.get("inventoryItem") or {}
        sku = clean_value(variant.get("sku") or inventory_item.get("sku")).upper()
        if sku:
            skus.add(sku)
    return skus


def _missing_expected_variant_skus(product_variant_rows, product_data):
    actual = _actual_variant_skus(product_data)
    return [sku for sku in _expected_variant_skus(product_variant_rows) if sku not in actual]


def _variant_update_payload_from_row(row, variant_id, fallback_price=None, fallback_compare_at_price=None):
    variant_id = clean_value(variant_id)
    if not variant_id:
        return None
    sku = clean_value(row.get("Variant SKU"))
    price = _valid_price(row.get("Variant Price")) or _valid_price(fallback_price)
    barcode = clean_value(row.get("Variant Barcode"))
    if not price and not sku and not barcode:
        return None
    payload = {
        "id": variant_id,
    }
    if price:
        payload["price"] = price
    compare_at_price = _valid_price(row.get("Variant Compare At Price")) or _valid_price(fallback_compare_at_price)
    if compare_at_price:
        payload["compareAtPrice"] = compare_at_price
    if barcode:
        payload["barcode"] = barcode
    if sku:
        payload["inventoryItem"] = {"sku": sku, "tracked": True}
    return payload


def _inventory_item_update_payload_from_row(row, inventory_item_id):
    inventory_item_id = clean_value(inventory_item_id)
    sku = clean_value(row.get("Variant SKU"))
    if not inventory_item_id or not sku:
        return None
    return {
        "id": inventory_item_id,
        "input": {"sku": sku, "tracked": True},
        "sku": sku,
    }


def _existing_variant_updates_from_shopify(product_variant_rows, product_data):
    if product_variant_rows.empty:
        return []
    requested_option_name = clean_value(product_variant_rows.iloc[0].get("Option1 Name")) or "Talla"
    size_option = _size_option_from_product_data(product_data, requested_option_name)
    option_name = clean_value(size_option.get("name")) or requested_option_name
    fallback_price, fallback_compare_at_price = _price_fallback_from_rows(product_variant_rows)
    expected_by_size = {}
    for _, row in product_variant_rows.iterrows():
        size = clean_value(row.get("Option1 Value")).upper()
        if size:
            _set_row_by_size_keys(expected_by_size, size, row)

    updates = []
    for variant in ((product_data.get("variants") or {}).get("nodes")) or []:
        size = _selected_option_value(variant, option_name).upper()
        row = _row_by_size_keys(expected_by_size, size)
        if row is None:
            continue
        expected_price = _valid_price(row.get("Variant Price")) or _valid_price(fallback_price)
        expected_barcode = clean_value(row.get("Variant Barcode"))
        expected_sku = clean_value(row.get("Variant SKU")).upper()
        current_price = _valid_price(variant.get("price"))
        current_barcode = clean_value(variant.get("barcode"))
        inventory_item = variant.get("inventoryItem") or {}
        current_inventory_sku = clean_value(inventory_item.get("sku")).upper()
        needs_update = (
            (expected_price and current_price != expected_price)
            or (expected_barcode and current_barcode != expected_barcode)
            or (expected_sku and current_inventory_sku != expected_sku)
        )
        if not needs_update:
            continue
        payload = _variant_update_payload_from_row(
            row,
            variant.get("id"),
            fallback_price=fallback_price,
            fallback_compare_at_price=fallback_compare_at_price,
        )
        if payload:
            updates.append(payload)
    return updates


def _existing_inventory_item_updates_from_shopify(product_variant_rows, product_data):
    if product_variant_rows.empty:
        return []
    requested_option_name = clean_value(product_variant_rows.iloc[0].get("Option1 Name")) or "Talla"
    size_option = _size_option_from_product_data(product_data, requested_option_name)
    option_name = clean_value(size_option.get("name")) or requested_option_name
    expected_by_size = {}
    for _, row in product_variant_rows.iterrows():
        size = clean_value(row.get("Option1 Value")).upper()
        if size:
            _set_row_by_size_keys(expected_by_size, size, row)

    updates = []
    for variant in ((product_data.get("variants") or {}).get("nodes")) or []:
        size = _selected_option_value(variant, option_name).upper()
        row = _row_by_size_keys(expected_by_size, size)
        if row is None:
            continue
        expected_sku = clean_value(row.get("Variant SKU")).upper()
        if not expected_sku:
            continue
        inventory_item = variant.get("inventoryItem") or {}
        current_sku = clean_value(inventory_item.get("sku")).upper()
        tracked = bool(inventory_item.get("tracked"))
        if current_sku == expected_sku and tracked:
            continue
        payload = _inventory_item_update_payload_from_row(row, inventory_item.get("id"))
        if payload:
            updates.append(payload)
    return updates


def _apply_inventory_item_updates(shopify_config, inventory_item_updates):
    inventory_ok = 0
    inventory_errors = []
    for inventory_update in inventory_item_updates or []:
        try:
            if inventory_item_update is None:
                raise RuntimeError("Falta actualizar shopify_api.py: no existe inventory_item_update.")
            inventory_item_update(
                shopify_config,
                inventory_update["id"],
                inventory_update["input"],
            )
            inventory_ok += 1
        except Exception as exc:
            inventory_errors.append(f"{inventory_update.get('sku')}: {exc}")
    return inventory_ok, inventory_errors


def _shopify_inventory_target_locations(shopify_config):
    configured = clean_value(
        shopify_config.get("inventory_location_ids")
        or shopify_config.get("inventory_locations")
        or shopify_config.get("location_ids")
    )
    if configured:
        requested = {clean_value(value) for value in re.split(r"[,;|\n]+", configured) if clean_value(value)}
        if not requested:
            return []
        if any(value.startswith("gid://shopify/Location/") for value in requested):
            return [
                {
                    "id": value if value.startswith("gid://shopify/Location/") else f"gid://shopify/Location/{value}",
                    "legacyResourceId": value.rsplit("/", 1)[-1],
                    "name": f"Location {value.rsplit('/', 1)[-1]}",
                }
                for value in sorted(requested)
            ]
        if fetch_locations is None:
            return [{"id": f"gid://shopify/Location/{value}", "legacyResourceId": value, "name": f"Location {value}"} for value in sorted(requested)]
    else:
        requested = set()

    if fetch_locations is None:
        raise RuntimeError(
            "No puedo leer sucursales porque falta fetch_locations. Configura inventory_location_ids en Secrets."
        )
    location_cache_key = f"shopify_locations_cache_{clean_value(shopify_config.get('shop_domain'))}"
    if st.session_state.get(location_cache_key) is not None:
        locations = st.session_state[location_cache_key]
    else:
        try:
            locations = fetch_locations(shopify_config)
            st.session_state[location_cache_key] = locations
        except Exception as exc:
            raise RuntimeError(
                "Shopify no permite leer locations con este token. Configura inventory_location_ids en Secrets "
                "con los IDs/GIDs de las sucursales a activar, o agrega permisos de lectura de locations al token. "
                f"Detalle: {exc}"
            ) from exc
    if configured:
        requested_lower = {value.lower() for value in requested}

        def matches(location):
            location_values = {
                clean_value(location.get("id")),
                clean_value(location.get("legacyResourceId")),
                clean_value(location.get("name")),
            }
            return any(value in requested or value.lower() in requested_lower for value in location_values if value)

        locations = [location for location in locations if matches(location)]
    return [location for location in locations if clean_value(location.get("id"))]


def _inventory_activation_rows_from_product_data(product_data):
    rows = []
    for variant in ((product_data.get("variants") or {}).get("nodes")) or []:
        inventory_item = variant.get("inventoryItem") or {}
        sku = clean_value(inventory_item.get("sku") or variant.get("sku"))
        inventory_item_id = clean_value(inventory_item.get("id"))
        if sku and inventory_item_id:
            rows.append(
                {
                    "SKU": sku,
                    "Variant ID": clean_value(variant.get("legacyResourceId")),
                    "Variant GID": clean_value(variant.get("id")),
                    "Inventory Item ID": clean_value(inventory_item.get("legacyResourceId")),
                    "Inventory Item GID": inventory_item_id,
                    "Tracked": bool(inventory_item.get("tracked")),
                }
            )
    return rows


def _inventory_activation_rows_from_products(shopify_products, only_codes=None, only_skus=None):
    only_codes = {clean_value(value).upper() for value in (only_codes or []) if clean_value(value)}
    only_skus = {clean_value(value).upper() for value in (only_skus or []) if clean_value(value)}
    rows = []
    for product in shopify_products or []:
        mod_col = clean_value(product.get("Mod-Col")).upper()
        handle = clean_value(product.get("Handle"))
        if only_codes and mod_col not in only_codes and handle.upper() not in only_codes:
            continue
        for variant in product.get("Variants") or []:
            sku = clean_value(variant.get("Variant SKU"))
            inventory_gid = clean_value(variant.get("Variant Inventory Item GID"))
            if not sku or not inventory_gid:
                continue
            if only_skus and sku.upper() not in only_skus:
                continue
            rows.append(
                {
                    "Handle": handle,
                    "Mod-Col": mod_col,
                    "SKU": sku,
                    "Variant ID": clean_value(variant.get("Variant ID")),
                    "Inventory Item ID": clean_value(variant.get("Variant Inventory Item ID")),
                    "Inventory Item GID": inventory_gid,
                }
            )
    return rows


def _inventory_activation_progress_path(shopify_config):
    site = clean_value(shopify_config.get("site_key") or shopify_config.get("shop_domain") or "shopify")
    safe_site = re.sub(r"[^a-zA-Z0-9_-]+", "_", site) or "shopify"
    return OUTPUT_DIR / f"inventory_activation_progress_{safe_site}.json"


def _load_inventory_activation_progress(shopify_config):
    path = _inventory_activation_progress_path(shopify_config)
    if not path.exists():
        return {"done": [], "errors": []}
    try:
        with path.open("r", encoding="utf-8") as progress_file:
            data = json.load(progress_file)
        if not isinstance(data, dict):
            return {"done": [], "errors": []}
        data.setdefault("done", [])
        data.setdefault("errors", [])
        return data
    except Exception:
        return {"done": [], "errors": []}


def _save_inventory_activation_progress(shopify_config, progress):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    progress = dict(progress or {})
    progress["updated_at"] = datetime.now(timezone.utc).isoformat()
    with _inventory_activation_progress_path(shopify_config).open("w", encoding="utf-8") as progress_file:
        json.dump(progress, progress_file, ensure_ascii=False, indent=2)


def _inventory_activation_pair_key(inventory_gid, location_id):
    return f"{clean_value(inventory_gid)}|{clean_value(location_id)}"


def _location_lookup_keys(location):
    values = {
        clean_value(location.get("id")),
        clean_value(location.get("legacyResourceId")),
        clean_value(location.get("name")),
    }
    return {value.lower() for value in values if value}


def _active_location_keys_for_inventory_item(shopify_config, inventory_gid):
    cache_key = f"inventory_active_locations_{clean_value(shopify_config.get('shop_domain'))}_{clean_value(inventory_gid)}"
    if st.session_state.get(cache_key) is not None:
        return st.session_state[cache_key]
    if inventory_item_active_locations is None:
        return set()
    locations = inventory_item_active_locations(shopify_config, inventory_gid)
    keys = set()
    for location in locations:
        keys.update(_location_lookup_keys(location))
    st.session_state[cache_key] = keys
    return keys


def _activate_inventory_with_retries(shopify_config, inventory_gid, location_id, available=None, max_attempts=3):
    last_error = None
    for attempt in range(1, max(1, max_attempts) + 1):
        try:
            return inventory_activate(shopify_config, inventory_gid, location_id, available=available)
        except Exception as exc:
            last_error = exc
            message = clean_value(exc)
            retryable = any(
                needle in message.lower()
                for needle in ("timeout", "timed out", "throttled", "429", "temporarily", "502", "503", "504")
            )
            fatal = any(
                needle in message
                for needle in ("ACCESS_DENIED", "Access denied", "Unauthorized", "Invalid API key", "@idempotent directive")
            )
            if fatal or not retryable or attempt >= max_attempts:
                raise
            time.sleep(min(2 * attempt, 8))
    if last_error:
        raise last_error
    return {}


def _inventory_activation_summary_df(result_df):
    if result_df is None or result_df.empty or "Resultado" not in result_df.columns:
        return pd.DataFrame()
    return pd.DataFrame(
        [
            {"Metrica": "Variantes revisadas", "Valor": safe_int_value(result_df["Inventory Item GID"].nunique()) if "Inventory Item GID" in result_df.columns else 0},
            {"Metrica": "Sucursales ya activas", "Valor": safe_int_value((result_df["Resultado"] == "ACTIVO").sum())},
            {"Metrica": "Sucursales activadas correctamente", "Valor": safe_int_value((result_df["Resultado"] == "OK").sum())},
            {"Metrica": "Errores", "Valor": safe_int_value((result_df["Resultado"] == "ERROR").sum())},
            {"Metrica": "Omitidas", "Valor": safe_int_value((result_df["Resultado"] == "OMITIDO").sum())},
        ]
    )


def _activate_inventory_items_in_locations(shopify_config, activation_rows, locations=None, available=None, batch_size=100, max_actions=None):
    if inventory_activate is None:
        raise RuntimeError("Falta actualizar shopify_api.py: no existe inventory_activate.")
    locations = locations if locations is not None else _shopify_inventory_target_locations(shopify_config)
    if not locations:
        return pd.DataFrame(
            [
                {
                    "SKU": "",
                    "Sucursal": "",
                    "Resultado": "ERROR",
                    "Mensaje": "No hay sucursales activas configuradas o disponibles.",
                }
            ]
        )
    unique_items = {}
    for row in activation_rows or []:
        inventory_gid = clean_value(row.get("Inventory Item GID") or row.get("inventoryItemId"))
        sku = clean_value(row.get("SKU") or row.get("sku"))
        if inventory_gid and sku:
            unique_items.setdefault(inventory_gid, {**row, "SKU": sku, "Inventory Item GID": inventory_gid})

    progress = _load_inventory_activation_progress(shopify_config)
    done_keys = set(progress.get("done") or [])
    max_actions = safe_int_value(
        max_actions
        or shopify_config.get("inventory_activation_max_actions")
        or st.session_state.get("inventory_activation_max_actions")
        or 1500
    )
    batch_size = max(1, safe_int_value(batch_size or shopify_config.get("inventory_activation_batch_size") or 100))
    results = []
    activated_this_run = 0
    for item in unique_items.values():
        try:
            active_location_keys = _active_location_keys_for_inventory_item(shopify_config, item["Inventory Item GID"])
        except Exception as exc:
            active_location_keys = set()
            results.append(
                {
                    "Handle": clean_value(item.get("Handle")),
                    "Mod-Col": clean_value(item.get("Mod-Col")),
                    "SKU": item["SKU"],
                    "Inventory Item GID": item["Inventory Item GID"],
                    "Sucursal": "",
                    "Location GID": "",
                    "Estado actual": "NO VALIDADO",
                    "Acción requerida": "REVISAR",
                    "Resultado": "ERROR",
                    "Mensaje": f"No se pudo consultar locations activas: {clean_value(exc)[:500]}",
                }
            )
            continue
        for location in locations:
            location_id = clean_value(location.get("id"))
            location_name = clean_value(location.get("name")) or location_id
            pair_key = _inventory_activation_pair_key(item["Inventory Item GID"], location_id)
            location_keys = _location_lookup_keys(location)
            if pair_key in done_keys:
                results.append(
                    {
                        "Handle": clean_value(item.get("Handle")),
                        "Mod-Col": clean_value(item.get("Mod-Col")),
                        "SKU": item["SKU"],
                        "Inventory Item GID": item["Inventory Item GID"],
                        "Sucursal": location_name,
                        "Location GID": location_id,
                        "Estado actual": "ACTIVO",
                        "Acción requerida": "OMITIR",
                        "Resultado": "ACTIVO",
                        "Mensaje": "Ya procesado en una ejecucion anterior",
                    }
                )
                continue
            if active_location_keys and location_keys.intersection(active_location_keys):
                done_keys.add(pair_key)
                results.append(
                    {
                        "Handle": clean_value(item.get("Handle")),
                        "Mod-Col": clean_value(item.get("Mod-Col")),
                        "SKU": item["SKU"],
                        "Inventory Item GID": item["Inventory Item GID"],
                        "Sucursal": location_name,
                        "Location GID": location_id,
                        "Estado actual": "ACTIVO",
                        "Acción requerida": "OMITIR",
                        "Resultado": "ACTIVO",
                        "Mensaje": "Inventory item ya estaba activo en sucursal",
                    }
                )
                continue
            if max_actions and activated_this_run >= max_actions:
                results.append(
                    {
                        "Handle": clean_value(item.get("Handle")),
                        "Mod-Col": clean_value(item.get("Mod-Col")),
                        "SKU": item["SKU"],
                        "Inventory Item GID": item["Inventory Item GID"],
                        "Sucursal": location_name,
                        "Location GID": location_id,
                        "Estado actual": "NO ACTIVO",
                        "Acción requerida": "ACTIVAR",
                        "Resultado": "PENDIENTE",
                        "Mensaje": f"Lote pausado al llegar a {max_actions:,} activaciones. Ejecuta nuevamente para continuar.",
                    }
                )
                continue
            try:
                _activate_inventory_with_retries(shopify_config, item["Inventory Item GID"], location_id, available=available)
                activated_this_run += 1
                done_keys.add(pair_key)
                results.append(
                    {
                        "Handle": clean_value(item.get("Handle")),
                        "Mod-Col": clean_value(item.get("Mod-Col")),
                        "SKU": item["SKU"],
                        "Inventory Item GID": item["Inventory Item GID"],
                        "Sucursal": location_name,
                        "Location GID": location_id,
                        "Estado actual": "NO ACTIVO",
                        "Acción requerida": "ACTIVAR",
                        "Resultado": "OK",
                        "Mensaje": "Inventario activo en sucursal",
                    }
                )
                if activated_this_run % batch_size == 0:
                    progress["done"] = sorted(done_keys)
                    _save_inventory_activation_progress(shopify_config, progress)
            except Exception as exc:
                error_message = clean_value(exc)
                if "ACCESS_DENIED" in error_message or "Access denied" in error_message or "nego activar inventario" in error_message:
                    error_message = (
                        "Shopify nego activar inventario. El token necesita permiso de escritura de inventario "
                        "(write_inventory / Inventory management). Actualiza los permisos del token o crea un token nuevo con ese scope."
                    )
                results.append(
                    {
                        "Handle": clean_value(item.get("Handle")),
                        "Mod-Col": clean_value(item.get("Mod-Col")),
                        "SKU": item["SKU"],
                        "Inventory Item GID": item["Inventory Item GID"],
                        "Sucursal": location_name,
                        "Location GID": location_id,
                        "Estado actual": "NO ACTIVO",
                        "Acción requerida": "ACTIVAR",
                        "Resultado": "ERROR",
                        "Mensaje": error_message[:500],
                    }
                )
                progress.setdefault("errors", []).append(
                    {
                        "sku": item["SKU"],
                        "inventory_item_gid": item["Inventory Item GID"],
                        "location_gid": location_id,
                        "location": location_name,
                        "error": error_message[:500],
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    }
                )
                progress["done"] = sorted(done_keys)
                _save_inventory_activation_progress(shopify_config, progress)
                fatal_error = (
                    "permiso de escritura de inventario" in error_message
                    or "ACCESS_DENIED" in error_message
                    or "Access denied" in error_message
                    or "@idempotent directive" in error_message
                    or "Invalid API key" in error_message
                    or "Unauthorized" in error_message
                )
                if fatal_error:
                    return pd.DataFrame(results)
    progress["done"] = sorted(done_keys)
    _save_inventory_activation_progress(shopify_config, progress)
    return pd.DataFrame(results)


def activate_product_inventory_locations(shopify_config, product_data):
    rows = _inventory_activation_rows_from_product_data(product_data)
    result_df = _activate_inventory_items_in_locations(shopify_config, rows)
    if result_df.empty or "Resultado" not in result_df.columns:
        return 0, []
    ok_count = safe_int_value((result_df["Resultado"] == "OK").sum())
    errors = (
        result_df.loc[result_df["Resultado"] == "ERROR", "Mensaje"].dropna().map(clean_value).tolist()
        if "Mensaje" in result_df.columns
        else []
    )
    return ok_count, errors


def inventory_activation_filters_from_input(df):
    if df is None or df.empty:
        return set(), set()
    normalized = coalesce_duplicate_columns(df)
    mod_col_col = first_existing_column(
        normalized,
        ["Mod-Col", "COD MOD COL", "Codigo Modelo Color", "Código Modelo Color", "Handle"],
    )
    sku_col = first_existing_column(normalized, ["SKU", "Variant SKU", "SKU de la variante", "CODINT_MA", "Cod Int"])
    codes = set()
    skus = set()
    if mod_col_col:
        codes = {clean_value(value).upper() for value in normalized[mod_col_col].dropna() if clean_value(value)}
    if sku_col:
        skus = {clean_value(value).upper() for value in normalized[sku_col].dropna() if clean_value(value)}
    return codes, skus


def _variant_option_values(product_variant_rows):
    values = []
    for _, row in product_variant_rows.iterrows():
        size = clean_value(row.get("Option1 Value"))
        if size and size not in values:
            values.append(size)
    return values


def _ordered_sizes_from_rows(product_variant_rows):
    if product_variant_rows.empty:
        return []
    ordered = []
    for _, row in product_variant_rows.iterrows():
        size = clean_value(row.get("Option1 Value"))
        if size and size not in ordered:
            ordered.append(size)
    return ordered


def _selected_option_value(variant, option_name):
    expected = clean_value(option_name).lower()
    for option in variant.get("selectedOptions") or []:
        if clean_value(option.get("name")).lower() == expected:
            return clean_value(option.get("value"))
    if (variant.get("selectedOptions") or []):
        return clean_value((variant.get("selectedOptions") or [{}])[0].get("value"))
    return ""


def _verify_shopify_variants(product_variant_rows, product_data):
    expected = {}
    for _, row in product_variant_rows.iterrows():
        sku = clean_value(row.get("Variant SKU")).upper()
        if not sku:
            continue
        expected[sku] = {
            "size": clean_value(row.get("Option1 Value")),
            "price": _valid_price(row.get("Variant Price")),
            "barcode": clean_value(row.get("Variant Barcode")),
        }
    if not expected:
        return ["No hay SKUs esperados para verificar."]

    actual = {}
    for variant in ((product_data.get("variants") or {}).get("nodes")) or []:
        inventory_item = variant.get("inventoryItem") or {}
        sku = clean_value(inventory_item.get("sku") or variant.get("sku")).upper()
        if sku:
            actual[sku] = variant

    problems = []

    def normalize_price(value):
        text = _valid_price(value)
        if not text:
            return ""
        try:
            return f"{float(text.replace(',', '.')):.2f}"
        except ValueError:
            return text

    for sku, expected_values in expected.items():
        variant = actual.get(sku)
        if not variant:
            problems.append(f"{sku}: no existe en Shopify")
            continue
        inventory_item = variant.get("inventoryItem") or {}
        inventory_sku = clean_value(inventory_item.get("sku")).upper()
        if inventory_sku != sku:
            problems.append(f"{sku}: inventory item SKU '{inventory_sku or 'vacio'}', esperado {sku}")
        price = _valid_price(variant.get("price"))
        barcode = clean_value(variant.get("barcode"))
        expected_price = normalize_price(expected_values["price"])
        actual_price = normalize_price(price)
        if expected_values["price"] and not price:
            problems.append(f"{sku}: creado sin precio")
        elif expected_price and actual_price and expected_price != actual_price:
            problems.append(f"{sku}: precio Shopify {actual_price}, esperado {expected_price}")
        if expected_values["barcode"] and not barcode:
            problems.append(f"{sku}: creado sin barcode")
        elif expected_values["barcode"] and barcode and barcode != expected_values["barcode"]:
            problems.append(f"{sku}: barcode Shopify {barcode}, esperado {expected_values['barcode']}")
    return problems


def _reorder_product_sizes(shopify_config, product_gid, product_variant_rows):
    ordered_sizes = _ordered_sizes_from_rows(product_variant_rows)
    if not product_gid or len(ordered_sizes) < 2:
        return ""

    product_data = fetch_product_options_and_variants(shopify_config, product_gid)
    options = product_data.get("options") or []
    variants = ((product_data.get("variants") or {}).get("nodes")) or []
    if not options or not variants:
        return ""

    option_name = clean_value(product_variant_rows.iloc[0].get("Option1 Name")) or "Talla"
    size_option = None
    for option in options:
        if clean_value(option.get("name")).lower() == option_name.lower():
            size_option = option
            break
    if size_option is None:
        size_option = options[0]
        option_name = clean_value(size_option.get("name")) or option_name

    existing_values = [
        clean_value(option_value.get("name"))
        for option_value in size_option.get("optionValues") or []
        if clean_value(option_value.get("name"))
    ]
    values_in_order = [size for size in ordered_sizes if size in existing_values]
    values_in_order.extend(value for value in existing_values if value not in values_in_order)

    variant_by_size = {}
    for variant in variants:
        size = _selected_option_value(variant, option_name)
        if size and size not in variant_by_size:
            variant_by_size[size] = variant.get("id")
    variant_order = [size for size in ordered_sizes if size in variant_by_size]
    variant_order.extend(size for size in variant_by_size if size not in variant_order)
    if len(variant_order) < len(ordered_sizes):
        missing_sizes = [size for size in ordered_sizes if size not in variant_by_size]
        raise ShopifyApiError(f"No se puede ordenar porque faltan variantes creadas: {', '.join(missing_sizes)}")
    positions = [
        {"id": variant_by_size[size], "position": position}
        for position, size in enumerate(variant_order, start=1)
        if clean_value(variant_by_size.get(size))
    ]
    if not positions:
        raise ShopifyApiError("No se encontraron variantes para ordenar.")
    product_variants_bulk_reorder(shopify_config, product_gid, positions)

    verified_product = fetch_product_options_and_variants(shopify_config, product_gid)
    verified_variants = ((verified_product.get("variants") or {}).get("nodes")) or []
    verified_sizes = [
        _selected_option_value(variant, option_name)
        for variant in verified_variants
        if _selected_option_value(variant, option_name)
    ]
    expected_prefix = [size for size in ordered_sizes if size in verified_sizes]
    if verified_sizes[: len(expected_prefix)] != expected_prefix:
        raise ShopifyApiError(
            f"Shopify no confirmo el orden. Esperado: {', '.join(expected_prefix)}. Actual: {', '.join(verified_sizes[:len(expected_prefix)])}"
        )
    return "orden obligatorio de variantes confirmado"


def apply_full_product_updates(shopify_config, matrixify_df, progress_callback=None, activate_inventory_locations=True):
    rows = []
    product_rows = _top_product_rows(matrixify_df)
    total_products = len(product_rows)
    metafield_columns = [
        column
        for column in matrixify_df.columns
        if clean_value(column).startswith("Metafield: ")
        and column not in (
            "Metafield: custom.guia_de_tallas [page_reference]",
        )
    ]
    brand_metafield_column = "Metafield: custom.marca [single_line_text_field]"
    if brand_metafield_column not in metafield_columns:
        metafield_columns.append(brand_metafield_column)

    for position, (_, row) in enumerate(product_rows.iterrows(), start=1):
        handle = clean_value(row.get("Handle"))
        product_id = clean_value(row.get("ID"))
        product_gid = _product_gid(product_id)
        product_messages = []
        product_status = "OK"
        product_started_at = time.perf_counter()
        if progress_callback:
            progress_callback(position, total_products, handle, "Preparando producto")
        product_variant_rows = _variant_rows_for_handle(matrixify_df, handle)
        product_variant_rows, dedupe_messages = _dedupe_product_variant_rows(product_variant_rows)
        expected_variant_skus = _expected_variant_skus(product_variant_rows)
        if dedupe_messages:
            product_status = "PARCIAL"
            product_messages.extend(dedupe_messages)
        variant_input_issues = _variant_validation_issues(product_variant_rows)
        was_new_product = not bool(product_gid)

        try:
            if variant_input_issues:
                product_status = "PARCIAL"
                product_messages.append("Variantes incompletas no se enviaran completas: " + " | ".join(variant_input_issues[:8]))
            if expected_variant_skus:
                product_messages.append(f"Variantes esperadas BigQuery: {len(expected_variant_skus)}")

            status = clean_value(row.get("Status")).upper()
            if status == "ACTIVE":
                shopify_status = "ACTIVE"
            elif status == "DRAFT":
                shopify_status = "DRAFT"
            else:
                shopify_status = None

            if product_gid:
                if progress_callback:
                    progress_callback(position, total_products, handle, "Actualizando producto base")
                product_update(
                    shopify_config,
                    product_gid,
                    title=clean_value(row.get("Title")) or None,
                    body_html=clean_value(row.get("Body HTML")) or None,
                    tags=_split_tags(row.get("Tags")) if clean_value(row.get("Tags")) else None,
                    vendor=clean_value(row.get("Vendor")) or None,
                    product_type=clean_value(row.get("Type")) or None,
                    status=shopify_status,
                )
                product_messages.append("Producto actualizado")
            else:
                if progress_callback:
                    progress_callback(position, total_products, handle, "Creando producto")
                created_product = product_create(
                    shopify_config,
                    title=clean_value(row.get("Title")) or handle,
                    handle=handle or None,
                    body_html=clean_value(row.get("Body HTML")) or None,
                    tags=_split_tags(row.get("Tags")) if clean_value(row.get("Tags")) else None,
                    vendor=clean_value(row.get("Vendor")) or None,
                    product_type=clean_value(row.get("Type")) or None,
                    status=shopify_status or "ACTIVE",
                    option_name=clean_value(row.get("Option1 Name")) or "Talla",
                    option_values=_variant_option_values(product_variant_rows),
                )
                product_gid = clean_value(created_product.get("id"))
                product_id = clean_value(created_product.get("legacyResourceId")) or product_id
                product_messages.append("Producto nuevo creado")

            publish_date = publication_date_from_row(row)
            if shopify_status != "DRAFT":
                try:
                    if progress_callback:
                        progress_callback(position, total_products, handle, "Publicando producto")
                    publishable_publish(shopify_config, product_gid, publish_date=publish_date)
                    if publish_date:
                        product_messages.append(f"Publicacion programada: {publish_date}")
                    else:
                        product_messages.append("Publicado en Online Store")
                except Exception as exc:
                    product_status = "PARCIAL"
                    product_messages.append(f"Error publicacion: {exc}")

            metafields = []
            skipped_metafields = []
            metafield_errors = []
            for column in metafield_columns:
                value = _product_metafield_value(row, product_variant_rows, column)
                if column == brand_metafield_column and value == "":
                    value = clean_value(row.get("Vendor"))
                if value == "":
                    continue
                namespace, key = _metafield_namespace_key(column)
                if not namespace or not key:
                    continue
                can_write, skip_reason = _metafield_can_write_direct(column)
                if not can_write:
                    skipped_metafields.append(skip_reason)
                    continue
                try:
                    api_value = _metafield_value_for_api(column, value, shopify_config)
                except Exception as exc:
                    metafield_errors.append(f"{namespace}.{key}: {exc}")
                    continue
                metafields.append(
                    {
                        "ownerId": product_gid,
                        "namespace": namespace,
                        "key": key,
                        "type": _metafield_type_from_column(column),
                        "value": api_value,
                    }
                )
            if metafields:
                if progress_callback:
                    progress_callback(position, total_products, handle, f"Actualizando {len(metafields)} metafields")
                metafield_ok = 0
                for metafield in metafields:
                    try:
                        metafields_set(shopify_config, [metafield])
                        metafield_ok += 1
                    except Exception as exc:
                        metafield_errors.append(f"{metafield['namespace']}.{metafield['key']}: {exc}")
                if metafield_ok:
                    product_messages.append(f"{metafield_ok} metafields actualizados")
                if metafield_errors:
                    product_status = "PARCIAL"
                    product_messages.append("Errores metafields: " + " | ".join(metafield_errors[:5]))
            elif metafield_errors:
                product_status = "PARCIAL"
                product_messages.append("Errores metafields: " + " | ".join(metafield_errors[:5]))
            if skipped_metafields:
                product_status = "PARCIAL" if product_status == "OK" else product_status
                product_messages.append("Metafields omitidos: " + " | ".join(dict.fromkeys(skipped_metafields)))

            raw_image_urls = [url.strip() for url in clean_value(row.get("Image Src")).split(";") if url.strip()]
            if raw_image_urls:
                try:
                    if progress_callback:
                        progress_callback(position, total_products, handle, f"Procesando {min(len(raw_image_urls), 10)} fotos")
                    existing_media_ids = [
                        media_id.strip()
                        for media_id in clean_value(row.get("Media IDs")).split(";")
                        if media_id.strip()
                    ]
                    if clean_value(row.get("Image Command")).upper() == "REPLACE" and existing_media_ids:
                        product_delete_media(shopify_config, product_gid, existing_media_ids)
                        product_messages.append(f"{len(existing_media_ids)} fotos anteriores eliminadas")
                    product_files = []
                    image_errors = []
                    for image_index, raw_image_url in enumerate(raw_image_urls[:10], start=1):
                        try:
                            image_bytes, mime_type, filename, source_url = _download_image_bytes(raw_image_url)
                            resource_url = staged_upload_image(shopify_config, filename, mime_type, image_bytes)
                            product_files.append(
                                {
                                    "originalSource": resource_url,
                                    "alt": clean_value(row.get("Image Alt Text")) or clean_value(row.get("Title")),
                                    "filename": filename,
                                    "contentType": "IMAGE",
                                    "duplicateResolutionMode": "APPEND_UUID",
                                }
                            )
                        except Exception as exc:
                            image_errors.append(f"foto {image_index}: {exc}")
                    if product_files:
                        try:
                            route = _product_set_files_with_fallback(shopify_config, product_gid, product_files)
                            product_messages.append(f"{len(product_files)} fotos enviadas por {route}")
                        except Exception as upload_exc:
                            direct_media = product_create_media(shopify_config, product_gid, raw_image_urls[:10])
                            if direct_media:
                                product_messages.append(
                                    f"{len(direct_media)} fotos enviadas por URL directa tras fallback. "
                                    f"Ruta staged fallo: {upload_exc}"
                                )
                            else:
                                raise
                    elif raw_image_urls:
                        direct_media = product_create_media(shopify_config, product_gid, raw_image_urls[:10])
                        if direct_media:
                            product_messages.append(f"{len(direct_media)} fotos enviadas por URL directa")
                    if image_errors:
                        product_status = "PARCIAL"
                        product_messages.append(
                            f"Fotos no cargadas: {len(image_errors)} de {min(len(raw_image_urls), 10)}. "
                            f"Detalle: {' | '.join(image_errors[:3])}"
                        )
                except Exception as exc:
                    product_status = "PARCIAL"
                    product_messages.append(f"Error fotos: {exc}")

            if progress_callback:
                progress_callback(position, total_products, handle, "Leyendo variantes Shopify")
            product_data_for_variants = fetch_product_options_and_variants(shopify_config, product_gid)
            existing_variant_updates = _existing_variant_updates_from_shopify(
                product_variant_rows,
                product_data_for_variants,
            )
            if existing_variant_updates:
                try:
                    if progress_callback:
                        progress_callback(position, total_products, handle, f"Actualizando {len(existing_variant_updates)} variantes")
                    if product_variants_bulk_update is None:
                        raise RuntimeError("Falta actualizar shopify_api.py: no existe product_variants_bulk_update.")
                    updated_variants = product_variants_bulk_update(
                        shopify_config,
                        product_gid,
                        existing_variant_updates,
                    )
                    product_messages.append(
                        f"{len(updated_variants)} variantes existentes actualizadas con SKU/precio/barcode"
                    )
                    product_data_for_variants = fetch_product_options_and_variants(shopify_config, product_gid)
                except Exception as exc:
                    product_status = "PARCIAL"
                    product_messages.append(f"Error actualizando variantes existentes: {exc}")

            inventory_item_updates = _existing_inventory_item_updates_from_shopify(
                product_variant_rows,
                product_data_for_variants,
            )
            if inventory_item_updates:
                if progress_callback:
                    progress_callback(position, total_products, handle, f"Actualizando {len(inventory_item_updates)} inventory items")
                inventory_ok, inventory_errors = _apply_inventory_item_updates(
                    shopify_config,
                    inventory_item_updates,
                )
                if inventory_ok:
                    product_messages.append(f"{inventory_ok} SKUs/tracking actualizados en inventory item")
                    product_data_for_variants = fetch_product_options_and_variants(shopify_config, product_gid)
                if inventory_errors:
                    product_status = "PARCIAL"
                    product_messages.append("Error SKU inventory item: " + " | ".join(inventory_errors[:5]))

            missing_variants = _missing_variant_inputs_from_shopify(
                product_variant_rows,
                product_data_for_variants,
            )
            if missing_variants:
                try:
                    if progress_callback:
                        progress_callback(position, total_products, handle, f"Creando {len(missing_variants)} variantes faltantes")
                    try:
                        created_variants = product_variants_bulk_create(
                            shopify_config,
                            product_gid,
                            missing_variants,
                            strategy="REMOVE_STANDALONE_VARIANT" if was_new_product else None,
                        )
                    except Exception as first_variant_exc:
                        fallback_missing_variants = _missing_variant_inputs_from_shopify(
                            product_variant_rows,
                            product_data_for_variants,
                            force_option_name=True,
                        )
                        if not fallback_missing_variants:
                            raise
                        created_variants = product_variants_bulk_create(
                            shopify_config,
                            product_gid,
                            fallback_missing_variants,
                            strategy="REMOVE_STANDALONE_VARIANT" if was_new_product else None,
                        )
                        product_messages.append(
                            f"Variantes creadas con fallback por nombre de opcion tras error inicial: {first_variant_exc}"
                        )
                        missing_variants = fallback_missing_variants
                    if created_variants:
                        product_messages.append(
                            f"{len(created_variants)} variantes creadas de {len(missing_variants)} faltantes"
                        )
                    if len(created_variants) < len(missing_variants):
                        product_status = "PARCIAL"
                        product_messages.append(
                            f"Shopify confirmo menos variantes de las enviadas: {len(created_variants)} de {len(missing_variants)}"
                        )
                except Exception as exc:
                    product_status = "PARCIAL"
                    product_messages.append(f"Error variantes: {exc}")

            if progress_callback:
                progress_callback(position, total_products, handle, "Verificando variantes creadas")
            product_data_for_variants = fetch_product_options_and_variants(shopify_config, product_gid)
            still_missing_skus = _missing_expected_variant_skus(product_variant_rows, product_data_for_variants)
            if still_missing_skus:
                retry_rows = product_variant_rows[
                    product_variant_rows["Variant SKU"].map(lambda value: clean_value(value).upper()).isin(still_missing_skus)
                ].copy()
                retry_missing_variants = _missing_variant_inputs_from_shopify(
                    retry_rows,
                    product_data_for_variants,
                    force_option_name=True,
                )
                if retry_missing_variants:
                    try:
                        retry_created = product_variants_bulk_create(
                            shopify_config,
                            product_gid,
                            retry_missing_variants,
                            strategy=None,
                        )
                        product_messages.append(
                            f"Reintento variantes: {len(retry_created)} creadas de {len(retry_missing_variants)} faltantes"
                        )
                        product_data_for_variants = fetch_product_options_and_variants(shopify_config, product_gid)
                        still_missing_skus = _missing_expected_variant_skus(product_variant_rows, product_data_for_variants)
                    except Exception as exc:
                        product_status = "PARCIAL"
                        product_messages.append(f"Error reintento variantes: {exc}")
                if still_missing_skus:
                    product_status = "PARCIAL"
                    product_messages.append(
                        "SKUs de variantes no creados en Shopify: " + ", ".join(still_missing_skus[:20])
                    )
            post_create_inventory_updates = _existing_inventory_item_updates_from_shopify(
                product_variant_rows,
                product_data_for_variants,
            )
            if post_create_inventory_updates:
                if progress_callback:
                    progress_callback(position, total_products, handle, f"Reforzando {len(post_create_inventory_updates)} inventory items")
                inventory_ok, inventory_errors = _apply_inventory_item_updates(
                    shopify_config,
                    post_create_inventory_updates,
                )
                if inventory_ok:
                    product_messages.append(f"{inventory_ok} SKUs/tracking reforzados despues de crear variantes")
                    product_data_for_variants = fetch_product_options_and_variants(shopify_config, product_gid)
                if inventory_errors:
                    product_status = "PARCIAL"
                    product_messages.append("Error SKU post-creacion: " + " | ".join(inventory_errors[:5]))

            if progress_callback:
                progress_callback(position, total_products, handle, "Confirmando producto en Shopify")
            verified_product_data = fetch_product_options_and_variants(shopify_config, product_gid)
            actual_variant_skus = _actual_variant_skus(verified_product_data)
            if expected_variant_skus:
                confirmed_count = len([sku for sku in expected_variant_skus if sku in actual_variant_skus])
                product_messages.append(
                    f"Variantes confirmadas Shopify: {confirmed_count}/{len(expected_variant_skus)}"
                )
            if activate_inventory_locations:
                try:
                    if progress_callback:
                        progress_callback(position, total_products, handle, "Activando inventario en sucursales")
                    activation_ok, activation_errors = activate_product_inventory_locations(
                        shopify_config,
                        verified_product_data,
                    )
                    if activation_ok:
                        product_messages.append(f"{activation_ok} activaciones de inventario en sucursales")
                    if activation_errors:
                        product_status = "PARCIAL"
                        product_messages.append("Error activando inventario: " + " | ".join(activation_errors[:5]))
                except Exception as exc:
                    product_status = "PARCIAL"
                    product_messages.append(f"Error activacion inventario sucursales: {exc}")
            else:
                product_messages.append("Activacion de sucursales omitida en esta sincronizacion; ejecutar carga parcial de inventario despues")

            variant_sync_problems = _verify_shopify_variants(product_variant_rows, verified_product_data)
            if variant_sync_problems:
                product_status = "PARCIAL"
                product_messages.append("Verificacion variantes: " + " | ".join(variant_sync_problems[:8]))

            try:
                if progress_callback:
                    progress_callback(position, total_products, handle, "Ordenando tallas")
                reorder_message = _reorder_product_sizes(shopify_config, product_gid, product_variant_rows)
                if reorder_message:
                    product_messages.append(reorder_message)
            except Exception as exc:
                product_status = "PARCIAL"
                product_messages.append(f"Error orden tallas: {exc}")

            rows.append(
                {
                    "Handle": handle,
                    "ID": product_id,
                    "Resultado": product_status,
                    "Duracion segundos": round(time.perf_counter() - product_started_at, 2),
                    "Mensaje": ". ".join(product_messages) or "Sin cambios aplicados",
                }
            )
            if progress_callback:
                progress_callback(position, total_products, handle, f"Finalizado {product_status}", " | ".join(product_messages[-3:]))
        except Exception as exc:
            rows.append(
                {
                    "Handle": handle,
                    "ID": product_id,
                    "Resultado": "ERROR",
                    "Duracion segundos": round(time.perf_counter() - product_started_at, 2),
                    "Mensaje": str(exc),
                }
            )
            if progress_callback:
                progress_callback(position, total_products, handle, "Error", str(exc))
    return pd.DataFrame(rows)


def _now_lima_text():
    return datetime.now(timezone(timedelta(hours=-5))).strftime("%Y-%m-%d %H:%M:%S")


def _sync_job_safe_id(value):
    text = clean_value(value).lower()
    text = re.sub(r"[^a-z0-9_-]+", "-", text).strip("-")
    return text or "job"


def _sync_job_path(job_id):
    return SYNC_JOB_DIR / f"{_sync_job_safe_id(job_id)}.json"


def _sync_job_data_path(job_id):
    return SYNC_JOB_DIR / f"{_sync_job_safe_id(job_id)}.pkl"


def _sync_job_product_key_series(df, mode="full"):
    if df is None or df.empty:
        return pd.Series(dtype=object)
    if clean_value(mode).startswith("partial"):
        handle = df.get("Handle", pd.Series("", index=df.index)).map(clean_value)
        mod_col = df.get("Mod-Col", pd.Series("", index=df.index)).map(clean_value)
        product_id = df.get("Product ID", pd.Series("", index=df.index)).map(clean_value)
        key = handle.where(handle != "", mod_col)
        key = key.where(key != "", product_id)
        key = key.where(key != "", pd.Series([f"fila-{idx}" for idx in df.index], index=df.index))
        return key.map(clean_value)
    if "Handle" not in df.columns:
        return pd.Series([f"fila-{idx}" for idx in df.index], index=df.index, dtype=object)
    handle = df["Handle"].map(clean_value)
    mod_col_column = "Metafield: custom.codigo_modelo_color [id]"
    mod_col = df.get(mod_col_column, pd.Series("", index=df.index)).map(clean_value)
    key = handle.where(handle != "", mod_col)
    key = key.replace("", pd.NA).ffill().fillna("")
    key = key.where(key != "", pd.Series([f"fila-{idx}" for idx in df.index], index=df.index))
    return key.map(clean_value)


def _sync_job_product_keys(df, mode="full"):
    if df is None or df.empty:
        return []
    keys = _sync_job_product_key_series(df, mode=mode)
    return list(dict.fromkeys([key for key in keys.tolist() if clean_value(key)]))


def _sync_job_subset_df(df, product_key, mode="full"):
    if df is None or df.empty:
        return pd.DataFrame()
    keys = _sync_job_product_key_series(df, mode=mode)
    subset = df.loc[keys == product_key].copy()
    return subset


def _save_sync_job(job):
    SYNC_JOB_DIR.mkdir(parents=True, exist_ok=True)
    job["updated_at"] = _now_lima_text()
    with _sync_job_path(job["id"]).open("w", encoding="utf-8") as job_file:
        json.dump(job, job_file, ensure_ascii=False, indent=2)


def _load_sync_job(job_id):
    path = _sync_job_path(job_id)
    if not path.exists():
        return None
    try:
        with path.open("r", encoding="utf-8") as job_file:
            return json.load(job_file)
    except Exception:
        return None


def _load_sync_job_df(job_id):
    path = _sync_job_data_path(job_id)
    if not path.exists():
        return pd.DataFrame()
    try:
        with path.open("rb") as data_file:
            return pickle.load(data_file)
    except Exception:
        return pd.DataFrame()


def _create_sync_job(site_key, mode, source_df, batch_size=20, activate_inventory_locations=True):
    SYNC_JOB_DIR.mkdir(parents=True, exist_ok=True)
    product_keys = _sync_job_product_keys(source_df, mode=mode)
    job_id = (
        f"{_sync_job_safe_id(site_key)}_{_sync_job_safe_id(mode)}_"
        f"{datetime.now(timezone(timedelta(hours=-5))).strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}"
    )
    batch_size = max(1, int(batch_size or 20))
    job = {
        "id": job_id,
        "site_key": clean_value(site_key),
        "mode": clean_value(mode),
        "status": "pending",
        "batch_size": batch_size,
        "total_products": len(product_keys),
        "processed_products": 0,
        "ok_products": 0,
        "partial_products": 0,
        "error_products": 0,
        "current_block": 0,
        "total_blocks": int((len(product_keys) + batch_size - 1) / batch_size) if product_keys else 0,
        "product_keys": product_keys,
        "completed_keys": [],
        "error_keys": [],
        "pending_keys": product_keys,
        "result_rows": [],
        "events": [],
        "activate_inventory_locations": bool(activate_inventory_locations),
        "created_at": _now_lima_text(),
        "updated_at": _now_lima_text(),
    }
    with _sync_job_data_path(job_id).open("wb") as data_file:
        pickle.dump(source_df.copy(), data_file)
    _save_sync_job(job)
    return job


def _latest_sync_job(site_key, mode):
    if not SYNC_JOB_DIR.exists():
        return None
    safe_site = clean_value(site_key)
    safe_mode = clean_value(mode)
    candidates = []
    for path in SYNC_JOB_DIR.glob("*.json"):
        job = _load_sync_job(path.stem)
        if not job:
            continue
        if clean_value(job.get("site_key")) == safe_site and clean_value(job.get("mode")) == safe_mode:
            candidates.append(job)
    if not candidates:
        return None
    candidates.sort(key=lambda item: clean_value(item.get("updated_at")), reverse=True)
    return candidates[0]


def _sync_job_result_df(job):
    rows = job.get("result_rows") or []
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def _sync_job_summary_df(job):
    total = int(job.get("total_products") or 0)
    processed = int(job.get("processed_products") or 0)
    errors = int(job.get("error_products") or 0)
    pending = max(total - processed, 0)
    return pd.DataFrame(
        [
            {"Indicador": "Job ID", "Valor": job.get("id")},
            {"Indicador": "Estado", "Valor": job.get("status")},
            {"Indicador": "Total productos", "Valor": total},
            {"Indicador": "Procesados", "Valor": processed},
            {"Indicador": "Pendientes", "Valor": pending},
            {"Indicador": "Sin observaciones", "Valor": int(job.get("ok_products") or 0)},
            {"Indicador": "Creados con observacion", "Valor": int(job.get("partial_products") or 0)},
            {"Indicador": "Errores", "Valor": errors},
            {"Indicador": "Bloque actual", "Valor": f"{int(job.get('current_block') or 0)} / {int(job.get('total_blocks') or 0)}"},
            {"Indicador": "Actualizado", "Valor": job.get("updated_at")},
        ]
    )


def _is_transient_sync_error(message):
    text = clean_value(message).lower()
    transient_terms = (
        "timeout",
        "timed out",
        "rate limit",
        "throttle",
        "throttled",
        "too many requests",
        "429",
        "502",
        "503",
        "504",
        "temporarily",
        "connection reset",
        "remote end closed",
    )
    return any(term in text for term in transient_terms)


def _sync_job_run_one_product(shopify_config, source_df, product_key, mode, activate_inventory_locations, progress_callback=None):
    product_df = _sync_job_subset_df(source_df, product_key, mode=mode)
    if product_df.empty:
        return pd.DataFrame(
            [
                {
                    "Handle": product_key,
                    "Resultado": "ERROR",
                    "Mensaje": "No se encontro el producto dentro del snapshot del job.",
                }
            ]
        )
    if clean_value(mode).startswith("partial"):
        return apply_shopify_preview(shopify_config, product_df, progress_callback=progress_callback)
    return apply_full_product_updates(
        shopify_config,
        product_df,
        progress_callback=progress_callback,
        activate_inventory_locations=activate_inventory_locations,
    )


def _append_sync_job_event(job, product_key, stage, detail=""):
    events = job.setdefault("events", [])
    events.append(
        {
            "Fecha": _now_lima_text(),
            "Producto": clean_value(product_key),
            "Etapa": clean_value(stage),
            "Detalle": clean_value(detail)[:500],
        }
    )
    if len(events) > 250:
        job["events"] = events[-250:]


def process_sync_job_next_block(job_id, shopify_config, max_retries=2, progress_callback=None):
    job = _load_sync_job(job_id)
    if not job:
        raise RuntimeError(f"No se encontro el job {job_id}")
    source_df = _load_sync_job_df(job_id)
    if source_df.empty:
        raise RuntimeError("No se encontro el snapshot de datos del job. Crea nuevamente el job.")
    pending_keys = [key for key in job.get("pending_keys", []) if key not in set(job.get("completed_keys", []))]
    if not pending_keys:
        job["status"] = "completed" if not job.get("error_keys") else "completed_with_errors"
        _save_sync_job(job)
        return job

    batch_size = max(1, int(job.get("batch_size") or 20))
    block_keys = pending_keys[:batch_size]
    job["status"] = "running"
    job["current_block"] = int(job.get("current_block") or 0) + 1
    _append_sync_job_event(job, "", "Inicio bloque", f"{len(block_keys)} productos")
    _save_sync_job(job)

    for block_position, product_key in enumerate(block_keys, start=1):
        last_error = ""
        product_result_df = pd.DataFrame()
        total_products = max(1, int(job.get("total_products") or len(job.get("product_keys") or []) or len(block_keys)))
        global_position = min(len(job.get("completed_keys", [])) + 1, total_products)
        current_block = int(job.get("current_block") or 0)
        total_blocks = int(job.get("total_blocks") or 0)

        def job_progress_callback(_current, _total, inner_handle, stage, message=""):
            if progress_callback:
                block_label = f"Bloque {current_block}/{total_blocks}" if total_blocks else f"Bloque {current_block}"
                progress_callback(
                    global_position,
                    total_products,
                    inner_handle or product_key,
                    f"{block_label} - {stage}",
                    message,
                )

        for attempt in range(1, max(1, int(max_retries or 1)) + 1):
            try:
                if progress_callback:
                    job_progress_callback(block_position, len(block_keys), product_key, f"Procesando intento {attempt}")
                product_result_df = _sync_job_run_one_product(
                    shopify_config,
                    source_df,
                    product_key,
                    job.get("mode"),
                    bool(job.get("activate_inventory_locations")),
                    progress_callback=job_progress_callback,
                )
                result_text = ""
                message_text = ""
                if product_result_df is not None and not product_result_df.empty:
                    result_text = " ".join(product_result_df.get("Resultado", pd.Series(dtype=object)).map(clean_value).tolist()).upper()
                    message_text = " | ".join(product_result_df.get("Mensaje", pd.Series(dtype=object)).map(clean_value).tolist())
                if "ERROR" not in result_text:
                    break
                last_error = message_text or "Resultado ERROR"
                if not _is_transient_sync_error(last_error) or attempt >= max_retries:
                    break
                time.sleep(min(2 * attempt, 8))
            except Exception as exc:
                last_error = clean_value(exc)
                if not _is_transient_sync_error(last_error) or attempt >= max_retries:
                    product_result_df = sync_error_result(product_key, exc)
                    break
                time.sleep(min(2 * attempt, 8))

        if product_result_df is None or product_result_df.empty:
            product_result_df = pd.DataFrame(
                [{"Handle": product_key, "Resultado": "ERROR", "Mensaje": last_error or "Sin resultado"}]
            )
        result_records = product_result_df.to_dict("records")
        job.setdefault("result_rows", []).extend(result_records)
        result_values = [clean_value(row.get("Resultado")).upper() for row in result_records]
        if any(value == "ERROR" for value in result_values):
            if product_key not in job["error_keys"]:
                job["error_keys"].append(product_key)
            job["error_products"] = int(job.get("error_products") or 0) + 1
            _append_sync_job_event(job, product_key, "Error", last_error)
        elif any(value == "PARCIAL" for value in result_values):
            job["partial_products"] = int(job.get("partial_products") or 0) + 1
            _append_sync_job_event(job, product_key, "Parcial", "")
        else:
            job["ok_products"] = int(job.get("ok_products") or 0) + 1
            _append_sync_job_event(job, product_key, "OK", "")

        if product_key not in job["completed_keys"]:
            job["completed_keys"].append(product_key)
        job["pending_keys"] = [key for key in job.get("pending_keys", []) if key != product_key]
        job["processed_products"] = len(job.get("completed_keys", []))
        _save_sync_job(job)

    job["status"] = "completed" if not job.get("pending_keys") and not job.get("error_keys") else ("completed_with_errors" if not job.get("pending_keys") else "pending")
    _append_sync_job_event(job, "", "Fin bloque", job["status"])
    _save_sync_job(job)
    return job


def _reset_sync_job_errors(job_id):
    job = _load_sync_job(job_id)
    if not job:
        return None
    retry_keys = list(dict.fromkeys(job.get("error_keys") or []))
    if retry_keys:
        job["pending_keys"] = list(dict.fromkeys(retry_keys + job.get("pending_keys", [])))
        job["completed_keys"] = [key for key in job.get("completed_keys", []) if key not in set(retry_keys)]
        job["error_keys"] = []
        job["error_products"] = 0
        job["processed_products"] = len(job.get("completed_keys", []))
        job["status"] = "pending"
        _append_sync_job_event(job, "", "Reintento", f"{len(retry_keys)} productos con error vuelven a pendiente")
        _save_sync_job(job)
    return job


def _update_sync_job_batch_size(job_id, batch_size):
    job = _load_sync_job(job_id)
    if not job:
        return None
    batch_size = max(1, int(batch_size or job.get("batch_size") or 20))
    pending_count = len(job.get("pending_keys") or [])
    current_block = int(job.get("current_block") or 0)
    remaining_blocks = int((pending_count + batch_size - 1) / batch_size) if pending_count else 0
    job["batch_size"] = batch_size
    job["total_blocks"] = current_block + remaining_blocks
    job["updated_at"] = _now_lima_text()
    _save_sync_job(job)
    return job


def render_persistent_sync_job_panel(
    shopify_config,
    brand_config,
    source_df,
    mode,
    label,
    activate_inventory_locations=True,
    session_key=None,
):
    if source_df is None or source_df.empty:
        st.warning("No hay datos para sincronizar.")
        return
    site_key = brand_config["site_key"]
    session_key = session_key or f"sync_job_{site_key}_{mode}"
    product_total = len(_sync_job_product_keys(source_df, mode=mode))
    st.markdown("#### Sincronizacion recuperable por bloques")
    st.caption(
        "La app procesa un bloque por ejecucion y guarda avance en disco. Si Streamlit se refresca, puedes continuar desde el ultimo pendiente."
    )
    c1, c2, c3 = st.columns([1, 1, 2])
    with c1:
        batch_choice = st.selectbox(
            "Productos por bloque",
            ["10", "20", "30", "50", "Otro", "Todos pendientes"],
            index=3,
            key=f"{session_key}_batch_choice_v2",
        )
        if batch_choice == "Otro":
            batch_size = int(
                st.number_input(
                    "Cantidad personalizada",
                    min_value=1,
                    max_value=max(product_total, 1),
                    value=min(max(product_total, 1), 100),
                    step=5,
                    key=f"{session_key}_custom_batch_size",
                )
            )
        elif batch_choice == "Todos pendientes":
            batch_size = max(product_total, 1)
            st.caption("Procesara todo lo pendiente del job en el siguiente bloque.")
        else:
            batch_size = int(batch_choice)
    with c2:
        st.metric("Productos detectados", f"{product_total:,}")
    with c3:
        latest_job = _latest_sync_job(site_key, mode)
        if latest_job and st.button("Retomar ultimo job", key=f"{session_key}_resume"):
            st.session_state[session_key] = latest_job["id"]

    job_id = st.session_state.get(session_key)
    job = _load_sync_job(job_id) if job_id else None
    if not job:
        if st.button("Crear job y procesar primer bloque", type="primary", key=f"{session_key}_create"):
            job = _create_sync_job(
                site_key,
                mode,
                source_df,
                batch_size=batch_size,
                activate_inventory_locations=activate_inventory_locations,
            )
            st.session_state[session_key] = job["id"]
            progress_callback = make_sync_progress_callback(label)
            job = process_sync_job_next_block(job["id"], shopify_config, progress_callback=progress_callback)
            clear_shopify_products_cache(site_key)
            st.success("Primer bloque procesado. Puedes continuar con el siguiente bloque o salir y retomarlo luego.")
        else:
            return

    total = int(job.get("total_products") or 0)
    processed = int(job.get("processed_products") or 0)
    pending = max(total - processed, 0)
    progress = processed / total if total else 0
    status = clean_value(job.get("status")) or "pending"
    st.progress(min(max(progress, 0), 1))
    status_cols = st.columns(5)
    status_cols[0].metric("Procesados", f"{processed:,}/{total:,}")
    status_cols[1].metric("Pendientes", f"{pending:,}")
    status_cols[2].metric("Sin observaciones", f"{int(job.get('ok_products') or 0):,}")
    status_cols[3].metric("Con observacion", f"{int(job.get('partial_products') or 0):,}")
    status_cols[4].metric("Errores", f"{int(job.get('error_products') or 0):,}")
    st.info(
        f"Job {job['id']} | estado: {status} | bloque {int(job.get('current_block') or 0)} de {int(job.get('total_blocks') or 0)} | actualizado: {job.get('updated_at')}"
    )
    job_events = job.get("events") or []
    last_product_event = next((event for event in reversed(job_events) if clean_value(event.get("Producto"))), {})
    last_product = clean_value(last_product_event.get("Producto"))
    last_stage = clean_value(last_product_event.get("Etapa"))
    pending_preview = [clean_value(key) for key in (job.get("pending_keys") or [])[:6] if clean_value(key)]
    if last_product or pending_preview:
        st.caption(
            (f"Ultimo producto registrado: {last_product}" + (f" ({last_stage})" if last_stage else "") if last_product else "")
            + (f" | Proximos pendientes: {', '.join(pending_preview)}" if pending_preview else "")
        )

    action_cols = st.columns([1, 1, 1, 2])
    if pending and action_cols[0].button("Continuar siguiente bloque", type="primary", key=f"{session_key}_continue"):
        job = _update_sync_job_batch_size(job["id"], min(max(1, int(batch_size or 20)), max(pending, 1))) or job
        progress_callback = make_sync_progress_callback(label)
        job = process_sync_job_next_block(job["id"], shopify_config, progress_callback=progress_callback)
        clear_shopify_products_cache(site_key)
        st.success("Bloque procesado y avance guardado.")
    if job.get("error_keys") and action_cols[1].button("Reintentar errores", key=f"{session_key}_retry"):
        job = _reset_sync_job_errors(job["id"])
        st.warning("Errores devueltos a pendiente. Presiona Continuar siguiente bloque para reintentarlos.")
    if action_cols[2].button("Crear nuevo job", key=f"{session_key}_new"):
        st.session_state.pop(session_key, None)
        st.rerun()

    result_df = _sync_job_result_df(job)
    if not result_df.empty:
        render_sync_result_summary(result_df, label)
        st.dataframe(result_df.tail(100), use_container_width=True)
    event_df = pd.DataFrame(job.get("events") or [])
    report_bytes = dataframe_to_excel_bytes(
        {
            "Resumen": _sync_job_summary_df(job),
            "Resultado": result_df,
            "Eventos": event_df,
        }
    )
    st.download_button(
        "Descargar reporte del job",
        data=report_bytes,
        file_name=f"resultado_job_{mode}_{site_key}_{job['id']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"{session_key}_download",
    )


SITE_UI_CONFIG = {
    "Columbia.pe": {
        "brand_name": "Columbia",
        "logo_path": "assets/brands/columbia.png",
        "primary_color": "#004B8D",
        "accent_color": "#009FE3",
        "shopify_store": "columbiape.myshopify.com",
    },
    "Rockford.pe": {
        "brand_name": "Rockford",
        "logo_path": "assets/brands/rockford.png",
        "primary_color": "#0B2345",
        "accent_color": "#B0895B",
        "shopify_store": "rockfordpe.myshopify.com",
    },
    "HushPuppies.pe": {
        "brand_name": "Hush Puppies",
        "logo_path": "assets/brands/hushpuppies.png",
        "primary_color": "#4B2E1F",
        "accent_color": "#C49A6C",
        "shopify_store": "hushpuppiespe.myshopify.com",
    },
    "Vans.pe": {
        "brand_name": "Vans",
        "logo_path": "assets/brands/vans.png",
        "primary_color": "#111827",
        "accent_color": "#D71920",
        "shopify_store": "vans-dev.myshopify.com",
    },
    "Patagonia.pe": {
        "brand_name": "Patagonia",
        "logo_path": "assets/brands/patagonia.png",
        "primary_color": "#1D4E89",
        "accent_color": "#F15A24",
        "shopify_store": "patagoniape.myshopify.com",
    },
    "Sorel.pe": {
        "brand_name": "Sorel",
        "logo_path": "assets/brands/sorel.png",
        "primary_color": "#111827",
        "accent_color": "#C2410C",
        "shopify_store": "sorelpe.myshopify.com",
    },
    "MountainHardwear.pe": {
        "brand_name": "Mountain Hardwear",
        "logo_path": "assets/brands/mountainhardwear.png",
        "primary_color": "#B91C1C",
        "accent_color": "#111827",
        "shopify_store": "mountainhardwearpe.myshopify.com",
    },
}


def image_data_uri(path):
    path = Path(path)
    if not path.exists():
        return ""
    suffix = path.suffix.lower().replace(".", "")
    mime_by_suffix = {
        "jpg": "jpeg",
        "jpeg": "jpeg",
        "png": "png",
        "webp": "webp",
        "gif": "gif",
    }
    mime = mime_by_suffix.get(suffix, "png")
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:image/{mime};base64,{encoded}"


def resolve_logo_path(path):
    path = Path(path)
    if path.exists():
        return str(path)

    folder = path.parent
    stem = path.stem
    aliases = {
        "mountainhardwear": ["mhw", "mountainhardwear"],
        "hushpuppies": ["hushpuppies", "hush_puppies"],
    }
    stems = [stem, f"logo_{stem}"]
    for alias in aliases.get(stem, []):
        stems.extend([alias, f"logo_{alias}"])

    for candidate_stem in dict.fromkeys(stems):
        for suffix in ("png", "jpg", "jpeg", "webp"):
            candidate = folder / f"{candidate_stem}.{suffix}"
            if candidate.exists():
                return str(candidate)

    wanted_keys = {re.sub(r"[^a-z0-9]+", "", stem.lower())}
    for alias in aliases.get(stem, []):
        wanted_keys.add(re.sub(r"[^a-z0-9]+", "", alias.lower()))
    if folder.exists():
        for candidate in folder.iterdir():
            if not candidate.is_file() or candidate.suffix.lower().replace(".", "") not in ("png", "jpg", "jpeg", "webp"):
                continue
            candidate_key = candidate.stem.lower()
            candidate_key = re.sub(r"^logo[_-]*", "", candidate_key)
            candidate_key = re.sub(r"\.(png|jpg|jpeg|webp)$", "", candidate_key)
            candidate_key = re.sub(r"[^a-z0-9]+", "", candidate_key)
            if candidate_key in wanted_keys:
                return str(candidate)
    return str(path)


def brand_logo_stem_for_name(brand_name):
    normalized = re.sub(r"[^a-z0-9]+", "", clean_value(brand_name).lower())
    logo_stems = {
        "columbia": "columbia",
        "rockford": "rockford",
        "patagonia": "patagonia",
        "sorel": "sorel",
        "mountainhardwear": "mountainhardwear",
        "hushpuppies": "hushpuppies",
        "hushpuppieskids": "hpk",
        "accesorioshp": "hushpuppies",
        "keds": "keds",
        "vans": "vans",
    }
    return logo_stems.get(normalized, normalized)


def brand_logo_path_for_name(brand_name):
    stem = brand_logo_stem_for_name(brand_name)
    return resolve_logo_path(f"assets/brands/{stem}.png")


def render_html(html, sidebar=False):
    html = repair_mojibake_text(html)
    target = st.sidebar if sidebar else st
    if hasattr(target, "html"):
        target.html(html)
    else:
        target.markdown(html, unsafe_allow_html=True)


def get_site_config(brand_config, shopify_config=None):
    if isinstance(brand_config, str):
        selected_site = brand_config
        site_key = next(
            (key for key, config in SITE_CONFIGS.items() if config["site_label"] == selected_site),
            selected_site,
        )
        brand_config = get_brand_config(site_key)
    ui_config = SITE_UI_CONFIG.get(brand_config["site_label"], {}).copy()
    ui_config.setdefault("brand_name", brand_config["label"])
    ui_config.setdefault("logo_path", ui_config.get("logo") or f"assets/brands/{brand_config['site_key']}.png")
    ui_config["logo"] = ui_config["logo_path"]
    ui_config.setdefault("primary_color", "#17269A")
    ui_config.setdefault("accent_color", "#009FE3")
    ui_config["site_label"] = brand_config["site_label"]
    ui_config["allowed_brands"] = brand_config.get("allowed_arti_brands", [])
    ui_config["output_file"] = brand_config.get("output_filename", "")
    ui_config["shopify_store"] = clean_value((shopify_config or {}).get("shop_domain")) or ui_config.get("shopify_store", "")
    ui_config["api_version"] = clean_value((shopify_config or {}).get("api_version")) or DEFAULT_API_VERSION
    return ui_config


APP_DIR = Path(__file__).resolve().parent
APP_CSS_PATH = APP_DIR / "assets" / "app.css"
APP_CSS_PREFIX = '\n        <style>\n'
APP_CSS_SUFFIX = '</style>\n        '
APP_CSS_PLACEHOLDER_RE = re.compile(
    r"__(?:BRAND_PRIMARY_COLOR|BRAND_ACCENT_COLOR|SITE_LOGO_CSS|SITE_LABEL_CSS)__"
)
_APP_CSS_CACHE = None


def load_app_css():
    global _APP_CSS_CACHE
    if _APP_CSS_CACHE is None:
        try:
            _APP_CSS_CACHE = APP_CSS_PATH.read_text(encoding="utf-8")
        except OSError as exc:
            st.error(
                f"No se pudo leer la hoja de estilos en {APP_CSS_PATH}. "
                f"La app sigue operativa pero sin estilos. Detalle: {exc}"
            )
            _APP_CSS_CACHE = ""
    return _APP_CSS_CACHE


def inject_custom_css(config):
    site_logo_src = image_data_uri(resolve_logo_path(config.get("logo_path") or config.get("logo", "")))
    site_logo_css = f'url("{site_logo_src}")' if site_logo_src else "none"
    site_label_css = clean_value(config.get("site_label")).replace("\\", "\\\\").replace('"', '\\"')
    reemplazos = {
        "__BRAND_PRIMARY_COLOR__": str(config["primary_color"]),
        "__BRAND_ACCENT_COLOR__": str(config["accent_color"]),
        "__SITE_LOGO_CSS__": site_logo_css,
        "__SITE_LABEL_CSS__": site_label_css,
    }
    cuerpo_css = APP_CSS_PLACEHOLDER_RE.sub(lambda match: reemplazos[match.group(0)], load_app_css())
    st.markdown(
        APP_CSS_PREFIX + cuerpo_css + APP_CSS_SUFFIX,
        unsafe_allow_html=True,
    )


def inject_styles(config=None):
    inject_custom_css(config or get_site_config(get_brand_config()))


def render_sidebar_brand():
    forus_src = image_data_uri(FORUS_LOGO_PATH)
    logo_html = (
        f'<img src="{forus_src}" alt="Forus" style="max-width:138px;max-height:54px;object-fit:contain;">'
        if forus_src
        else '<div class="forus-logo">FORUS</div><div class="forus-tagline">CONSUMER FANATIC</div>'
    )
    render_html(f'<div class="forus-sidebar">{logo_html}</div>', sidebar=True)


def render_sidebar_brand_card(config):
    brand_src = image_data_uri(resolve_logo_path(config.get("logo_path") or config.get("logo", "")))
    brand_html = (
        f'<img src="{brand_src}" alt="{config["brand_name"]}">'
        if brand_src
        else f'<p class="sidebar-brand-name">{config["brand_name"]}</p>'
    )
    render_html(
        f"""
        <div class="sidebar-brand-card">
            <div class="sidebar-brand-logo">{brand_html}</div>
            <p class="sidebar-brand-caption">{config["site_label"]}</p>
        </div>
        """,
        sidebar=True,
    )


def render_allowed_brands_card(brand_config):
    allowed_brands = list(brand_config["allowed_arti_brands"])
    primary_brand = brand_config["label"].upper()
    ordered_allowed = [primary_brand] + [brand for brand in allowed_brands if brand != primary_brand]
    chips = []
    rendered_logo_stems = set()
    for index, brand in enumerate(ordered_allowed):
        clean_brand = clean_value(brand)
        logo_stem = brand_logo_stem_for_name(clean_brand)
        if logo_stem in rendered_logo_stems:
            continue
        rendered_logo_stems.add(logo_stem)
        brand_src = image_data_uri(brand_logo_path_for_name(clean_brand))
        brand_label = escape(clean_brand.title())
        visual = f'<img src="{brand_src}" alt="{brand_label}">' if brand_src else f"<span>{brand_label}</span>"
        primary_class = " primary" if not chips else ""
        chips.append(
            f"""
            <div class="allowed-logo-chip{primary_class}" title="{brand_label}">
                {visual}
            </div>
            """
        )
    render_html(
        f"""
        <p class="sidebar-label">Marca(s) permitidas</p>
        <div class="allowed-logo-grid">
            {''.join(chips)}
        </div>
        """,
        sidebar=True,
    )


def render_sidebar_status(config, shopify_config, bigquery_ready, input_loaded=False):
    input_state = "Cargado" if input_loaded else "Pendiente"
    render_html(
        f"""
        <div class="sidebar-card">
            <p class="sidebar-label">Estado operativo</p>
            <p class="sidebar-value">Marca activa: {config["brand_name"]}</p>
            <p class="sidebar-value">Input comercial: {input_state}</p>
            <p class="sidebar-value">Salida: {config["output_file"]}</p>
        </div>
        """,
        sidebar=True,
    )


def render_sidebar_shopify_card(config, shopify_config):
    configured = is_shopify_configured(shopify_config)
    state = "OK" if configured else "Pend."
    domain = clean_value(shopify_config.get("shop_domain")) or config.get("shopify_store") or "No configurado"
    render_html(
        f"""
        <div class="shopify-card-head">
            <h3>Shopify API</h3>
            <span class="status-badge">{state}</span>
        </div>
        <div class="shopify-config-box">Configurado:<br>{domain}</div>
        <p class="shopify-meta">Admin API {config["api_version"]} &middot; Token en Secrets</p>
        """,
        sidebar=True,
    )


def render_top_header(config):
    brand_src = image_data_uri(resolve_logo_path(config.get("logo_path") or config.get("logo", "")))
    shopify_src = image_data_uri(SHOPIFY_LOGO_PATH)
    brand_html = (
        f'<img src="{brand_src}" alt="{config["brand_name"]}">'
        if brand_src
        else f'<div class="brand-fallback">{config["brand_name"]}</div>'
    )
    shopify_html = (
        f'<img src="{shopify_src}" alt="Shopify" style="max-width:44px;max-height:44px;object-fit:contain;">'
        if shopify_src
        else '<div class="shopify-bag">S</div>'
    )
    render_html(
        f"""
        <div class="top-header">
            <div class="brand-lockup">
                <div class="brand-logo-card">{brand_html}</div>
                <div>
                    <p class="header-eyebrow">Catálogo Control Center</p>
                    <h1 class="header-title">{config["site_label"]} &rarr; Shopify</h1>
                    <p class="header-subtitle">Gestiona el catálogo con datos de BigQuery y sincronización directa con Shopify.</p>
                </div>
            </div>
            <div class="shopify-lockup">
                <span class="status-badge blue">BigQuery activo</span>
                <span class="status-badge">Shopify conectado</span>
                {shopify_html}
            </div>
        </div>
        """,
    )


def render_stepper(config, current_step=1):
    steps = [
        ("Input", "Archivo comercial"),
        ("BigQuery", "Fuente maestra"),
        ("Validación", "Reglas y cruces"),
        ("Shopify", "Sincronización final"),
    ]
    items = []
    for index, (title, caption) in enumerate(steps, start=1):
        current = " current" if index == current_step else ""
        status = "Actual" if index == 1 else ("OK" if index == 2 else ("Revisar" if index == 3 else "Pend."))
        tone = "blue" if index == 1 else ("" if index == 2 else (" warn" if index == 3 else " blue"))
        items.append(
            f"""
            <div class="step-card{current}">
                <span class="step-index">{index}</span>
                <div style="min-width:0;flex:1;">
                    <p class="step-title">{title}</p>
                    <p class="step-caption">{caption}</p>
                </div>
                <span class="status-badge{tone}">{status}</span>
            </div>
            """
        )
    render_html(f'<div class="matrix-stepper">{"".join(items)}</div>')


def current_flow_step():
    if (
        st.session_state.get("complete_apply_result_df") is not None
        or st.session_state.get("shopify_apply_result_df") is not None
        or st.session_state.get("inventory_activation_result_df") is not None
    ):
        return 4
    if (
        st.session_state.get("complete_matrixify_df") is not None
        or st.session_state.get("shopify_preview_df") is not None
        or st.session_state.get("inventory_activation_preview_df") is not None
    ):
        return 3
    if st.session_state.get("input_loaded") or st.session_state.get("input") is not None or st.session_state.get("input_row_count"):
        return 2
    return 1


def clear_complete_load_state():
    for key in (
        "complete_matrixify_df",
        "complete_summary_df",
        "complete_issues_df",
        "complete_type_warnings_df",
        "complete_skipped_df",
        "complete_sial_df",
        "complete_centry_df",
        "complete_centry_issues_df",
        "complete_apply_result_df",
        "complete_analysis_message",
        "complete_input_df",
        "complete_template_df",
        "complete_arti_df",
        "complete_template_source",
        "complete_detected_brands",
        "complete_data_context",
        "complete_excel_bytes",
    ):
        st.session_state.pop(key, None)


def reset_load_workspace():
    keep_keys = {"authenticated", "auth_user", "site_picker"}
    for key in list(st.session_state.keys()):
        if key not in keep_keys:
            st.session_state.pop(key, None)
    st.session_state["operation_area_choice"] = "KPIs de catálogo"
    st.session_state["operation_mode_choice"] = "Carga completa"
    st.session_state["load_reset_nonce"] = int(datetime.now(timezone.utc).timestamp())
    st.session_state["load_reset_message"] = "Listo. La app esta limpia y lista para una nueva carga."


def uploaded_file_fingerprint(uploaded_file):
    if not uploaded_file:
        return ""
    return "|".join(
        clean_value(value)
        for value in (
            getattr(uploaded_file, "name", ""),
            getattr(uploaded_file, "size", ""),
        )
    )


def render_sources_card(config, bigquery_ready, arti_source="", template_source="Shopify API", input_count=0, shopify_count=0, arti_count=0):
    bigquery_config = get_bigquery_config()
    project = clean_value(bigquery_config.get("project_id"))
    if not project and isinstance(bigquery_config.get("service_account_info"), dict):
        project = clean_value(bigquery_config["service_account_info"].get("project_id"))
    dataset = clean_value(bigquery_config.get("dataset"))
    table = clean_value(bigquery_config.get("table")) or "ARTI"
    table_label = table if table.count(".") == 2 else ".".join(part for part in [project, dataset, table] if part)
    input_text = f"{input_count:,} productos detectados" if input_count else "Pendiente de carga"
    shopify_text = f"{shopify_count:,} productos sincronizados" if shopify_count else (config["shopify_store"] or template_source)
    arti_text = f"{arti_count:,} filas BigQuery" if arti_count else "Tabla central enlazada"
    render_html(
        f"""
        <div>
            <div>
                <h2>Archivos y fuentes cargadas</h2>
                <p>Resumen limpio de lo que la app usara para preparar la carga.</p>
            </div>
            <div class="source-grid">
                <div class="source-card" style="background:#EFF6FF;border-color:#BFDBFE;"><b>Input productos</b><span>{input_text}</span></div>
                <div class="source-card" style="background:#ECFDF5;border-color:#BBF7D0;"><b>Shopify API</b><span>{shopify_text}</span></div>
                <div class="source-card"><b>ARTI BigQuery</b><span>{arti_text}</span></div>
            </div>
        </div>
        """,
    )


def render_operational_status(config, shopify_config, bigquery_ready, input_loaded):
    render_html(
        f"""
        <div class="section-card">
            <h2>Estado operativo</h2>
            <div class="check-item">Shopify API: {"Conectado" if is_shopify_configured(shopify_config) else "Pendiente"}</div>
            <div class="check-item">BigQuery: {"Activo" if bigquery_ready else "Respaldo local"}</div>
            <div class="check-item">Marca activa: {config["brand_name"]}</div>
            <div class="check-item">Input comercial: {"Cargado" if input_loaded else "Pendiente"}</div>
        </div>
        """,
    )


def render_summary_metrics(metrics):
    items = "".join(
        f'<div class="metric-card"><span>{label}</span><strong>{value}</strong></div>'
        for label, value in metrics
    )
    render_html(f'<div class="section-card"><h2>Resumen bases</h2><p>Datos principales</p><div class="metric-grid">{items}</div></div>')


def render_preview_table(input_df):
    total = len(input_df) if input_df is not None else 0
    shown = min(total, 20)
    render_html(
        f"""
        <div class="section-card">
            <div style="display:flex;align-items:center;justify-content:space-between;gap:16px;">
                <div>
                    <h2>Vista previa del input</h2>
                    <p>Primeras filas detectadas antes de analizar. Mostrando {shown} de {total} productos.</p>
                </div>
                <span class="status-badge blue">Preview</span>
            </div>
        </div>
        """,
    )
    if input_df is not None and not input_df.empty:
        st.dataframe(input_df.head(20), use_container_width=True, height=330)


def render_validations_card():
    render_html(
        """
        <div class="wide-checklist">
            <h2>Checklist</h2>
            <p>Estado de preparacion</p>
            <div class="wide-checklist-grid">
                <div class="wide-checklist-item">SKU, barcode y talla obligatorios.</div>
                <div class="wide-checklist-item">Vendor validado contra marcas permitidas.</div>
                <div class="wide-checklist-item">Cruce automatico con BigQuery.</div>
                <div class="wide-checklist-item">Reporte de errores antes de exportar.</div>
            </div>
        </div>
        """,
    )


def render_analyze_card(config):
    render_html(
        f"""
        <div class="section-card action-card" style="background:var(--forus-blue);border-color:var(--forus-blue);">
            <p style="color:#BFDBFE;font-size:12px;font-weight:900;letter-spacing:.22em;text-transform:uppercase;margin:0 0 10px;">Siguiente acción</p>
            <h2>Analizar y preparar carga</h2>
            <p>Cuando el input este correcto, genera la estructura Matrixify y la hoja Carga Sial.</p>
        </div>
        """,
    )


def render_matrixify_result_card(ready=False):
    state = "Listo para descargar" if ready else "Pendiente de analisis"
    tone = "" if ready else " warn"
    render_html(
        f"""
        <div class="section-card">
            <h2>Archivo Matrixify</h2>
            <p>La estructura queda lista para revisar, descargar y sincronizar con Shopify.</p>
            <span class="status-badge{tone}">{state}</span>
        </div>
        """,
    )


def render_base_status_card(setup_rows):
    cards = []
    for row in setup_rows:
        status = clean_value(row.get("Estado"))
        tone = "" if status.upper().startswith("OK") else " warn"
        cards.append(
            f"""
            <div class="base-status-item">
                <div style="display:flex;align-items:center;justify-content:space-between;gap:8px;">
                    <b>{row.get("Base", "")}</b>
                    <span class="status-badge{tone}">{status}</span>
                </div>
                <span>{row.get("Ruta", "")}</span>
            </div>
            """
        )
    render_html(
        f"""
        <div class="base-status-card">
            <div class="base-status-head">
                <h3>Estado de bases</h3>
                <span class="status-badge blue">Fuentes listas</span>
            </div>
            <div class="base-status-grid">{"".join(cards)}</div>
        </div>
        """
    )


def format_kpi_number(value):
    if isinstance(value, str):
        return value
    try:
        return f"{int(value):,}".replace(",", ".")
    except (TypeError, ValueError):
        return clean_value(value)


def render_kpi_cards(kpis):
    primary_cards = [
        ("Modelos con stock eComm", kpis["modelos_con_stock"], "blue", "&#9633;"),
        ("Creados con stock", kpis["modelos_creados_con_stock"], "green", "&#9679;"),
        ("Pendientes por crear", kpis["modelos_pendientes"], "orange", "!"),
        ("Cobertura stock eComm", f"{kpis['cobertura_shopify']:.0%}", "purple", "%"),
        ("Visibles en web", kpis["modelos_visibles_web"], "green", "&#9711;"),
        ("No visibles en web", kpis["modelos_no_visibles_web"], "orange", "&#9676;"),
        ("Sync stock Shopify", f"{kpis['sincronizacion_stock_shopify']:.0%}", "purple", "%"),
    ]
    def cards_html(cards):
        return "".join(
            f"""
            <div class="kpi-card {tone}">
                <div class="kpi-icon">{icon}</div>
                <div><span>{label}</span><strong>{format_kpi_number(value)}</strong></div>
            </div>
            """
            for label, value, tone, icon in cards
        )

    render_html(
        f"""
        <div class="kpi-section-label">KPIs principales según stock eComm</div>
        <div class="kpi-card-grid">{cards_html(primary_cards)}</div>
        """
    )


def short_problem_label(value):
    text = clean_value(value)
    mapping = {
        "Con stock no visible": "No activo",
        "No activo en Shopify": "No activo",
        "Modelo con stock no creado": "No creado",
        "Modelo con stock sin foto": "Sin foto",
        "Creado con stock sin precio": "Sin precio",
        "Modelo con stock BQ sin stock Shopify": "Sin stock Shopify",
        "Modelo con stock eComm sin stock Shopify": "Sin stock Shopify",
        "Sin stock visible": "Visible sin stock",
    }
    return mapping.get(text, text[:18])


def render_brand_summary_table(brand_summary):
    if brand_summary is None or brand_summary.empty:
        return
    rows = []
    for index, row in brand_summary.reset_index(drop=True).iterrows():
        brand = clean_value(row.get("Marca"))
        logo_src = image_data_uri(brand_logo_path_for_name(brand))
        logo_html = f'<img src="{logo_src}" alt="{escape(brand)}">' if logo_src else ""
        coverage = safe_float_value(row.get("Cobertura"))
        coverage_pct = max(0, min(100, coverage * 100))
        rows.append(
            f"""
            <tr>
                <td>{index + 1}</td>
                <td><div class="brand-cell">{logo_html}<span>{escape(brand.title())}</span></div></td>
                <td>{format_kpi_number(row.get("Modelos_con_stock"))}</td>
                <td style="color:#16A34A;font-weight:950;">{format_kpi_number(row.get("Creados_Shopify"))}</td>
                <td style="color:#EA580C;font-weight:950;">{format_kpi_number(row.get("Pendientes_creacion"))}</td>
                <td>{format_kpi_number(row.get("Stock_total"))}</td>
                <td>
                    <div class="coverage-cell">
                        <div class="coverage-track"><div class="coverage-bar" style="width:{coverage_pct:.0f}%;"></div></div>
                        <span>{coverage_pct:.1f}%</span>
                    </div>
                </td>
            </tr>
            """
        )
    render_html(
        f"""
        <div class="kpi-panel">
            <h3>Resumen por marca</h3>
            <table class="brand-kpi-table">
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Marca</th>
                        <th>Modelos con stock</th>
                        <th>Creados Shopify</th>
                        <th>Pendientes creacion</th>
                        <th>Stock total</th>
                        <th>Cobertura</th>
                    </tr>
                </thead>
                <tbody>{''.join(rows)}</tbody>
            </table>
        </div>
        """
    )


def render_kpi_bar_chart(title, rows, icon="&#9661;", purple=False):
    rows = list(rows or [])
    max_value = max([safe_float_value(row.get("value")) for row in rows] or [1]) or 1
    bars = []
    for row in rows:
        value = safe_float_value(row.get("value"))
        width = max(2, min(100, (value / max_value) * 100)) if value else 2
        label = escape(clean_value(row.get("label")))
        short_label = escape(clean_value(row.get("short")) or clean_value(row.get("label")))
        value_text = format_kpi_number(value)
        bar_class = "bar-fill purple" if purple else "bar-fill"
        bars.append(
            f"""
            <div class="bar-item" tabindex="0" aria-label="{label}: {value_text}">
                <div class="bar-tooltip">{label}<strong>{value_text}</strong></div>
                <div class="bar-label">
                    <span class="bar-label-icon">{row.get("icon", "")}</span>
                    <span>{short_label}</span>
                </div>
                <div class="bar-track"><div class="{bar_class}" style="width:{width:.1f}%;"></div></div>
                <div class="bar-value">{value_text}</div>
            </div>
            """
        )
    return f"""
    <div class="chart-card">
        <div class="chart-head">
            <div>
                <div class="chart-title"><span>{icon}</span><span>{escape(title)}</span></div>
                <p class="chart-subtitle">Ranking operativo para priorizar acciones y revisar proporciones.</p>
            </div>
        </div>
        <div class="bar-stage">{''.join(bars)}</div>
    </div>
    """


def render_kpi_chart_grid(funnel_rows, pareto_rows):
    render_html(
        f"""
        <div class="kpi-chart-grid">
            {render_kpi_bar_chart("Funnel de catálogo", funnel_rows, icon="&#9661;")}
            {render_kpi_bar_chart("Pareto de problemas", pareto_rows, icon="&#9638;", purple=True)}
        </div>
        """
    )


def render_non_visible_combo_table(combo_df):
    combo_df = combo_df.copy() if isinstance(combo_df, pd.DataFrame) else pd.DataFrame()
    if combo_df.empty:
        st.success("No hay modelos creados con stock bloqueados para web.")
        return combo_df

    def commercial_status(row):
        state = clean_value(row.get("Estado operativo"))
        blockers = clean_value(row.get("Bloqueos"))
        return pd.Series(
            {
                "Stock": "Sin stock Shopify" not in state and "Sin stock Shopify" not in blockers,
                "Imagen": "Sin foto" not in state and "Sin foto" not in blockers,
                "Precio": "Sin precio" not in state and "Sin precio" not in blockers,
            }
        )

    status_df = combo_df.apply(commercial_status, axis=1)
    commercial_df = pd.concat([combo_df, status_df], axis=1)
    operational_total = safe_int_value(combo_df["Modelos"].sum()) if "Modelos" in combo_df.columns else 0
    ready_mask = commercial_df["Stock"] & commercial_df["Precio"] & commercial_df["Imagen"]
    ready_to_publish = safe_int_value(commercial_df.loc[ready_mask, "Modelos"].sum())
    combo_view = (
        commercial_df.groupby(["Stock", "Precio", "Imagen"], as_index=False)
        .agg(Modelos=("Modelos", "sum"))
        .sort_values("Modelos", ascending=False)
    )
    combo_view = combo_view[~(combo_view["Stock"] & combo_view["Precio"] & combo_view["Imagen"])].copy()
    if combo_view.empty:
        render_html(
            f"""
            <div class="combo-card">
                <div class="combo-card-head">
                    <div>
                        <div class="combo-title"><span class="combo-title-icon">&#9678;</span> Checklist comercial web</div>
                        <p>Lectura ejecutiva del bloqueo de venta web y las oportunidades listas para activar.</p>
                    </div>
                    <div class="combo-chip">{format_kpi_number(ready_to_publish)} listos para prender</div>
                </div>
                <div class="commercial-flow">
                    <div class="flow-total" title="Total de modelo-color creados con stock eComm que hoy no se ven en la web.">
                        <span>No visibles en web</span>
                        <strong>{format_kpi_number(operational_total)}</strong>
                        <small>Universo que requiere accion comercial u operativa</small>
                    </div>
                    <div class="flow-split">
                        <div class="flow-node danger" title="No hay faltantes de stock Shopify, precio o imagen.">
                            <span>Bloqueados comercialmente</span>
                            <strong>0</strong>
                            <small>Sin bloqueo de data comercial</small>
                        </div>
                        <div class="flow-node ready" title="Tienen stock, precio e imagen. Requieren activar/publicar en Shopify para prender en web.">
                            <span>Listos para prender</span>
                            <strong>{format_kpi_number(ready_to_publish)}</strong>
                            <small>Activar/publicar en Shopify</small>
                        </div>
                    </div>
                </div>
                <div class="flow-actions">
                    <b>Accion recomendada</b>
                    <ul><li>Activar/publicar {format_kpi_number(ready_to_publish)} listos para prender.</li></ul>
                </div>
            </div>
            """
        )
        return combo_df
    blocked_total = safe_int_value(combo_view["Modelos"].sum())
    total_models = safe_float_value(blocked_total)
    total_models_safe = total_models or 1
    other_blocked = max(0, operational_total - blocked_total - ready_to_publish)
    stock_missing = safe_int_value(combo_view.loc[~combo_view["Stock"], "Modelos"].sum())
    price_missing = safe_int_value(combo_view.loc[~combo_view["Precio"], "Modelos"].sum())
    image_missing = safe_int_value(combo_view.loc[~combo_view["Imagen"], "Modelos"].sum())

    def percent(value):
        return max(0, min(100, (safe_float_value(value) / total_models_safe) * 100))

    def status_tile(label, ok):
        state_class = "ok" if ok else "bad"
        icon = "&#10003;" if ok else "&#10005;"
        text = "OK" if ok else "Falta"
        return f"""
        <div class="commercial-status-tile {state_class}">
            <span>{escape(label)}</span>
            <b>{icon}</b>
            <small>{text}</small>
        </div>
        """

    action_items = [
        f"Corregir stock, imagen y precio en los {format_kpi_number(blocked_total)} bloqueados comercialmente.",
        f"Activar/publicar los {format_kpi_number(ready_to_publish)} listos para prender.",
    ]
    if other_blocked:
        action_items.append(f"Revisar {format_kpi_number(other_blocked)} en otros estados operativos.")
    action_html = "".join(f"<li>{escape(item)}</li>" for item in action_items)

    rows = []
    for _, row in combo_view.iterrows():
        models = safe_int_value(row.get("Modelos"))
        pct_value = percent(models)
        rows.append(
            f"""
            <tr>
                <td>
                    <div class="commercial-status-grid">
                        {status_tile("Stock", bool(row.get("Stock")))}
                        {status_tile("Precio", bool(row.get("Precio")))}
                        {status_tile("Imagen", bool(row.get("Imagen")))}
                    </div>
                </td>
                <td>
                    <div class="combo-model-metric">
                        <strong>{format_kpi_number(models)}</strong>
                        <span>{pct_value:.1f}% del bloqueo comercial</span>
                        <i><b style="width:{max(2, min(100, pct_value)):.1f}%"></b></i>
                    </div>
                </td>
            </tr>
            """
        )

    render_html(
        f"""
        <div class="combo-card">
            <div class="combo-card-head">
                <div>
                    <div class="combo-title"><span class="combo-title-icon">&#9678;</span> Checklist comercial web</div>
                    <p>Lectura ejecutiva del bloqueo de venta web y las oportunidades listas para activar.</p>
                </div>
                <div class="combo-chip">{format_kpi_number(blocked_total)} bloqueo comercial</div>
            </div>
            <div class="commercial-flow">
                <div class="flow-total" title="Total de modelo-color creados con stock eComm que hoy no se ven en la web.">
                    <span>No visibles en web</span>
                    <strong>{format_kpi_number(operational_total)}</strong>
                    <small>Universo que requiere accion comercial u operativa</small>
                </div>
                <div class="flow-split">
                    <div class="flow-node danger" title="Faltan stock Shopify, precio o imagen. Requiere correccion de data comercial.">
                        <span>Bloqueados comercialmente</span>
                        <strong>{format_kpi_number(blocked_total)}</strong>
                        <small>Falta stock, precio o imagen</small>
                    </div>
                    <div class="flow-node ready" title="Tienen stock, precio e imagen. Requieren activar/publicar en Shopify para prender en web.">
                        <span>Listos para prender</span>
                        <strong>{format_kpi_number(ready_to_publish)}</strong>
                        <small>Activar/publicar en Shopify</small>
                    </div>
                </div>
            </div>
            <div class="commercial-subtitle">Causas del bloqueo comercial</div>
            <div class="flow-cause-grid">
                <div class="flow-cause"><span>Stock faltante</span><strong>{format_kpi_number(stock_missing)}</strong></div>
                <div class="flow-cause" title="Modelo-color creado, con stock eComm, sin foto Shopify dentro del universo de venta web. No es el total bruto de Shopify sin imagen."><span>Sin foto web</span><strong>{format_kpi_number(image_missing)}</strong></div>
                <div class="flow-cause"><span>Precio faltante</span><strong>{format_kpi_number(price_missing)}</strong></div>
            </div>
            <div class="flow-actions">
                <b>Accion recomendada</b>
                <ul>{action_html}</ul>
            </div>
            <div class="combo-table-wrap compact">
                <table class="combo-table compact">
                    <colgroup>
                        <col style="width:68%;">
                        <col style="width:32%;">
                    </colgroup>
                    <thead>
                        <tr>
                            <th>Bases comerciales</th>
                            <th>Modelos</th>
                        </tr>
                    </thead>
                    <tbody>{''.join(rows)}</tbody>
                    <tfoot>
                        <tr>
                            <td><b>Totales</b><span>{len(combo_view)} combinaciones comerciales</span></td>
                            <td>{format_kpi_number(safe_int_value(total_models))}<small>100%</small></td>
                        </tr>
                    </tfoot>
                </table>
            </div>
        </div>
        """
    )
    return combo_df


def render_actions_table(actions_df, key_prefix):
    actions_df = actions_df.copy() if isinstance(actions_df, pd.DataFrame) else pd.DataFrame()
    render_html(
        """
        <div class="kpi-table-card">
            <div class="kpi-table-head">
                <div class="kpi-table-title"><span>&#9635;</span><span>Pendientes accionables</span></div>
            </div>
        </div>
        """
    )
    if actions_df.empty:
        st.success("No hay pendientes accionables con la regla actual.")
        return actions_df

    control_left, control_problem, control_brand, control_right = st.columns([1.8, 1.1, 1.0, 0.7])
    with control_left:
        search = st.text_input(
            "Buscar pendientes",
            placeholder="Buscar por Mod-Col, marca o problema...",
            label_visibility="collapsed",
            key=f"{key_prefix}_actions_search",
        )
    with control_problem:
        problems = ["Todos"] + sorted(actions_df["Problema"].dropna().map(clean_value).unique().tolist())
        selected_problem = st.selectbox(
            "Filtrar",
            problems,
            label_visibility="collapsed",
            key=f"{key_prefix}_actions_filter",
        )
    with control_brand:
        brands = ["Todas"] + sorted(actions_df["Marca"].dropna().map(clean_value).unique().tolist())
        selected_brand = st.selectbox(
            "Marca",
            brands,
            label_visibility="collapsed",
            key=f"{key_prefix}_actions_brand",
        )
    with control_right:
        page_size = st.selectbox(
            "Filas",
            [12, 25, 50, 100, "Todos"],
            label_visibility="collapsed",
            key=f"{key_prefix}_actions_rows",
        )

    filtered = actions_df.copy()
    if search:
        needle = clean_value(search).lower()
        filtered = filtered[
            filtered.apply(lambda row: needle in " ".join(clean_value(value).lower() for value in row.values), axis=1)
        ].copy()
    if selected_problem != "Todos":
        filtered = filtered[filtered["Problema"].map(clean_value) == selected_problem].copy()
    if selected_brand != "Todas":
        filtered = filtered[filtered["Marca"].map(clean_value) == selected_brand].copy()

    page_key = f"{key_prefix}_actions_page"
    if page_key not in st.session_state:
        st.session_state[page_key] = 1
    if page_size == "Todos":
        page_size_int = len(filtered) or 1
        total_pages = 1
        st.session_state[page_key] = 1
    else:
        page_size_int = int(page_size)
        total_pages = max(1, (len(filtered) + page_size_int - 1) // page_size_int)
        st.session_state[page_key] = min(max(1, int(st.session_state[page_key])), total_pages)
    start = (st.session_state[page_key] - 1) * page_size_int
    visible = filtered.iloc[start : start + page_size_int].copy()
    rows = []
    for index, row in visible.reset_index(drop=True).iterrows():
        rows.append(
            f"""
            <tr>
                <td><span class="row-index">{start + index + 1}</span></td>
                <td><strong>{escape(clean_value(row.get("Mod-Col")))}</strong></td>
                <td>{escape(clean_value(row.get("Marca")))}</td>
                <td><span class="problem-dot"></span>{escape(clean_value(row.get("Problema")))}</td>
                <td><span class="action-chip">{escape(first_non_empty(row.get("Acción sugerida"), row.get("Accion sugerida")))}</span></td>
                <td style="text-align:center;"><span class="stock-badge">{format_kpi_number(row.get("Stock total"))}</span></td>
            </tr>
            """
        )
    render_html(
        f"""
        <div class="kpi-table-card" style="margin-top:0;">
            <table class="kpi-table">
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Mod-Col</th>
                        <th>Marca</th>
                        <th>Problema</th>
                        <th>Acción sugerida</th>
                        <th>Stock total</th>
                    </tr>
                </thead>
                <tbody>{''.join(rows)}</tbody>
            </table>
        </div>
        """
    )
    pager_left, pager_mid, pager_right = st.columns([1.2, 1.8, 1.2])
    with pager_left:
        st.caption(f"Mostrando {len(visible)} de {len(filtered)} resultados filtrados.")
    with pager_mid:
        c1, c2, c3 = st.columns([1.3, 0.9, 1.3])
        with c1:
            if st.button("Anterior", key=f"{key_prefix}_actions_prev", disabled=st.session_state[page_key] <= 1):
                st.session_state[page_key] -= 1
                st.rerun()
        with c2:
            st.markdown(
                f"<div style='text-align:center;color:#172554;font-weight:950;padding-top:8px;'>"
                f"{st.session_state[page_key]} / {total_pages}</div>",
                unsafe_allow_html=True,
            )
        with c3:
            if st.button("Siguiente", key=f"{key_prefix}_actions_next", disabled=st.session_state[page_key] >= total_pages):
                st.session_state[page_key] += 1
                st.rerun()
    return filtered


def render_visibility_audit_table(audit_df):
    audit_df = audit_df.copy() if isinstance(audit_df, pd.DataFrame) else pd.DataFrame()
    if audit_df.empty:
        st.warning("No se genero auditoria de visibilidad.")
        return
    rows = []
    for index, row in audit_df.reset_index(drop=True).iterrows():
        rows.append(
            f"""
            <tr>
                <td><span class="row-index">{index + 1}</span></td>
                <td><strong>{escape(clean_value(row.get("Indicador")))}</strong></td>
                <td style="text-align:center;"><span class="stock-badge">{format_kpi_number(row.get("Valor"))}</span></td>
                <td>{escape(clean_value(row.get("Lectura")))}</td>
            </tr>
            """
        )
    render_html(
        f"""
        <div class="kpi-table-card" style="margin-top:0;">
            <div class="kpi-table-head">
                <div class="kpi-table-title"><span>&#9635;</span><span>Resumen de visibilidad</span></div>
            </div>
            <table class="kpi-table">
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Indicador</th>
                        <th>Valor</th>
                        <th>Lectura</th>
                    </tr>
                </thead>
                <tbody>{''.join(rows)}</tbody>
            </table>
        </div>
        """
    )


def render_missing_models_audit_table(missing_models_input_df, key_prefix):
    df = missing_models_input_df.copy() if isinstance(missing_models_input_df, pd.DataFrame) else pd.DataFrame()
    if df.empty:
        return df
    search = st.text_input(
        "Buscar modelo no creado",
        placeholder="Buscar por Mod-Col, nombre, color, tipo o genero...",
        label_visibility="collapsed",
        key=f"{key_prefix}_missing_input_search",
    )
    filtered = df.copy()
    if search:
        needle = clean_value(search).lower()
        filtered = filtered[
            filtered.apply(lambda row: needle in " ".join(clean_value(value).lower() for value in row.values), axis=1)
        ].copy()
    visible = filtered.head(12).copy()
    rows = []
    for index, row in visible.reset_index(drop=True).iterrows():
        rows.append(
            f"""
            <tr>
                <td><span class="row-index">{index + 1}</span></td>
                <td><strong>{escape(clean_value(row.get("Mod-Col")))}</strong></td>
                <td>{escape(first_non_empty(row.get("Nombre web sugerido"), row.get("Title")))}</td>
                <td>{escape(clean_value(row.get("Marca")))}</td>
                <td>{escape(clean_value(row.get("Tipo de prenda")))}</td>
                <td>{escape(clean_value(row.get("Genero")))}</td>
                <td>{escape(first_non_empty(row.get("Color web"), row.get("Color")))}</td>
                <td style="text-align:center;"><span class="stock-badge">{format_kpi_number(row.get("Stock disponible"))}</span></td>
            </tr>
            """
        )
    render_html(
        f"""
        <div class="kpi-table-card" style="margin-top:16px;">
            <div class="kpi-table-head">
                <div class="kpi-table-title"><span>&#9635;</span><span>Modelos no creados con input sugerido</span></div>
            </div>
            <table class="kpi-table">
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Mod-Col</th>
                        <th>Nombre sugerido</th>
                        <th>Marca</th>
                        <th>Tipo</th>
                        <th>Genero</th>
                        <th>Color web</th>
                        <th>Stock</th>
                    </tr>
                </thead>
                <tbody>{''.join(rows)}</tbody>
            </table>
        </div>
        """
    )
    st.caption(f"Mostrando {len(visible)} de {len(filtered)} modelos no creados. Descarga el Excel para ver el detalle completo.")
    return filtered


def render_missing_variants_table(missing_variants_df, key_prefix):
    if missing_variants_df is None or missing_variants_df.empty:
        return pd.DataFrame()
    df = missing_variants_df.copy()
    control_left, control_mid, control_right = st.columns([2.0, 1.1, 0.8])
    with control_left:
        search = st.text_input(
            "Buscar variantes",
            placeholder="Buscar por Mod-Col, marca, talla o SKU...",
            label_visibility="collapsed",
            key=f"{key_prefix}_variants_search",
        )
    with control_mid:
        brands = ["Todas"] + sorted(df["MARCA_MA"].dropna().map(clean_value).unique().tolist())
        selected_brand = st.selectbox(
            "Filtrar marca",
            brands,
            label_visibility="collapsed",
            key=f"{key_prefix}_variants_brand",
        )
    with control_right:
        page_size = st.selectbox(
            "Filas",
            [12, 25, 50, 100, "Todas"],
            label_visibility="collapsed",
            key=f"{key_prefix}_variants_rows",
        )

    filtered = df.copy()
    if search:
        needle = clean_value(search).lower()
        filtered = filtered[
            filtered.apply(lambda row: needle in " ".join(clean_value(value).lower() for value in row.values), axis=1)
        ].copy()
    if selected_brand != "Todas":
        filtered = filtered[filtered["MARCA_MA"].map(clean_value) == selected_brand].copy()

    page_key = f"{key_prefix}_variants_page"
    if page_key not in st.session_state:
        st.session_state[page_key] = 1
    if page_size == "Todas":
        page_size_int = len(filtered) or 1
        total_pages = 1
        st.session_state[page_key] = 1
    else:
        page_size_int = int(page_size)
        total_pages = max(1, (len(filtered) + page_size_int - 1) // page_size_int)
        st.session_state[page_key] = min(max(1, int(st.session_state[page_key])), total_pages)
    start = (st.session_state[page_key] - 1) * page_size_int
    visible = filtered.iloc[start : start + page_size_int].copy()
    rows = []
    for index, row in visible.reset_index(drop=True).iterrows():
        rows.append(
            f"""
            <tr>
                <td><span class="row-index">{start + index + 1}</span></td>
                <td><strong>{escape(clean_value(row.get("Mod-Col")))}</strong></td>
                <td>{escape(clean_value(row.get("MARCA_MA")))}</td>
                <td>{escape(clean_value(row.get("Talla")))}</td>
                <td>{escape(clean_value(row.get("SKU")))}</td>
                <td>{escape(clean_value(row.get("Motivo web")))}</td>
                <td style="text-align:center;"><span class="stock-badge">{format_kpi_number(row.get("Stock total"))}</span></td>
            </tr>
            """
        )
    render_html(
        f"""
        <div class="kpi-table-card">
            <div class="kpi-table-head">
                <div class="kpi-table-title"><span>&#9635;</span><span>Detalle de variantes no visibles por stock</span></div>
            </div>
            <table class="kpi-table">
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Mod-Col</th>
                        <th>Marca</th>
                        <th>Talla</th>
                        <th>SKU</th>
                        <th>Motivo web</th>
                        <th>Stock total</th>
                    </tr>
                </thead>
                <tbody>{''.join(rows)}</tbody>
            </table>
        </div>
        """
    )
    pager_left, pager_mid, pager_right = st.columns([1.2, 1.8, 1.2])
    with pager_left:
        st.caption(f"Mostrando {len(visible)} de {len(filtered)} variantes de Mod-Col no visibles por stock.")
    with pager_mid:
        c1, c2, c3 = st.columns([1.3, 0.9, 1.3])
        with c1:
            if st.button("Anterior", key=f"{key_prefix}_variants_prev", disabled=st.session_state[page_key] <= 1):
                st.session_state[page_key] -= 1
                st.rerun()
        with c2:
            st.markdown(
                f"<div style='text-align:center;color:#172554;font-weight:950;padding-top:8px;'>"
                f"{st.session_state[page_key]} / {total_pages}</div>",
                unsafe_allow_html=True,
            )
        with c3:
            if st.button("Siguiente", key=f"{key_prefix}_variants_next", disabled=st.session_state[page_key] >= total_pages):
                st.session_state[page_key] += 1
                st.rerun()
    return filtered


def render_catalog_kpi_dashboard(ui_config, brand_config, shopify_config, bigquery_ready):
    render_html(
        """
        <div class="kpi-hero">
            <div class="kpi-title">
                <h2>Dashboard Shopify</h2>
                <p>Visibilidad ejecutiva del catálogo, stock disponible y estado comercial en Shopify.</p>
            </div>
        </div>
        """
    )

    if not bigquery_ready:
        st.error("BigQuery no esta configurado. Para KPIs se necesita leer stock actual y ARTI.")
        return
    if not is_shopify_configured(shopify_config):
        st.error("Shopify API no esta configurado para este sitio.")
        return

    run_key = f"kpi_result_{brand_config['site_key']}"
    result = st.session_state.get(run_key)
    if result is not None and not is_current_kpi_result(result):
        st.session_state.pop(run_key, None)
        result = None
    if result is None:
        result = load_cached_catalog_kpi_result(brand_config["site_key"])
        if result is not None and is_current_kpi_result(result):
            st.session_state[run_key] = result
        else:
            result = None
    if result is None:
        cached_before_refresh = load_cached_catalog_kpi_result(brand_config["site_key"])
        spinner_text = (
            "Cargando dashboard actualizado..."
        )
        with st.spinner(spinner_text):
            try:
                result = load_catalog_kpi_result(brand_config, shopify_config)
                st.session_state[run_key] = result
                save_cached_catalog_kpi_result(brand_config["site_key"], result)
            except Exception as exc:
                if cached_before_refresh is not None and is_current_kpi_result(cached_before_refresh):
                    result = cached_before_refresh
                    st.session_state[run_key] = result
                render_dashboard_refresh_error(exc, result)

    if not result:
        st.info("Dashboard pendiente de datos validos. Presiona Actualizar cuando BigQuery vuelva a entregar stock.")
        return

    meta = result.get("meta", {}) if isinstance(result, dict) else {}
    refreshed_at = parse_iso_datetime(meta.get("refreshed_at"))
    refreshed_label = format_datetime_lima(meta.get("refreshed_at"))
    toolbar_left, toolbar_right = st.columns([0.82, 0.18], vertical_alignment="center")
    with toolbar_left:
        if refreshed_label:
            st.caption(f"Última actualización: {refreshed_label}")
            if is_stale_kpi_result(result):
                st.caption("Dashboard en caché. Presiona Actualizar para recalcular BigQuery y Shopify.")
        if is_current_kpi_result(result):
            st.caption(
                "Filtro eComm: "
                f"{safe_int_value(meta.get('ecomm_bodegas_count'))} bodegas configuradas | "
                f"fecha corte stock: {clean_value(meta.get('fecha_corte')) or 'sin fecha'} | "
                f"filas BigQuery: {format_kpi_number(meta.get('stock_raw_rows'))} -> "
                f"{format_kpi_number(meta.get('stock_cutoff_rows', meta.get('stock_raw_rows')))} ultimo corte | "
                f"tallas eComm con stock: {format_kpi_number(meta.get('stock_filtered_rows'))}"
            )
        else:
            st.warning("Dashboard con metadatos antiguos. Presiona Actualizar para recalcular el filtro eComm.")
    with toolbar_right:
        manual_refresh = st.button(
            "Actualizar",
            type="primary",
            help="Recalcular ahora BigQuery y Shopify. La app muestra el caché hasta que decidas actualizar.",
            key=f"{brand_config['site_key']}_refresh_kpis",
        )
    if manual_refresh:
        with st.spinner("Actualizando dashboard..."):
            try:
                result = load_catalog_kpi_result(brand_config, shopify_config)
                st.session_state[run_key] = result
                save_cached_catalog_kpi_result(brand_config["site_key"], result)
                st.rerun()
            except Exception as exc:
                render_dashboard_refresh_error(exc, result)

    kpis = result["kpis"]
    combo_summary_df = result.get("non_visible_combo_summary", pd.DataFrame())
    ecomm_match_df = result.get("ecomm_stock_match", pd.DataFrame())
    if ecomm_match_df is None or ecomm_match_df.empty:
        st.warning("No se genero auditoria de bodegas eComm para este sitio.")
    else:
        matched_stores = safe_int_value((ecomm_match_df["Aparece en query"] == "Si").sum())
        total_stores = len(ecomm_match_df)
        missing_stores = ecomm_match_df[ecomm_match_df["Aparece en query"] != "Si"]["Bodega"].astype(str).tolist()
        if missing_stores:
            st.warning(
                f"Bodegas eComm configuradas sin stock en BigQuery: {matched_stores}/{total_stores} | "
                f"Sin match: {', '.join(missing_stores[:8])}"
            )
    render_kpi_cards(kpis)
    render_non_visible_combo_table(combo_summary_df)

    actions_df = result["actions"]
    non_visible_web_df = result.get("non_visible_web", pd.DataFrame())
    problem_counts = (
        non_visible_web_df["Bloqueos"].value_counts().rename_axis("Problema").reset_index(name="Casos")
        if non_visible_web_df is not None and not non_visible_web_df.empty and "Bloqueos" in non_visible_web_df.columns
        else pd.DataFrame({"Problema": ["Sin observaciones"], "Casos": [0]})
    )
    funnel_rows = [
        {"label": "Modelos con stock eComm", "short": "Stock eComm", "value": kpis["modelos_con_stock"], "icon": "&#9633;"},
        {"label": "Creados con stock", "short": "Creados stock", "value": kpis["modelos_creados_con_stock"], "icon": "&#9635;"},
        {"label": "Pendientes de creacion", "short": "Pendientes", "value": kpis["modelos_pendientes"], "icon": "!"},
        {"label": "Visibles en web", "short": "Visibles web", "value": kpis["modelos_visibles_web"], "icon": "&#9711;"},
        {"label": "No visibles en web", "short": "No visibles web", "value": kpis["modelos_no_visibles_web"], "icon": "&#9676;"},
        {"label": "Causa principal: sin stock Shopify", "short": "Causa stock", "value": kpis["no_visible_sin_stock_shopify"], "icon": "S"},
        {"label": "Causa principal: sin foto", "short": "Causa foto", "value": kpis["no_visible_sin_foto"], "icon": "&#9673;"},
        {"label": "Causa principal: sin precio", "short": "Causa precio", "value": kpis["no_visible_sin_precio"], "icon": "$"},
        {"label": "Causa principal: no activo", "short": "Causa activo", "value": kpis["no_visible_no_activo"], "icon": "&#9676;"},
        {"label": "Causa principal: no publicado", "short": "No publicado", "value": kpis.get("no_visible_no_publicado", 0), "icon": "&#9676;"},
        {"label": "Variantes con stock eComm sin stock Shopify", "short": "Revisar sucursal", "value": kpis.get("variantes_stock_ecomm_sin_stock_shopify", 0), "icon": "&#8635;"},
    ]
    pareto_rows = [
        {
            "label": clean_value(row.get("Problema")),
            "short": short_problem_label(row.get("Problema")),
            "value": safe_int_value(row.get("Casos")),
            "icon": "&#9679;",
        }
        for _, row in problem_counts.head(6).iterrows()
    ]

    brand_summary = result.get("brand_summary", pd.DataFrame())
    if brand_summary is not None and not brand_summary.empty and len(brand_summary) > 1:
        render_brand_summary_table(brand_summary)

    render_kpi_chart_grid(funnel_rows, pareto_rows)
    audit_df = result.get("kpi_audit", pd.DataFrame())
    missing_models_input_df = result.get("missing_models_input", pd.DataFrame())
    missing_models_variants_df = result.get("missing_models_variants", pd.DataFrame())
    missing_models_fields_df = result.get("missing_models_fields", pd.DataFrame())
    with st.expander("Auditoria de visibilidad Shopify", expanded=False):
        st.caption(
            "Visible real web = producto creado en Shopify, status ACTIVE, publicado en Online Store, con stock Shopify, precio y foto."
        )
        if audit_df is not None and not audit_df.empty:
            render_visibility_audit_table(audit_df)
        stock_activation_audit_df = result.get("stock_location_activation_audit", pd.DataFrame())
        if stock_activation_audit_df is not None and not stock_activation_audit_df.empty:
            st.warning(
                "Hay SKUs con stock eComm en BigQuery y variante creada en Shopify, pero Shopify reporta stock 0. "
                "Esto suele indicar falta de activacion de sucursales o sincronizacion de stock pendiente."
            )
            st.download_button(
                "Descargar SKUs para revisar sucursales / stock Shopify",
                data=dataframe_to_excel_bytes({"Auditoria sucursales stock": stock_activation_audit_df}),
                file_name=f"auditoria_sucursales_stock_{brand_config['site_key']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        if missing_models_input_df is not None and not missing_models_input_df.empty:
            st.download_button(
                "Descargar input sugerido: modelos no creados en Shopify",
                data=dataframe_to_excel_bytes(
                    {
                        "Input sugerido": missing_models_input_df,
                        "Variantes a crear": missing_models_variants_df if missing_models_variants_df is not None else pd.DataFrame(),
                        "Campos por completar": missing_models_fields_df if missing_models_fields_df is not None else pd.DataFrame(),
                    }
                ),
                file_name=f"input_modelos_no_creados_{brand_config['site_key']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Archivo prellenado desde ARTI/BigQuery para que la marca complete solo lo manual antes de crear productos.",
            )
            render_missing_models_audit_table(missing_models_input_df, f"{brand_config['site_key']}_audit")
        else:
            st.success("No hay modelos con stock pendientes de creacion en Shopify para este sitio.")
    filtered_actions_df = render_actions_table(actions_df, f"{brand_config['site_key']}_kpi")
    if filtered_actions_df is not None and not filtered_actions_df.empty:
        st.download_button(
            "Descargar pendientes filtrados",
            data=dataframe_to_excel_bytes({"Pendientes filtrados": filtered_actions_df}),
            file_name=f"pendientes_filtrados_{brand_config['site_key']}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    missing_variants_df = result["missing_stock_variants"]
    filtered_variants_df = missing_variants_df
    if missing_variants_df is not None and not missing_variants_df.empty:
        filtered_variants_df = render_missing_variants_table(missing_variants_df, f"{brand_config['site_key']}_kpi")
        if filtered_variants_df is not None and not filtered_variants_df.empty:
            st.download_button(
                "Descargar variantes filtradas",
                data=dataframe_to_excel_bytes({"Variantes filtradas": filtered_variants_df}),
                file_name=f"variantes_filtradas_{brand_config['site_key']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    kpi_excel_key = f"kpi_excel_bytes_{brand_config['site_key']}"
    kpi_excel_context_key = f"{kpi_excel_key}_context"
    kpi_excel_context = "|".join(
        [
            clean_value(meta.get("refreshed_at")),
            clean_value(meta.get("fecha_corte")),
            clean_value(kpis.get("modelos_con_stock")),
            clean_value(kpis.get("modelos_creados_con_stock")),
            clean_value(st.session_state.get(f"{brand_config['site_key']}_kpi_actions_search")),
            clean_value(st.session_state.get(f"{brand_config['site_key']}_kpi_actions_filter")),
            clean_value(st.session_state.get(f"{brand_config['site_key']}_kpi_actions_brand")),
            clean_value(st.session_state.get(f"{brand_config['site_key']}_kpi_variants_search")),
            clean_value(st.session_state.get(f"{brand_config['site_key']}_kpi_variants_brand")),
            clean_value(len(filtered_actions_df) if filtered_actions_df is not None else 0),
            clean_value(len(filtered_variants_df) if filtered_variants_df is not None else 0),
        ]
    )
    if st.session_state.get(kpi_excel_context_key) == kpi_excel_context and st.session_state.get(kpi_excel_key) is not None:
        excel_bytes = st.session_state[kpi_excel_key]
    else:
        excel_bytes = dataframe_to_excel_bytes(
            {
                "Auditoria KPIs": result.get("kpi_audit", pd.DataFrame()),
                "Resumen modelos": result["model_stock"],
                "Resumen por marca": result.get("brand_summary", pd.DataFrame()),
                "Match bodegas eComm": result.get("ecomm_stock_match", pd.DataFrame()),
                "Pendientes accionables": filtered_actions_df if filtered_actions_df is not None else pd.DataFrame(),
                "Detalle variantes stock": filtered_variants_df if filtered_variants_df is not None else pd.DataFrame(),
                "Resumen bloqueos web": result.get("non_visible_combo_summary", pd.DataFrame()),
                "No visibles web": result.get("non_visible_web", pd.DataFrame()),
                "Input modelos no creados": result.get("missing_models_input", pd.DataFrame()),
                "Variantes modelos no creados": result.get("missing_models_variants", pd.DataFrame()),
                "Campos por completar": result.get("missing_models_fields", pd.DataFrame()),
                "Auditoria sucursales stock": result.get("stock_location_activation_audit", pd.DataFrame()),
                "Sin stock Shopify": result.get("no_shopify_stock_models", pd.DataFrame()),
                "Sin precio": result["no_price_models"],
                "Sin foto stock eComm": result["no_photo_models"],
                "Sin foto total Shopify": result.get("shopify_no_photo_all", pd.DataFrame()),
            }
        )
        st.session_state[kpi_excel_key] = excel_bytes
        st.session_state[kpi_excel_context_key] = kpi_excel_context
    st.download_button(
        "Descargar diagnostico KPIs",
        data=excel_bytes,
        file_name=f"kpis_catalogo_{brand_config['site_key']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _normalize_auth_username(value):
    return clean_value(value).strip().casefold()


TICKET_OPERATOR_USERS = (
    "hugo.camara@forus.pe",
    "luis.nunez@forus.pe",
)
# Usuarios comerciales: solo ven "Input comercial" y "Mis solicitudes".
# Estar en esta lista fuerza ROLE_BRAND aunque no haya rol configurado en Secrets,
# lo que evita que caigan en el ROLE_ADMIN por defecto de auth_access_scope().
COMMERCIAL_INPUT_ONLY_USERS = {
    "comercial@forus.pe",
    "alejandro.mosqueira@forus.pe",
    "clara.gallastegui@forus.pe",
    "natalia.ludowieg@forus.pe",
    "daniela.ballon@forus.pe",
    "mario.biggio@forus.pe",
    "nicolas.rodriguez@forus.pe",
    "alejandro.espinoza@forus.pe",
}

AUTH_ROLE_LABELS = {
    ROLE_ADMIN: "Administrador",
    ROLE_OPERATOR: "Operaciones",
    ROLE_BRAND: "Comercial",
}


def auth_display_name(username):
    """Nombre legible a partir del correo: alejandro.mosqueira@forus.pe -> Alejandro Mosqueira."""
    normalized = _normalize_auth_username(username)
    if not normalized:
        return "Usuario"
    known = ticket_operator_display_name(normalized)
    if known and known != normalized:
        return known
    local = normalized.split("@", 1)[0]
    partes = [p for p in re.split(r"[._-]+", local) if p]
    return " ".join(p.capitalize() for p in partes) or normalized


def auth_role_label(scope):
    return AUTH_ROLE_LABELS.get(clean_value(scope).casefold(), "Usuario")


def is_ticket_operator_user(username):
    return _normalize_auth_username(username) in set(TICKET_OPERATOR_USERS)


def ticket_operator_users():
    return list(TICKET_OPERATOR_USERS)


def ticket_operator_display_name(username):
    labels = {
        "hugo.camara@forus.pe": "Hugo Camara",
        "luis.nunez@forus.pe": "Luis Nunez",
    }
    normalized = _normalize_auth_username(username)
    return labels.get(normalized, normalized)


def auth_access_scope(username):
    normalized_username = _normalize_auth_username(username)
    if normalized_username in COMMERCIAL_INPUT_ONLY_USERS:
        return ROLE_BRAND
    if normalized_username in set(TICKET_OPERATOR_USERS):
        # Hugo y Luis administran toda la aplicacion. Su permiso especial de
        # bandeja de solicitudes se determina por identidad en current_ticket_actor.
        return ROLE_ADMIN
    try:
        auth_config = dict(st.secrets.get("app_auth", {}))
    except Exception:
        auth_config = {}
    try:
        configured_roles = dict(auth_config.get("roles", {}))
    except (TypeError, ValueError):
        configured_roles = {}
    normalized_roles = {
        _normalize_auth_username(user): clean_value(role).strip().casefold()
        for user, role in configured_roles.items()
        if _normalize_auth_username(user)
    }
    configured_scope = normalized_roles.get(normalized_username)
    if configured_scope in {ROLE_BRAND, "marca"}:
        return ROLE_BRAND
    if configured_scope in {ROLE_OPERATOR, "operador", "operaciones"}:
        return ROLE_OPERATOR
    if configured_scope in {ROLE_ADMIN, "administrator", "administrador", "full"}:
        return ROLE_ADMIN
    if configured_scope in {"commercial", "comercial", "input_comercial", "commercial_input"}:
        return ROLE_BRAND
    return ROLE_ADMIN


def auth_allowed_brands(username, role=None):
    normalized_username = _normalize_auth_username(username)
    role = clean_value(role or auth_access_scope(username)).casefold()
    if role == ROLE_ADMIN:
        return configured_commercial_brands()
    try:
        auth_config = dict(st.secrets.get("app_auth", {}))
        configured = dict(auth_config.get("brands", {}))
    except (TypeError, ValueError, Exception):
        configured = {}
    raw = configured.get(normalized_username, configured.get(username, []))
    if isinstance(raw, str):
        values = re.split(r"[,;|]", raw)
    elif isinstance(raw, (list, tuple, set)):
        values = list(raw)
    else:
        values = []
    allowed = []
    configured_labels = configured_commercial_brands()
    if normalized_username in COMMERCIAL_INPUT_ONLY_USERS:
        return configured_labels
    configured_by_key = {_input_norm_key(label): label for label in configured_labels}
    for value in values:
        label = configured_by_key.get(_input_norm_key(value))
        if label and label not in allowed:
            allowed.append(label)
    if role == ROLE_OPERATOR and not allowed:
        return configured_labels
    return allowed


def current_ticket_actor():
    username = clean_value(st.session_state.get("auth_user"))
    role = ROLE_OPERATOR if is_ticket_operator_user(username) else ROLE_BRAND
    return TicketService.actor(username, role, auth_allowed_brands(username, role))


def get_ticket_service():
    try:
        config = dict(st.secrets.get("ticketing", {}))
    except Exception:
        config = {}
    backend = clean_value(config.get("backend") or os.getenv("CATALOG_TICKETS_BACKEND") or "local").casefold()
    if backend == "github":
        repository = clean_value(config.get("repository") or os.getenv("CATALOG_TICKETS_REPOSITORY"))
        owner = clean_value(config.get("owner"))
        repo = clean_value(config.get("repo"))
        if repository and "/" in repository:
            owner, repo = repository.split("/", 1)
        token = clean_value(config.get("token") or os.getenv("CATALOG_TICKETS_GITHUB_TOKEN"))
        store = GitHubTicketStore(
            owner=owner,
            repo=repo,
            token=token,
            branch=clean_value(config.get("branch")) or "catalog-tickets",
            prefix=clean_value(config.get("prefix")) or "catalog_tickets",
        )
        persistent_backend = "github"
    else:
        root = clean_value(config.get("local_path")) or "outputs/catalog_tickets"
        store = LocalTicketStore(root)
        persistent_backend = "local"
    sla = {}
    try:
        sla = {clean_value(key).casefold(): float(value) for key, value in dict(config.get("sla_hours", {})).items()}
    except (TypeError, ValueError):
        sla = {}
    service = TicketService(
        store,
        notifier=MockNotificationAdapter(),
        jobs=MockJobAdapter(),
        sla_hours=sla,
        operator_users=ticket_operator_users(),
    )
    return service, persistent_backend


def get_audit_service():
    """Servicio de auditoria. Reutiliza la configuracion [ticketing] de Secrets."""
    try:
        config = dict(st.secrets.get("ticketing", {}))
    except Exception:
        config = {}
    backend = clean_value(config.get("backend") or os.getenv("CATALOG_TICKETS_BACKEND") or "local").casefold()
    if backend == "github":
        repository = clean_value(config.get("repository") or os.getenv("CATALOG_TICKETS_REPOSITORY"))
        owner = clean_value(config.get("owner"))
        repo = clean_value(config.get("repo"))
        if repository and "/" in repository:
            owner, repo = repository.split("/", 1)
        token = clean_value(config.get("token") or os.getenv("CATALOG_TICKETS_GITHUB_TOKEN"))
        prefix = (clean_value(config.get("prefix")) or "catalog_tickets") + "/audit"
        try:
            store = GitHubAuditStore(
                owner=owner, repo=repo, token=token,
                branch=clean_value(config.get("branch")) or "catalog-tickets",
                prefix=prefix,
            )
            return AuditService(store)
        except AuditError:
            pass
    root = clean_value(config.get("local_path")) or "outputs/catalog_tickets"
    return AuditService(LocalAuditStore(Path(root) / "audit"))


def audit_record(accion, **kwargs):
    """Registra una accion del usuario activo. Nunca interrumpe el flujo."""
    usuario = clean_value(st.session_state.get("auth_user"))
    if not usuario:
        return None
    kwargs.setdefault("rol", auth_role_label(st.session_state.get("auth_scope")))
    kwargs.setdefault("nombre", auth_display_name(usuario))
    return get_audit_service().record(accion, usuario, **kwargs)


def render_storage_badge():
    """Indica si las solicitudes y la auditoria se guardan de forma persistente."""
    _, backend = get_ticket_service()
    persistente = backend == "github"
    texto = "Almacenamiento persistente" if persistente else "Almacenamiento temporal"
    detalle = (
        "Solicitudes y auditoria se guardan en GitHub."
        if persistente
        else "Se borra en cada redespliegue. Configura [ticketing] en Secrets."
    )
    clase = "storage-ok" if persistente else "storage-warn"
    st.markdown(
        f'<div class="storage-badge {clase}" title="{escape(detalle)}">'
        f"<strong>{escape(texto)}</strong><small>{escape(detalle)}</small></div>",
        unsafe_allow_html=True,
    )


def get_auth_users():
    try:
        auth_config = dict(st.secrets.get("app_auth", {}))
    except Exception:
        auth_config = {}
    users = auth_config.get("users", {})
    try:
        configured_users = dict(users)
    except (TypeError, ValueError):
        configured_users = {}
    if configured_users:
        return {
            _normalize_auth_username(user): clean_value(password)
            for user, password in configured_users.items()
            if _normalize_auth_username(user) and clean_value(password)
        }
    username = _normalize_auth_username(auth_config.get("username"))
    password = clean_value(auth_config.get("password"))
    if username and password:
        return {username: password}
    return {
        "admin": "forus2026",
        "hugo.camara@forus.pe": "forus2026",
        "luis.nunez@forus.pe": "Forus2026*",
    }


def render_login_styles():
    st.markdown(
        """
        <style>
        [data-testid="stSidebar"] { display:none; }
        [data-testid="stToolbar"], .stDeployButton { display:none !important; }
        [data-testid="stAppViewContainer"] {
            background:#152238;
        }
        .main .block-container {
            padding-top:38px;
            max-width:620px;
        }
        .st-key-login_card {
            width:min(448px, calc(100vw - 32px));
            margin:0 auto;
            overflow:hidden;
            border-radius:16px;
            background:#FFFFFF;
            box-shadow:0 28px 80px rgba(0,0,0,.28);
            color-scheme:light;
        }
        .login-head {
            padding:32px 32px 34px;
            text-align:center;
            background:linear-gradient(180deg, #2367FF 0%, #1757EF 100%);
            color:#FFFFFF;
        }
        .login-logo-row {
            display:flex;
            align-items:center;
            justify-content:center;
            gap:22px;
            margin-bottom:24px;
        }
        .login-forus-logo {
            min-width:178px;
            height:64px;
            border-radius:10px;
            background:#FFFFFF;
            display:grid;
            place-items:center;
            padding:8px 14px;
            box-sizing:border-box;
        }
        .login-forus-logo img {
            max-width:100%;
            max-height:48px;
            object-fit:contain;
        }
        .login-forus-fallback {
            color:#14306B;
            font-size:34px;
            line-height:1;
            font-weight:950;
            letter-spacing:.02em;
        }
        .login-forus-fallback small {
            display:block;
            margin-top:2px;
            color:#14306B;
            font-size:8px;
            letter-spacing:.22em;
            font-weight:900;
        }
        .login-divider {
            width:1px;
            height:48px;
            background:rgba(255,255,255,.62);
        }
        .login-shopify-logo {
            width:52px;
            height:52px;
            border-radius:10px;
            background:#FFFFFF;
            display:grid;
            place-items:center;
            box-shadow:0 10px 22px rgba(15,23,42,.12);
        }
        .login-shopify-logo img {
            max-width:38px;
            max-height:38px;
            object-fit:contain;
        }
        .login-shopify-fallback {
            color:#16A34A;
            font-size:30px;
            font-weight:950;
        }
        .login-head h1 {
            margin:0;
            font-size:30px;
            line-height:1.12;
            font-weight:950;
        }
        .login-head p {
            margin:10px 0 0;
            color:#EAF2FF;
            font-size:16px;
            font-weight:750;
        }
        .st-key-login_form_area {
            padding:24px 32px 28px;
            background:#FFFFFF;
            color-scheme:light;
        }
        .st-key-login_form_area label {
            color:#1E293B !important;
            font-weight:850 !important;
        }
        .st-key-login_form_area .stTextInput input {
            border-radius:12px;
            min-height:48px;
            background:#F8FAFC !important;
            border:1px solid #CBD5E1 !important;
            font-size:15px;
            color:#0F172A !important;
            caret-color:#0F172A !important;
            -webkit-text-fill-color:#0F172A !important;
            opacity:1 !important;
            color-scheme:light !important;
        }
        .st-key-login_form_area .stTextInput input::placeholder {
            color:#64748B !important;
            -webkit-text-fill-color:#64748B !important;
            opacity:1 !important;
        }
        .st-key-login_form_area div[data-baseweb="input"],
        .st-key-login_form_area div[data-baseweb="base-input"] {
            background:#F8FAFC !important;
            color:#0F172A !important;
            color-scheme:light !important;
        }
        .st-key-login_form_area .stTextInput input:-webkit-autofill,
        .st-key-login_form_area .stTextInput input:-webkit-autofill:hover,
        .st-key-login_form_area .stTextInput input:-webkit-autofill:focus {
            -webkit-text-fill-color:#0F172A !important;
            caret-color:#0F172A !important;
            -webkit-box-shadow:0 0 0 1000px #F8FAFC inset !important;
            box-shadow:0 0 0 1000px #F8FAFC inset !important;
            transition:background-color 9999s ease-out 0s;
        }
        .st-key-login_form_area .stTextInput button,
        .st-key-login_form_area .stTextInput svg {
            color:#475569 !important;
            fill:currentColor !important;
        }
        .st-key-login_form_area .stButton button {
            width:100%;
            min-height:48px;
            border-radius:12px;
            background:#2367FF;
            font-weight:950;
            white-space:nowrap !important;
            word-break:keep-all !important;
            overflow-wrap:normal !important;
            line-height:1.12 !important;
        }
        .st-key-login_form_area .stButton button * {
            white-space:nowrap !important;
            word-break:keep-all !important;
            overflow-wrap:normal !important;
            line-height:1.12 !important;
        }
        button[data-testid^="stBaseButton"] {
            white-space:nowrap !important;
            word-break:keep-all !important;
            overflow-wrap:normal !important;
            line-height:1.12 !important;
        }
        button[data-testid^="stBaseButton"] * {
            white-space:nowrap !important;
            word-break:keep-all !important;
            overflow-wrap:normal !important;
            line-height:1.12 !important;
        }
        .login-note {
            padding:0 32px 32px;
            text-align:center;
            color:#64748B;
            font-size:13px;
            font-weight:750;
        }
        .login-foot {
            margin:26px auto 0;
            width:min(448px, calc(100vw - 32px));
            text-align:center;
            color:#FFFFFF;
            font-size:14px;
            line-height:1.7;
            font-weight:750;
        }
        .login-foot strong {
            display:block;
            margin-bottom:6px;
            font-weight:850;
        }
        @media (max-width: 560px) {
            .main .block-container { padding-top:20px; }
            .login-head { padding:26px 20px 28px; }
            .st-key-login_form_area { padding:22px 22px 26px; }
            .login-forus-logo { min-width:152px; }
            .login-head h1 { font-size:26px; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def require_login():
    if st.session_state.get("authenticated"):
        return True
    render_login_styles()
    forus_src = image_data_uri(FORUS_LOGO_PATH)
    shopify_src = image_data_uri(SHOPIFY_LOGO_PATH)
    forus_logo = (
        f'<img src="{forus_src}" alt="FORUS">'
        if forus_src
        else '<div class="login-forus-fallback">FORUS<small>CONSUMER FANATIC</small></div>'
    )
    shopify_logo = (
        f'<img src="{shopify_src}" alt="Shopify">'
        if shopify_src
        else '<div class="login-shopify-fallback">S</div>'
    )
    with st.container(key="login_card"):
        render_html(
            f"""
            <div class="login-head">
                <div class="login-logo-row">
                    <div class="login-forus-logo">{forus_logo}</div>
                    <div class="login-divider"></div>
                    <div class="login-shopify-logo">{shopify_logo}</div>
                </div>
                <h1>Catálogo Control Center</h1>
                <p>Sistema de gestion de productos</p>
            </div>
            """
        )
        with st.container(key="login_form_area"):
            with st.form("login_form"):
                username = st.text_input("Correo electronico", placeholder="hugo.camara@forus.pe")
                password = st.text_input("Contrasena", type="password", placeholder="********")
                submitted = st.form_submit_button("Ingresar", type="primary")
        render_html(
            """
            <div class="login-note">Sistema exclusivo para personal autorizado</div>
            """
        )
    render_html(
        """
        <div class="login-foot">
            <strong>Gestión de catálogos para múltiples marcas</strong>
            Columbia &bull; Hush Puppies &bull; Vans &bull; Patagonia &bull; Mas
        </div>
        """
    )
    if submitted:
        users = get_auth_users()
        normalized_username = _normalize_auth_username(username)
        expected = users.get(normalized_username)
        if expected and hmac.compare_digest(clean_value(password), expected):
            st.session_state["authenticated"] = True
            st.session_state["auth_user"] = normalized_username
            st.session_state["auth_scope"] = auth_access_scope(normalized_username)
            st.rerun()
        st.error("Usuario o contrasena incorrectos.")
    return False


def sidebar_nav_button(label, state_key, value, button_key, extra_state=None):
    selected = st.session_state.get(state_key) == value
    if selected:
        st.markdown(
            f"""
            <style>
            div.st-key-{button_key} button {{
                background:#EFF6FF !important;
                border-color:#60A5FA !important;
                box-shadow:0 0 0 1px #BFDBFE, 0 12px 24px rgba(37,99,235,0.10) !important;
                color:#0B1B46 !important;
            }}
            div.st-key-{button_key} button::before {{
                background-color:#DBEAFE !important;
                box-shadow:inset 0 0 0 1px #BFDBFE !important;
            }}
            </style>
            """,
            unsafe_allow_html=True,
        )
    if st.button(label, key=button_key, use_container_width=True):
        st.session_state[state_key] = value
        for extra_key, extra_value in (extra_state or {}).items():
            st.session_state[extra_key] = extra_value
        st.rerun()


def render_ticket_styles():
    st.markdown(
        """
        <style>
        .ticket-hero{padding:20px 24px;border:1px solid #D9E2EF;background:#fff;border-radius:14px;margin:0 0 16px}
        .ticket-hero p{margin:0 0 5px;color:#2563EB;font-size:11px;font-weight:900;text-transform:uppercase;letter-spacing:0}
        .ticket-hero h1{margin:0;color:#0B1B46;font-size:26px;line-height:1.15}
        .ticket-hero span{display:block;margin-top:7px;color:#64748B;font-size:15px}
        .ticket-kpi-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:12px;margin:0 0 16px}
        .ticket-kpi-card{min-height:88px;padding:13px 15px;border:1px solid #D9E2EF;border-radius:12px;background:#fff;box-shadow:0 12px 25px rgba(15,23,42,.05)}
        .ticket-kpi-card small{display:block;min-height:30px;color:#40516E;font-size:12px;font-weight:800;line-height:1.2}
        .ticket-kpi-card strong{display:block;margin-top:7px;color:#0B1B46;font-size:27px;line-height:1;font-weight:900}
        .ticket-kpi-card.blue{border-color:#BFDBFE;background:#F8FBFF}.ticket-kpi-card.blue strong{color:#2563EB}
        .ticket-kpi-card.amber{border-color:#FDE6BD;background:#FFFDF8}.ticket-kpi-card.amber strong{color:#C56A00}
        .ticket-kpi-card.red{border-color:#FECACA;background:#FFF9F9}.ticket-kpi-card.red strong{color:#DC2626}
        .ticket-kpi-card.green{border-color:#BBF7D0;background:#F7FFF9}.ticket-kpi-card.green strong{color:#15803D}
        .ticket-kpi-card.slate{background:#fff}.ticket-filter-panel{padding:14px 16px 4px;border:1px solid #D9E2EF;border-radius:12px;background:#fff;margin:0 0 16px}
        .ticket-filter-title{margin:0 0 8px;color:#0B1B46;font-size:15px;line-height:1.2}
        .ticket-filter-panel h3{margin:0 0 8px;color:#0B1B46;font-size:15px}
        .ticket-state{display:inline-flex;align-items:center;min-height:28px;padding:4px 10px;border-radius:999px;font-size:12px;font-weight:850;border:1px solid #CBD5E1;background:#F8FAFC;color:#334155}
        .ticket-state.blue{background:#EFF6FF;border-color:#BFDBFE;color:#1D4ED8}
        .ticket-state.yellow{background:#FFFBEB;border-color:#FDE68A;color:#A16207}
        .ticket-state.green{background:#ECFDF5;border-color:#A7F3D0;color:#047857}
        .ticket-state.red{background:#FEF2F2;border-color:#FECACA;color:#B91C1C}
        .ticket-state.gray{background:#F1F5F9;border-color:#CBD5E1;color:#475569}
        .ticket-detail-header{display:flex;align-items:center;justify-content:space-between;gap:14px;padding:18px 20px 16px;border:1px solid #D9E2EF;border-radius:14px 14px 0 0;background:#fff}
        .ticket-detail-header h2{margin:0;color:#0B1B46;font-size:22px;line-height:1.15}
        .ticket-detail-header p{margin:5px 0 0;color:#64748B;font-size:13px}
        .ticket-detail-shell{padding:0 20px 20px;border:1px solid #D9E2EF;border-top:0;border-radius:0 0 14px 14px;background:#fff;margin:0 0 16px}
        .ticket-summary{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:0;padding:16px 0}
        .ticket-summary>div{padding:13px 14px;border:1px solid #E2E8F0;border-radius:10px;background:#F8FAFC;min-height:80px}
        .ticket-summary small{display:block;color:#64748B;font-size:11px;font-weight:850;text-transform:uppercase;margin-bottom:7px}
        .ticket-summary strong{display:block;color:#0B1B46;font-size:18px;line-height:1.15}
        .ticket-summary span{display:block;margin-top:5px;color:#475569;font-size:12px;line-height:1.35;overflow-wrap:anywhere}
        .ticket-section{padding:18px 20px;border:1px solid #D9E2EF;border-radius:14px;background:#fff;margin:16px 0}
        .ticket-section h3{margin:0 0 4px;color:#0B1B46;font-size:18px}
        .ticket-section > p{margin:0 0 14px;color:#64748B;font-size:13px}
        .ticket-section-label{margin:0 0 10px;color:#2563EB;font-size:11px;font-weight:900;text-transform:uppercase}
        .ticket-info-line{margin:0;padding:10px 12px;border-radius:9px;background:#F8FAFC;color:#64748B;font-size:12px;line-height:1.4}
        .ticket-action-note{padding:12px 14px;border:1px solid #DBEAFE;border-radius:10px;background:#F8FBFF;color:#334155;font-size:13px;line-height:1.45}
        .ticket-comments-empty{padding:14px;border:1px dashed #CBD5E1;border-radius:10px;background:#F8FAFC;color:#64748B;font-size:13px}
        .ticket-event{padding:10px 12px;border-left:3px solid #93C5FD;background:#F8FAFC;margin:7px 0;border-radius:0 8px 8px 0}
        .ticket-event strong{color:#0B1B46}.ticket-event small{color:#64748B}
        @media(max-width:1100px){.ticket-kpi-grid{grid-template-columns:repeat(3,minmax(0,1fr))}}
        @media(max-width:900px){.ticket-summary,.ticket-kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}.ticket-detail-header{align-items:flex-start;flex-direction:column}}
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_ticket_kpi_grid(items):
    """Render a compact, dashboard-style operational summary for catalog tickets."""
    cards = []
    for label, value, tone in items:
        cards.append(
            f'<div class="ticket-kpi-card {escape(tone)}"><small>{escape(label)}</small>'
            f'<strong>{int(value):,}</strong></div>'
        )
    st.markdown(f'<div class="ticket-kpi-grid">{"".join(cards)}</div>', unsafe_allow_html=True)


def _ticket_state_color(status):
    if status in {STATE_COMPLETED, STATE_COMPLETED_OBS, STATE_APPROVED}:
        return "green"
    if status in {STATE_FAILED, STATE_REJECTED}:
        return "red"
    if status in {STATE_REVIEW, STATE_OBSERVED, STATE_CORRECTED}:
        return "yellow"
    if status in {STATE_PENDING, STATE_ASSIGNED, STATE_LOADING}:
        return "blue"
    return "gray"


def _ticket_summary_value(summary_df, indicator):
    if not isinstance(summary_df, pd.DataFrame) or summary_df.empty:
        return 0
    matches = summary_df[summary_df["Indicador"].astype(str).eq(indicator)]
    return int(matches.iloc[0].get("Valor", 0) or 0) if not matches.empty else 0


def _ticket_table(tickets):
    rows = []
    for ticket in tickets:
        rows.append(
            {
                "Ticket": ticket.get("code"),
                "Fecha": clean_value(ticket.get("created_at")).replace("T", " ")[:16],
                "Marca": ticket.get("brand"),
                "Solicitante": ticket.get("requester"),
                "Carga": "Completa" if ticket.get("load_type") == "complete" else "Parcial",
                "Sitios": ", ".join(ticket.get("sites", [])),
                "Productos": int(ticket.get("summary", {}).get("products", 0)),
                "Prioridad": PRIORITY_LABELS.get(ticket.get("priority"), ticket.get("priority")),
                "Responsable": ticket.get("assignee") or "Sin asignar",
                "Antigüedad": f"{ticket_age_hours(ticket):.0f} h",
                "Estado": STATE_LABELS.get(ticket.get("status"), ticket.get("status")),
                "Vencido": "Sí" if ticket_is_overdue(ticket) else "No",
                "Acción": "Abrir solicitud",
            }
        )
    return pd.DataFrame(rows)


FULL_LOAD_TICKET_STATES = {
    STATE_APPROVED,
    STATE_PREPARING,
    STATE_DRY_RUN,
    STATE_READY_EXECUTE,
    STATE_LOADING,
    STATE_VALIDATING,
    STATE_FAILED,
}


def _full_load_site_tokens(value):
    token = _input_norm_key(value)
    tokens = {token} if token else set()
    if token.endswith("pe") and len(token) > 2:
        tokens.add(token[:-2])
    if token.endswith("myshopifycom"):
        tokens.add(token[: -len("myshopifycom")])
    return {item for item in tokens if item}


def _ticket_matches_active_site(ticket, brand_config):
    active_tokens = set()
    for key in ("site_label", "site_key", "brand_name", "brand", "shop_domain"):
        active_tokens.update(_full_load_site_tokens(brand_config.get(key)))
    ticket_tokens = set()
    for value in list(ticket.get("sites") or []) + [ticket.get("brand")]:
        ticket_tokens.update(_full_load_site_tokens(value))
    return bool(active_tokens & ticket_tokens)


def _render_full_load_ticket_close(service, actor, ticket, latest_version):
    code = clean_value(ticket.get("code"))
    summary = ticket.get("summary") or {}
    st.markdown("#### Registrar resultado de la carga")
    st.caption(
        "Cuando termines la carga y la revisión en Shopify, cierra la solicitud aquí. "
        "El equipo comercial verá inmediatamente el resultado."
    )
    close_metrics = st.columns(2)
    processed_count = close_metrics[0].number_input(
        "Productos procesados",
        min_value=0,
        value=max(0, safe_int_value(summary.get("products"), 0)),
        step=1,
        key=f"full_load_close_processed_{code}",
    )
    error_count = close_metrics[1].number_input(
        "Productos con error",
        min_value=0,
        value=0,
        step=1,
        key=f"full_load_close_errors_{code}",
    )
    close_note = st.text_area(
        "Resultado u observaciones",
        key=f"full_load_close_note_{code}",
        placeholder="Ejemplo: carga completada y revisada en Shopify.",
    )
    close_confirmed = st.checkbox(
        "Confirmo que la carga terminó y revisé el resultado en Shopify.",
        key=f"full_load_close_confirmed_{code}",
    )
    close_result = {
        "processed": int(processed_count),
        "errors": int(error_count),
        "message": clean_value(close_note) or "Carga completada y validada en Shopify.",
        "detail": clean_value(close_note),
        "closed_by": actor.get("user"),
        "filename": latest_version.get("filename") or ticket.get("filename"),
        "file_version": latest_version.get("number", 1),
        "file_hash": latest_version.get("hash") or ticket.get("file_hash"),
    }
    close_cols = st.columns(3)
    if close_cols[0].button(
        "Finalizar carga",
        type="primary",
        key=f"full_load_complete_{code}",
        disabled=not close_confirmed,
        use_container_width=True,
    ):
        try:
            service.record_job_result(actor, code, success=True, result=close_result)
            st.success(f"La solicitud {code} quedó finalizada.")
            st.rerun()
        except TicketError as exc:
            st.error(str(exc))
    if close_cols[1].button(
        "Finalizar con observaciones",
        key=f"full_load_complete_obs_{code}",
        disabled=not close_confirmed,
        use_container_width=True,
    ):
        try:
            service.record_job_result(actor, code, success=True, observations=True, result=close_result)
            st.rerun()
        except TicketError as exc:
            st.error(str(exc))
    if close_cols[2].button(
        "Registrar incidencia",
        key=f"full_load_fail_{code}",
        disabled=not close_confirmed,
        use_container_width=True,
    ):
        try:
            service.record_job_result(
                actor,
                code,
                success=False,
                result=close_result,
                error=clean_value(close_note) or "Incidencia registrada por Operaciones.",
            )
            st.rerun()
        except TicketError as exc:
            st.error(str(exc))


def render_full_load_ticket_queue(brand_config):
    actor = current_ticket_actor()
    if actor.get("role") not in {ROLE_OPERATOR, ROLE_ADMIN}:
        return
    try:
        service, _ = get_ticket_service()
        all_full_load_tickets = [
            ticket
            for ticket in service.list_tickets(actor)
            if clean_value(ticket.get("load_type")).casefold() == "complete"
            and ticket.get("status") in FULL_LOAD_TICKET_STATES
        ]
        tickets = [ticket for ticket in all_full_load_tickets if _ticket_matches_active_site(ticket, brand_config)]
        other_site_tickets = [
            ticket for ticket in all_full_load_tickets if not _ticket_matches_active_site(ticket, brand_config)
        ]
    except TicketError as exc:
        st.warning(f"No se pudo consultar la bandeja de cargas pendientes: {exc}")
        return
    except Exception as exc:
        st.warning(f"La bandeja de cargas pendientes no está disponible: {exc}")
        return

    tickets.sort(key=lambda item: clean_value(item.get("created_at")), reverse=True)
    other_site_tickets.sort(key=lambda item: clean_value(item.get("created_at")), reverse=True)
    with st.container(border=True):
        header_cols = st.columns([4, 1])
        header_cols[0].markdown("### Cargas pendientes")
        header_cols[0].caption(
            "Descarga el input aprobado, ejecuta la carga con el flujo inferior y registra el cierre sin salir de esta pantalla."
        )
        header_cols[1].metric("Pendientes", len(tickets))
        if not tickets:
            if other_site_tickets:
                site_rows = []
                for ticket in other_site_tickets:
                    site_rows.append(
                        {
                            "Ticket": ticket.get("code"),
                            "Sitio": ", ".join(ticket.get("sites") or []) or "Sin sitio",
                            "Marca": ticket.get("brand"),
                            "Estado": STATE_LABELS.get(ticket.get("status"), ticket.get("status")),
                            "Responsable": ticket.get("assignee") or "Sin asignar",
                            "Productos": safe_int_value((ticket.get("summary") or {}).get("products"), 0),
                        }
                    )
                st.warning(
                    "No hay solicitudes pendientes para el sitio activo. Hay solicitudes pendientes para otra web; "
                    "cambia el Sitio activo a la web indicada para prepararlas, cargarlas y cerrarlas."
                )
                st.dataframe(pd.DataFrame(site_rows), use_container_width=True, hide_index=True)
            else:
                st.info("No hay solicitudes pendientes para el sitio activo.")
            with st.expander("Como cerrar una solicitud de carga"):
                st.markdown(
                    "1. Selecciona el sitio correcto y descarga el input validado.\n"
                    "2. Usa **Preparar para carga** para validar el archivo. Esta prueba no cierra la solicitud.\n"
                    "3. Usa **Marcar carga iniciada** antes de ejecutar la carga real en Shopify.\n"
                    "4. Cuando la carga real termine, vuelve a esta sección y registra **Finalizar carga**, "
                    "**Finalizar con observaciones** o **Registrar incidencia**."
                )
            return

        with st.expander("Como cerrar una solicitud de carga"):
            st.markdown(
                "1. Descarga el input validado y usa **Preparar para carga**. Esta prueba valida, pero no cierra el ticket.\n"
                "2. Usa **Marcar carga iniciada** antes de ejecutar la carga real.\n"
                "3. Ejecuta y verifica la carga en Shopify.\n"
                "4. Registra **Finalizar carga**, **Finalizar con observaciones** o **Registrar incidencia**."
            )

        queue_rows = []
        for ticket in tickets:
            queue_rows.append(
                {
                    "Ticket": ticket.get("code"),
                    "Marca": ticket.get("brand"),
                    "Solicitante": ticket.get("requester"),
                    "Productos": safe_int_value((ticket.get("summary") or {}).get("products"), 0),
                    "Estado": STATE_LABELS.get(ticket.get("status"), ticket.get("status")),
                    "Prioridad": PRIORITY_LABELS.get(ticket.get("priority"), ticket.get("priority")),
                    "Responsable": ticket.get("assignee") or "Sin asignar",
                }
            )
        st.dataframe(pd.DataFrame(queue_rows), use_container_width=True, hide_index=True)

        ticket_codes = [clean_value(ticket.get("code")) for ticket in tickets]
        selected_code = st.selectbox(
            "Solicitud para cargar",
            ticket_codes,
            format_func=lambda value: next(
                (
                    f"{value} · {STATE_LABELS.get(item.get('status'), item.get('status'))} · "
                    f"{safe_int_value((item.get('summary') or {}).get('products'), 0)} productos"
                    for item in tickets
                    if clean_value(item.get("code")) == value
                ),
                value,
            ),
            key=f"full_load_ticket_{brand_config.get('site_key')}",
        )
        try:
            ticket = service.get_ticket(actor, selected_code)
        except TicketError as exc:
            st.error(str(exc))
            return

        latest_version = (ticket.get("versions") or [{}])[-1]
        detail_cols = st.columns(4)
        detail_cols[0].metric("Ticket", ticket.get("code"))
        detail_cols[1].metric("Productos", safe_int_value((ticket.get("summary") or {}).get("products"), 0))
        detail_cols[2].metric("Estado", STATE_LABELS.get(ticket.get("status"), ticket.get("status")))
        detail_cols[3].metric("Responsable", ticket.get("assignee") or "Sin asignar")

        download_cols = st.columns(2)
        if latest_version.get("input_path"):
            try:
                download_cols[0].download_button(
                    "Descargar input validado",
                    data=service.store.get_artifact(latest_version["input_path"]),
                    file_name=latest_version.get("filename") or ticket.get("filename") or "input_validado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"full_load_download_input_{selected_code}",
                    use_container_width=True,
                )
            except (TicketError, OSError) as exc:
                download_cols[0].warning(f"No se pudo descargar el input: {exc}")
        if latest_version.get("report_path"):
            try:
                download_cols[1].download_button(
                    "Descargar validación",
                    data=service.store.get_artifact(latest_version["report_path"]),
                    file_name=f"validacion_{selected_code}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"full_load_download_report_{selected_code}",
                    use_container_width=True,
                )
            except (TicketError, OSError) as exc:
                download_cols[1].warning(f"No se pudo descargar el reporte: {exc}")

        actor_user = _normalize_auth_username(actor.get("user"))
        assignee = _normalize_auth_username(ticket.get("assignee"))
        can_manage = actor.get("role") == ROLE_ADMIN or assignee == actor_user
        if not assignee:
            if st.button(
                "Asignarme esta carga",
                key=f"full_load_assign_me_{selected_code}",
                use_container_width=True,
            ):
                try:
                    service.assign(actor, selected_code, actor.get("user"))
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
            st.info("Asígnate la solicitud para preparar, iniciar y cerrar la carga.")
            return
        if not can_manage:
            st.info(f"Esta solicitud está asignada a {ticket.get('assignee')}. Puedes descargar los archivos para consulta.")
            return

        status = ticket.get("status")
        if status in {STATE_APPROVED, STATE_PREPARING, STATE_FAILED}:
            if st.button(
                "Preparar para carga",
                type="primary",
                key=f"full_load_prepare_{selected_code}",
                use_container_width=True,
            ):
                try:
                    service.run_dry_run(actor, selected_code)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
        elif status == STATE_DRY_RUN:
            st.info("La solicitud está ejecutando su validación previa.")
        elif status == STATE_READY_EXECUTE:
            if st.button(
                "Marcar carga iniciada",
                type="primary",
                key=f"full_load_start_{selected_code}",
                use_container_width=True,
            ):
                try:
                    service.start_load(actor, selected_code)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
        elif status in {STATE_LOADING, STATE_VALIDATING}:
            _render_full_load_ticket_close(service, actor, ticket, latest_version)


def render_ticket_inbox(service, actor, brand_view=False):
    render_ticket_styles()
    widget_key = f"ticket_open_{actor.get('role')}"
    deleted_code = clean_value(st.session_state.pop("_catalog_ticket_deleted", ""))
    if deleted_code:
        st.session_state.pop("selected_catalog_ticket", None)
        st.session_state.pop(widget_key, None)

    title = "Mis solicitudes de catálogo" if brand_view else "Centro de solicitudes de catálogo"
    subtitle = (
        "Consulta tus archivos, observaciones y resultados."
        if brand_view
        else "Prioriza, asigna y controla cada solicitud desde su recepción hasta el cierre."
    )
    st.markdown(
        f'<div class="ticket-hero"><p>Flujo controlado</p><h1>{escape(title)}</h1><span>{escape(subtitle)}</span></div>',
        unsafe_allow_html=True,
    )
    if deleted_code:
        st.success(f"La solicitud {deleted_code} fue eliminada y ya no aparece en la bandeja.")
    try:
        all_tickets = service.list_tickets(actor)
    except TicketError as exc:
        st.error(f"No se pudo leer la bandeja: {exc}")
        return
    actor_user = clean_value(actor.get("user")).casefold()
    recent_notifications = []
    for ticket in all_tickets:
        for notification in ticket.get("notifications", []):
            recipients = {clean_value(value).casefold() for value in notification.get("recipients", [])}
            if not recipients or actor_user in recipients or actor.get("role") == ROLE_ADMIN:
                recent_notifications.append({
                    "Fecha": clean_value(notification.get("created_at")).replace("T", " ")[:16],
                    "Ticket": ticket.get("code"),
                    "Mensaje": notification.get("message"),
                    "Canal": "Interna" if notification.get("channel") == "internal" else notification.get("channel"),
                })
    if recent_notifications:
        recent_notifications = sorted(recent_notifications, key=lambda item: item["Fecha"], reverse=True)[:8]
        with st.expander(f"Notificaciones recientes ({len(recent_notifications)})", expanded=False):
            st.dataframe(pd.DataFrame(recent_notifications), use_container_width=True, hide_index=True)
    if brand_view:
        # KPIs de las solicitudes que envio este usuario comercial.
        # Se agrupan por familia de estado para no exponer los 19 estados internos.
        estados = [ticket.get("status") for ticket in all_tickets]
        en_proceso = {
            STATE_ASSIGNED, STATE_REVIEW, STATE_APPROVED, STATE_PREPARING,
            STATE_DRY_RUN, STATE_READY_EXECUTE, STATE_LOADING, STATE_VALIDATING,
            STATE_CORRECTED,
        }
        observadas = {STATE_OBSERVED, STATE_WAITING_BRAND}
        finalizadas = {STATE_COMPLETED, STATE_COMPLETED_OBS}
        pendientes = {STATE_PENDING, STATE_REQUEST_RECEIVED, STATE_DRAFT}
        render_ticket_kpi_grid([
            ("Total enviadas", len(all_tickets), "blue"),
            ("Pendientes", sum(1 for e in estados if e in pendientes), "amber"),
            ("En proceso", sum(1 for e in estados if e in en_proceso), "blue"),
            ("Observadas", sum(1 for e in estados if e in observadas), "red"),
            ("Finalizadas", sum(1 for e in estados if e in finalizadas), "green"),
        ])
    if not brand_view:
        estados = [ticket.get("status") for ticket in all_tickets]
        en_proceso = {
            STATE_ASSIGNED, STATE_REVIEW, STATE_APPROVED, STATE_PREPARING,
            STATE_DRY_RUN, STATE_READY_EXECUTE, STATE_LOADING, STATE_VALIDATING,
            STATE_CORRECTED,
        }
        observadas = {STATE_OBSERVED, STATE_WAITING_BRAND}
        finalizadas = {STATE_COMPLETED, STATE_COMPLETED_OBS}
        vencidos = sum(ticket_is_overdue(item) for item in all_tickets)
        fallidos = estados.count(STATE_FAILED)
        kpis = [
            ("Sin asignar", sum(1 for i in all_tickets if not i.get("assignee") and i.get("status") not in finalizadas), "amber"),
            ("Asignados a mí", sum(1 for i in all_tickets if i.get("assignee") == actor.get("user")), "blue"),
            ("En proceso", sum(1 for e in estados if e in en_proceso), "blue"),
            ("Observadas", sum(1 for e in estados if e in observadas), "red"),
            ("Finalizadas", sum(1 for e in estados if e in finalizadas), "green"),
        ]
        # Vencidos y fallidos solo aparecen si hay algo que atender.
        if vencidos:
            kpis.append(("Vencidos", vencidos, "red"))
        if fallidos:
            kpis.append(("Fallidos", fallidos, "red"))
        render_ticket_kpi_grid(kpis)
    with st.container(border=True):
        st.markdown("<h3 class=\"ticket-filter-title\">Buscar y filtrar solicitudes</h3>", unsafe_allow_html=True)
        filter_cols = st.columns([1.2, 1, 1, 1, 1.2])
        brands = sorted({clean_value(item.get("brand")) for item in all_tickets if clean_value(item.get("brand"))})
        statuses = sorted({item.get("status") for item in all_tickets if item.get("status")})
        assignees = sorted({clean_value(item.get("assignee")) for item in all_tickets if clean_value(item.get("assignee"))})
        with filter_cols[0]:
            search = st.text_input("Buscar", placeholder="Ticket, Mod-Col, archivo o usuario", key=f"ticket_search_{actor.get('role')}")
        with filter_cols[1]:
            brand_filter = st.selectbox("Marca", ["Todas"] + brands, key=f"ticket_brand_filter_{actor.get('role')}")
        with filter_cols[2]:
            state_label_options = ["Todos"] + [STATE_LABELS.get(value, value) for value in statuses]
            state_label = st.selectbox("Estado", state_label_options, key=f"ticket_state_filter_{actor.get('role')}")
        with filter_cols[3]:
            priority_label = st.selectbox("Prioridad", ["Todas"] + [PRIORITY_LABELS[key] for key in PRIORITIES], key=f"ticket_priority_filter_{actor.get('role')}")
        with filter_cols[4]:
            assignee_filter = st.selectbox("Responsable", ["Todos"] + assignees, key=f"ticket_assignee_filter_{actor.get('role')}")
        secondary_filters = st.columns([1, 1, 1, 1])
        sites = sorted({site for item in all_tickets for site in item.get("sites", []) if clean_value(site)})
        load_types = sorted({clean_value(item.get("load_type")) for item in all_tickets if clean_value(item.get("load_type"))})
        with secondary_filters[0]:
            site_filter = st.selectbox("Sitio", ["Todos"] + sites, key=f"ticket_site_filter_{actor.get('role')}")
        with secondary_filters[1]:
            load_type_labels = {"complete": "Completa", "partial": "Parcial"}
            load_type_label = st.selectbox(
                "Tipo de carga",
                ["Todas"] + [load_type_labels.get(value, value.title()) for value in load_types],
                key=f"ticket_load_filter_{actor.get('role')}",
            )
        with secondary_filters[2]:
            date_from = st.date_input("Desde", value=None, key=f"ticket_date_from_{actor.get('role')}")
        with secondary_filters[3]:
            date_to = st.date_input("Hasta", value=None, key=f"ticket_date_to_{actor.get('role')}")
    state_filter = ""
    if state_label != "Todos":
        state_filter = next((key for key, value in STATE_LABELS.items() if value == state_label), "")
    priority_filter = ""
    if priority_label != "Todas":
        priority_filter = next((key for key, value in PRIORITY_LABELS.items() if value == priority_label), "")
    load_type_filter = ""
    if load_type_label != "Todas":
        load_type_filter = next((key for key, value in load_type_labels.items() if value == load_type_label), "")
    filters = {
        "brand": "" if brand_filter == "Todas" else brand_filter,
        "status": state_filter,
        "priority": priority_filter,
        "assignee": "" if assignee_filter == "Todos" else assignee_filter,
        "site": "" if site_filter == "Todos" else site_filter,
        "load_type": load_type_filter,
        "date_from": date_from.isoformat() if date_from else "",
        "date_to": date_to.isoformat() if date_to else "",
    }
    try:
        tickets = service.list_tickets(actor, filters=filters, search=search)
    except TicketError as exc:
        st.error(str(exc))
        return
    if not tickets:
        st.info("No hay solicitudes que coincidan con los filtros.")
        return
    table_df = _ticket_table(tickets)
    st.dataframe(table_df, use_container_width=True, hide_index=True)
    ticket_codes = [ticket.get("code") for ticket in tickets]
    if st.session_state.get(widget_key) not in ticket_codes:
        st.session_state.pop(widget_key, None)
    if st.session_state.get("selected_catalog_ticket") not in ticket_codes:
        st.session_state.pop("selected_catalog_ticket", None)
    selected_default = st.session_state.get("selected_catalog_ticket")
    selected_index = ticket_codes.index(selected_default) if selected_default in ticket_codes else 0
    selected_code = st.selectbox("Abrir solicitud", ticket_codes, index=selected_index, key=widget_key)
    st.session_state["selected_catalog_ticket"] = selected_code
    render_ticket_detail(service, actor, selected_code)


def render_ticket_detail(service, actor, code):
    try:
        ticket = service.get_ticket(actor, code)
    except TicketError as exc:
        st.error(str(exc))
        return
    status = ticket.get("status")
    status_label = STATE_LABELS.get(status, status)
    status_color = _ticket_state_color(status)
    summary = ticket.get("summary", {})
    st.markdown(
        f"""
        <div class="ticket-detail-header">
          <div><p>Solicitud de catálogo</p><h2>{escape(ticket.get('code', ''))}</h2></div>
          <span class="ticket-state {status_color}">{escape(status_label)}</span>
        </div>
        <div class="ticket-detail-shell">
          <div class="ticket-summary">
            <div><small>Marca y sitios</small><strong>{escape(ticket.get('brand',''))}</strong><span>{escape(', '.join(ticket.get('sites', [])))}</span></div>
            <div><small>Productos</small><strong>{int(summary.get('products',0))}</strong><span>{int(summary.get('model_colors',0))} modelo-color</span></div>
            <div><small>Solicitante</small><strong style="font-size:14px">{escape(ticket.get('requester',''))}</strong><span>{escape(ticket.get('filename',''))}</span></div>
            <div><small>Responsable</small><strong style="font-size:14px">{escape(ticket.get('assignee') or 'Sin asignar')}</strong><span>{escape(PRIORITY_LABELS.get(ticket.get('priority'), ticket.get('priority','')))}</span></div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    latest_version = (ticket.get("versions") or [{}])[-1]
    st.markdown(
        f'<p class="ticket-info-line">Versión validada {latest_version.get("number", 1)} de {len(ticket.get("versions", []))} · '
        f'Hash {escape(clean_value(latest_version.get("hash"))[:12])} · Plantilla {escape(clean_value(ticket.get("template_version")))}</p>',
        unsafe_allow_html=True,
    )
    with st.expander("Vista previa de la solicitud", expanded=False):
        preview_cols = st.columns(3)
        preview_cols[0].metric("Productos nuevos", int(summary.get("new_products", 0)))
        preview_cols[1].metric("Productos a actualizar", int(summary.get("updated_products", 0)))
        preview_cols[2].metric("Variantes detectadas", int(summary.get("variants", 0)))
        model_colors = [clean_value(value) for value in ticket.get("model_colors", []) if clean_value(value)]
        if model_colors:
            st.dataframe(
                pd.DataFrame({"Código modelo-color": model_colors}),
                use_container_width=True,
                hide_index=True,
                height=min(360, 38 + (len(model_colors) * 35)),
            )
        else:
            st.caption("El detalle completo está disponible en el archivo de validación descargable.")
    with st.container(border=True):
        st.markdown('<p class="ticket-h">Archivos</p>', unsafe_allow_html=True)
        download_cols = st.columns(3)
        try:
            if latest_version.get("input_path"):
                download_cols[0].download_button(
                    "Descargar input validado",
                    service.store.get_artifact(latest_version["input_path"]),
                    file_name=latest_version.get("filename") or ticket.get("filename") or "input.xlsx",
                    key=f"ticket_input_{code}_{latest_version.get('number', 1)}",
                )
            if latest_version.get("report_path"):
                download_cols[1].download_button(
                    "Descargar validación",
                    service.store.get_artifact(latest_version["report_path"]),
                    file_name=f"{code}_validacion_v{latest_version.get('number',1)}.xlsx",
                    key=f"ticket_report_{code}_{latest_version.get('number', 1)}",
                )
        except TicketError as exc:
            st.warning(f"No se pudo descargar un adjunto: {exc}")
    if ticket.get("warnings"):
        with st.expander(f"Advertencias ({len(ticket['warnings'])})"):
            for warning in ticket["warnings"][:100]:
                st.write(f"- {warning}")
    if ticket.get("observations"):
        with st.expander(f"Observaciones activas ({len(ticket['observations'])})", expanded=status == STATE_OBSERVED):
            observations_df = pd.DataFrame(ticket["observations"])
            if not observations_df.empty:
                st.dataframe(observations_df, use_container_width=True, hide_index=True)
    if ticket.get("job") and status not in {STATE_COMPLETED, STATE_COMPLETED_OBS}:
        # Solo se muestra mientras el proceso sigue vivo. Al cerrar, el bloque
        # de resultado ya cuenta lo que paso y esto solo agregaba ruido.
        job = ticket.get("job", {})
        progress = max(0, min(100, safe_int_value(job.get("progress"), 0)))
        st.markdown('<p class="ticket-h">Estado del proceso</p>', unsafe_allow_html=True)
        st.progress(progress / 100.0)
        st.caption(clean_value(status_label))
    if ticket.get("result"):
        result = ticket.get("result", {})
        public_result = ticket.get("public_result", {})
        result_df = pd.DataFrame([result])
        result_status = clean_value(public_result.get("status")) or status_label
        result_message = clean_value(public_result.get("message")) or clean_value(result.get("message"))
        st.markdown('<p class="ticket-h">Resultado de la carga</p>', unsafe_allow_html=True)
        if status == STATE_COMPLETED:
            st.success(
                f"{result_status}. El archivo quedó registrado como cargado"
                + (f": {result_message}" if result_message else ".")
            )
        elif status == STATE_COMPLETED_OBS:
            st.warning(
                f"{result_status}. La carga terminó y conserva observaciones"
                + (f": {result_message}" if result_message else ".")
            )
        procesados = safe_int_value(public_result.get("processed"), safe_int_value(result.get("processed"), 0))
        errores = safe_int_value(public_result.get("errors"), safe_int_value(result.get("errors"), 0))
        version_cargada = clean_value(result.get("file_version")) or latest_version.get("number", 1)
        cerrado_por = clean_value(result.get("closed_by")) or "Operaciones"
        archivo = (
            clean_value(result.get("filename"))
            or latest_version.get("filename")
            or ticket.get("filename")
            or "Sin nombre"
        )
        # Cifras compactas: un digito no necesita 48px. El correo del responsable
        # no es una metrica, va como texto con su nombre legible.
        st.markdown(
            '<div class="close-grid">'
            f'<div class="close-stat"><small>Procesados</small><strong>{procesados:,}</strong></div>'
            f'<div class="close-stat"><small>Errores</small><strong>{errores:,}</strong></div>'
            f'<div class="close-stat"><small>Versión</small><strong>v{escape(str(version_cargada))}</strong></div>'
            "</div>"
            '<div class="close-meta">'
            f'<span><b>Archivo</b> {escape(archivo)}</span>'
            f'<span><b>Cerrada por</b> {escape(auth_display_name(cerrado_por))}</span>'
            "</div>",
            unsafe_allow_html=True,
        )
        with st.expander("Detalle técnico del cierre"):
            st.dataframe(result_df, use_container_width=True, hide_index=True)
        download_cols[2].download_button(
            "Descargar reporte final",
            dataframe_to_excel_bytes({"Resultado": result_df}),
            file_name=f"{code}_resultado_final.xlsx",
            key=f"ticket_final_report_{code}",
        )
    role = actor.get("role")
    if role == ROLE_BRAND and status == STATE_OBSERVED:
        st.markdown('<p class="ticket-h">Enviar una nueva versión</p>'
                    '<p class="ticket-hint">La versión anterior se conserva para mantener la trazabilidad.</p>',
                    unsafe_allow_html=True)
        st.warning("Operaciones solicitó una corrección.")
        correction = st.file_uploader("Archivo corregido", type=["xlsx", "xls"], key=f"ticket_correction_{code}")
        correction_comment = st.text_area("Respuesta a la observación", key=f"ticket_correction_comment_{code}")
        if correction is not None and st.button("Validar y enviar corrección", type="primary", key=f"submit_correction_{code}"):
            preview_df, report_df, summary_df = validate_brand_commercial_input(correction, ticket.get("brand"))
            blocked = _ticket_summary_value(summary_df, "Registros bloqueados")
            if blocked or preview_df.empty:
                st.error(f"La corrección aún tiene {blocked} bloqueos o no contiene productos válidos.")
            else:
                try:
                    saved = service.add_correction_version(
                        actor,
                        code,
                        correction.name,
                        correction.getvalue(),
                        dataframe_to_excel_bytes({"Resumen": summary_df, "Vista previa": preview_df, "Errores": report_df}),
                        {
                            "products": _ticket_summary_value(summary_df, "Filas analizadas"),
                            "model_colors": _ticket_summary_value(summary_df, "Modelos-color"),
                            "blocked": blocked,
                        },
                        correction_comment,
                    )
                    st.success(f"Versión {len(saved.get('versions', []))} enviada.")
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
    if role in {ROLE_OPERATOR, ROLE_ADMIN}:
        st.markdown('<p class="ticket-h">Acciones internas</p>', unsafe_allow_html=True)
        current_priority = ticket.get("priority", "normal")
        priority_options = list(PRIORITIES)
        priority_col, priority_action = st.columns([2, 1])
        selected_priority = priority_col.selectbox(
            "Prioridad y SLA",
            priority_options,
            index=priority_options.index(current_priority) if current_priority in priority_options else 1,
            format_func=lambda value: PRIORITY_LABELS.get(value, value),
            key=f"ticket_priority_{code}",
        )
        if priority_action.button("Actualizar prioridad", key=f"save_ticket_priority_{code}"):
            try:
                service.set_priority(actor, code, selected_priority)
                st.rerun()
            except TicketError as exc:
                st.error(str(exc))
        if status in {STATE_PENDING, STATE_ASSIGNED, STATE_REVIEW, STATE_OBSERVED, STATE_CORRECTED, STATE_APPROVED, STATE_FAILED}:
            assign_cols = st.columns([2, 1])
            operator_options = ticket_operator_users()
            current_assignee = _normalize_auth_username(ticket.get("assignee"))
            assignee_index = operator_options.index(current_assignee) if current_assignee in operator_options else 0
            assignee = assign_cols[0].selectbox(
                "Responsable de carga",
                operator_options,
                index=assignee_index,
                format_func=ticket_operator_display_name,
                key=f"assign_user_{code}",
            )
            if assign_cols[1].button("Guardar responsable", key=f"assign_other_{code}"):
                try:
                    service.assign(actor, code, assignee)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
        if status in {STATE_PENDING, STATE_ASSIGNED, STATE_CORRECTED}:
            if st.button("Iniciar revisión", key=f"review_{code}"):
                try:
                    service.start_review(actor, code)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
        if status in {STATE_REVIEW, STATE_CORRECTED}:
            review_comment = st.text_area("Observación o decisión", key=f"review_comment_{code}")
            with st.expander("Agregar observación por producto o campo"):
                observation_cols = st.columns(2)
                observed_product = observation_cols[0].text_input("Producto / Mod-Col", key=f"observed_product_{code}")
                observed_field = observation_cols[1].text_input("Campo", key=f"observed_field_{code}")
                found_value = observation_cols[0].text_input("Valor encontrado", key=f"observed_found_{code}")
                recommendation = observation_cols[1].text_input("Corrección recomendada", key=f"observed_recommendation_{code}")
            structured_observations = []
            if any(clean_value(value) for value in [observed_product, observed_field, found_value, recommendation]):
                structured_observations.append(
                    {
                        "Producto": clean_value(observed_product),
                        "Campo": clean_value(observed_field),
                        "Valor encontrado": clean_value(found_value),
                        "Corrección recomendada": clean_value(recommendation),
                    }
                )
            action_cols = st.columns(3)
            if action_cols[0].button("Solicitar corrección", key=f"observe_{code}"):
                try:
                    service.request_correction(actor, code, review_comment, structured_observations)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
            if action_cols[1].button("Aprobar", type="primary", key=f"approve_{code}"):
                try:
                    service.approve(actor, code, review_comment)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
            if action_cols[2].button("Rechazar", key=f"reject_{code}"):
                try:
                    service.reject(actor, code, review_comment)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
        if status == STATE_APPROVED:
            run_cols = st.columns(2)
            if run_cols[0].button("Ejecutar simulación", key=f"dry_run_{code}"):
                try:
                    service.run_dry_run(actor, code)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
            if ticket.get("dry_run", {}).get("status") == "completed" and run_cols[1].button("Marcar carga iniciada", type="primary", key=f"start_load_{code}"):
                try:
                    service.start_load(actor, code)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
        if status == STATE_FAILED:
            if st.button("Reintentar carga", type="primary", key=f"retry_load_{code}"):
                try:
                    service.start_load(actor, code)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
        if status in {STATE_LOADING, STATE_VALIDATING}:
            st.markdown('<p class="ticket-h">Cerrar carga</p>'
                        '<p class="ticket-hint">Registra el resultado cuando termines la revisión en Shopify.</p>',
                        unsafe_allow_html=True)
            close_metrics = st.columns(2)
            processed_count = close_metrics[0].number_input(
                "Productos procesados",
                min_value=0,
                value=max(0, safe_int_value(summary.get("products"), 0)),
                step=1,
                key=f"close_ticket_processed_{code}",
            )
            error_count = close_metrics[1].number_input(
                "Productos con error",
                min_value=0,
                value=0,
                step=1,
                key=f"close_ticket_errors_{code}",
            )
            close_note = st.text_area(
                "Resultado u observaciones de la carga",
                key=f"close_ticket_note_{code}",
                placeholder="Ejemplo: carga completada y revisada en Shopify.",
            )
            close_confirmed = st.checkbox(
                "Confirmo que la carga terminó y que revisé el resultado en Shopify.",
                key=f"close_ticket_confirmed_{code}",
            )
            close_result = {
                "processed": int(processed_count),
                "errors": int(error_count),
                "message": clean_value(close_note) or "Carga completada y validada en Shopify.",
                "detail": clean_value(close_note),
                "closed_by": actor.get("user"),
                "filename": latest_version.get("filename") or ticket.get("filename"),
                "file_version": latest_version.get("number", 1),
                "file_hash": latest_version.get("hash") or ticket.get("file_hash"),
            }
            close_cols = st.columns(3)
            if close_cols[0].button(
                "Finalizar carga",
                type="primary",
                key=f"complete_load_{code}",
                disabled=not close_confirmed,
            ):
                try:
                    service.record_job_result(actor, code, success=True, result=close_result)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
            if close_cols[1].button(
                "Finalizar con observaciones",
                key=f"complete_obs_load_{code}",
                disabled=not close_confirmed,
            ):
                try:
                    service.record_job_result(actor, code, success=True, observations=True, result=close_result)
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
            if close_cols[2].button(
                "Registrar incidencia",
                key=f"fail_load_{code}",
                disabled=not close_confirmed,
            ):
                try:
                    service.record_job_result(
                        actor,
                        code,
                        success=False,
                        result=close_result,
                        error=clean_value(close_note) or "Incidencia registrada por Operaciones.",
                    )
                    st.rerun()
                except TicketError as exc:
                    st.error(str(exc))
        if is_ticket_operator_user(actor.get("user")):
            with st.expander("Eliminar solicitud", expanded=False):
                delete_reason = st.text_area("Motivo de eliminación", key=f"delete_ticket_reason_{code}")
                delete_confirm = st.checkbox("Confirmo que deseo eliminar esta solicitud", key=f"delete_ticket_confirm_{code}")
                if st.button("Eliminar solicitud", key=f"delete_ticket_{code}", disabled=not delete_confirm):
                    try:
                        service.cancel_ticket(actor, code, delete_reason)
                        # El selectbox ya fue instanciado en esta ejecución. La
                        # limpieza se difiere al inicio del siguiente rerun para
                        # evitar que Streamlit restaure el ticket eliminado.
                        st.session_state["_catalog_ticket_deleted"] = code
                        st.rerun()
                    except TicketError as exc:
                        st.error(str(exc))
    st.markdown('<p class="ticket-h">Comentarios</p>', unsafe_allow_html=True)
    comment = st.text_area("Agregar comentario", key=f"ticket_comment_{code}", label_visibility="collapsed", placeholder="Escribe un comentario para el equipo...")
    if st.button("Publicar comentario", key=f"add_ticket_comment_{code}"):
        try:
            service.add_comment(actor, code, comment)
            st.rerun()
        except TicketError as exc:
            st.error(str(exc))
    comments = list(reversed(ticket.get("comments", [])))
    if not comments:
        st.markdown('<div class="ticket-comments-empty">Aún no hay comentarios para esta solicitud.</div>', unsafe_allow_html=True)
    for item in comments:
        st.markdown(f"**{escape(item.get('user',''))}** · {escape(item.get('created_at',''))}<br>{escape(item.get('message',''))}", unsafe_allow_html=True)
    with st.expander("Historial y auditoría", expanded=False):
        for event in reversed(ticket.get("events", [])):
            state_text = STATE_LABELS.get(event.get("to_state"), event.get("to_state", ""))
            st.markdown(
                f'<div class="ticket-event"><strong>{escape(state_text or event.get("action", ""))}</strong><br><small>{escape(event.get("created_at", ""))} · {escape(event.get("user", ""))}</small><br>{escape(event.get("detail", ""))}</div>',
                unsafe_allow_html=True,
            )


def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="XL", layout="wide")
    if not require_login():
        return
    bigquery_config = get_bigquery_config()
    bigquery_ready = is_bigquery_configured(bigquery_config)

    render_sidebar_brand()
    auth_user = clean_value(st.session_state.get("auth_user"))
    auth_scope = auth_access_scope(auth_user)
    st.session_state["auth_scope"] = auth_scope
    commercial_input_only = auth_scope == "commercial_input"
    ticket_actor = current_ticket_actor()
    ticket_operator = ticket_actor.get("role") == ROLE_OPERATOR
    with st.sidebar.container(key="logout_card"):
        nombre_sesion = auth_display_name(auth_user)
        rol_sesion = auth_role_label(auth_scope)
        iniciales = "".join(p[0] for p in nombre_sesion.split()[:2]).upper() or "U"
        st.markdown(
            '<div class="session-card">'
            f'<span class="session-avatar">{escape(iniciales)}</span>'
            '<span class="session-text">'
            f'<strong>{escape(nombre_sesion)}</strong>'
            f'<small>{escape(rol_sesion)}</small>'
            "</span></div>",
            unsafe_allow_html=True,
        )
        if st.button("Cerrar sesion"):
            audit_record("logout")
            st.session_state.pop("authenticated", None)
            st.session_state.pop("auth_user", None)
            st.session_state.pop("auth_scope", None)
            st.rerun()
    site_options = {config["site_label"]: key for key, config in SITE_CONFIGS.items()}
    current_site_label = clean_value(st.session_state.get("site_picker")) or next(iter(site_options))
    if current_site_label not in site_options:
        current_site_label = next(iter(site_options))
    current_site_key = site_options[current_site_label]
    current_site_config = get_site_config(get_brand_config(current_site_key), get_shopify_config(current_site_key))
    current_logo_src = image_data_uri(resolve_logo_path(current_site_config.get("logo_path") or current_site_config.get("logo", "")))
    current_brand_name = escape(clean_value(current_site_config.get("brand_name")) or current_site_label)
    current_site_name = escape(clean_value(current_site_config.get("site_label")) or current_site_label)
    current_logo_html = (
        f'<img src="{current_logo_src}" alt="{current_brand_name}">'
        if current_logo_src
        else f"<span>{current_brand_name[:2].upper()}</span>"
    )
    selected_site_label = current_site_label
    if not commercial_input_only and auth_scope not in {ROLE_BRAND, ROLE_OPERATOR}:
        with st.sidebar.container(key="site_picker_card"):
            logo_column, selector_column = st.columns([0.32, 0.68], gap="small", vertical_alignment="center")
            with logo_column:
                st.markdown(
                    f'<div class="site-picker-logo" aria-hidden="true">{current_logo_html}</div>',
                    unsafe_allow_html=True,
                )
            with selector_column:
                selected_site_label = st.selectbox(
                    "Sitio activo",
                    list(site_options),
                    index=list(site_options).index(current_site_label),
                    key="site_picker",
                )
    selected_site_key = site_options[selected_site_label]
    brand_config = get_brand_config(selected_site_key)
    shopify_config = get_shopify_config(selected_site_key)
    ui_config = get_site_config(brand_config, shopify_config)
    inject_styles(ui_config)
    if commercial_input_only:
        with st.sidebar.container(key="commercial_limited_access_card"):
            st.markdown('<p class="sidebar-label">Acceso comercial</p>', unsafe_allow_html=True)
            st.info("Descarga de formatos habilitada")
        render_commercial_input_center(download_only=True)
        return
    if auth_scope == ROLE_BRAND:
        allowed_brands = auth_allowed_brands(auth_user, ROLE_BRAND)
        st.sidebar.markdown('<p class="sidebar-label">Portal Brand</p>', unsafe_allow_html=True)
        with st.sidebar.container(key="brand_ticket_navigation"):
            if st.session_state.get("brand_portal_view") not in {"Input comercial", "Mis solicitudes"}:
                st.session_state["brand_portal_view"] = "Input comercial"
            sidebar_nav_button("Input comercial", "brand_portal_view", "Input comercial", "brand_portal_input")
            sidebar_nav_button("Mis solicitudes", "brand_portal_view", "Mis solicitudes", "brand_portal_tickets")
        service, backend = get_ticket_service()
        if backend == "local":
            st.caption("Entorno de prueba local. Persistencia productiva pendiente de GitHub.")
        if st.session_state.get("brand_portal_view") == "Mis solicitudes":
            render_ticket_inbox(service, ticket_actor, brand_view=True)
        else:
            render_commercial_input_center(forced_brands=allowed_brands, actor=ticket_actor)
        return
    render_allowed_brands_card(brand_config)
    with st.sidebar.container(key="storage_badge_card"):
        render_storage_badge()
    nav_options = ["KPIs de catálogo", "Input comercial", "Solicitudes", "Carga de catálogo"]
    if st.session_state.get("operation_area_choice") not in nav_options:
        st.session_state["operation_area_choice"] = nav_options[0]
    st.sidebar.markdown('<p class="sidebar-label">Operaciones</p>', unsafe_allow_html=True)
    with st.sidebar.container(key="operation_nav"):
        sidebar_nav_button("KPIs de catálogo", "operation_area_choice", "KPIs de catálogo", "operation_nav_kpis")
        sidebar_nav_button("Input comercial", "operation_area_choice", "Input comercial", "operation_nav_input")
        sidebar_nav_button("Solicitudes", "operation_area_choice", "Solicitudes", "operation_nav_tickets")
        operation_area = st.session_state.get("operation_area_choice", nav_options[0])
    operation_mode = "Carga completa"
    load_options = ["Carga completa", "Carga parcial"]
    if st.session_state.get("operation_mode_choice") not in load_options:
        st.session_state["operation_mode_choice"] = load_options[0]
    st.sidebar.markdown('<p class="sidebar-label">Modo de carga</p>', unsafe_allow_html=True)
    with st.sidebar.container(key="load_mode_nav"):
        sidebar_nav_button(
            "Carga completa",
            "operation_mode_choice",
            "Carga completa",
            "load_mode_complete",
            extra_state={"operation_area_choice": "Carga de catálogo"},
        )
        sidebar_nav_button(
            "Carga parcial",
            "operation_mode_choice",
            "Carga parcial",
            "load_mode_partial",
            extra_state={"operation_area_choice": "Carga de catálogo"},
        )
        operation_mode = st.session_state.get("operation_mode_choice", load_options[0])
    st.sidebar.markdown('<p class="sidebar-label">Acciones</p>', unsafe_allow_html=True)
    with st.sidebar.container(key="sidebar_actions"):
        st.button(
            "Nueva carga / refrescar",
            key="reset_load_workspace",
            help="Limpia archivos cargados, previews y resultados para empezar otra carga.",
            on_click=reset_load_workspace,
        )
        if is_shopify_configured(shopify_config):
            if st.button("Probar conexión Shopify", key="test_shopify_connection"):
                try:
                    shop = test_connection(shopify_config)
                    st.success(f"Conectado a {shop.get('name', brand_config['site_label'])}")
                    st.caption(shop.get("myshopifyDomain") or shopify_config["shop_domain"])
                    st.caption(f"Origen token: {shop.get('token_source', '')}")
                except ShopifyApiError as exc:
                    st.error(str(exc))
    with st.sidebar.container(key="shopify_sidebar_card"):
        render_sidebar_shopify_card(ui_config, shopify_config)
        if not is_shopify_configured(shopify_config):
            st.warning("API no configurada para este sitio.")
            st.code(
                f"""[shopify_sites.{selected_site_key}]
shop_domain = "tienda.myshopify.com"
client_id = "..."
client_secret = "..."
admin_access_token = "..."
api_version = "{DEFAULT_API_VERSION}"
""",
                language="toml",
            )

    render_top_header(ui_config)
    load_reset_message = st.session_state.pop("load_reset_message", "")
    if load_reset_message:
        st.success(load_reset_message)
    if operation_area == "KPIs de catálogo":
        render_catalog_kpi_dashboard(ui_config, brand_config, shopify_config, bigquery_ready)
        return
    if operation_area == "Input comercial":
        render_commercial_input_center(actor=ticket_actor)
        return
    if operation_area == "Solicitudes":
        service, backend = get_ticket_service()
        if backend == "local":
            st.warning("Modo local de prueba: configura el backend GitHub antes de habilitarlo para varios usuarios.")
        render_ticket_inbox(service, ticket_actor, brand_view=not ticket_operator)
        return

    render_stepper(ui_config, current_step=current_flow_step())

    if operation_mode == "Carga parcial":
        operation_labels = {
            "Centry": "centry",
            "Tags": "tags",
            "Fotos 10 vistas": "photos",
            "Siblings": "siblings",
            "Titulo": "title",
            "Guías de talla": "size_guides",
            "Mantención tecnologías": "technologies",
            "Mantención Body HTML": "body",
            "Activar inventario en sucursales": "inventory_locations",
        }
        st.markdown('<div class="section-card"><h2>Carga parcial</h2>', unsafe_allow_html=True)
        update_label = st.selectbox(
            "Que quieres actualizar",
            list(operation_labels),
            index=0,
            key=f"partial_operation_select_{brand_config['site_key']}_v3",
        )
        update_operation = operation_labels[update_label]
        update_source = st.radio(
            "Fuente de datos actuales",
            ["Shopify API", "Respaldo Excel"],
            index=0 if is_shopify_configured(shopify_config) else 1,
            help="Shopify API es la referencia operativa. El respaldo Excel solo se usa si la API no esta disponible.",
        )
        template_file = None
        if update_source == "Respaldo Excel":
            template_file = st.file_uploader(
                f"1. Subir respaldo operativo de {brand_config['site_label']}",
                type=["xlsx", "xls"],
                key="template_update",
                help="Se usa como respaldo para encontrar ID, Handle, tags actuales, fotos actuales y codigo modelo color.",
            )

        update_file = None
        tag_mode = "merge"
        image_mode = "replace"
        only_missing_images = True
        body_mode = "from_input"

        if update_operation == "centry":
            update_file = st.file_uploader(
                "2. Subir Excel con codigos modelo-color faltantes",
                type=["xlsx", "xls"],
                key="update_centry_codes",
                help="Puede venir con columna Mod-Col, Codigo Modelo Color, Cod Mod Col, SKU, o una sola columna con los codigos.",
            )
            st.caption("La app toma esos codigos, cruza contra Shopify y devuelve el Excel Centry listo para enviar.")
        elif update_operation == "tags":
            tag_mode = st.radio("Como aplicar tags", ["merge", "replace"], format_func=lambda v: "Agregar a los tags actuales" if v == "merge" else "Reemplazar todos los tags")
            update_file = st.file_uploader("2. Subir archivo con Mod-Col y Tags", type=["xlsx", "xls"], key="update_tags")
        elif update_operation == "photos":
            image_mode = st.radio("Comando de fotos", ["replace", "merge"], format_func=lambda v: "Reemplazar fotos del producto" if v == "replace" else "Agregar/mezclar fotos")
            if image_mode == "replace":
                only_missing_images = False
                st.caption("REPLACE procesa productos aunque ya tengan fotos: elimina las actuales y sube las 10 vistas nuevas por API.")
            else:
                only_missing_images = st.checkbox("Solo productos sin foto en el catálogo", value=False)
            update_file = st.file_uploader("2. Opcional: subir lista con Mod-Col a corregir", type=["xlsx", "xls"], key="update_photos")
            st.caption("Si no subes lista, revisa el catálogo completo. Siempre genera 10 URLs por producto.")
        elif update_operation == "siblings":
            st.caption("Recalcula siblings para todo el catálogo: todos los productos con el mismo código modelo quedan separados por comas.")
        elif update_operation == "title":
            update_file = st.file_uploader("2. Subir archivo con Mod-Col y Title", type=["xlsx", "xls"], key="update_title")
        elif update_operation == "body":
            st.info(
                "Mantención Body HTML: puedes subir un Excel con Mod-Col, Body HTML, Material y Cuidado "
                "para reemplazar el HTML de esos productos. Si no subes archivo, la app corrige solamente "
                "los Body HTML actuales donde Materiales y Cuidados estén mezclados o mal estructurados."
            )
            if update_source == "Shopify API":
                update_file = st.file_uploader(
                    "2. Opcional: subir Excel con Mod-Col, Body HTML, Material y Cuidado",
                    type=["xlsx", "xls"],
                    key="update_body_html",
                    help="Si subes archivo, la app busca cada Mod-Col en Shopify y arma el Body HTML con las columnas recibidas.",
                )
                body_mode = "from_input" if update_file else "fix_catalog"
                st.caption(
                    "Con archivo: reemplaza Body HTML de los Mod-Col indicados. "
                    "Sin archivo: revisa el catálogo actual y corrige solo HTML mal estructurado."
                )
            else:
                body_mode = "from_input" if template_file else "fix_catalog"
                st.caption(
                    "El Excel cargado arriba se interpretará como input de mantenimiento Body HTML. "
                    "Debe incluir Mod-Col y columnas como Body HTML, Material, Composición, Cuidado o Cuidados."
                )
        elif update_operation == "size_guides":
            if update_source == "Shopify API":
                update_file = st.file_uploader(
                    "2. Opcional: subir Excel con Mod-Col y Guía de tallas",
                    type=["xlsx", "xls"],
                    key="update_size_guides",
                    help=(
                        "Si subes archivo, se usa como propuesta. Si no subes archivo, la app revisa el catálogo "
                        "actual y valida la guía contra categoría, tipo, género y reglas TOP/BOTTOM."
                    ),
                )
                st.success("Opcional: sin archivo se audita el catálogo actual de Shopify.")
            else:
                st.success("El respaldo Excel cargado arriba se usará como input de guías de talla.")
            st.caption(
                "Guías de talla: valida vacíos, valores inválidos e incompatibilidades como calzado con guía de vestuario "
                "o vestuario/calzado con talla única. No modifica precio, stock, fotos, variantes ni otros metacampos."
            )
        elif update_operation == "technologies":
            if update_source == "Shopify API":
                update_file = st.file_uploader(
                    "2. Opcional: subir Excel con Mod-Col y Tecnologia",
                    type=["xlsx", "xls"],
                    key="update_technologies",
                    help=(
                        "Si subes archivo, la app usa sus columnas Cod Mod Col/Mod-Col y Tecnologia para actualizar "
                        "custom.tecnologia y custom.logo. Si no subes archivo, analiza el catálogo actual de Shopify."
                    ),
                )
                st.success("Opcional: sin archivo se analiza el catálogo actual de Shopify.")
            else:
                st.success("El respaldo Excel cargado arriba se usará como input de tecnologías.")
            st.caption(
                "Mantención tecnologías: detecta tecnologías principalmente desde Tags, y también desde título, "
                "descripción y metacampos disponibles. Luego actualiza custom.tecnologia y custom.logo sin borrar "
                "otros metacampos."
            )
        elif update_operation == "inventory_locations":
            update_file = st.file_uploader(
                "2. Opcional: subir lista de SKUs, Mod-Col o Handles",
                type=["xlsx", "xls"],
                key="update_inventory_locations",
                help="Si no subes archivo, se revisan todas las variantes con SKU del catálogo Shopify.",
            )
            configured_locations = clean_value(
                shopify_config.get("inventory_location_ids")
                or shopify_config.get("inventory_locations")
                or shopify_config.get("location_ids")
            )
            if configured_locations:
                configured_count = len([value for value in re.split(r"[,;|\n]+", configured_locations) if clean_value(value)])
                st.success(f"Locations configuradas en Secrets detectadas: {configured_count}")
            else:
                st.warning("No detecto inventory_location_ids en Secrets; intentare leer locations desde Shopify.")
            st.caption(
                "Activa cada inventory item con SKU en las sucursales de Shopify. "
                "Si el token no puede leer locations, configura inventory_location_ids en Secrets."
            )
        st.markdown("</div>", unsafe_allow_html=True)

        update_ready = (
            update_file
            or update_operation in ("photos", "siblings", "technologies", "size_guides", "inventory_locations")
            or body_mode == "fix_catalog"
            or (update_operation == "body" and update_source == "Respaldo Excel" and template_file)
            or (update_operation == "size_guides" and update_source == "Respaldo Excel" and template_file)
        )
        partial_context = "|".join(
            [
                clean_value(brand_config.get("site_key")),
                clean_value(update_source),
                clean_value(update_operation),
                clean_value(tag_mode),
                clean_value(image_mode),
                clean_value(body_mode),
                clean_value(only_missing_images),
                uploaded_file_fingerprint(update_file),
                uploaded_file_fingerprint(template_file),
            ]
        )
        if st.session_state.get("partial_context") != partial_context:
            for state_key in (
                "shopify_preview_df",
                "shopify_preview_issues_df",
                "shopify_preview_matrixify_df",
                "shopify_preview_diagnostic_df",
                "shopify_preview_operation",
                "shopify_apply_result_df",
                "inventory_activation_preview_df",
                "inventory_activation_rows",
                "inventory_activation_locations",
                "inventory_activation_issues_df",
                "inventory_activation_result_df",
                "centry_maintainer_df",
                "centry_maintainer_sial_df",
                "centry_maintainer_issues_df",
                "centry_maintainer_codes",
                "centry_maintainer_excel_bytes",
                f"shopify_preview_excel_{brand_config['site_key']}_{update_operation}",
            ):
                st.session_state.pop(state_key, None)
            st.session_state["partial_context"] = partial_context
        effective_update_source = update_source
        body_backup_input_file = None
        technology_backup_input_file = None
        size_guides_backup_input_file = None
        if update_operation == "body" and update_source == "Respaldo Excel" and template_file:
            body_backup_input_file = template_file
            if is_shopify_configured(shopify_config):
                effective_update_source = "Shopify API"
        if update_operation == "technologies" and update_source == "Respaldo Excel" and template_file:
            technology_backup_input_file = template_file
            if is_shopify_configured(shopify_config):
                effective_update_source = "Shopify API"
        if update_operation == "size_guides" and update_source == "Respaldo Excel" and template_file:
            size_guides_backup_input_file = template_file
            if is_shopify_configured(shopify_config):
                effective_update_source = "Shopify API"

        if effective_update_source == "Shopify API" and not is_shopify_configured(shopify_config):
            st.error("Este sitio no tiene Shopify API configurada en Secrets.")
            update_ready = False
        if update_operation == "inventory_locations" and effective_update_source != "Shopify API":
            st.error("La activación de inventario en sucursales solo se puede ejecutar con Shopify API.")
            update_ready = False

        if effective_update_source == "Shopify API" and update_ready:
            try:
                effective_update_file = (
                    body_backup_input_file
                    if body_backup_input_file is not None
                    else (
                        technology_backup_input_file
                        if technology_backup_input_file is not None
                        else (size_guides_backup_input_file if size_guides_backup_input_file is not None else update_file)
                    )
                )
                update_df = read_uploaded_excel_cached(
                    effective_update_file,
                    f"partial_{brand_config['site_key']}_{update_operation}",
                ) if effective_update_file else None
                if update_df is not None and update_operation != "centry":
                    _, detected_brands, blocked_brands = input_brand_report(update_df, brand_config)
                    if blocked_brands:
                        st.error(
                            f"El archivo tiene marcas no permitidas para {brand_config['site_label']}: "
                            f"{', '.join(blocked_brands)}."
                        )
                        st.stop()

                if update_operation == "centry":
                    if st.button("Generar Centry", type="primary"):
                        for state_key in (
                            "centry_maintainer_df",
                            "centry_maintainer_sial_df",
                            "centry_maintainer_issues_df",
                            "centry_maintainer_codes",
                        ):
                            st.session_state.pop(state_key, None)
                        codes = model_codes_from_excel(update_df)
                        if not codes:
                            st.error("El Excel no tiene codigos modelo-color reconocibles. Usa una columna como Mod-Col, Codigo Modelo Color, Cod Mod Col, SKU, o una sola columna con los codigos.")
                        else:
                            with st.spinner("Leyendo Shopify, BigQuery y armando Centry..."):
                                shopify_products = session_shopify_products(brand_config["site_key"], shopify_config)
                                shopify_matrixify_df = shopify_products_to_matrixify_df(shopify_products)
                                arti_df, arti_source = session_arti_for_app(brand_config)
                                centry_matrixify_df, master_issues_df = build_centry_matrixify_from_master(
                                    codes,
                                    shopify_matrixify_df,
                                    arti_df,
                                    brand_config,
                                )
                                centry_df, centry_issues_df = build_centry_from_matrixify(
                                    centry_matrixify_df,
                                    brand_config,
                                    only_codes=codes,
                                    arti_df=arti_df,
                                )
                                centry_sial_df = build_centry_sial_from_matrixify(centry_matrixify_df, brand_config)
                                centry_issues_df = pd.concat(
                                    [master_issues_df, centry_issues_df],
                                    ignore_index=True,
                                ).drop_duplicates()
                            st.caption(f"Base maestra usada: {arti_source}")
                            st.session_state["centry_maintainer_df"] = centry_df
                            st.session_state["centry_maintainer_sial_df"] = centry_sial_df
                            st.session_state["centry_maintainer_issues_df"] = centry_issues_df
                            st.session_state["centry_maintainer_codes"] = codes
                            st.session_state["centry_maintainer_excel_bytes"] = dataframe_to_excel_bytes(
                                {
                                    "Centry": centry_df,
                                    "Carga Sial": centry_sial_df if centry_sial_df is not None else pd.DataFrame(),
                                    "Revision Centry": centry_issues_df if centry_issues_df is not None else pd.DataFrame(),
                                }
                            )

                    centry_df = st.session_state.get("centry_maintainer_df")
                    centry_sial_df = st.session_state.get("centry_maintainer_sial_df", pd.DataFrame())
                    centry_issues_df = st.session_state.get("centry_maintainer_issues_df", pd.DataFrame())
                    codes = st.session_state.get("centry_maintainer_codes", [])
                    if centry_df is not None:
                        if centry_df.empty:
                            st.warning("No se encontraron filas Centry para los codigos indicados.")
                        else:
                            st.success(f"Centry generado con {len(centry_df):,} filas para {len(codes):,} codigos consultados.")
                            render_centry_preview(centry_df, centry_issues_df, "Centry generado")
                            if centry_sial_df is not None and not centry_sial_df.empty:
                                st.write("Vista previa Carga Sial Centry")
                                st.dataframe(centry_sial_df.head(100), use_container_width=True, height=320)
                            missing_ean_count = centry_missing_ean_count(centry_df)
                            st.download_button(
                                "Descargar Centry",
                                data=st.session_state.get("centry_maintainer_excel_bytes")
                                or dataframe_to_excel_bytes(
                                    {
                                        "Centry": centry_df,
                                        "Carga Sial": centry_sial_df if centry_sial_df is not None else pd.DataFrame(),
                                        "Revision Centry": centry_issues_df if centry_issues_df is not None else pd.DataFrame(),
                                    }
                                ),
                                file_name=f"centry_{brand_config['site_key']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_centry_partial_{brand_config['site_key']}",
                            )
                            if missing_ean_count:
                                st.warning(
                                    f"Faltan {missing_ean_count:,} EAN. "
                                    "Revisa la hoja Revision Centry para ver columnas EAN detectadas desde BigQuery."
                                )
                    return

                if update_operation == "inventory_locations":
                    if st.button("Analizar activación de inventario", type="primary"):
                        with st.spinner("Leyendo Shopify y preparando activación por sucursal..."):
                            shopify_products = session_shopify_products(brand_config["site_key"], shopify_config)
                            codes, skus = inventory_activation_filters_from_input(update_df)
                            activation_rows = _inventory_activation_rows_from_products(
                                shopify_products,
                                only_codes=codes,
                                only_skus=skus,
                            )
                            issues = []
                            try:
                                locations = _shopify_inventory_target_locations(shopify_config)
                            except Exception as exc:
                                locations = []
                                issues.append({"Tipo": "Sucursales", "Detalle": str(exc)})
                            preview_rows = []
                            for item in activation_rows:
                                preview_rows.append(
                                    {
                                        "Handle": item.get("Handle"),
                                        "Mod-Col": item.get("Mod-Col"),
                                        "SKU": item.get("SKU"),
                                        "Inventory Item GID": item.get("Inventory Item GID"),
                                        "Sucursales objetivo": len(locations),
                                        "Acción": "Validar diferencial al ejecutar",
                                    }
                                )
                            preview_df = pd.DataFrame(preview_rows)
                            if update_df is not None and not codes and not skus:
                                issues.append(
                                    {
                                        "Tipo": "Input",
                                        "Detalle": "El archivo no tiene columnas reconocibles de SKU, Variant SKU, Mod-Col, COD MOD COL o Handle.",
                                    }
                                )
                            if not locations:
                                issues.append({"Tipo": "Sucursales", "Detalle": "No se encontraron locations activas en Shopify."})
                            if not activation_rows:
                                issues.append({"Tipo": "Variantes", "Detalle": "No se encontraron variantes con SKU e Inventory Item GID."})
                            st.session_state["inventory_activation_preview_df"] = preview_df
                            st.session_state["inventory_activation_rows"] = activation_rows
                            st.session_state["inventory_activation_locations"] = locations
                            st.session_state["inventory_activation_issues_df"] = pd.DataFrame(issues)

                    preview_df = st.session_state.get("inventory_activation_preview_df")
                    issues_df = st.session_state.get("inventory_activation_issues_df", pd.DataFrame())
                    activation_rows = st.session_state.get("inventory_activation_rows", [])
                    locations = st.session_state.get("inventory_activation_locations", [])
                    if preview_df is not None:
                        if preview_df.empty:
                            st.warning("No se genero vista previa de activación.")
                        else:
                            estimated_pairs = len(activation_rows) * len(locations)
                            st.success(
                                f"Vista previa lista: {len(activation_rows):,} variantes con SKU x "
                                f"{len(locations):,} sucursales = {estimated_pairs:,} pares potenciales. "
                                "La ejecución solo activará las sucursales faltantes."
                            )
                            st.dataframe(preview_df.head(200), use_container_width=True, height=360)
                        if issues_df is not None and not issues_df.empty:
                            st.warning(f"Hay {len(issues_df):,} observaciones.")
                            st.dataframe(issues_df, use_container_width=True)
                        st.download_button(
                            "Descargar vista previa activación",
                            data=dataframe_to_excel_bytes(
                                {
                                    "Activaciones": preview_df if preview_df is not None else pd.DataFrame(),
                                    "Revision": issues_df if issues_df is not None else pd.DataFrame(),
                                }
                            ),
                            file_name=f"activacion_inventario_{brand_config['site_key']}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                        can_apply_activation = bool(activation_rows) and bool(locations)
                        if can_apply_activation:
                            default_batch_limit = safe_int_value(st.session_state.get("inventory_activation_max_actions") or 1500)
                            activation_limit = st.number_input(
                                "Máximo de activaciones por ejecución",
                                min_value=50,
                                max_value=10000,
                                value=max(50, default_batch_limit),
                                step=50,
                                help="La app guarda avance. Si quedan pendientes, vuelve a ejecutar para continuar sin repetir lo ya activo.",
                            )
                            st.session_state["inventory_activation_max_actions"] = activation_limit
                            confirm_activation = st.checkbox("Confirmo activar estas variantes en las sucursales Shopify")
                            if confirm_activation and st.button("Activar inventario en Shopify", type="primary"):
                                with st.spinner("Activando inventory items en sucursales..."):
                                    result_df = _activate_inventory_items_in_locations(
                                        shopify_config,
                                        activation_rows,
                                        locations=locations,
                                        max_actions=activation_limit,
                                    )
                                clear_shopify_products_cache(brand_config["site_key"])
                                st.session_state["inventory_activation_result_df"] = result_df
                                summary_df = _inventory_activation_summary_df(result_df)
                                if summary_df is not None and not summary_df.empty:
                                    st.dataframe(summary_df, use_container_width=True, hide_index=True)
                                st.dataframe(result_df, use_container_width=True, height=360)
                                st.download_button(
                                    "Descargar reporte de activación",
                                    data=dataframe_to_excel_bytes(
                                        {
                                            "Resumen": summary_df if summary_df is not None else pd.DataFrame(),
                                            "Resultado": result_df,
                                            "Errores": result_df[result_df["Resultado"] == "ERROR"] if "Resultado" in result_df.columns else pd.DataFrame(),
                                            "Pendientes": result_df[result_df["Resultado"] == "PENDIENTE"] if "Resultado" in result_df.columns else pd.DataFrame(),
                                        }
                                    ),
                                    file_name=f"resultado_activacion_inventario_{brand_config['site_key']}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                )
                        saved_result_df = st.session_state.get("inventory_activation_result_df")
                        if saved_result_df is not None and not saved_result_df.empty:
                            st.markdown("#### Último resultado de activación")
                            saved_summary_df = _inventory_activation_summary_df(saved_result_df)
                            if saved_summary_df is not None and not saved_summary_df.empty:
                                st.dataframe(saved_summary_df, use_container_width=True, hide_index=True)
                            st.dataframe(saved_result_df, use_container_width=True, height=360)
                    return

                if st.button(f"Analizar carga parcial: {update_label}", type="primary"):
                    with st.spinner("Leyendo productos actuales desde Shopify..."):
                        shopify_products = session_shopify_products(brand_config["site_key"], shopify_config)
                    preview_arti_df = None
                    if update_operation == "body":
                        with st.spinner("Leyendo BigQuery/ARTI para recuperar materiales..."):
                            preview_arti_df, _ = session_arti_for_app(brand_config)
                    preview_df, issues_df, matrixify_df = build_shopify_update_preview(
                        shopify_products,
                        update_df,
                        update_operation,
                        brand_config,
                        shopify_config=shopify_config,
                        arti_df=preview_arti_df,
                        tag_mode=tag_mode,
                        image_mode=image_mode,
                        only_missing_images=only_missing_images,
                        body_mode=body_mode,
                    )
                    st.session_state["shopify_preview_df"] = preview_df
                    st.session_state["shopify_preview_issues_df"] = issues_df
                    st.session_state["shopify_preview_matrixify_df"] = matrixify_df
                    if update_operation in ("body", "photos", "size_guides"):
                        st.session_state["shopify_preview_diagnostic_df"] = build_partial_diagnostic_table(
                            preview_df,
                            issues_df,
                            update_operation,
                        )
                    else:
                        st.session_state["shopify_preview_diagnostic_df"] = pd.DataFrame()
                    st.session_state["shopify_preview_operation"] = update_operation
                    st.session_state.pop(f"shopify_preview_excel_{brand_config['site_key']}_{update_operation}", None)

                preview_df = st.session_state.get("shopify_preview_df")
                issues_df = st.session_state.get("shopify_preview_issues_df", pd.DataFrame())
                matrixify_df = st.session_state.get("shopify_preview_matrixify_df", pd.DataFrame())
                diagnostic_df = st.session_state.get("shopify_preview_diagnostic_df", pd.DataFrame())
                if preview_df is not None:
                    if preview_df.empty:
                        st.warning("No se genero ninguna fila de vista previa.")
                    else:
                        st.success(f"Vista previa generada con {len(preview_df):,} cambios.")
                        if update_operation in ("body", "photos", "size_guides"):
                            render_partial_diagnostic_panel(diagnostic_df, update_operation)
                        else:
                            summary_df = partial_preview_summary(preview_df, issues_df)
                            st.write("Resumen de carga parcial")
                            st.dataframe(summary_df, use_container_width=True, hide_index=True)
                            st.dataframe(preview_df.head(100), use_container_width=True)
                    if issues_df is not None and not issues_df.empty:
                        st.warning(f"Hay {len(issues_df):,} observaciones.")
                        st.dataframe(issues_df, use_container_width=True)

                    excel_key = f"shopify_preview_excel_{brand_config['site_key']}_{update_operation}"
                    excel_bytes = st.session_state.get(excel_key)
                    if excel_bytes is None:
                        sheets = {
                            "Resumen": partial_preview_summary(preview_df, issues_df),
                            "Vista previa": preview_df if preview_df is not None else pd.DataFrame(),
                            "Revision": issues_df if issues_df is not None else pd.DataFrame(),
                            "Matrixify fotos": matrixify_df if matrixify_df is not None else pd.DataFrame(),
                        }
                        if update_operation in ("body", "photos", "size_guides"):
                            sheets["Resumen Diagnostico"] = partial_diagnostic_summary(diagnostic_df, update_operation)
                            sheets["Diagnostico"] = diagnostic_df if diagnostic_df is not None else pd.DataFrame()
                            sheets["Listos para aplicar"] = filter_preview_by_diagnostic_ready(preview_df, diagnostic_df)
                        excel_bytes = dataframe_to_excel_bytes(sheets)
                        st.session_state[excel_key] = excel_bytes
                    st.download_button(
                        "Descargar estructura Matrixify",
                        data=excel_bytes,
                        file_name=f"vista_previa_{update_operation}_{brand_config['site_key']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                    writable_preview_df = preview_df
                    if update_operation in ("body", "photos", "size_guides"):
                        writable_preview_df = filter_preview_by_diagnostic_ready(preview_df, diagnostic_df)
                    can_apply = update_operation in ("tags", "title", "body", "siblings", "photos", "technologies") and writable_preview_df is not None and not writable_preview_df.empty
                    if update_operation == "photos":
                        st.info("REPLACE elimina las fotos actuales del producto y sube las 10 URLs nuevas. MERGE agrega las URLs nuevas sin borrar las actuales.")
                    if update_operation == "technologies":
                        from generate_columbia_matrixify import site_uses_technology_logo_metaobjects

                        if site_uses_technology_logo_metaobjects(brand_config):
                            st.info("Se actualizaran custom.tecnologia y custom.logo. Los logos se resuelven contra los metaobjetos definidos en Columbia.")
                        else:
                            st.info("Se actualizara custom.tecnologia con el tipo definido para este sitio. No se enviaran logos de Columbia.")
                    if update_operation == "size_guides":
                        st.warning(
                            "Las guías de talla son page_reference. La vista previa y el Excel quedan listos, "
                            "pero la escritura directa por API se mantiene bloqueada para no enviar texto donde Shopify espera un ID de página."
                        )
                    if can_apply:
                        confirm_apply = st.checkbox("Confirmo que revise la vista previa y quiero aplicar en Shopify")
                        if confirm_apply:
                            render_persistent_sync_job_panel(
                                shopify_config,
                                brand_config,
                                writable_preview_df,
                                mode=f"partial_{update_operation}",
                                label="Carga parcial Shopify",
                                activate_inventory_locations=False,
                                session_key=f"shopify_partial_job_{brand_config['site_key']}_{update_operation}",
                            )
            except Exception as exc:
                st.error("No pude generar o aplicar la carga parcial con Shopify API.")
                st.exception(exc)
        elif update_source == "Respaldo Excel" and template_file and update_ready:
            try:
                template_df = read_uploaded_excel_sheet_or_first(
                    template_file,
                    f"partial_template_{brand_config['site_key']}_{update_operation}",
                    preferred_sheet="Products",
                )
                if "Vendor" in template_df.columns:
                    catalog_vendors = {
                        clean_value(value).lower()
                        for value in template_df["Vendor"].dropna()
                        if clean_value(value)
                    }
                    expected_vendors = expected_catalog_vendors(brand_config)
                    if catalog_vendors and catalog_vendors.isdisjoint(expected_vendors):
                        st.error(
                            f"El respaldo Excel cargado no parece ser de {brand_config['site_label']}. "
                            f"Vendors esperados: {', '.join(sorted(expected_vendors))}. Vendors encontrados: {', '.join(sorted(catalog_vendors))}."
                        )
                        st.stop()

                update_df = read_uploaded_excel_cached(
                    update_file,
                    f"partial_backup_{brand_config['site_key']}_{update_operation}",
                ) if update_file else None
                if update_df is not None:
                    _, detected_brands, blocked_brands = input_brand_report(update_df, brand_config)
                    if blocked_brands:
                        st.error(
                            f"El archivo tiene marcas no permitidas para {brand_config['site_label']}: "
                            f"{', '.join(blocked_brands)}."
                        )
                        st.stop()

                if st.button(f"Analizar carga parcial: {update_label}", type="primary"):
                    update_arti_df = None
                    if update_operation == "photos":
                        try:
                            update_arti_df, _ = session_arti_for_app(brand_config)
                        except Exception:
                            update_arti_df = None
                    matrixify_df, issues_df = build_matrixify_updates(
                        template_df,
                        update_input_df=update_df,
                        arti=update_arti_df,
                        operation=update_operation,
                        brand_config=brand_config,
                        tag_mode=tag_mode,
                        image_mode=image_mode,
                        only_missing_images=only_missing_images,
                        body_mode=body_mode,
                    )
                    diagnostic_df = pd.DataFrame()
                    if matrixify_df.empty:
                        st.warning("No se genero ninguna fila de carga parcial. Revisa la hoja Revision.")
                    else:
                        st.success(f"Carga parcial generada con {len(matrixify_df):,} productos.")
                        if update_operation in ("body", "photos", "size_guides"):
                            diagnostic_df = build_partial_diagnostic_table(matrixify_df, issues_df, update_operation)
                            render_partial_diagnostic_panel(diagnostic_df, update_operation)
                        else:
                            st.dataframe(matrixify_df.head(100), use_container_width=True)
                    if issues_df is not None and not issues_df.empty:
                        st.warning(f"Hay {len(issues_df):,} observaciones.")
                        st.dataframe(issues_df, use_container_width=True)
                    if update_operation in ("body", "photos", "size_guides"):
                        excel_bytes = dataframe_to_excel_bytes(
                            {
                                "Products": matrixify_df,
                                "Resumen Diagnostico": partial_diagnostic_summary(diagnostic_df, update_operation),
                                "Diagnostico": diagnostic_df,
                                "Revision": issues_df if issues_df is not None else pd.DataFrame(),
                            }
                        )
                    else:
                        excel_bytes = update_to_excel_bytes(matrixify_df, issues_df)
                    st.download_button(
                        "Descargar estructura Matrixify",
                        data=excel_bytes,
                        file_name=f"carga_parcial_{update_operation}_{brand_config['site_key']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
            except Exception as exc:
                st.error("No pude generar la carga parcial.")
                st.exception(exc)
        else:
            st.info("Sube los archivos requeridos para generar la carga parcial.")
        return

    render_full_load_ticket_queue(brand_config)

    with st.container(key="sources_upload_panel"):
        render_sources_card(
            ui_config,
            bigquery_ready,
            input_count=int(st.session_state.get("input_row_count") or 0),
            shopify_count=int(st.session_state.get("shopify_product_count") or 0),
            arti_count=int(st.session_state.get("arti_row_count") or 0),
        )
        complete_source = st.radio(
            "Fuente de datos actuales",
            ["Shopify API", "Respaldo Excel"],
            index=0 if is_shopify_configured(shopify_config) else 1,
            help="Shopify API es la referencia operativa. El respaldo Excel solo se usa si la API no esta disponible.",
        )
        template_file = None
        upload_nonce = int(st.session_state.get("load_reset_nonce") or 0)
        input_upload_key = f"input_{upload_nonce}"
        template_upload_key = f"template_{upload_nonce}"
        upload_left, upload_right = st.columns([3.5, 1.5], gap="large")
        input_file = st.session_state.get(input_upload_key)
        with upload_left:
            if complete_source == "Respaldo Excel" and input_file:
                st.caption("Input cargado. Ahora sube el Catálogo Matrixify para conservar IDs.")
            elif complete_source == "Respaldo Excel":
                st.caption("Primero carga el input comercial. Luego este mismo espacio pedirá el Catálogo Matrixify.")
        with upload_right:
            if complete_source == "Respaldo Excel" and input_file:
                with st.container(key="catalog_upload_slot"):
                    template_file = st.file_uploader(
                        "Subir Catálogo Matrixify",
                        type=["xlsx", "xls"],
                        key=template_upload_key,
                        label_visibility="collapsed",
                        help="Este archivo conserva Product ID y Variant ID cuando no usas Shopify API.",
                    )
            else:
                with st.container(key="input_upload_slot"):
                    input_file = st.file_uploader("Cargar input", type=["xlsx", "xls"], key=input_upload_key, label_visibility="collapsed")
        st.session_state["input_loaded"] = bool(input_file)
        complete_context = "|".join(
            [
                clean_value(brand_config.get("site_key")),
                clean_value(complete_source),
                uploaded_file_fingerprint(input_file),
                uploaded_file_fingerprint(template_file),
            ]
        )
        if st.session_state.get("complete_context") != complete_context:
            clear_complete_load_state()
            st.session_state["complete_context"] = complete_context

    setup_rows = [
        {
            "Base": "Datos actuales Shopify",
            "Ruta": "Shopify API" if complete_source == "Shopify API" else f"Respaldo Excel de {brand_config['site_label']}",
            "Estado": "OK API" if complete_source == "Shopify API" and is_shopify_configured(shopify_config) else ("Obligatorio" if complete_source == "Respaldo Excel" else "Falta API"),
        },
        {
            "Base": "ARTI",
            "Ruta": "BigQuery" if bigquery_ready else f"{DEFAULT_ARTI_ZIP_PATH} / {DEFAULT_ARTI_CSV_PATH} / {DEFAULT_ARTI_PATH}",
            "Estado": "OK BigQuery"
            if bigquery_ready
            else (
                "OK ZIP"
                if Path(DEFAULT_ARTI_ZIP_PATH).exists()
                else ("OK CSV" if Path(DEFAULT_ARTI_CSV_PATH).exists() else ("OK XLSX" if Path(DEFAULT_ARTI_PATH).exists() else "Falta"))
            ),
        },
        {
            "Base": "Tipos Shopify",
            "Ruta": "data/tipos_shopify.xlsx",
            "Estado": "OK" if Path("data/tipos_shopify.xlsx").exists() else "Opcional",
        },
    ]
    render_base_status_card(setup_rows)

    can_process_complete = input_file and (complete_source == "Shopify API" or template_file)
    if can_process_complete:
        try:
            data_ready = (
                st.session_state.get("complete_data_context") == complete_context
                and st.session_state.get("complete_input_df") is not None
                and st.session_state.get("complete_template_df") is not None
                and st.session_state.get("complete_arti_df") is not None
            )
            if data_ready:
                input_df = st.session_state["complete_input_df"]
                template_df = st.session_state["complete_template_df"]
                arti_df = st.session_state["complete_arti_df"]
                template_source = st.session_state.get("complete_template_source", "")
                detected_brands = st.session_state.get("complete_detected_brands", [])
            else:
                if complete_source == "Shopify API":
                    if not is_shopify_configured(shopify_config):
                        st.error("Este sitio no tiene Shopify API configurada en Secrets.")
                        st.stop()
                    with st.spinner("Leyendo productos y variantes actuales desde Shopify..."):
                        shopify_products = session_shopify_products(brand_config["site_key"], shopify_config)
                    st.session_state["shopify_product_count"] = len(shopify_products)
                    st.session_state["complete_shopify_products"] = shopify_products
                    template_df = shopify_products_to_matrixify_df(shopify_products)
                    template_source = f"Shopify API ({len(shopify_products):,} productos)"
                else:
                    template_df = read_uploaded_excel_cached(
                        template_file,
                        f"complete_template_{brand_config['site_key']}",
                        sheet_name="Products",
                    )
                    template_source = f"respaldo operativo cargado para {brand_config['site_label']}"
                    if "Vendor" in template_df.columns:
                        catalog_vendors = {
                            clean_value(value).lower()
                            for value in template_df["Vendor"].dropna()
                            if clean_value(value)
                        }
                        expected_vendors = expected_catalog_vendors(brand_config)
                        if catalog_vendors and catalog_vendors.isdisjoint(expected_vendors):
                            st.error(
                                f"El respaldo Excel cargado no parece ser de {brand_config['site_label']}. "
                                f"Vendors esperados: {', '.join(sorted(expected_vendors))}. Vendors encontrados: {', '.join(sorted(catalog_vendors))}."
                            )
                            st.stop()

                input_df = read_uploaded_excel_cached(input_file, f"complete_input_{brand_config['site_key']}")
                st.session_state["input_row_count"] = len(input_df)
                brand_column, detected_brands, blocked_brands = input_brand_report(input_df, brand_config)
                if blocked_brands:
                    st.error(
                        f"El input tiene marcas no permitidas para {brand_config['site_label']}: "
                        f"{', '.join(blocked_brands)}. Marcas permitidas: {', '.join(brand_config['allowed_arti_brands'])}."
                    )
                    st.stop()

                try:
                    arti_df, arti_source = session_arti_for_app(brand_config)
                except FileNotFoundError:
                    st.error(
                        "Falta configurar BigQuery o dejar un respaldo local de ARTI: "
                        f"{DEFAULT_ARTI_ZIP_PATH}, {DEFAULT_ARTI_CSV_PATH} o {DEFAULT_ARTI_PATH}"
                    )
                    st.stop()
                except Exception as exc:
                    st.error("No se pudo leer el ARTI desde BigQuery.")
                    st.exception(exc)
                    st.stop()

                st.session_state["arti_row_count"] = len(arti_df)
                st.session_state["complete_input_df"] = input_df
                st.session_state["complete_template_df"] = template_df
                st.session_state["complete_arti_df"] = arti_df
                st.session_state["complete_template_source"] = template_source
                st.session_state["complete_detected_brands"] = detected_brands
                st.session_state["complete_data_context"] = complete_context
            left_col, right_col = st.columns([2, 1], gap="large")
            analyze_clicked = False
            with left_col:
                render_preview_table(input_df)
                with st.container(key="action_panel"):
                    render_analyze_card(ui_config)
                    analyze_clicked = st.button("Analizar input", type="primary", key=f"analyze_input_{brand_config['site_key']}")
                render_validations_card()
            with right_col:
                render_summary_metrics(
                    [
                        ("Columnas base", len(template_df.columns)),
                        ("Filas ARTI BigQuery", len(arti_df)),
                        ("Productos input", len(input_df)),
                        ("Marcas detectadas", len(detected_brands)),
                    ]
                )
                render_operational_status(ui_config, shopify_config, bigquery_ready, input_loaded=True)

            if analyze_clicked:
                with st.spinner("Analizando input y cruzando contra Shopify/BigQuery..."):
                    matrixify_df, summary_df, issues_df, type_warnings_df, skipped_df, sial_df = build_columbia_matrixify(
                        input_df, arti_df, template_df, brand_config=brand_config
                    )
                matrixify_df = coalesce_duplicate_columns(matrixify_df)
                summary_df = coalesce_duplicate_columns(summary_df)
                issues_df = coalesce_duplicate_columns(issues_df)
                type_warnings_df = coalesce_duplicate_columns(type_warnings_df)
                skipped_df = coalesce_duplicate_columns(skipped_df)
                sial_df = coalesce_duplicate_columns(sial_df)
                if complete_source == "Shopify API":
                    matrixify_df = apply_shopify_siblings_to_matrixify(
                        matrixify_df,
                        st.session_state.get("complete_shopify_products", []),
                    )
                centry_df, centry_issues_df = build_centry_from_matrixify(matrixify_df, brand_config, arti_df=arti_df)
                st.session_state["complete_matrixify_df"] = matrixify_df
                st.session_state["complete_summary_df"] = summary_df
                st.session_state["complete_issues_df"] = issues_df
                st.session_state["complete_type_warnings_df"] = type_warnings_df
                st.session_state["complete_skipped_df"] = skipped_df
                st.session_state["complete_sial_df"] = sial_df
                st.session_state["complete_centry_df"] = centry_df
                st.session_state["complete_centry_issues_df"] = centry_issues_df
                st.session_state["complete_analysis_message"] = (
                    f"Analisis terminado: {len(matrixify_df):,} filas Matrixify, "
                    f"{len(issues_df):,} observaciones, {len(skipped_df):,} omitidos sin cambios."
                )
                st.session_state["complete_excel_bytes"] = columbia_to_excel_bytes(
                    matrixify_df, summary_df, issues_df, type_warnings_df, skipped_df, sial_df, centry_df, centry_issues_df
                )

            matrixify_df = st.session_state.get("complete_matrixify_df")
            if matrixify_df is not None:
                summary_df = st.session_state.get("complete_summary_df", pd.DataFrame())
                issues_df = st.session_state.get("complete_issues_df", pd.DataFrame())
                type_warnings_df = st.session_state.get("complete_type_warnings_df", pd.DataFrame())
                skipped_df = st.session_state.get("complete_skipped_df", pd.DataFrame())
                sial_df = st.session_state.get("complete_sial_df", pd.DataFrame())
                centry_df = st.session_state.get("complete_centry_df", pd.DataFrame())
                centry_issues_df = st.session_state.get("complete_centry_issues_df", pd.DataFrame())
                matrixify_df = coalesce_duplicate_columns(matrixify_df)
                summary_df = coalesce_duplicate_columns(summary_df)
                issues_df = coalesce_duplicate_columns(issues_df)
                type_warnings_df = coalesce_duplicate_columns(type_warnings_df)
                skipped_df = coalesce_duplicate_columns(skipped_df)
                sial_df = coalesce_duplicate_columns(sial_df)
                centry_df = coalesce_duplicate_columns(centry_df)
                centry_issues_df = coalesce_duplicate_columns(centry_issues_df)

                analysis_message = st.session_state.get("complete_analysis_message")
                if analysis_message:
                    st.info(analysis_message)
                render_matrixify_result_card(ready=not matrixify_df.empty)
                if matrixify_df.empty:
                    st.error("No se pudo generar ninguna fila Matrixify.")
                    if issues_df is not None and not issues_df.empty:
                        st.warning(f"Hay {len(issues_df):,} observaciones para revisar.")
                        st.dataframe(issues_df, use_container_width=True)
                    if skipped_df is not None and not skipped_df.empty:
                        st.info(f"{len(skipped_df):,} productos fueron omitidos porque no presentaban cambios.")
                        st.dataframe(skipped_df, use_container_width=True)
                else:
                    st.success(f"Vista previa generada con {len(matrixify_df):,} variantes.")
                    matrixify_tab, centry_tab, revision_tab = st.tabs(["Matrixify", "Centry", "Revision"])
                    with matrixify_tab:
                        st.dataframe(summary_df, use_container_width=True)
                        st.dataframe(matrixify_df.head(100), use_container_width=True, height=360)
                    with centry_tab:
                        if centry_df is not None and not centry_df.empty:
                            render_centry_preview(centry_df, centry_issues_df)
                        else:
                            st.warning("No se genero vista previa Centry. Revisa si Matrixify tiene filas validas con SKU, Vendor y talla.")
                    with revision_tab:
                        if issues_df is not None and not issues_df.empty:
                            st.warning(f"Hay {len(issues_df):,} observaciones para revisar.")
                            st.dataframe(issues_df, use_container_width=True)
                        else:
                            st.success("Sin observaciones Matrixify.")
                        if type_warnings_df is not None and not type_warnings_df.empty:
                            st.warning(
                                "Tipos de prenda nuevos detectados. Revisa la hoja Tipos nuevos y la Revision antes de sincronizar: "
                                "hay que confirmar si el Type existe en la web destino o agregarlo al diccionario por web."
                            )
                            st.dataframe(type_warnings_df, use_container_width=True)
                        if skipped_df is not None and not skipped_df.empty:
                            st.info(f"{len(skipped_df):,} productos fueron omitidos porque no presentaban cambios.")
                            st.dataframe(skipped_df, use_container_width=True)
                        if sial_df is not None and not sial_df.empty:
                            st.write("Vista previa Carga Sial")
                            st.dataframe(sial_df.head(100), use_container_width=True, height=320)

                excel_bytes = st.session_state.get("complete_excel_bytes")
                if excel_bytes is None:
                    excel_bytes = columbia_to_excel_bytes(
                        matrixify_df, summary_df, issues_df, type_warnings_df, skipped_df, sial_df, centry_df, centry_issues_df
                    )
                    st.session_state["complete_excel_bytes"] = excel_bytes
                st.download_button(
                    "Descargar estructura Matrixify",
                    data=excel_bytes,
                    file_name=brand_config["output_filename"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                if complete_source == "Shopify API" and is_shopify_configured(shopify_config) and not matrixify_df.empty:
                    st.info(
                        "Sincronizacion directa habilitada: titulo, descripcion, vendor, tipo, tags, metafields, fotos, variantes, precios y SKUs de inventory item."
                        " Ninguna variante se envia por API sin SKU."
                    )
                    confirm_complete = st.checkbox("Confirmo que revise la vista previa y quiero sincronizar productos existentes en Shopify")
                    if confirm_complete:
                        render_persistent_sync_job_panel(
                            shopify_config,
                            brand_config,
                            matrixify_df,
                            mode="complete",
                            label="Carga completa Shopify",
                            activate_inventory_locations=True,
                            session_key=f"shopify_complete_job_{brand_config['site_key']}",
                        )
            st.markdown("</div>", unsafe_allow_html=True)
        except Exception as exc:
            st.error("No pude procesar los archivos.")
            st.exception(exc)
    else:
        st.info("Carga el input comercial para comenzar. Si no usas Shopify API, tambien sube el respaldo Excel del sitio.")

    st.markdown(
        """
        <div class="benefits-wrap">
        <div class="benefits">
            <div class="benefit"><b>Actualiza con IDs</b><p>Usa Shopify API como referencia operativa para conservar IDs de producto y variante.</p></div>
            <div class="benefit"><b>Variantes por talla</b><p>Lee ARTI y genera SKUs, barcodes, precios y tallas ordenadas.</p></div>
            <div class="benefit"><b>Estructura controlada</b><p>Entrega siempre las hojas y columnas necesarias para carga Matrixify.</p></div>
        </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def run_app():
    try:
        main()
    except Exception as exc:
        st.error("La app no pudo iniciar correctamente en este entorno.")
        st.warning(
            "Esto suele deberse a Secrets, permisos, dependencias o una respuesta externa de BigQuery/Shopify."
        )
        st.code(f"{type(exc).__name__}: {exc}")
        # El traceback es lo unico que permite ubicar el archivo y la linea.
        # Sin esto solo se ve el tipo de error y no hay forma de diagnosticarlo.
        detalle = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
        with st.expander("Ver detalle tecnico del error (traceback)"):
            st.caption(
                "Copia este bloque completo si necesitas reportar el problema. "
                "Tambien aparece en Manage app > Logs en Streamlit Cloud."
            )
            st.code(detalle, language="text")


if __name__ == "__main__":
    run_app()
