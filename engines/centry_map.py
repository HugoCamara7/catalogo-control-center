"""Motor de generacion Centry a partir de la plantilla oficial.

Sin dependencias de Streamlit.

Que hace
--------
1. Lee `data/plantilla_centry_productos.xlsx` como UNICA fuente de verdad de
   columnas y valores permitidos. Si Centry cambia la plantilla, se reemplaza el
   Excel y no hay que tocar Python.
2. Decide la familia del producto (superior, inferior, calzado, accesorios) a
   partir del tipo de prenda y la clase que ya trae el catalogo.
3. Arma las columnas del producto: las de la hoja SIEMPRE mas las de su familia.
4. Resuelve la categoria Centry con la tabla Marca|Tipo|Genero|Categoria,
   dando prioridad a la marca concreta sobre "Todos".
5. Valida los valores contra el diccionario de cada columna.

Que NO hace
-----------
No inventa valores. Cuando una equivalencia es dudosa devuelve un pendiente con
el motivo, para que la carga avise en vez de publicar algo inventado.
"""

import re
import unicodedata
from pathlib import Path

RUTA_PLANTILLA = Path(__file__).resolve().parents[1] / "data" / "plantilla_centry_productos.xlsx"

HOJA_SIEMPRE = "SIEMPRE"
HOJA_CATEGORIAS = "Categorias"

SUPERIOR = "superior"
INFERIOR = "inferior"
CALZADO = "calzado"
ACCESORIOS = "accesorios"

HOJA_POR_FAMILIA = {
    SUPERIOR: "Vestuario parte superior",
    INFERIOR: "Vestuario parte inferior",
    CALZADO: "Calzado",
    ACCESORIOS: "Accesorios",
}

ETIQUETA_FAMILIA = {
    SUPERIOR: "Vestuario parte superior",
    INFERIOR: "Vestuario parte inferior",
    CALZADO: "Calzado",
    ACCESORIOS: "Accesorios",
}


def _texto(valor):
    return "" if valor is None else str(valor).strip()


def normalizar(valor):
    """Minusculas, sin tildes, sin puntuacion y sin espacios repetidos.

    Es la clave con la que se comparan tipos, generos y valores de diccionario.
    "MOCASÍN", "Mocasin" y "mocasines" tienen que caer en la misma entrada.
    """
    texto = _texto(valor).casefold()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^\w\s/-]", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _singular(valor):
    """Quita el plural mas comun para emparejar 'Camisas' con 'Camisa'."""
    texto = normalizar(valor)
    if len(texto) > 3 and texto.endswith("es"):
        return texto[:-2]
    if len(texto) > 3 and texto.endswith("s"):
        return texto[:-1]
    return texto


# --- familia del producto ------------------------------------------------
# Se busca por palabra dentro del tipo de prenda. El orden importa: gana la
# primera familia cuyo termino aparezca, y calzado se evalua antes que vestuario
# porque "zapatilla de running" tambien contiene "running".
TERMINOS_FAMILIA = [
    (CALZADO, (
        "zapatilla", "zapato", "bota", "botin", "bototo", "sandalia", "mocasin",
        "alpargata", "ballerina", "pantufla", "chala", "hojota", "ojota",
        "calzado", "sueco", "zueco", "slipper", "nautico", "tacon", "taco",
    )),
    (ACCESORIOS, (
        "mochila", "bolso", "cartera", "billetera", "morral", "maleta", "canguro",
        "gorro", "gorra", "sombrero", "jockey", "bufanda", "panuelo", "guante",
        "cinturon", "correa", "media", "calcetin", "calcetine", "llavero",
        "collera", "anteojo", "lente", "botella", "termo", "bastone", "baston",
        "accesorio", "riñonera", "rinonera", "banano", "neceser", "estuche",
        "paraguas", "toalla", "manta", "saco de dormir", "carpa", "casco",
        # Tipos del diccionario de la app que no estaban cubiertos.
        "chullo", "pasamontana", "cuellera", "cooler", "funda para lata",
        "cuchilla", "maletin", "cantimplora", "silbato", "brujula", "linterna",
        "taza", "vaso", "jarro", "mug",
    )),
    (INFERIOR, (
        "pantalon", "short", "bermuda", "jogger", "falda", "legging", "calza",
        "buzo inferior", "pantaloneta", "capri", "jean", "cargo",
    )),
    (SUPERIOR, (
        "polo", "polera", "camisa", "camiseta", "casaca", "chaqueta", "chaleco",
        "abrigo", "sweater", "chompa", "poleron", "hoodie", "canguro superior",
        "cortavientos", "parka", "blusa", "top", "bividi", "musculosa", "vestido",
        "enterizo", "buzo", "primera capa", "polar", "hoody", "hoodie",
        "impermeable", "rompeviento", "anorak",
    )),
]

# La clase que ya trae el catalogo (Calzado / Vestuario / Accesorios) sirve de
# respaldo cuando el tipo de prenda no coincide con ningun termino.
FAMILIA_POR_CLASE = {
    "calzado": CALZADO,
    "accesorio": ACCESORIOS,
    "accesorios": ACCESORIOS,
}


def resolver_tipo(tipo_prenda):
    """El tipo pasa PRIMERO por el diccionario de la app.

    Devuelve (tipo_canonico, clase, aviso). Si el diccionario no lo conoce se
    acepta el tipo tal cual y se devuelve un aviso: no se descarta el producto
    ni se inventa un tipo. Asi la carga sale y la advertencia queda escrita.
    """
    texto = _texto(tipo_prenda)
    if not texto:
        return "", "", "el producto no trae tipo de prenda"
    try:
        from engines import garment_types
    except ImportError:
        return texto, "", ""
    canonico = garment_types.tipo_canonico(texto)
    if canonico:
        return canonico, garment_types.clase_de(texto), ""
    return texto, "", f'"{texto}" no está en el diccionario de tipos de prenda'


def detectar_familia(tipo_prenda="", clase="", categoria=""):
    """Devuelve (familia, motivo). familia vacia = no se pudo decidir.

    No adivina entre superior e inferior: si la clase dice "Vestuario" pero el
    tipo no coincide con ningun termino conocido, se devuelve vacio para que la
    carga lo marque como pendiente de revision.
    """
    for texto in (tipo_prenda, categoria):
        clave = normalizar(texto)
        if not clave:
            continue
        for familia, terminos in TERMINOS_FAMILIA:
            for termino in terminos:
                if re.search(rf"\b{re.escape(termino)}", clave):
                    return familia, f'"{_texto(texto)}" contiene "{termino}"'
    clave_clase = normalizar(clase)
    for nombre, familia in FAMILIA_POR_CLASE.items():
        if nombre in clave_clase:
            return familia, f'clase "{_texto(clase)}"'
    return "", (
        f'no se pudo deducir la familia de "{_texto(tipo_prenda) or _texto(clase)}"'
    )


# --- lectura de la plantilla ---------------------------------------------
_PLANTILLA = {}


def _leer_hoja(libro, nombre):
    """(columnas, {columna: [valores permitidos]}) de una hoja de la plantilla.

    Una columna repetida se queda con la aparicion que SI trae diccionario:
    en Calzado, "Forma de la punta" sale dos veces y la primera esta vacia.
    """
    hoja = libro[nombre]
    filas = list(hoja.iter_rows(values_only=True))
    if not filas:
        return [], {}
    cabeceras = [_texto(valor) for valor in filas[0]]
    columnas = []
    diccionarios = {}
    for indice, columna in enumerate(cabeceras):
        if not columna:
            continue
        valores = []
        for fila in filas[1:]:
            if indice >= len(fila):
                continue
            valor = _texto(fila[indice])
            if valor and valor not in valores:
                valores.append(valor)
        if columna in diccionarios:
            if valores and not diccionarios[columna]:
                diccionarios[columna] = valores
            continue
        columnas.append(columna)
        diccionarios[columna] = valores
    return columnas, diccionarios


def _leer_categorias(libro):
    """Tabla Marca|Tipo|Genero -> [categorias].

    Se guardan TODAS las categorias de una clave: 28 claves traen mas de una y
    la plantilla no dice cual manda. La eleccion se hace en resolver_categoria.
    """
    filas = list(libro[HOJA_CATEGORIAS].iter_rows(values_only=True))
    tabla = {}
    for fila in filas[1:]:
        if not fila or not _texto(fila[0]):
            continue
        marca, tipo, genero, categoria = (_texto(fila[indice]) if indice < len(fila) else "" for indice in range(4))
        if not categoria:
            continue
        clave = (normalizar(marca), _singular(tipo), normalizar(genero))
        tabla.setdefault(clave, [])
        if categoria not in tabla[clave]:
            tabla[clave].append(categoria)
    return tabla


def cargar_plantilla(ruta=None, recargar=False):
    """Lee la plantilla una sola vez y la deja en memoria."""
    global _PLANTILLA
    ruta = Path(ruta or RUTA_PLANTILLA)
    clave = str(ruta)
    if not recargar and _PLANTILLA.get("__clave__") == clave:
        return _PLANTILLA
    import openpyxl

    libro = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    # OJO: las filas de SIEMPRE son dos productos de EJEMPLO, no un diccionario
    # de valores permitidos. Tomarlas como diccionario rechazaria todo producto
    # real ("Casaca Lino Hombre..." no estaria "entre los valores permitidos").
    # Solo las cuatro hojas de tipo traen listas cerradas.
    siempre, ejemplos = _leer_hoja(libro, HOJA_SIEMPRE)
    datos = {
        "__clave__": clave,
        "siempre": siempre,
        "ejemplos": ejemplos,
        "diccionarios": {},
        "familias": {},
        "categorias": _leer_categorias(libro),
    }
    for familia, hoja in HOJA_POR_FAMILIA.items():
        columnas, diccionarios = _leer_hoja(libro, hoja)
        datos["familias"][familia] = columnas
        datos["diccionarios"].update(diccionarios)
    libro.close()
    _PLANTILLA = datos
    return datos


# Claves operativas que NO vienen en la plantilla de Centry pero la carga
# necesita para cruzar contra el maestro. Van siempre al final, en este orden.
COLUMNAS_CLAVE = ("COD MOD", "COD COL", "TALLA")
COLUMNA_ADVERTENCIA = "Advertencias"


def columnas_de(familia, ruta=None):
    """SIEMPRE + las de la familia + las claves + la advertencia, en ese orden.

    Con la familia sin determinar se devuelven solo las de SIEMPRE mas la cola:
    el producto igual sale, y la advertencia dice que le faltan los atributos
    de marketplace.
    """
    plantilla = cargar_plantilla(ruta)
    propias = plantilla["familias"].get(familia, [])
    cola = list(COLUMNAS_CLAVE) + [COLUMNA_ADVERTENCIA]
    return list(dict.fromkeys(plantilla["siempre"] + propias + cola))


def valores_permitidos(columna, ruta=None):
    """Diccionario de la columna. Lista vacia = texto libre, no se valida."""
    return list(cargar_plantilla(ruta)["diccionarios"].get(_texto(columna), []))


def valor_valido(columna, valor, ruta=None):
    """(valor_normalizado, ok). Empareja sin tildes ni mayusculas.

    Devuelve el valor tal y como lo escribe la plantilla, para que el archivo
    salga con la ortografia que Centry espera.
    """
    texto = _texto(valor)
    permitidos = valores_permitidos(columna, ruta)
    if not texto:
        return "", True
    if not permitidos:
        return texto, True
    for permitido in permitidos:
        if normalizar(permitido) == normalizar(texto):
            return permitido, True
    return texto, False


# --- categoria ------------------------------------------------------------
# Equivalencias de genero entre el catalogo y la plantilla.
GENEROS = {
    "hombre": "Hombre", "masculino": "Hombre", "varon": "Hombre", "caballero": "Hombre",
    "mujer": "Mujer", "femenino": "Mujer", "dama": "Mujer",
    "unisex": "Unisex Adultos", "unisex adultos": "Unisex Adultos",
    "unisex adulto": "Unisex Adultos", "adulto": "Unisex Adultos",
    "unisex ninos": "Unisex Niños", "unisex nino": "Unisex Niños",
    "nino": "Niños", "ninos": "Niños", "junior": "Niños",
    "nina": "Niñas", "ninas": "Niñas",
    "bebe": "Unisex Niños", "infantil": "Unisex Niños", "kids": "Unisex Niños",
}


def normalizar_genero(valor):
    """Devuelve el genero tal como lo escribe la plantilla, o "" si no se sabe."""
    clave = normalizar(valor)
    if not clave:
        return ""
    if clave in GENEROS:
        return GENEROS[clave]
    for nombre, salida in GENEROS.items():
        if re.search(rf"\b{re.escape(nombre)}\b", clave):
            return salida
    return ""


def _categoria_mas_general(opciones):
    """La de menos niveles. Entre "Botas" y "Botas / Botas de Invierno" gana la
    primera: es la que vale para cualquier producto de ese tipo."""
    return min(opciones, key=lambda texto: (texto.count("/"), len(texto)))


def resolver_categoria(marca="", tipo="", genero="", familia="", ruta=None):
    """(categoria, aviso). categoria vacia = no se encontro y queda pendiente.

    La marca concreta gana sobre "Todos". Si la clave trae varias categorias se
    elige la mas general y se avisa; si las opciones estan al mismo nivel (el
    caso de Guantes/Unisex Adultos, que duda entre Deporte Masculino y Femenino)
    no se elige ninguna: eso lo tiene que definir una persona.
    """
    plantilla = cargar_plantilla(ruta)
    tabla = plantilla["categorias"]
    genero_normalizado = normalizar_genero(genero)
    if not genero_normalizado:
        if not _texto(genero):
            return "", "el producto no trae género (ni en Shopify ni en BigQuery/ARTI)"
        return "", f'género "{_texto(genero)}" no reconocido'
    clave_tipo = _singular(tipo)
    if not clave_tipo:
        return "", "el producto no trae tipo de prenda"

    for marca_clave in (normalizar(marca), "todos"):
        if not marca_clave:
            continue
        opciones = tabla.get((marca_clave, clave_tipo, normalizar(genero_normalizado)))
        if not opciones:
            continue
        origen = "" if marca_clave != "todos" else " (regla Todos)"
        if len(opciones) == 1:
            return opciones[0], ""
        niveles = {texto.count("/") for texto in opciones}
        if len(niveles) == 1:
            # Mismo nivel: son alternativas de verdad, no general vs especifica.
            return "", (
                f'"{_texto(tipo)}" / {genero_normalizado} tiene {len(opciones)} categorias '
                f'igual de especificas en la plantilla: {" | ".join(opciones)}'
            )
        elegida = _categoria_mas_general(opciones)
        descartadas = [texto for texto in opciones if texto != elegida]
        return elegida, (
            f'se eligio la categoria general{origen}; revisar si corresponde '
            f'{" o ".join(descartadas)}'
        )
    # 2) No esta en la tabla: se deduce del mismo tipo en otro genero.
    deducida, origen = deducir_categoria(marca, tipo, genero, ruta)
    if deducida:
        return deducida, f"categoría deducida: {origen}. Revisar."

    # 3) Ultima red: la categoria base de su familia y genero, para que el
    #    producto no se quede sin categoria y Centry lo acepte igual.
    if not familia:
        # La CLASE del diccionario entra aqui tambien: "Slip Ons" no coincide
        # con ningun termino, pero el diccionario dice que es Calzado.
        _, clase_tipo, _ = resolver_tipo(tipo)
        familia = detectar_familia(tipo, clase_tipo)[0]
        if not familia and "vestuario" in normalizar(clase_tipo):
            # Superior e inferior comparten la misma base de categoria
            # ("Vestuario / Ropa Femenina"), asi que para la generica da igual.
            familia = SUPERIOR
    if familia:
        generica = categoria_generica(familia, genero_normalizado, ruta)
        if not generica:
            # Ese genero no tiene datos para esa familia: se traduce el de otro
            # genero del mismo bloque.
            for otro in _generos_cercanos(genero_normalizado):
                base = categoria_generica(familia, otro, ruta)
                if base:
                    generica = _traducir_categoria(base, genero_normalizado)
                    break
        if generica:
            return generica, (
                f'categoría genérica de {ETIQUETA_FAMILIA.get(familia, familia)} / '
                f'{genero_normalizado}: no hay regla para "{_texto(tipo)}". Revisar.'
            )
    return "", (
        f'no hay categoria para tipo "{_texto(tipo)}" y genero {genero_normalizado}'
    )


# --- construccion y validacion -------------------------------------------
# Campos de SIEMPRE sin los que Centry rechaza el producto.
OBLIGATORIAS = (
    "Nombre del Producto", "Marca", "Categoría", "SKU del producto",
    "SKU de la variante", "Color", "Talla", "URL imagen principal",
)

# Valores fijos que la plantilla trae con una sola opcion: no hay nada que
# elegir, se rellenan solos.
def _valores_fijos(familia, ruta=None):
    fijos = {}
    for columna in columnas_de(familia, ruta):
        permitidos = valores_permitidos(columna, ruta)
        if len(permitidos) == 1:
            fijos[columna] = permitidos[0]
    return fijos


def construir_producto(datos, marca="", tipo="", genero="", clase="", mod_col="", talla="", ruta=None):
    """Arma la fila Centry de un producto.

    Devuelve {"familia", "columnas", "fila", "pendientes"}.

    El producto SIEMPRE sale, aunque no se reconozca el tipo de prenda: en ese
    caso lleva las columnas de SIEMPRE y las claves, sin los atributos de
    marketplace, y la columna Advertencias explica que falta. Antes se devolvia
    vacio y el producto desaparecia del archivo.
    """
    pendientes = []
    familia, motivo = detectar_familia(tipo, clase, (datos or {}).get("Categoría", ""))
    if not familia:
        pendientes.append({
            "campo": "Familia",
            "problema": f"{motivo}; sale sin los atributos de Falabella/MercadoLibre",
        })

    columnas = columnas_de(familia, ruta)
    fila = {columna: "" for columna in columnas}
    if familia:
        fila.update(_valores_fijos(familia, ruta))

    for columna, valor in (datos or {}).items():
        if columna not in fila:
            continue
        limpio, ok = valor_valido(columna, valor, ruta)
        fila[columna] = limpio
        if not ok:
            pendientes.append({
                "campo": columna,
                "problema": f'"{_texto(valor)}" no está entre los valores permitidos',
                "permitidos": ", ".join(valores_permitidos(columna, ruta)[:8]),
            })

    if not _texto(fila.get("Categoría")):
        categoria, aviso = resolver_categoria(marca, tipo, genero, ruta)
        fila["Categoría"] = categoria
        if aviso:
            pendientes.append({"campo": "Categoría", "problema": aviso})

    genero_plantilla = normalizar_genero(genero)
    if genero_plantilla and not _texto(fila.get("Género")):
        fila["Género"] = genero_plantilla

    # Claves operativas: el modelo y el color salen del Mod-Col.
    modelo, color = _partir_mod_col(mod_col)
    fila["COD MOD"] = modelo
    fila["COD COL"] = color
    fila["TALLA"] = _texto(talla) or _texto(fila.get("Talla"))
    for columna in COLUMNAS_CLAVE:
        if not _texto(fila.get(columna)):
            pendientes.append({"campo": columna, "problema": "clave vacía; revisar el Mod-Col"})

    for columna in OBLIGATORIAS:
        if columna in fila and not _texto(fila[columna]):
            pendientes.append({"campo": columna, "problema": "obligatorio y está vacío"})

    fila[COLUMNA_ADVERTENCIA] = " | ".join(
        f'{p["campo"]}: {p["problema"]}' for p in pendientes
    )
    return {"familia": familia, "columnas": columnas, "fila": fila, "pendientes": pendientes}


def _partir_mod_col(mod_col):
    """"RK202011432-645" -> ("RK202011432", "645"). Sin guion, todo es modelo."""
    texto = _texto(mod_col).upper()
    if "-" not in texto:
        return texto, ""
    modelo, color = texto.rsplit("-", 1)
    return modelo.strip(), color.strip()


def validar_productos(productos, ruta=None):
    """Recorre una lista de productos ya construidos y junta sus pendientes.

    Cada producto es {"sku": ..., "resultado": <lo que devuelve construir_producto>}.
    Devuelve filas listas para la hoja de revision.
    """
    avisos = []
    for producto in productos or []:
        sku = _texto(producto.get("sku"))
        resultado = producto.get("resultado") or {}
        for pendiente in resultado.get("pendientes", []):
            avisos.append({
                "SKU": sku,
                "Familia": ETIQUETA_FAMILIA.get(resultado.get("familia"), "Sin determinar"),
                "Campo": pendiente.get("campo", ""),
                "Problema": pendiente.get("problema", ""),
                "Valores permitidos": pendiente.get("permitidos", ""),
            })
    return avisos


def resumen_plantilla(ruta=None):
    """Para diagnostico: cuantas columnas y diccionarios trae cada familia."""
    plantilla = cargar_plantilla(ruta)
    filas = [{
        "Hoja": HOJA_SIEMPRE,
        "Columnas": len(plantilla["siempre"]),
        "Con diccionario": sum(1 for c in plantilla["siempre"] if plantilla["diccionarios"].get(c)),
    }]
    for familia, columnas in plantilla["familias"].items():
        filas.append({
            "Hoja": HOJA_POR_FAMILIA[familia],
            "Columnas": len(columnas) + len(plantilla["siempre"]),
            "Con diccionario": sum(1 for c in columnas if plantilla["diccionarios"].get(c)),
        })
    filas.append({"Hoja": HOJA_CATEGORIAS, "Columnas": len(plantilla["categorias"]), "Con diccionario": ""})
    return filas


# --- categoria deducida ---------------------------------------------------
# Segmentos que cambian con el genero. Cada fila son equivalentes entre si, en
# el orden de GENEROS_EQUIVALENTES. Salen de la propia tabla de la plantilla.
GENEROS_EQUIVALENTES = ("Hombre", "Mujer", "Niños", "Niñas", "Unisex Adultos", "Unisex Niños")

SEGMENTOS_EQUIVALENTES = [
    ("Ropa Masculina", "Ropa Femenina", "Niños", "Niña", "Ropa Masculina", "Niños"),
    ("Calzados Masculinos", "Calzados Femeninos", "Calzado de Niño", "Calzado de Niña",
     "Calzados Masculinos", "Calzado de Niño"),
    ("Deporte Masculino", "Deporte Femenino", "Deporte Masculino", "Deporte Femenino",
     "Deporte Masculino", "Deporte Masculino"),
    ("Masculinos", "Femeninos", "Infantiles", "Infantiles", "Masculinos", "Infantiles"),
]

_INDICE_SEGMENTO = {}
for _fila in SEGMENTOS_EQUIVALENTES:
    for _pos, _segmento in enumerate(_fila):
        _INDICE_SEGMENTO.setdefault(normalizar(_segmento), {})[GENEROS_EQUIVALENTES[_pos]] = _fila


def _traducir_categoria(categoria, genero_destino):
    """Cambia el segmento de genero de una categoria. "" si no se puede.

    "Vestuario / Ropa Masculina / Blusas" con destino Mujer da
    "Vestuario / Ropa Femenina / Blusas". Si la categoria no lleva ningun
    segmento de genero (Accesorios / Bolsos... / Mochilas) vale para cualquiera
    y se devuelve igual.
    """
    if genero_destino not in GENEROS_EQUIVALENTES:
        return ""
    destino = GENEROS_EQUIVALENTES.index(genero_destino)
    partes = [parte.strip() for parte in _texto(categoria).split("/")]
    if not partes:
        return ""
    cambiadas = 0
    salida = []
    for parte in partes:
        fila = None
        for candidata in SEGMENTOS_EQUIVALENTES:
            if any(normalizar(seg) == normalizar(parte) for seg in candidata):
                fila = candidata
                break
        if fila is None:
            salida.append(parte)
            continue
        salida.append(fila[destino])
        cambiadas += 1
    if cambiadas:
        return " / ".join(salida)
    # Sin segmento de genero: la misma categoria sirve para todos.
    return " / ".join(salida)


def deducir_categoria(marca, tipo, genero, ruta=None):
    """Categoria para un (tipo, genero) que no esta en la tabla.

    Busca el MISMO tipo bajo otro genero y traduce el segmento de genero. Es lo
    que haria una persona: si "Blusas / Hombre" es "Vestuario / Ropa Masculina /
    Blusas", entonces "Blusas / Mujer" es "Vestuario / Ropa Femenina / Blusas".

    Devuelve (categoria, origen). origen dice de donde se dedujo, para que la
    advertencia lo diga y se pueda revisar.
    """
    plantilla = cargar_plantilla(ruta)
    tabla = plantilla["categorias"]
    genero_destino = normalizar_genero(genero)
    clave_tipo = _singular(tipo)
    if not genero_destino or not clave_tipo:
        return "", ""

    # Primero la misma marca, luego Todos, y dentro de cada una el genero mas
    # parecido: del mismo "bloque" (adulto o infantil) antes que del otro.
    orden_generos = _generos_cercanos(genero_destino)
    for marca_clave in (normalizar(marca), "todos"):
        if not marca_clave:
            continue
        for genero_origen in orden_generos:
            opciones = tabla.get((marca_clave, clave_tipo, normalizar(genero_origen)))
            if not opciones:
                continue
            base = _categoria_mas_general(opciones)
            traducida = _traducir_categoria(base, genero_destino)
            if traducida:
                marca_txt = "la misma marca" if marca_clave != "todos" else "Todos"
                return traducida, f'deducida de "{tipo}" / {genero_origen} ({marca_txt})'
    return "", ""


def _generos_cercanos(genero):
    """Los otros generos del MISMO bloque (adulto o infantil).

    No se cruza entre bloques a proposito: deducir "Cortavientos / Mujer" desde
    "Niñas" daba "Vestuario / Infantil / Ropa Femenina / Cortavientos", que es
    una categoria que no existe. Sin fuente del mismo bloque se deja pendiente.
    """
    adultos = ["Hombre", "Mujer", "Unisex Adultos"]
    ninos = ["Niños", "Niñas", "Unisex Niños"]
    bloque = adultos if genero in adultos else ninos
    return [g for g in bloque if g != genero]


# --- red final: categoria por familia + genero ---------------------------
_GENERICAS = {}


def categoria_generica(familia, genero, ruta=None):
    """La categoria base mas usada para esa familia y ese genero.

    Es la ultima red: cuando el tipo no esta en la tabla ni se puede deducir de
    otro genero, el producto igual necesita una categoria para que Centry lo
    acepte. Se calcula CONTANDO la propia tabla de la plantilla (no es una lista
    escrita a mano): para cada entrada se mira a que familia pertenece su tipo y
    se guarda el prefijo de dos niveles de su categoria; gana el mas repetido.

    Siempre queda marcado en la advertencia como categoria generica.
    """
    global _GENERICAS
    plantilla = cargar_plantilla(ruta)
    clave_cache = plantilla["__clave__"]
    if _GENERICAS.get("__clave__") != clave_cache:
        conteo = {}
        for (_marca, tipo, gen), categorias in plantilla["categorias"].items():
            fam, _ = detectar_familia(tipo)
            if not fam:
                continue
            for categoria in categorias:
                partes = [p.strip() for p in categoria.split("/") if p.strip()]
                if len(partes) < 2:
                    continue
                prefijo = " / ".join(partes[:2])
                conteo.setdefault((fam, gen), {}).setdefault(prefijo, 0)
                conteo[(fam, gen)][prefijo] += 1
        _GENERICAS = {"__clave__": clave_cache}
        for clave, opciones in conteo.items():
            _GENERICAS[clave] = max(opciones.items(), key=lambda par: (par[1], -len(par[0])))[0]

    genero_normalizado = normalizar_genero(genero)
    if not familia or not genero_normalizado:
        return ""
    return _GENERICAS.get((familia, normalizar(genero_normalizado)), "")


# --- atributos desde las caracteristicas ---------------------------------
# El input comercial trae las caracteristicas como "Etiqueta : Valor" separadas
# por |. Ahi viven muchos atributos que Centry pide en columnas propias. Se
# mapea la etiqueta a su columna y el valor se valida contra el diccionario de
# esa columna: si no esta permitido, no se escribe y queda el aviso.
ETIQUETAS_A_COLUMNA = {
    "forro": ("Material del forro - Calzado (Falabella GSC Perú)",
              "Materiales del interior(MercadoLibre Perú)"),
    "suela": ("Material de la suela - Calzado (Falabella GSC Perú)",
              "Material de la suela (MercadoLibre Perú)"),
    "capellada": ("Material principal - Calzado (Falabella GSC Perú)",
                  "Material del calzado (MercadoLibre Perú)"),
    "plantilla": ("Material de la plantilla - Calzado (Falabella GSC Perú)",),
    "ajuste": ("Tipo de ajuste - Tipo de ajuste - Calzado (Falabella GSC Perú)",),
    "altura de cana": ("Tipo de caña - Calzado (Falabella GSC Perú)",
                       "Tipo de caña MPE68c (MercadoLibre Perú)"),
    "tipo de cana": ("Tipo de caña - Calzado (Falabella GSC Perú)",
                     "Tipo de caña MPE68c (MercadoLibre Perú)"),
    "forma de la punta": ("Forma de la punta - Calzado (Falabella GSC Perú)",
                          "Forma de la punta MPE6a3 (MercadoLibre Perú)"),
    "altura plataforma": ("Altura de la plataforma - Calzado (Falabella GSC Perú)",),
    "altura de plataforma": ("Altura de la plataforma - Calzado (Falabella GSC Perú)",),
    "horma": ("Horma - Calzado (Falabella GSC Perú)",),
    "material": ("Material de vestuario - Ropa y accesorios (Falabella GSC Perú)",
                 "Material del accesorio - Ropa y accesorios (Falabella GSC Perú)",
                 "Material principal (MercadoLibre Perú)"),
    "materialidad": ("Material de vestuario - Ropa y accesorios (Falabella GSC Perú)",
                     "Material principal (MercadoLibre Perú)"),
    "composicion": ("Composición - Ropa y accesorios (Falabella GSC Perú)",
                    "Composición (MercadoLibre Perú)"),
    "tipo de cierre": ("Tipo de cierre - Ropa y accesorios (Falabella GSC Perú)",),
    "cierre": ("Tipo de cierre - Ropa y accesorios (Falabella GSC Perú)",),
    "largo de manga": ("Largo de mangas - Ropa y accesorios (Falabella GSC Perú)",
                       "Tipo de manga (MercadoLibre Perú)"),
    "manga": ("Largo de mangas - Ropa y accesorios (Falabella GSC Perú)",
              "Tipo de manga (MercadoLibre Perú)"),
    "cuello": ("Tipo de cuello - Tipo de cuello - Ropa y accesorios (Falabella GSC Perú)",),
    "fit": ("Fit - Ropa y accesorios (Falabella GSC Perú)",
            "Fit prenda superior - Ropa y accesorios (Falabella GSC)",
            "Fit prenda inferior - Ropa y accesorios (Falabella GSC)"),
    "tiro": ("Tiro para prendas - Ropa y accesorios (Falabella GSC Perú)",
             "Tiro del pantalón (MercadoLibre Perú)"),
    "largo de la prenda": ("Largo de la prenda - Ropa y accesorios (Falabella GSC Perú)",),
    "disciplina": ("Disciplina - Ropa y accesorios (Falabella GSC Perú)",
                   "Disciplina zapatillas - Calzado (Falabella GSC Perú)"),
}


def pares_de_caracteristicas(texto):
    """Extrae los pares "Etiqueta : Valor" de un listado separado por |."""
    pares = []
    for trozo in re.split(r"[|\n]+", _texto(texto)):
        if ":" not in trozo:
            continue
        etiqueta, _, valor = trozo.partition(":")
        etiqueta, valor = _texto(etiqueta), _texto(valor)
        if etiqueta and valor:
            pares.append((etiqueta, valor))
    return pares


def atributos_desde_caracteristicas(texto, familia, ruta=None):
    """{columna: valor} para las columnas de esa familia. Y los no aplicados.

    Solo escribe cuando el valor esta en el diccionario de la columna. Un
    "Altura De Taco: No Aplica" no ensucia el archivo: se ignora y se informa.
    """
    columnas_familia = set(columnas_de(familia, ruta)) if familia else set()
    aplicados = {}
    ignorados = []
    for etiqueta, valor in pares_de_caracteristicas(texto):
        destinos = ETIQUETAS_A_COLUMNA.get(normalizar(etiqueta))
        if not destinos:
            continue
        colocado = False
        for columna in destinos:
            if columna not in columnas_familia or aplicados.get(columna):
                continue
            limpio, ok = valor_valido(columna, valor, ruta)
            if ok and limpio:
                aplicados[columna] = limpio
                colocado = True
        if not colocado:
            ignorados.append(f"{etiqueta}: {valor}")
    return aplicados, ignorados


def caracteristicas_del_body(body_html):
    """Los bullets de la seccion Caracteristicas del Body HTML.

    Ahi viven los pares "Forro: 100% Cuero" que Centry pide en columnas
    propias. El listado de Sial no sirve: ese trae otra cosa (Tipo De Producto,
    Genero, Color, Marca).
    """
    texto = _texto(body_html)
    if not texto:
        return ""
    bloque = re.search(
        r'nweb__Caracteristicas.*?<ul>(.*?)</ul>', texto, flags=re.IGNORECASE | re.DOTALL
    )
    if not bloque:
        return ""
    items = re.findall(r"<li[^>]*>(.*?)</li>", bloque.group(1), flags=re.IGNORECASE | re.DOTALL)
    limpios = [_sin_etiquetas(item) for item in items]
    return "|".join(item for item in limpios if item)


def _sin_etiquetas(html):
    """Texto plano de un fragmento, sin depender del modulo del generador."""
    texto = re.sub(r"<[^>]+>", " ", _texto(html))
    return re.sub(r"\s+", " ", texto).strip()


def columnas_de_tipo(familia, ruta=None):
    """Columnas de la familia que piden el TIPO de prenda/calzado.

    Son las que empiezan por "Tipo" y traen diccionario cerrado: "Tipo -
    Calzado", "Tipo de prenda para la parte superior", "Tipo de prenda
    (MercadoLibre)"... Se detectan por la plantilla, no por una lista fija.
    """
    salida = []
    for columna in columnas_de(familia, ruta):
        if not normalizar(columna).startswith("tipo"):
            continue
        if valores_permitidos(columna, ruta):
            salida.append(columna)
    return salida


def tipo_para_columnas(tipo, familia, ruta=None):
    """{columna: valor} escribiendo el tipo en las columnas donde SI encaja.

    El tipo del catalogo ("Casacas") se compara contra el diccionario de cada
    columna de tipo. Solo se escribe donde el valor esta permitido: asi "Tipo de
    prenda para la parte superior" recibe "Casacas" y "Tipo de chaqueta/chaleco"
    no recibe nada si ese valor no esta en su lista.

    Devuelve tambien las columnas de tipo que quedaron sin poder completarse.
    """
    aplicados = {}
    sin_valor = []
    clave = _singular(tipo)
    if not clave:
        return aplicados, columnas_de_tipo(familia, ruta)
    for columna in columnas_de_tipo(familia, ruta):
        # Se comparan las dos partes en singular: el catalogo dice "Canguro" y
        # la plantilla "Canguros". Sin esto no cruzaba ninguno de los dos.
        encontrado = ""
        for permitido in valores_permitidos(columna, ruta):
            if _singular(permitido) == clave:
                encontrado = permitido
                break
        if encontrado:
            aplicados[columna] = encontrado
        else:
            sin_valor.append(columna)
    return aplicados, sin_valor
