from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "outputs" / "INPUT_COMERCIAL_HUSH_PUPPIES_FORMATO_EXPORT.xlsx"

ALLOWED_BRANDS = [
    "HUSH PUPPIES",
    "HUSH PUPPIES KIDS",
    "ACCESORIOS HP",
    "KEDS",
    "ROCKFORD",
]

INPUT_COLUMNS = [
    "Mod-Col",
    "Marca",
    "Handle Input",
    "Title",
    "Body HTML",
    "Type",
    "Color Comercial",
    "Color Web",
    "Tags",
    "Metafield: custom.materialidad [single_line_text_field]",
    "Metafield: custom.pais_de_fabricacion [single_line_text_field]",
    "Metafield: custom.marca [single_line_text_field]",
    "Metafield: custom.color_forus [single_line_text_field]",
    "Metafield: custom.siblings_color [single_line_text_field]",
    "Metafield: custom.grupo_color [single_line_text_field]",
    "Metafield: custom.genero [single_line_text_field]",
    "Metafield: custom.tipo [single_line_text_field]",
    "Metafield: custom.descripcion_corta [single_line_text_field]",
    "Metafield: custom.nombre_corto [single_line_text_field]",
    "Metafield: custom.codigo_modelo_color [id]",
    "Metafield: custom.sub_categoria [single_line_text_field]",
    "Metafield: custom.categoria [single_line_text_field]",
    "Metafield: custom.guia_de_tallas [page_reference]",
    "Metafield: custom.tecnologia [list.single_line_text_field]",
    "Caracteristicas",
    "Material",
    "Cuidado",
]


def add_header_style(ws, width_overrides=None):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    width_overrides = width_overrides or {}

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        cell = ws.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[letter].width = width_overrides.get(letter, 28)

    for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, 101), max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 42


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Input"
    ws.append(INPUT_COLUMNS)

    example = {
        "Mod-Col": "HP202011248963-SMV",
        "Marca": "HUSH PUPPIES",
        "Handle Input": "mocasines-mujer-natalia-hush-puppies-hp202011248963-smv",
        "Title": "Mocasin Natalia Cuero Mujer",
        "Body HTML": "Mocasin para mujer cuero Natalia azul Hush Puppies. Garantiza estilo y comodidad para uso diario.",
        "Type": "Mocasines",
        "Color Comercial": "Azul",
        "Color Web": "Azul",
        "Tags": "Azul,Calzado,Casual,Cuero,Hush Puppies,Mocasines,Mujer,HP202011248963-SMV",
        "Metafield: custom.materialidad [single_line_text_field]": "Cuero",
        "Metafield: custom.pais_de_fabricacion [single_line_text_field]": "Brasil",
        "Metafield: custom.marca [single_line_text_field]": "Hush Puppies",
        "Metafield: custom.color_forus [single_line_text_field]": "Azul",
        "Metafield: custom.siblings_color [single_line_text_field]": "Azul",
        "Metafield: custom.grupo_color [single_line_text_field]": "Azul",
        "Metafield: custom.genero [single_line_text_field]": "Mujer",
        "Metafield: custom.tipo [single_line_text_field]": "Mocasines",
        "Metafield: custom.descripcion_corta [single_line_text_field]": "Mocasines Mujer",
        "Metafield: custom.nombre_corto [single_line_text_field]": "Mocasin Natalia Cuero Mujer",
        "Metafield: custom.codigo_modelo_color [id]": "HP202011248963-SMV",
        "Metafield: custom.sub_categoria [single_line_text_field]": "Mocasines",
        "Metafield: custom.categoria [single_line_text_field]": "Calzado",
        "Metafield: custom.guia_de_tallas [page_reference]": "hpp_mujer_francesa",
        "Metafield: custom.tecnologia [list.single_line_text_field]": "",
        "Caracteristicas": "Tipo Producto: Mocasin\nGenero: Mujer\nColor: Azul\nMarca: Hush Puppies\nOcasion: Casual\nMaterial: Cuero",
        "Material": "Capellada: Cuero\nForro: Cuero\nSuela: Goma",
        "Cuidado": "Limpiar con pano seco.\nNo lavar.\nNo usar blanqueador.",
    }
    ws.append([example.get(column, "") for column in INPUT_COLUMNS])
    for _ in range(99):
        ws.append(["" for _ in INPUT_COLUMNS])

    add_header_style(
        ws,
        {
            "A": 20,
            "B": 22,
            "C": 54,
            "D": 36,
            "E": 70,
            "F": 18,
            "G": 18,
            "H": 18,
            "I": 58,
            "Y": 58,
            "Z": 42,
            "AA": 42,
        },
    )
    ws.auto_filter.ref = f"A1:{get_column_letter(len(INPUT_COLUMNS))}101"
    ws.row_dimensions[2].height = 118

    marca_col = INPUT_COLUMNS.index("Marca") + 1
    validation = DataValidation(type="list", formula1='"' + ",".join(ALLOWED_BRANDS) + '"', allow_blank=False)
    ws.add_data_validation(validation)
    validation.add(f"{get_column_letter(marca_col)}2:{get_column_letter(marca_col)}500")

    instructions = wb.create_sheet("Instrucciones")
    instructions.append(["Regla", "Detalle"])
    rules = [
        ("Sitio destino", "Seleccionar HushPuppies.pe en la app antes de cargar este input."),
        ("Marca", "Usar la lista permitida: HUSH PUPPIES, HUSH PUPPIES KIDS, ACCESORIOS HP, KEDS o ROCKFORD."),
        ("custom.marca", "Usar el nombre comercial visible, por ejemplo Hush Puppies."),
        ("Mod-Col", "Debe coincidir con ARTI y con custom.codigo_modelo_color. Ejemplo: HP202011248963-SMV."),
        ("Una fila por producto-color", "No incluir tallas. Las tallas y variantes salen desde ARTI."),
        ("Type / tipo / sub_categoria", "Mantener categorias como el export: Zapatillas, Sandalias, Zapatos, Botines, Mocasines, Medias, Carteras."),
        ("categoria", "Usar Calzado, Accesorios o Vestuario segun corresponda."),
        ("guia_de_tallas", "Usar referencias como hpp_mujer_francesa, hpp_hombre_francesa, hpp_calzado_relaxshoe si aplica."),
        ("Body HTML", "Puede ser descripcion simple. La app agrega secciones de Caracteristicas, Material y Cuidado."),
        ("Fotos", "La app genera URLs por Mod-Col desde carpeta HUSH PUPPIES."),
        ("Fila 2", "Es ejemplo basado en el export; borrar o reemplazar con data real."),
    ]
    for row in rules:
        instructions.append(list(row))
    add_header_style(instructions, {"A": 32, "B": 120})

    lists = wb.create_sheet("Listas Hush")
    lists.append(["Marcas permitidas", "Categorias frecuentes", "Guias de talla frecuentes"])
    list_rows = [
        ("HUSH PUPPIES", "Calzado", "hpp_mujer_francesa"),
        ("HUSH PUPPIES KIDS", "Accesorios", "hpp_hombre_francesa"),
        ("ACCESORIOS HP", "Vestuario", "hpp_mujer_novi"),
        ("KEDS", "", "hpp_calzado_relaxshoe"),
        ("ROCKFORD", "", "hpp_mujer_americana"),
        ("", "", "hpp_hombre_americana"),
    ]
    for row in list_rows:
        lists.append(list(row))
    add_header_style(lists, {"A": 28, "B": 24, "C": 36})

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
