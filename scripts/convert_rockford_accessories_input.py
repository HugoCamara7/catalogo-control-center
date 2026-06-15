from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\hcamara\OneDrive - Peru Forus S.A\Documentos\Rockford Accesorios 03-06-2026.xlsx")
OUTPUT = ROOT / "outputs" / "INPUT_ROCKFORD_ACCESORIOS_03-06-2026.xlsx"

INPUT_COLUMNS = [
    "Mod-Col",
    "Marca",
    "Handle Input",
    "Title",
    "Body HTML",
    "Type",
    "Color Comercial",
    "Color Web",
    "Tags",
    "Metafield: custom.materialidad [single_line_text_field]",
    "Metafield: custom.pais_de_fabricacion [single_line_text_field]",
    "Metafield: custom.marca [single_line_text_field]",
    "Metafield: custom.color_forus [single_line_text_field]",
    "Metafield: custom.siblings_color [single_line_text_field]",
    "Metafield: custom.grupo_color [single_line_text_field]",
    "Metafield: custom.genero [single_line_text_field]",
    "Metafield: custom.tipo [single_line_text_field]",
    "Metafield: custom.descripcion_corta [single_line_text_field]",
    "Metafield: custom.nombre_corto [single_line_text_field]",
    "Metafield: custom.codigo_modelo_color [id]",
    "Metafield: custom.sub_categoria [single_line_text_field]",
    "Metafield: custom.categoria [single_line_text_field]",
    "Metafield: custom.guia_de_tallas [page_reference]",
    "Metafield: custom.tecnologia [list.single_line_text_field]",
    "Caracteristicas",
    "Material",
    "Cuidado",
]


def clean(value):
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def title_case(value):
    text = clean(value).lower()
    if not text:
        return ""
    return " ".join(part.capitalize() for part in text.split())


def slug(value):
    text = clean(value).lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
        "ü": "u",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def join_unique(values, separator=","):
    result = []
    seen = set()
    for value in values:
        text = clean(value)
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(text)
    return separator.join(result)


def bullets(row):
    pairs = [
        ("Tipo Producto", title_case(row.get("Tipo Producto"))),
        ("Genero", title_case(row.get("Género") or row.get("Genero"))),
        ("Color", title_case(row.get("Color Primario"))),
        ("Marca", "Rockford"),
        ("Ocasion", title_case(row.get("Ocasión"))),
        ("Material", title_case(row.get("Material"))),
        ("Producto Sustentable", title_case(row.get("Producto Sustentable"))),
        ("Tipo de Ajuste", title_case(row.get("Tipo de Ajuste"))),
        ("Alto", clean(row.get("Alto"))),
        ("Ancho", clean(row.get("Ancho "))),
        ("Profundidad", clean(row.get("Profundidad"))),
        ("Capacidad Litros", clean(row.get("Capacitadad Litros"))),
        ("Compartimentos", clean(row.get("Compartimentos"))),
        ("N Compartimentos", clean(row.get("N° Compartimentos"))),
    ]
    return "\n".join(f"{label}: {value}" for label, value in pairs if value)


def technology(row):
    return join_unique([row.get("Tecnología"), row.get("Tecnología 2"), row.get("Tecnología 3")])


def convert():
    df = pd.read_excel(SOURCE, sheet_name=0, dtype=object).dropna(how="all")
    rows = []
    for _, row in df.iterrows():
        mod_col = clean(row.get("Mod-Col")).upper()
        product_type = title_case(row.get("Tipo Producto"))
        gender = title_case(row.get("Género") or row.get("Genero"))
        color = title_case(row.get("Color Primario") or row.get("Color Comercial (Bullet)"))
        name = clean(row.get("Nombre")) or " ".join(part for part in [product_type, gender, color, "Rockford"] if part)
        model_name = clean(row.get("Nombre del Modelo")) or clean(row.get("Modelo"))
        material = title_case(row.get("Material"))
        description = clean(row.get("Descripción"))
        tech = technology(row)
        tags = join_unique(
            [
                "Rockford",
                "Accesorios",
                product_type,
                gender,
                color,
                material,
                title_case(row.get("Ocasión")),
                title_case(row.get("Producto Sustentable")),
                tech,
                mod_col,
            ],
            separator=",",
        )
        handle = slug(f"{name} {mod_col}")
        short_description = " ".join(part for part in [product_type, gender] if part)
        rows.append(
            {
                "Mod-Col": mod_col,
                "Marca": "ROCKFORD",
                "Handle Input": handle,
                "Title": name,
                "Body HTML": description,
                "Type": product_type,
                "Color Comercial": color,
                "Color Web": color,
                "Tags": tags,
                "Metafield: custom.materialidad [single_line_text_field]": material,
                "Metafield: custom.pais_de_fabricacion [single_line_text_field]": "",
                "Metafield: custom.marca [single_line_text_field]": "Rockford",
                "Metafield: custom.color_forus [single_line_text_field]": color,
                "Metafield: custom.siblings_color [single_line_text_field]": color,
                "Metafield: custom.grupo_color [single_line_text_field]": color,
                "Metafield: custom.genero [single_line_text_field]": gender,
                "Metafield: custom.tipo [single_line_text_field]": product_type,
                "Metafield: custom.descripcion_corta [single_line_text_field]": short_description,
                "Metafield: custom.nombre_corto [single_line_text_field]": model_name or name,
                "Metafield: custom.codigo_modelo_color [id]": mod_col,
                "Metafield: custom.sub_categoria [single_line_text_field]": product_type,
                "Metafield: custom.categoria [single_line_text_field]": "Accesorios",
                "Metafield: custom.guia_de_tallas [page_reference]": "",
                "Metafield: custom.tecnologia [list.single_line_text_field]": tech,
                "Caracteristicas": clean(row.get("Bullets")) or bullets(row),
                "Material": material,
                "Cuidado": clean(row.get("Cuidado Lavado")),
            }
        )
    return pd.DataFrame(rows, columns=INPUT_COLUMNS)


def format_workbook(path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 42

    ws = wb["Input"]
    widths = {"A": 20, "B": 18, "C": 58, "D": 58, "E": 82, "F": 18, "G": 18, "H": 18, "I": 68}
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(letter, 30)
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(path)


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    output_df = convert()
    instructions = pd.DataFrame(
        [
            ("Sitio destino", "Seleccionar Rockford.pe en la app."),
            ("Marca", "Se dejo ROCKFORD para pasar validacion y usar carpeta de imagenes ROCKFORD."),
            ("Una fila por producto-color", "No incluir tallas. Las variantes salen desde ARTI por Mod-Col."),
            ("Categoria", "Se mapeo como Accesorios para todos los productos del archivo."),
            ("Fotos", "La app generara URLs usando el Mod-Col y la carpeta ROCKFORD."),
            ("Revision", "Revisar especialmente pais de fabricacion y guia de tallas si aplica."),
        ],
        columns=["Regla", "Detalle"],
    )
    mapping = pd.DataFrame(
        [
            ("Nombre", "Title"),
            ("Descripción", "Body HTML"),
            ("Tipo Producto", "Type / custom.tipo / custom.sub_categoria"),
            ("Color Primario", "Color Comercial / Color Web / grupo_color"),
            ("Material", "custom.materialidad / Material"),
            ("Género", "custom.genero"),
            ("Tecnología 1-3", "custom.tecnologia"),
            ("Cuidado Lavado", "Cuidado"),
        ],
        columns=["Origen", "Destino input"],
    )
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, sheet_name="Input")
        instructions.to_excel(writer, index=False, sheet_name="Instrucciones")
        mapping.to_excel(writer, index=False, sheet_name="Mapeo")
    format_workbook(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
