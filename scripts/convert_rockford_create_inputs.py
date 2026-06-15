from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]

JOBS = [
    {
        "source": Path(r"C:\Users\hcamara\OneDrive - Peru Forus S.A\Documentos\Crear Vestuario 03-06-2026.xlsx"),
        "output": ROOT / "outputs" / "INPUT_CREAR_VESTUARIO_03-06-2026.xlsx",
        "category": "Vestuario",
        "site": "Rockford.pe",
    },
    {
        "source": Path(r"C:\Users\hcamara\OneDrive - Peru Forus S.A\Documentos\Crear Calzado Rockford 03-06-2026.xlsx"),
        "output": ROOT / "outputs" / "INPUT_CREAR_CALZADO_ROCKFORD_03-06-2026.xlsx",
        "category": "Calzado",
        "site": "Rockford.pe",
    },
]

INPUT_COLUMNS = [
    "Mod-Col",
    "Marca",
    "Handle Input",
    "Title",
    "Body HTML",
    "Type",
    "Color Comercial",
    "Color Web",
    "Tags",
    "Metafield: custom.materialidad [single_line_text_field]",
    "Metafield: custom.pais_de_fabricacion [single_line_text_field]",
    "Metafield: custom.marca [single_line_text_field]",
    "Metafield: custom.color_forus [single_line_text_field]",
    "Metafield: custom.siblings_color [single_line_text_field]",
    "Metafield: custom.grupo_color [single_line_text_field]",
    "Metafield: custom.genero [single_line_text_field]",
    "Metafield: custom.tipo [single_line_text_field]",
    "Metafield: custom.descripcion_corta [single_line_text_field]",
    "Metafield: custom.nombre_corto [single_line_text_field]",
    "Metafield: custom.codigo_modelo_color [id]",
    "Metafield: custom.sub_categoria [single_line_text_field]",
    "Metafield: custom.categoria [single_line_text_field]",
    "Metafield: custom.guia_de_tallas [page_reference]",
    "Metafield: custom.tecnologia [list.single_line_text_field]",
    "Caracteristicas",
    "Material",
    "Cuidado",
]


def clean(value):
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def pick(row, *names):
    for name in names:
        value = clean(row.get(name))
        if value:
            return value
    return ""


def title_case(value):
    text = clean(value).lower()
    if not text:
        return ""
    return normalize_text_tokens(" ".join(part[:1].upper() + part[1:] for part in text.split()))


def normalize_text_tokens(value):
    text = clean(value)
    replacements = {
        "M/l": "M/L",
        "M/c": "M/C",
        "O/s": "O/S",
        "Rkf": "RKF",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


def slug(value):
    text = clean(value).lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
        "ü": "u",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def join_unique(values, separator=","):
    result = []
    seen = set()
    for value in values:
        text = clean(value)
        if not text:
            continue
        key = text.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(text)
    return separator.join(result)


def build_technology(row):
    return join_unique(
        [
            pick(row, "Tecnología", "Tecnologia"),
            pick(row, "Tecnología 2", "Tecnologia 2"),
            pick(row, "Tecnología 3", "Tecnologia 3"),
        ]
    )


def build_features(row, category):
    source_bullets = clean(row.get("Bullets"))
    if source_bullets:
        return source_bullets

    common = [
        ("Tipo producto", title_case(row.get("Tipo Producto"))),
        ("Genero", title_case(pick(row, "Género", "Genero"))),
        ("Color", title_case(pick(row, "Color Primario", "Color Comercial (Bullet)"))),
        ("Ocasion", title_case(pick(row, "Ocasión", "Ocasion"))),
        ("Actividad", title_case(row.get("Actividad"))),
        ("Tecnologia", build_technology(row)),
        ("Material", title_case(row.get("Material"))),
        ("Composicion", clean(row.get("Composición"))),
        ("Ajuste", title_case(row.get("Tipo de Ajuste"))),
        ("Producto sustentable", title_case(row.get("Producto Sustentable"))),
    ]
    apparel = [
        ("Tipo manga", title_case(row.get("Tipo Manga"))),
        ("Calce", title_case(row.get("Calce"))),
        ("Tipo cuello", title_case(row.get("Tipo Cuello"))),
        ("Forro", title_case(row.get("Forro"))),
        ("Bolsillos", title_case(row.get("Bolsillos"))),
        ("Longitud", title_case(row.get("Longitud"))),
    ]
    footwear = [
        ("Forro", title_case(pick(row, "FORRO", "Forro"))),
        ("Suela", title_case(row.get("SUELA"))),
        ("Tipo taco", title_case(row.get("Tipo Taco"))),
        ("Altura taco", clean(row.get("ALTURA_TACO (CMS)"))),
        ("Altura cana", clean(row.get("ALTURA_CANA (CMS)"))),
        ("Plataforma", title_case(row.get("PLATAFORMA"))),
        ("Punta", title_case(row.get("FORMA DE LA PUNTA"))),
    ]

    pairs = common + (footwear if category == "Calzado" else apparel)
    return "\n".join(f"{label}: {value}" for label, value in pairs if value)


def convert_file(job):
    df = pd.read_excel(job["source"], sheet_name=0, dtype=object).dropna(how="all")
    rows = []

    for _, row in df.iterrows():
        mod_col = clean(row.get("Mod-Col")).upper()
        if not mod_col:
            continue

        product_type = title_case(row.get("Tipo Producto"))
        gender = title_case(pick(row, "Género", "Genero"))
        color = title_case(pick(row, "Color Primario", "Color Comercial (Bullet)"))
        material = title_case(row.get("Material"))
        model_name = clean(row.get("Nombre del Modelo")) or clean(row.get("Modelo"))
        name = normalize_text_tokens(clean(row.get("Nombre"))) or " ".join(
            part for part in [product_type, "Para", gender, material, model_name, color, "Rockford"] if part
        )
        description = clean(row.get("Descripción"))
        tech = build_technology(row)
        short_description = " ".join(part for part in [product_type, gender, material] if part)
        tags = join_unique(
            [
                "Rockford",
                job["category"],
                product_type,
                gender,
                color,
                material,
                title_case(pick(row, "Ocasión", "Ocasion")),
                title_case(row.get("Actividad")),
                title_case(row.get("Producto Sustentable")),
                tech,
                mod_col,
            ]
        )

        rows.append(
            {
                "Mod-Col": mod_col,
                "Marca": "ROCKFORD",
                "Handle Input": slug(f"{name} {mod_col}"),
                "Title": name,
                "Body HTML": description,
                "Type": product_type,
                "Color Comercial": color,
                "Color Web": color,
                "Tags": tags,
                "Metafield: custom.materialidad [single_line_text_field]": material,
                "Metafield: custom.pais_de_fabricacion [single_line_text_field]": "",
                "Metafield: custom.marca [single_line_text_field]": "Rockford",
                "Metafield: custom.color_forus [single_line_text_field]": color,
                "Metafield: custom.siblings_color [single_line_text_field]": color,
                "Metafield: custom.grupo_color [single_line_text_field]": color,
                "Metafield: custom.genero [single_line_text_field]": gender,
                "Metafield: custom.tipo [single_line_text_field]": product_type,
                "Metafield: custom.descripcion_corta [single_line_text_field]": short_description,
                "Metafield: custom.nombre_corto [single_line_text_field]": model_name or name,
                "Metafield: custom.codigo_modelo_color [id]": mod_col,
                "Metafield: custom.sub_categoria [single_line_text_field]": product_type,
                "Metafield: custom.categoria [single_line_text_field]": job["category"],
                "Metafield: custom.guia_de_tallas [page_reference]": "",
                "Metafield: custom.tecnologia [list.single_line_text_field]": tech,
                "Caracteristicas": build_features(row, job["category"]),
                "Material": material,
                "Cuidado": clean(row.get("Cuidado Lavado")),
            }
        )

    return pd.DataFrame(rows, columns=INPUT_COLUMNS)


def format_workbook(path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 42
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    ws = wb["Input"]
    widths = {"A": 22, "B": 16, "C": 58, "D": 62, "E": 86, "F": 20, "G": 18, "H": 18, "I": 72}
    for idx in range(1, ws.max_column + 1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(letter, 30)

    wb.save(path)


def write_output(job):
    job["output"].parent.mkdir(parents=True, exist_ok=True)
    output_df = convert_file(job)

    instructions = pd.DataFrame(
        [
            ("Sitio destino", f"Seleccionar {job['site']} en la app."),
            ("Marca", "Se dejo ROCKFORD para validar contra el sitio Rockford.pe."),
            ("Unidad de carga", "Una fila por codigo modelo-color. No incluir tallas."),
            ("Tallas", "La app debe cruzar variantes desde ARTI/SIAL por Mod-Col."),
            ("Fotos", "No se incluyen URLs en el input comercial."),
            ("Categoria", f"Se mapeo como {job['category']} para todos los productos de este archivo."),
        ],
        columns=["Regla", "Detalle"],
    )
    mapping = pd.DataFrame(
        [
            ("Mod-Col", "Mod-Col / custom.codigo_modelo_color"),
            ("Nombre", "Title"),
            ("Descripción", "Body HTML"),
            ("Tipo Producto", "Type / custom.tipo / custom.sub_categoria"),
            ("Color Primario", "Color Comercial / Color Web / grupo_color"),
            ("Material", "custom.materialidad / Material"),
            ("Género", "custom.genero"),
            ("Tecnología 1-3", "custom.tecnologia"),
            ("Cuidado Lavado", "Cuidado"),
        ],
        columns=["Origen", "Destino input"],
    )

    with pd.ExcelWriter(job["output"], engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, sheet_name="Input")
        instructions.to_excel(writer, index=False, sheet_name="Instrucciones")
        mapping.to_excel(writer, index=False, sheet_name="Mapeo")
    format_workbook(job["output"])
    return job["output"], len(output_df)


def main():
    for job in JOBS:
        output, rows = write_output(job)
        print(f"{output} | filas: {rows}")


if __name__ == "__main__":
    main()
